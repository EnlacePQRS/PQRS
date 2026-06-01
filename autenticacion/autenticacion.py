"""Sistema de Gestión de PQRS para Empresas Públicas - Sprint 1: Registro de Ciudadanos"""
import re
from datetime import datetime, date, timedelta
import random
import bcrypt
import base64
import json
import uuid
import os
import shutil
from pathlib import Path
from urllib.parse import quote
import reflex as rx
from .usuario_model import Usuario, Solicitud
from .solicitud_estado_historial_model import SolicitudEstadoHistorial
from sqlmodel import select, SQLModel, create_engine, text, Session
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from starlette.staticfiles import StaticFiles
from dotenv import load_dotenv
from reflex.components import recharts as rc
from notificaciones import (
    notificar_solicitud_creada,
    notificar_cambio_estado,
    notificar_respuesta_final,
    enviar_correo_smtp,
    formatear_nota_documento,
    get_app_base_url,
)

# Carpeta donde se guardarán los archivos subidos por los usuarios
BASE_DIR = Path(__file__).resolve().parent.parent
UPLOAD_DIR = BASE_DIR / "assets" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
# Segunda carpeta: la que sirve el frontend Vite en /uploads/
WEB_UPLOAD_DIR = BASE_DIR / ".web" / "public" / "uploads"
WEB_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
from typing import Any

# Almacenamiento temporal para descargas (limpieza automática después de acceso)
TEMP_DOWNLOADS = {}

# Cargar variables de entorno
load_dotenv()
# Ruta para configuración de correo almacenada por la app (opcional)
EMAIL_CONFIG_PATH = BASE_DIR / "email_config.json"


def load_email_config() -> dict:
    """Carga configuración de correo desde `email_config.json` si existe y
    aplica valores a `os.environ` cuando sea apropiado.
    """
    try:
        if EMAIL_CONFIG_PATH.exists():
            with open(EMAIL_CONFIG_PATH, "r", encoding="utf-8") as f:
                cfg = json.load(f)
            # Aplicar solo si no están en env ya (permite override por .env)
            for k, v in cfg.items():
                if v is None:
                    continue
                os.environ.setdefault(k, str(v))
            return cfg
    except Exception as e:
        print("No se pudo cargar email_config.json:", e)
    return {}


def save_email_config(cfg: dict) -> None:
    try:
        with open(EMAIL_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f)
        # Also ensure it's available in current process env for immediate use
        for k, v in cfg.items():
            if v is None:
                continue
            os.environ[k] = str(v)
    except Exception as e:
        print("No se pudo guardar email_config.json:", e)


# Load persisted email config on startup (if any)
_EMAIL_CONFIG_CACHE = load_email_config()
DEFAULT_DATABASE_PATH = BASE_DIR / "reflex.db"
DEFAULT_DATABASE_URL = f"sqlite:///{DEFAULT_DATABASE_PATH.as_posix()}"
DATABASE_URL = os.getenv("DATABASE_URL", DEFAULT_DATABASE_URL)
engine = create_engine(DATABASE_URL, echo=False)
SQLModel.metadata.create_all(engine)

from sqlalchemy import inspect

# Asegura que las columnas necesarias existan en la tabla usuario
inspector = inspect(engine)
try:
    columnas_usuario = [col['name'] for col in inspector.get_columns('usuario')]
except Exception:
    columnas_usuario = []

with engine.connect() as conn:
    if columnas_usuario:
        if 'etnia' not in columnas_usuario:
            conn.execute(text("ALTER TABLE usuario ADD COLUMN etnia TEXT"))
        if 'persona_vulnerable' not in columnas_usuario:
            conn.execute(text("ALTER TABLE usuario ADD COLUMN persona_vulnerable TEXT"))
        if 'acepta_notificaciones' not in columnas_usuario:
            conn.execute(text("ALTER TABLE usuario ADD COLUMN acepta_notificaciones INTEGER DEFAULT 0"))
        if 'acepta_politica_datos' not in columnas_usuario:
            conn.execute(text("ALTER TABLE usuario ADD COLUMN acepta_politica_datos INTEGER DEFAULT 0"))
        conn.commit()

# Asegura que la columna persona_vulnerable exista en la tabla solicitud cuando se añada al modelo
try:
    columnas_solicitud = [col['name'] for col in inspector.get_columns('solicitud')]
except Exception:
    columnas_solicitud = []

with engine.connect() as conn:
    if columnas_solicitud:
        if 'persona_vulnerable' not in columnas_solicitud:
            conn.execute(text("ALTER TABLE solicitud ADD COLUMN persona_vulnerable TEXT"))
        if 'fecha_respuesta' not in columnas_solicitud:
            conn.execute(text("ALTER TABLE solicitud ADD COLUMN fecha_respuesta TIMESTAMP"))
        conn.commit()

    # Backfill: si una solicitud ya está cerrada, usar el último cambio de estado como fecha_respuesta.
    try:
        conn.execute(
            text(
                """
                UPDATE solicitud
                SET fecha_respuesta = (
                    SELECT MAX(h.fecha_cambio)
                    FROM solicitudestadohistorial AS h
                    WHERE h.solicitud_id = solicitud.id
                      AND lower(coalesce(h.estado_nuevo,'')) IN ('cerrada','resuelta','finalizada','respondida')
                )
                WHERE fecha_respuesta IS NULL
                  AND lower(coalesce(estado,'')) IN ('cerrada','resuelta','finalizada','respondida')
                """
            )
        )
        conn.commit()
    except Exception as e:
        print("No se pudo backfillear fecha_respuesta:", e)

def tiene_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def confirmar_contraseña(contraseña: str, hashed_password: str) -> bool:
    try:
        return bcrypt.checkpw(contraseña.encode('utf-8'), hashed_password.encode('utf-8'))
    except Exception:
        return False


def normalizar_texto(valor: Any) -> str:
    """Quita espacios sobrantes y aplica formato título para campos de registro."""
    if valor is None:
        return ""
    texto = re.sub(r"\s+", " ", str(valor).strip())
    return texto.title() if texto else ""


def validar_correo(correo: str) -> bool:
    # Solo rechaza si:
    # 1. No tiene @
    # 2. No tiene .
    # 3. La extensión es menor a 2 caracteres (ej: "com", "es", "co" son válidos, pero "c" no)
    if "@" not in correo:
        return False
    if "." not in correo:
        return False
    
    # Validar que después del punto hay al menos 2 caracteres
    partes = correo.split(".")
    if partes[-1].strip() and len(partes[-1].strip()) >= 2:
        return True
    return False

def cantida_minima_contraseña(contraseña: str) -> bool:
    # Requiere: al menos 8 caracteres, una mayúscula, una minúscula,
    # un número y al menos un carácter especial (cualquier signo de puntuación).
    return (
        len(contraseña) >= 8
        and re.search(r'[A-Z]', contraseña)
        and re.search(r'[a-z]', contraseña)
        and re.search(r'[0-9]', contraseña)
        and re.search(r'[^\w\s]', contraseña) is not None
    )

def sanitizar_nombre_archivo(nombre: str) -> str:
    """Sanitiza un nombre de archivo para evitar problemas de seguridad."""
    # Remover caracteres peligrosos
    nombre = re.sub(r'[^\w\s\-\.]', '', nombre)
    # Limitar la longitud
    nombre = nombre[:255]
    return nombre or "archivo"


REFLEX_UPLOAD_DIRS = (
    BASE_DIR / ".web" / "uploaded_files",
    BASE_DIR / ".web" / "backend" / "uploaded_files",
    BASE_DIR / "uploaded_files",
)


def _resolver_ruta_archivo_existente(ruta: str) -> str:
    """Devuelve ruta absoluta si el archivo existe en disco."""
    if not ruta:
        return ""
    candidato = Path(ruta)
    if candidato.is_file():
        return str(candidato.resolve())

    candidatos_busqueda = [
        candidato,
        BASE_DIR / ruta,
        BASE_DIR / ".web" / ruta,
        BASE_DIR / ".web" / "uploaded_files" / candidato.name,
    ]
    for base in REFLEX_UPLOAD_DIRS:
        candidatos_busqueda.extend([base / candidato.name, base / ruta.lstrip("/\\")])
    candidatos_busqueda.append(UPLOAD_DIR / candidato.name)

    for path in candidatos_busqueda:
        if path.is_file():
            return str(path.resolve())
    return ""


EXTENSIONES_IMAGEN = {".png", ".jpg", ".jpeg", ".gif", ".webp"}


def _es_imagen(nombre: str) -> bool:
    return Path(str(nombre or "")).suffix.lower() in EXTENSIONES_IMAGEN


def asegurar_espejo_web(nombre_archivo: str) -> None:
    """Copia el archivo a .web/public/uploads para servirlo en /uploads/ del frontend."""
    if not nombre_archivo:
        return
    nombre = sanitizar_nombre_archivo(os.path.basename(str(nombre_archivo).replace("\\", "/")))
    origen = UPLOAD_DIR / nombre
    if not origen.is_file():
        resuelto = _resolver_ruta_archivo_existente(nombre_archivo)
        if not resuelto:
            return
        origen = Path(resuelto)
        nombre = origen.name
    os.makedirs(WEB_UPLOAD_DIR, exist_ok=True)
    destino = WEB_UPLOAD_DIR / nombre
    try:
        if not destino.is_file() or destino.stat().st_mtime < origen.stat().st_mtime:
            shutil.copy2(origen, destino)
    except Exception:
        pass


def url_archivo_adjunto(nombre_archivo: str) -> str:
    """URL relativa para ver o descargar un adjunto persistido."""
    nombre = sanitizar_nombre_archivo(os.path.basename(str(nombre_archivo or "").replace("\\", "/")))
    if not nombre:
        return ""
    asegurar_espejo_web(nombre)
    return f"/uploads/{quote(nombre)}"


def construir_metadata_adjunto(path_o_nombre: str, basename: str = "") -> dict[str, Any]:
    """Metadatos de adjunto para enlaces y vista previa en la UI."""
    raw = str(path_o_nombre or "").replace("\\", "/")
    nombre = sanitizar_nombre_archivo(basename or os.path.basename(raw))
    if not nombre:
        nombre = sanitizar_nombre_archivo(raw)
    ruta_resuelta = _resolver_ruta_archivo_existente(raw) or _resolver_ruta_archivo_existente(nombre)
    if ruta_resuelta:
        nombre = Path(ruta_resuelta).name
        asegurar_espejo_web(nombre)
    href = url_archivo_adjunto(nombre) if nombre and ruta_resuelta else ""
    es_imagen = _es_imagen(nombre)
    return {
        "basename": nombre,
        "href": href,
        "es_imagen": es_imagen,
        "preview_src": href if es_imagen and href else "",
        "existe": bool(ruta_resuelta),
    }


def _ultimo_adjunto_historial(solicitud_id: int) -> str:
    """�?ltimo nombre de adjunto registrado en el historial de la solicitud."""
    if not solicitud_id:
        return ""
    try:
        with Session(engine) as session:
            rows = session.exec(
                select(SolicitudEstadoHistorial)
                .where(SolicitudEstadoHistorial.solicitud_id == solicitud_id)
                .where(SolicitudEstadoHistorial.documento_adjunto.isnot(None))
                .order_by(SolicitudEstadoHistorial.fecha_cambio.desc())
            ).all()
        for row in rows:
            nombre = str(row.documento_adjunto or "").strip()
            if nombre:
                return nombre
    except Exception:
        pass
    return ""


def persistir_archivo_en_uploads(
    item: Any,
    nombre_preferido: str = "",
    prefijo: str = "adjunto",
) -> str:
    """Guarda un archivo subido en UPLOAD_DIR y devuelve la ruta absoluta."""
    if item is None:
        return ""

    if isinstance(item, list):
        if not item:
            return ""
        item = item[0]

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(WEB_UPLOAD_DIR, exist_ok=True)

    def _guardar_bytes(data: bytes, nombre: str) -> str:
        nombre_limpio = sanitizar_nombre_archivo(nombre)
        destino = UPLOAD_DIR / nombre_limpio
        with open(destino, "wb") as f:
            f.write(data)
        # Espejo en .web/public/uploads/ para que el frontend lo sirva en /uploads/
        destino_web = WEB_UPLOAD_DIR / nombre_limpio
        with open(destino_web, "wb") as f:
            f.write(data)
        return str(destino.resolve())

    if isinstance(item, str):
        if item.startswith("data:"):
            header, b64 = item.split(",", 1)
            mime = header.split(";")[0].split(":")[1] if ":" in header else ""
            ext = mime.split("/")[-1] if "/" in mime else "bin"
            nombre = nombre_preferido or f"{prefijo}_{uuid.uuid4().hex}.{ext}"
            if not os.path.splitext(nombre)[1] and ext not in ("", "bin"):
                nombre = f"{nombre}.{ext}"
            return _guardar_bytes(base64.b64decode(b64), nombre)

        existente = _resolver_ruta_archivo_existente(item)
        if existente:
            destino = UPLOAD_DIR / Path(existente).name
            if Path(existente).resolve() != destino.resolve():
                shutil.copy2(existente, destino)
            destino_web = WEB_UPLOAD_DIR / Path(existente).name
            shutil.copy2(str(destino), str(destino_web))
            return str(destino.resolve())
        return ""

    if isinstance(item, dict):
        nombre = sanitizar_nombre_archivo(
            item.get("name") or item.get("filename") or nombre_preferido or f"{prefijo}_{uuid.uuid4().hex}"
        )
        contenido = item.get("content") or item.get("data") or item.get("file")
        if contenido:
            if isinstance(contenido, str) and contenido.startswith("data:"):
                _, b64 = contenido.split(",", 1)
                data = base64.b64decode(b64)
            elif isinstance(contenido, str):
                data = base64.b64decode(contenido)
            else:
                data = bytes(contenido)
            return _guardar_bytes(data, nombre)

        for clave in ("path", "filepath", "full_path", "tmp_path", "file_path"):
            existente = _resolver_ruta_archivo_existente(str(item.get(clave) or ""))
            if existente:
                destino = UPLOAD_DIR / Path(existente).name
                if Path(existente).resolve() != destino.resolve():
                    shutil.copy2(existente, destino)
                destino_web = WEB_UPLOAD_DIR / Path(existente).name
                shutil.copy2(str(destino), str(destino_web))
                return str(destino.resolve())
        return ""

    nombre_obj = (
        nombre_preferido
        or getattr(item, "filename", None)
        or getattr(item, "name", None)
        or f"{prefijo}_{uuid.uuid4().hex}"
    )
    nombre_final = sanitizar_nombre_archivo(str(nombre_obj))
    for attr in ("path", "file_path", "filepath", "full_path"):
        existente = _resolver_ruta_archivo_existente(str(getattr(item, attr, "") or ""))
        if existente:
            destino = UPLOAD_DIR / nombre_final
            if Path(existente).resolve() != destino.resolve():
                shutil.copy2(existente, destino)
            destino_web = WEB_UPLOAD_DIR / nombre_final
            shutil.copy2(str(destino), str(destino_web))
            return str(destino.resolve())

    file_obj = getattr(item, "file", None)
    if file_obj is not None:
        try:
            if hasattr(file_obj, "seek"):
                file_obj.seek(0)
            data = file_obj.read()
            if data:
                return _guardar_bytes(data, nombre_final)
        except Exception:
            pass

    return ""


def enviar_correo_bienvenida(email_destinatario: str, email_usuario: str):
    """Envía un correo de bienvenida después de un registro exitoso.

    El servicio se activa desde `_crear_usuario()` cuando un ciudadano o funcionario
    se registra correctamente. Usa SMTP (`EMAIL_SENDER`, `EMAIL_PASSWORD`, `SMTP_SERVER`, `SMTP_PORT`).
    Si falla, guarda el correo en local para reintento.
    """
    try:
        email_sender = os.getenv("EMAIL_SENDER", "enlacepqrs1755@gmail.com")
        empresa_nombre = os.getenv("EMPRESA_NOMBRE", "Sistema de Gestión de PQRS")

        # Construir mensaje HTML
        mensaje = MIMEMultipart("alternative")
        mensaje["Subject"] = f"¡Bienvenido a {empresa_nombre}!"
        mensaje["From"] = email_sender
        mensaje["To"] = email_destinatario

        html = f"""
        <html>
            <body style="font-family: Arial, sans-serif; background-color: #f5f5f5; padding: 20px;">
                <div style="max-width: 600px; margin: 0 auto; background-color: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                    <h1 style="color: #1e40af; text-align: center;">¡Bienvenido!</h1>
                    <p style="color: #333; font-size: 16px;">Hola,</p>
                    <p style="color: #333; font-size: 16px;">Tu registro en <strong>{empresa_nombre}</strong> ha sido exitoso. A continuación, encontrarás tus datos de acceso:</p>
                    
                    <div style="background-color: #f0f7ff; padding: 15px; border-left: 4px solid #1e40af; margin: 20px 0; border-radius: 5px;">
                        <p style="margin: 5px 0;"><strong>�??� Correo:</strong> <code>{email_usuario}</code></p>
                    </div>
                    
                    <p style="color: #333; font-size: 16px;">Para iniciar sesión, ingresa a:</p>
                    <p style="text-align: center; margin: 20px 0;">
                        <a href="{get_app_base_url() or 'http://localhost:3000'}/login" style="background-color: #1e40af; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; font-weight: bold;">Ir a Iniciar Sesión</a>
                    </p>
                    
                    <hr style="border: 1px solid #ddd; margin: 20px 0;">
                    <p style="color: #666; font-size: 14px;"><strong>Recuerda:</strong> Nunca compartas tu contraseña con terceros. El equipo de soporte nunca te pedirá tu contraseña.</p>
                    <p style="color: #666; font-size: 14px;">Si tienes preguntas o problemas, contacta a nuestro equipo de soporte.</p>
                    <p style="text-align: center; color: #999; font-size: 12px; margin-top: 30px;">© 2026 {empresa_nombre}. Todos los derechos reservados.</p>
                </div>
            </body>
        </html>
        """

        if enviar_correo_smtp(email_destinatario, mensaje["Subject"], html):
            print(f"�?? Correo enviado exitosamente a {email_destinatario} vía SMTP")
            return True

        print("�? No se pudo enviar el correo de bienvenida por SMTP. Revisa EMAIL_SENDER y EMAIL_PASSWORD.")

        # Registrar correo fallido en disco para reintento manual
        failed_path = BASE_DIR / "failed_emails.log"
        try:
            with open(failed_path, "a", encoding="utf-8") as f:
                f.write(f"{datetime.utcnow().isoformat()} | {email_destinatario} | subject: {mensaje['Subject']}\n{html}\n\n---\n")
            print(f"�?�️ Correo no enviado. Guardado en {failed_path}")
        except Exception as e:
            print("�? No se pudo guardar el correo fallido:", e)

        return False
    except Exception as e:
        print(f"�? Error inesperado al preparar correo: {e}")
        return False

def enviar_correo_notificacion(email_destinatario: str, asunto: str, cuerpo: str) -> bool:
    """Envía una notificación por correo electrónico al ciudadano sobre actualizaciones en su solicitud."""
    try:
        empresa_nombre = os.getenv("EMPRESA_NOMBRE", "Sistema de Gestión de PQRS")
        html = (
            f"<html><body>"
            f"<pre style='font-family:Arial,sans-serif'>{cuerpo}</pre>"
            f"<p style='color:#666;font-size:12px'>© 2026 {empresa_nombre}</p>"
            f"</body></html>"
        )
        if enviar_correo_smtp(email_destinatario, asunto, html):
            print(f"�?? Notificación enviada a {email_destinatario} (SMTP)")
            return True
        print(f"�?�️ No se pudo enviar notificación a {email_destinatario}")
        return False
    except Exception as e:
        print(f"�? Error inesperado al enviar notificación: {e}")
        return False


def _enviar_notificacion_estado_solicitud(
    datos_notificacion: dict[str, Any],
    documentos_respuesta: list[str],
) -> None:
    """Envía correo de cambio de estado en segundo plano (no bloquear la UI)."""
    if not datos_notificacion or not datos_notificacion.get("correo"):
        return
    try:
        adjuntos_correo: list[str] = []
        notas_documento: list[str] = []
        for doc_path in documentos_respuesta or []:
            if not doc_path:
                continue
            doc_name = os.path.basename(doc_path)
            nota_doc, adj = formatear_nota_documento(doc_name, doc_path)
            if nota_doc:
                notas_documento.append(nota_doc)
            adjuntos_correo.extend(adj)
        nota_total = "".join(notas_documento)
        if datos_notificacion["estado_nuevo"].lower() == "cerrada":
            descripcion_respuesta = datos_notificacion["respuesta"] or "Su solicitud ha sido cerrada."
            descripcion_respuesta += nota_total
            notificar_respuesta_final(
                nombre_solicitante=datos_notificacion["nombre"],
                correo_solicitante=datos_notificacion["correo"],
                numero_solicitud=datos_notificacion["radicado"],
                tipo_pqrs=datos_notificacion["tipo"],
                fecha_respuesta=datetime.now().strftime("%d/%m/%Y %H:%M"),
                descripcion_respuesta=descripcion_respuesta,
                adjuntos=adjuntos_correo or None,
            )
        else:
            area_en_obs = datos_notificacion.get("area_asignada", "")
            observaciones = datos_notificacion["respuesta"] or None
            if nota_total:
                observaciones = (observaciones or "") + nota_total
            notificar_cambio_estado(
                nombre_solicitante=datos_notificacion["nombre"],
                correo_solicitante=datos_notificacion["correo"],
                numero_solicitud=datos_notificacion["radicado"],
                estado_anterior=datos_notificacion["estado_anterior"],
                estado_nuevo=datos_notificacion["estado_nuevo"],
                fecha_cambio=datetime.now().strftime("%d/%m/%Y %H:%M"),
                observaciones=observaciones,
                adjuntos=adjuntos_correo or None,
                area_asignada=area_en_obs or None,
            )
    except Exception as e:
        print(f"Error enviando notificación: {e}")


def _slug_filtro_excel(filtro_tipo: str) -> str:
    slug = (filtro_tipo or "todos").strip().lower()
    slug = (
        slug.replace("ó", "o")
        .replace("í", "i")
        .replace("é", "e")
        .replace(" ", "_")
    )
    if slug in {"todos", "todas"}:
        return "general"
    return slug


def _nombre_archivo_excel_reporte(filtro_tipo: str) -> str:
    return f"reporte_pqrs_{_slug_filtro_excel(filtro_tipo)}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"


def _valor_celda_excel(valor: Any) -> Any:
    """Convierte valores de estado Reflex (MutableProxy, listas vacías, etc.) a tipos Excel."""
    if valor is None:
        return ""
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, int):
        return valor
    if isinstance(valor, float):
        return valor
    if isinstance(valor, str):
        return valor
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(valor, date):
        return valor.isoformat()

    tipo = type(valor).__name__
    if "Proxy" in tipo:
        try:
            if hasattr(valor, "keys"):
                valor = dict(valor)
            else:
                valor = list(valor)
        except Exception:
            return str(valor)

    if isinstance(valor, dict):
        if not valor:
            return ""
        return ", ".join(f"{k}: {_valor_celda_excel(v)}" for k, v in valor.items())

    if isinstance(valor, (list, tuple, set)):
        if not valor:
            return ""
        return ", ".join(str(_valor_celda_excel(v)) for v in valor)

    return str(valor)


def _sanitizar_datos_exportacion(data: list[Any]) -> list[dict[str, Any]]:
    """Convierte solicitudes del estado Reflex a dicts con valores nativos para exportación."""
    limpios: list[dict[str, Any]] = []
    for item in data or []:
        try:
            if hasattr(item, "keys"):
                raw = dict(item)
            elif isinstance(item, dict):
                raw = item
            else:
                continue
        except Exception:
            continue
        limpios.append({str(k): _valor_celda_excel(v) for k, v in raw.items()})
    return limpios


def _clasificar_semaforo_solicitud(solicitud: dict[str, Any]) -> str | None:
    """Devuelve 'Verde', 'Amarillo', 'Rojo' para solicitudes activas, o None si está cerrada."""
    closed_states = {
        "resuelta", "cerrada", "respondida", "finalizada",
        "respondido", "cerrado", "finalizado",
    }
    if str(solicitud.get("estado") or "").strip().lower() in closed_states:
        return None
    if solicitud.get("semaforo_expired"):
        return "Rojo"
    fill = str(solicitud.get("semaforo_fill") or "").lower()
    if fill in {"#ef4444", "red", "rojo"}:
        return "Rojo"
    if fill in {"#f59e0b", "orange", "yellow", "amarillo", "#eab308"}:
        return "Amarillo"
    if fill in {"#10b981", "green", "verde"}:
        return "Verde"
    remaining = solicitud.get("semaforo_remaining")
    if remaining is not None:
        try:
            rem = int(float(str(remaining)))
            if rem <= 0:
                return "Rojo"
            if rem <= 5:
                return "Amarillo"
            return "Verde"
        except Exception:
            pass
    return None


def generar_excel_filtrado_con_grafica(
    data: list[dict[str, Any]],
    filtro_tipo: str = "Todos",
) -> bytes:
    """Genera Excel con hoja de datos y hoja 'Gráfica' con charts nativos de Excel."""
    import io
    from collections import Counter

    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font

    data = _sanitizar_datos_exportacion(list(data or []))

    wb = Workbook()
    ws_graf = wb.active
    ws_graf.title = "Gráfica"
    ws_data = wb.create_sheet("Solicitudes")

    if data:
        cols_preferidas = [
            "radicado", "tipo_solicitud", "estado", "asunto", "descripcion",
            "creado_por", "fecha", "area_responsable", "fecha_respuesta",
        ]
        todas = list(data[0].keys())
        cols = [c for c in cols_preferidas if c in todas] + [c for c in todas if c not in cols_preferidas]
        ws_data.append(cols)
        for fila in data:
            ws_data.append([fila.get(c, "") for c in cols])
    else:
        ws_data.append(["Sin registros"])

    closed_states = {
        "resuelta", "cerrada", "respondida", "finalizada",
        "respondido", "cerrado", "finalizado",
    }
    estados_counter: Counter[str] = Counter()
    semaforo = {"Verde": 0, "Amarillo": 0, "Rojo": 0}
    for solicitud in data or []:
        est = str(solicitud.get("estado") or "Sin estado").strip()
        estados_counter[est] += 1
        color = _clasificar_semaforo_solicitud(solicitud)
        if color:
            semaforo[color] += 1

    filtro_label = (filtro_tipo or "Todos").strip()
    total = len(data or [])
    cerradas = sum(
        c for e, c in estados_counter.items()
        if e.lower() in closed_states
    )
    no_cerradas = max(0, total - cerradas)
    pct = int((cerradas / total) * 100) if total else 0

    ws_graf["A1"] = f"Reporte PQRS �?? Filtro: {filtro_label}"
    ws_graf["A1"].font = Font(bold=True, size=16, color="1E3A8A")
    ws_graf.merge_cells("A1:H1")
    ws_graf["A2"] = f"Registros exportados: {total}  |  Cumplimiento: {pct}%"
    ws_graf["A2"].font = Font(size=11, color="475569")

    ws_graf["A4"] = "Estado"
    ws_graf["B4"] = "Cantidad"
    for celda in ("A4", "B4"):
        ws_graf[celda].font = Font(bold=True)

    fila = 5
    for estado, cantidad in sorted(estados_counter.items(), key=lambda x: (-x[1], x[0])):
        ws_graf[f"A{fila}"] = estado
        ws_graf[f"B{fila}"] = cantidad
        fila += 1
    fin_estados = fila - 1

    if fin_estados >= 5:
        bar_estados = BarChart()
        bar_estados.type = "col"
        bar_estados.title = f"Distribución por estado �?? {filtro_label}"
        bar_estados.y_axis.title = "Cantidad"
        bar_estados.style = 10
        bar_estados.height = 14
        bar_estados.width = 24
        bar_estados.add_data(
            Reference(ws_graf, min_col=2, min_row=4, max_row=fin_estados),
            titles_from_data=True,
        )
        bar_estados.set_categories(
            Reference(ws_graf, min_col=1, min_row=5, max_row=fin_estados)
        )
        ws_graf.add_chart(bar_estados, "D4")

    fila_pie = max(fila + 2, 4)
    ws_graf[f"A{fila_pie}"] = "Nivel de cumplimiento"
    ws_graf[f"A{fila_pie}"].font = Font(bold=True, size=12)
    hdr = fila_pie + 1
    ws_graf[f"A{hdr}"] = "Categoría"
    ws_graf[f"B{hdr}"] = "Cantidad"
    ws_graf[f"A{hdr}"].font = Font(bold=True)
    ws_graf[f"B{hdr}"].font = Font(bold=True)
    ws_graf[f"A{hdr + 1}"] = "Cerradas / Resueltas"
    ws_graf[f"B{hdr + 1}"] = cerradas
    ws_graf[f"A{hdr + 2}"] = "Pendientes"
    ws_graf[f"B{hdr + 2}"] = no_cerradas

    if total > 0 and (cerradas > 0 or no_cerradas > 0):
        pie = PieChart()
        pie.title = f"Cumplimiento: {pct}%"
        pie.height = 12
        pie.width = 16
        pie.add_data(
            Reference(ws_graf, min_col=2, min_row=hdr + 1, max_row=hdr + 2)
        )
        pie.set_categories(
            Reference(ws_graf, min_col=1, min_row=hdr + 1, max_row=hdr + 2)
        )
        ws_graf.add_chart(pie, f"D{fila_pie}")

    sem_total = sum(semaforo.values())
    if sem_total > 0:
        fila_sem = hdr + 5
        ws_graf[f"A{fila_sem}"] = "Semáforo de vencimiento (activas)"
        ws_graf[f"A{fila_sem}"].font = Font(bold=True, size=12)
        sh = fila_sem + 1
        ws_graf[f"A{sh}"] = "Estado"
        ws_graf[f"B{sh}"] = "Cantidad"
        ws_graf[f"A{sh + 1}"] = "Verde"
        ws_graf[f"B{sh + 1}"] = semaforo["Verde"]
        ws_graf[f"A{sh + 2}"] = "Amarillo"
        ws_graf[f"B{sh + 2}"] = semaforo["Amarillo"]
        ws_graf[f"A{sh + 3}"] = "Rojo"
        ws_graf[f"B{sh + 3}"] = semaforo["Rojo"]
        bar_sem = BarChart()
        bar_sem.type = "col"
        bar_sem.title = "Estado de vencimiento"
        bar_sem.height = 12
        bar_sem.width = 18
        bar_sem.add_data(
            Reference(ws_graf, min_col=2, min_row=sh + 1, max_row=sh + 3)
        )
        bar_sem.set_categories(
            Reference(ws_graf, min_col=1, min_row=sh + 1, max_row=sh + 3)
        )
        ws_graf.add_chart(bar_sem, "G4")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# quitar prints de prueba

class State(rx.State):
    # --- Modal de Vencimiento de Reportes ---
    vencimiento_modal_abierto: bool = False
    rango_vencimiento_seleccionado: str = ""
    solicitudes_vencimiento_filtradas: list[dict] = []
    detalle_solicitud_modal_abierto: bool = False
    # --- Modal confirmación de cierre de sesión ---
    show_logout_confirm: bool = False

    def abrir_detalle_solicitud(self, solicitud_id: int):
        for s in self.solicitudes:
            if s.get("id") == solicitud_id:
                self.solicitud_consultada = dict(s)
                self.cargar_historial_solicitud(solicitud_id)
                self.detalle_solicitud_modal_abierto = True
                break

    def cerrar_detalle_solicitud(self):
        self.detalle_solicitud_modal_abierto = False

    def abrir_vencimiento_modal(self, data: Any):
        import logging
        print(f"[LOG] abrir_vencimiento_modal llamado con data: {data} (tipo: {type(data)})")
        
        rango = ""
        # 1. Si es un diccionario directo
        if isinstance(data, dict):
            # Podría venir en data['name'] o data['activeLabel']
            rango = data.get("name") or data.get("activeLabel") or ""
        # 2. Si viene de una lista de payload (en algunos charts de Recharts)
        elif isinstance(data, list) and len(data) > 0:
            item = data[0]
            if isinstance(item, dict):
                rango = item.get("name") or item.get("payload", {}).get("name", "")
        # 3. Si es un string directo
        elif isinstance(data, str):
            rango = data
            
        print(f"[LOG] Rango determinado: '{rango}'")
        if not rango:
            print(f"[LOG] Rango vacío, no se hace nada.")
            return

        self.rango_vencimiento_seleccionado = rango
        self.solicitudes_vencimiento_filtradas = []
        
        filtered = []
        for s in (self.solicitudes_filtradas or []):
            rem = s.get("semaforo_remaining")
            
            match = False
            if rem is None:
                if rango == ">10 días":
                    match = True
            else:
                if rango == "Vencidas" and rem <= 0:
                    match = True
                elif rango == "1-5 días" and 1 <= rem <= 5:
                    match = True
                elif rango == "6-10 días" and 6 <= rem <= 10:
                    match = True
                elif rango == ">10 días" and rem > 10:
                    match = True
                
            if match:
                filtered.append({
                    "id": s.get("id"),
                    "radicado": s.get("radicado") or f"ID-{s.get('id')}",
                    "tipo_solicitud": s.get("tipo_solicitud") or s.get("tipo_pqrs") or "N/A",
                    "asunto": s.get("asunto") or "Sin asunto",
                    "creado_por": s.get("creado_por") or "N/A",
                    "estado": s.get("estado") or "Radicada",
                    "area_responsable": s.get("area_responsable") or "N/A",
                    "semaforo_remaining": s.get("semaforo_remaining"),
                    "semaforo_fill": s.get("semaforo_fill") or "gray",
                    "is_expired": bool(rem is not None and rem <= 0),
                    "remaining_str": "Vencida" if (rem is not None and rem <= 0) else (f"{int(rem)} días" if rem is not None else "N/A")
                })
        
        print(f"[LOG] Encontradas {len(filtered)} solicitudes para el rango '{rango}'")
        self.solicitudes_vencimiento_filtradas = filtered
        self.vencimiento_modal_abierto = True

    def abrir_semaforo_activas_modal(self, color: str):
        """Abre el modal de vencimiento filtrando solo solicitudes activas (misma base que Excel)."""
        color = str(color or "").strip()
        labels = {
            "Verde": "Verde (>5 días)",
            "Amarillo": "Amarillo (1-5 días)",
            "Rojo": "Rojo (Vencidas)",
        }
        if color not in labels:
            return

        self.rango_vencimiento_seleccionado = labels[color]
        filtered = []
        for s in (self.solicitudes_abiertas or []):
            if _clasificar_semaforo_solicitud(s) != color:
                continue
            rem = s.get("semaforo_remaining")
            filtered.append({
                "id": s.get("id"),
                "radicado": s.get("radicado") or f"ID-{s.get('id')}",
                "tipo_solicitud": s.get("tipo_solicitud") or s.get("tipo_pqrs") or "N/A",
                "asunto": s.get("asunto") or "Sin asunto",
                "creado_por": s.get("creado_por") or "N/A",
                "estado": s.get("estado") or "Radicada",
                "area_responsable": s.get("area_responsable") or "N/A",
                "semaforo_remaining": rem,
                "semaforo_fill": s.get("semaforo_fill") or "gray",
                "is_expired": bool(rem is not None and rem <= 0),
                "remaining_str": "Vencida" if (rem is not None and rem <= 0) else (f"{int(rem)} días" if rem is not None else "N/A"),
            })

        self.solicitudes_vencimiento_filtradas = filtered
        self.vencimiento_modal_abierto = True

    def cerrar_vencimiento_modal(self):
        self.vencimiento_modal_abierto = False
        self.rango_vencimiento_seleccionado = ""
        self.solicitudes_vencimiento_filtradas = []

    # --- Historial de estados ---
    historial_modal_abierto: bool = False
    historial_solicitud_id: int = 0
    historial_estados: list[dict[str, Any]] = []

    def abrir_historial(self, solicitud_id: int):
        """Carga el historial de estados de una solicitud y abre el modal."""
        self.historial_solicitud_id = solicitud_id
        self.historial_estados = []
        from .solicitud_estado_historial_model import SolicitudEstadoHistorial
        with Session(engine) as session:
            rows = session.exec(
                select(SolicitudEstadoHistorial)
                .where(SolicitudEstadoHistorial.solicitud_id == solicitud_id)
                .order_by(SolicitudEstadoHistorial.fecha_cambio)
            ).all()
            estados: list[dict[str, Any]] = []
            for h in rows:
                obs_raw = h.observaciones or ""
                doc = h.documento_adjunto or ""
                if not doc:
                    match = re.search(r"\[DOCUMENTO(?: ADJUNTO)?:\s*([^\]\|]+)\]", obs_raw)
                    if match:
                        doc = match.group(1).strip()
                        obs_raw = re.sub(r"\s*\[DOCUMENTO(?: ADJUNTO)?:[^\]]+\]", "", obs_raw).strip()
                doc_meta = construir_metadata_adjunto(doc) if doc else {}
                estados.append(
                    {
                        "fecha": h.fecha_cambio.strftime("%Y-%m-%d %H:%M"),
                        "anterior": h.estado_anterior,
                        "nuevo": h.estado_nuevo,
                        "obs": obs_raw,
                        "documento_adjunto": doc_meta.get("basename", doc),
                        "documento_href": doc_meta.get("href", ""),
                        "documento_preview_src": doc_meta.get("preview_src", ""),
                        "documento_existe": doc_meta.get("existe", False),
                    }
                )
            self.historial_estados = estados
        self.historial_modal_abierto = True

    def cerrar_historial(self):
        self.historial_modal_abierto = False
        self.historial_solicitud_id = 0
        self.historial_estados = []


    # Dentro de class State, agrega estas variables:
    toast_mensaje: str = ""
    toast_tipo: str = ""  # "success" o "error"
    toast_visible: bool = False

    "En esta clase se define el estado de la aplicación, es decir, las variables que se van a usar en la aplicación y sus valores iniciales."
    state_auto_setters = True
    contraseña: str = ""
    confirmar_contraseña: str = ""
    correo: str = ""
    confirmar_correo: str = ""
    # Campos adicionales para registro extendido
    tipo_identificacion: str = ""
    numero_identificacion: str = ""
    apellidos: str = ""
    sexo: str = ""
    direccion: str = ""
    telefono: str = ""
    departamento: str = ""
    ciudad: str = ""
    etnia: str = ""
    persona_vulnerable_registro: str = ""
    # Estados de validación UX
    correo_validado: bool = False
    confirmar_correo_match: bool = False
    numero_identificacion_valid: bool = False
    nombres_valid: bool = False
    apellidos_valid: bool = False
    telefono_valid: bool = False
    departamento_valid: bool = False
    ciudad_valid: bool = False
    
    # Diccionario de departamentos y ciudades para dropdowns dinámicos
    departamentos_ciudades =  {
    "Amazonas": ["Leticia", "Puerto Nariño", "La Chorrera", "Tarapacá", "Puerto Santander", "Mirití-Paraná", "Puerto Alegría", "Puerto Arica", "La Victoria"],
    "Antioquia": ["Medellín", "Envigado", "Sabaneta", "Copacabana", "Girardota", "Barbosa", "Itagüí", "Bello", "Caldas", "La Estrella", "Rionegro", "La Ceja", "Apartadó", "Turbo", "Caucasia", "Santa Rosa de Osos"],
    "Arauca": ["Arauca", "Arauquita", "Cravo Norte", "Saravena", "Tame"],
    "Atlántico": ["Barranquilla", "Soledad", "Malambo", "Puerto Colombia", "Sabanalarga", "Baranoa", "Galapa"],
    "Bogotá D.C.": ["Bogotá D.C."],
    "Bolívar": ["Cartagena", "Turbaco", "Magangué", "Arjona", "El Carmen de Bolívar", "Mompox"],
    "Boyacá": ["Tunja", "Duitama", "Sogamoso", "Paipa", "Chiquinquirá", "Villa de Leyva", "Puerto Boyacá"],
    "Caldas": ["Manizales", "La Dorada", "Riosucio", "Chinchiná", "Villamaría", "Anserma"],
    "Caquetá": ["Florencia", "San Vicente del Caguán", "Puerto Rico", "Currillo"],
    "Casanare": ["Yopal", "Aguazul", "Paz de Ariporo", "Tauramena", "Maní"],
    "Cauca": ["Popayán", "Guachené", "Corinto", "Santander de Quilichao", "Puerto Tejada", "Patía"],
    "Cesar": ["Valledupar", "Aguachica", "Agustín Codazzi", "Bosconia", "Curumaní"],
    "Chocó": ["Quibdó", "Istmina", "Condoto", "Acandí", "Bahía Solano"],
    "Córdoba": ["Montería", "Cereté", "Sahagún", "Lorica", "Montelíbano", "Planeta Rica"],
    "Cundinamarca": ["Soacha", "Chía", "Sopó", "Tausa", "Tenjo", "Tena", "Tocaima", "Tocancipá", "Zipaquirá", "Fúquene", "Pacho", "�?tica", "Villapinzón", "Villeta", "Facatativá", "Girardot", "Fusagasugá"],
    "Guainía": ["Inírida", "Barrancominas"],
    "Guaviare": ["San José del Guaviare", "Calamar", "El Retorno", "Miraflores"],
    "Huila": ["Neiva", "Pitalito", "Garzón", "La Plata", "Campoalegre", "San Agustín"],
    "La Guajira": ["Riohacha", "Maicao", "Uribia", "San Juan del Cesar", "Fonseca"],
    "Magdalena": ["Santa Marta", "Ciénaga", "Fundación", "El Banco", "Plato"],
    "Meta": ["Villavicencio", "Acacías", "Granada", "Puerto López", "Cumaral"],
    "Nariño": ["Pasto", "Ipiales", "Tumaco", "Sandoná", "Túquerres", "La Unión"],
    "Norte de Santander": ["Cúcuta", "Ocaña", "Pamplona", "Villa del Rosario", "Los Patios", "Tibú"],
    "Putumayo": ["Mocoa", "Puerto Asís", "Orito", "Valle del Guamuez", "Sibundoy"],
    "Quindío": ["Armenia", "Calarcá", "Filandia", "Circasia", "Montenegro", "Quimbaya"],
    "Risaralda": ["Pereira", "Dosquebradas", "Santa Rosa de Cabal", "La Virginia", "Belén de Umbría"],
    "San Andrés y Providencia": ["San Andrés", "Providencia"],
    "Santander": ["Bucaramanga", "Floridablanca", "Girón", "Piedecuesta", "Barrancabermeja", "San Gil", "Socorro"],
    "Sucre": ["Sincelejo", "Corozal", "Tolú", "San Marcos", "Sampués"],
    "Tolima": ["Ibagué", "Espinal", "Melgar", "Mariquita", "Honda", "Líbano"],
    "Valle del Cauca": ["Cali", "Palmira", "Yumbo", "Cartago", "Buenaventura", "Tuluá", "Buga", "Jamundí"],
    "Vaupés": ["Mitú", "Carurú", "Taraira"],
    "Vichada": ["Puerto Carreño", "La Primavera", "Santa Rosalía", "Cumaribo"]
}
    
    # Habeas data / autorizaciones
    acepta_notificaciones: bool = False
    acepta_politica_datos: bool = False
    # Para el formulario de solicitudes
    acepta_politica_solicitud: bool = False
    area_responsable: str = ""
    area_otro: str = ""
    tipo_solicitud: str = ""
    persona_vulnerable: str = ""
    asunto: str = ""
    descripcion: str = ""
    ubicacion: str = ""
    documento: str = ""
    documentos: list[dict[str, str | int]] = []
    documento_nombres: list[str] = []
    documento_nombre: str = ""
    documento_previews: list[dict[str, str]] = []
    descripcion_len: int = 0
    query_solicitud: str = ""
    filter_tipo_solicitud: str = "Todos"
    filter_estado_solicitud: str = "Todos"
    filter_dias_restantes: str = "Todos"
    solicitudes: list[dict[str, Any]] = []
    editar_solicitud_id: int = 0
    eliminar_solicitud_id: int = 0
    solicitud_mensaje: str = ""
    pqrs_contacto_email: str = ""
    mantenimiento_modal_abierto: bool = False

    def abrir_mantenimiento(self):
        self.mantenimiento_modal_abierto = True

    def cerrar_mantenimiento(self):
        self.mantenimiento_modal_abierto = False


    error_de_registro: str = ""
    succes: str = ""
    error_de_contraseña: str = ""
    succes2: str = ""
    
    id_usuario: str = rx.Cookie("0")
    es_autentica: str = rx.Cookie("false")
    email_actual: str = rx.Cookie("")
    correo_usuario: str = rx.Cookie("")
    rol_usuario: str = rx.Cookie("")
    nombres: str = rx.Cookie("")
    show_password: bool = False
    # Campos para cambiar contraseña
    current_password: str = ""
    new_password: str = ""
    confirm_new_password: str = ""
    change_pw_message: str = ""
    # Campos para cambiar rol de ciudadano a funcionario
    cambiar_rol_email: str = ""
    cambiar_rol_mensaje: str = ""
    confirmar_promocion_rol: bool = False
    usuarios_registrados: list[dict[str, Any]] = []
    # Campos para editar estado de solicitud
    editar_estado_id: int = 0
    nuevo_estado: str = ""
    respuesta_solicitud: str = ""
    mensaje_actualizar_estado: str = ""
    respuesta_documento: str = ""
    respuesta_documento_nombre: str = ""
    respuesta_documento_content: str = ""
    respuesta_documento_preview_src: str = ""
    respuesta_documento_es_imagen: bool = False
    respuesta_documento_error: str = ""
    respuesta_documentos: list[dict[str, Any]] = []
    # HU9 y HU11: historial, calificación, estados
    historial_solicitud: list[dict[str, Any]] = []
    calificacion_seleccionada: int = 0
    calificacion_opcion: str = ""
    comentario_calificacion: str = ""
    calificacion_mensaje: str = ""
    ESTADOS_VALIDOS: list[str] = ["Radicada", "Asignada a Area", "En Gestion de Area", "Solucionada", "Reabierta", "Cerrada"]
    AREAS_DISPONIBLES: list[str] = ["Atención al Ciudadano", "Infraestructura", "Servicios Públicos", "Medio Ambiente", "Salud Pública", "Educación", "Secretaría Jurídica", "Planeación", "Hacienda"]
    area_para_asignar: str = ""
    # Variables para modal de política y validaciones
    modal_politica_visible: bool = False
    archivo_error_mensaje: str = ""
    correo_confirmacion_visible: bool = False
    correo_confirmacion_mensaje: str = ""
    ayuda_seccion_abierta: str = "estados"
    registro_seccion_abierta: str = "cuenta"
    registro_paso_habilitado: int = 1
    
    # Campos para asignación de área con mensaje
    asignar_area_id: int = 0
    asignar_area_mensaje: str = ""
    asignar_area_nombre: str = ""
    asignar_area_seleccionada: str = ""
    mensaje_asignacion: str = ""
    
    # Campos para consultar estado de solicitud
    consulta_radicado: str = ""
    solicitud_consultada: dict[str, Any] = {}
    consulta_mensaje: str = ""
    # Enlace generado tras exportar reportes (archivo descargable)
    export_href: str = ""
    export_filename: str = ""
    mostrar_menu_descarga: bool = False

    @rx.var
    def id_usuario_num(self) -> int:
        """Convierte el id de cookie a entero seguro."""
        try:
            return int(str(self.id_usuario or "0"))
        except Exception:
            return 0

    @rx.var
    def es_autenticada(self) -> bool:
        """Normaliza la cookie de sesión a booleano."""
        return str(self.es_autentica or "").strip().lower() in {"true", "1", "yes", "si", "sí"}

    @rx.var
    def home_url(self) -> str:
        """Retorna la URL de inicio dependiendo del estado de autenticación."""
        if self.es_autenticada:
            if self.rol_usuario == "funcionario":
                return "/dashboard-funcionario"
            return "/dashboard"
        return "/"

    def redirigir_si_autenticado(self):
        """Redirige automáticamente al usuario a su panel si ya tiene una sesión activa."""
        if self.es_autenticada:
            return rx.redirect(self.home_url)

    @rx.var
    def ciudades_disponibles(self) -> list[str]:
        """Retorna las ciudades del departamento seleccionado."""
        if self.departamento in self.departamentos_ciudades:
            return self.departamentos_ciudades[self.departamento]
        return []

    @rx.var
    def data_grafica_tipo(self) -> list[dict]:
        counts = self.estadisticas_por_tipo
        selected = (self.filter_tipo_solicitud or "Todos").strip()
        if selected.lower() in {"todas", "todos"}:
            return [
                {"name": "Petición", "cantidad": counts.get("Petición", 0)},
                {"name": "Queja", "cantidad": counts.get("Queja", 0)},
                {"name": "Reclamo", "cantidad": counts.get("Reclamo", 0)},
                {"name": "Sugerencia", "cantidad": counts.get("Sugerencia", 0)},
            ]

        normalized = self._normalize_tipo_solicitud(selected)
        if normalized:
            return [{"name": normalized, "cantidad": counts.get(normalized, 0)}]

        return []
    
    @rx.var
    def data_grafica_estado(self) -> list[dict]:
        return [
            {"name": "Radicada", "cantidad": int(self.numero_solicitudes_radicadas)},
            {"name": "En Proceso", "cantidad": int(self.numero_solicitudes_actualizadas)},
            {"name": "Cerrada", "cantidad": int(self.numero_solicitudes_cerradas)},
        ]

    def mostrar_toast(self, mensaje: str, tipo: str = "success"):
        self.toast_mensaje = mensaje
        self.toast_tipo = tipo
        self.toast_visible = True
        # Ocultar automáticamente después de 2.5 segundos
        import threading
        threading.Timer(2.5, lambda: setattr(self, 'toast_visible', False)).start()

    def export_reportes_csv(self):
        """Genera un CSV en memoria desde `solicitudes_filtradas` para descarga respetando filtros."""
        try:
            import csv
            import io
            from datetime import datetime

            data = _sanitizar_datos_exportacion(self.solicitudes_filtradas or [])
            if not data:
                self.mostrar_toast("No hay datos para exportar.", "warning")
                self.mostrar_menu_descarga = False
                return
            
            # Generar CSV en memoria
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=list(data[0].keys()))
            writer.writeheader()
            writer.writerows(data)
            
            # Guardar en almacenamiento temporal
            csv_bytes = output.getvalue().encode('utf-8')
            filename = f"reportes_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.csv"
            download_id = str(uuid.uuid4())
            
            TEMP_DOWNLOADS[download_id] = {
                "data": csv_bytes,
                "filename": filename,
                "mime": "text/csv; charset=utf-8"
            }
            
            self.export_filename = filename
            self.mostrar_toast(f"�?? CSV generado. {len(data)} registros. Descargando...", "success")
            self.mostrar_menu_descarga = False
            self.export_href = f"/api/download/{download_id}"
            
        except Exception as e:
            print(f"ERROR en export_reportes_csv: {type(e).__name__}: {e}")
            self.mostrar_toast(f"Error: {str(e)[:100]}", "error")
            self.mostrar_menu_descarga = False

    def descargar_excel_y_abrir(self):
        """Genera Excel en memoria con datos filtrados, gráfica nativa y descarga."""
        try:
            data = _sanitizar_datos_exportacion(self.solicitudes_filtradas or [])
            if not data:
                self.mostrar_toast("No hay datos para exportar.", "warning")
                return

            excel_bytes = generar_excel_filtrado_con_grafica(
                data,
                self.filter_tipo_solicitud,
            )
            filename = _nombre_archivo_excel_reporte(self.filter_tipo_solicitud)
            self.mostrar_menu_descarga = False
            self.mostrar_toast(
                f"Excel con gráfica generado ({len(data)} registros �?? {self.filter_tipo_solicitud}).",
                "success",
            )
            return rx.download(
                data=excel_bytes,
                filename=filename,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except (ImportError, ModuleNotFoundError):
            self.mostrar_toast("No se pudo generar Excel. Descargando CSV en su lugar.", "warning")
            return self.descargar_csv_y_abrir()
        except Exception as e:
            print(f"ERROR en descargar_excel_y_abrir: {e}")
            self.mostrar_toast(f"Error exportando Excel: {str(e)[:100]}", "error")

    def descargar_csv_y_abrir(self):
        """Genera CSV en memoria con datos FILTRADOS y dispara la descarga."""
        try:
            import csv
            import io
            from datetime import datetime

            data = _sanitizar_datos_exportacion(self.solicitudes_filtradas or [])
            if not data:
                self.mostrar_toast("No hay datos para exportar.", "warning")
                return

            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=list(data[0].keys()))
            writer.writeheader()
            writer.writerows(data)

            filename = f"reportes_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.csv"
            return rx.download(
                data=output.getvalue().encode('utf-8'),
                filename=filename,
                mime_type="text/csv; charset=utf-8"
            )
        except Exception as e:
            print(f"ERROR en descargar_csv_y_abrir: {e}")
            self.mostrar_toast(f"Error exportando CSV: {str(e)[:100]}", "error")


    def ocultar_toast(self):
        self.toast_visible = False
        self.toast_mensaje = ""

    def toggle_menu_descarga(self):
        """Alterna la visibilidad del menú de descarga."""
        self.mostrar_menu_descarga = not self.mostrar_menu_descarga
        
     
    @rx.var
    def numero_solicitudes(self) -> str:
        return str(len(self.solicitudes_filtradas or []))
    
    @rx.var
    def numero_solicitudes_radicadas(self) -> str:
        return str(sum(1 for solicitud in (self.solicitudes_filtradas or []) if (str(solicitud.get('estado') or '').strip().lower()) == 'radicada'))
    
    @rx.var
    def numero_solicitudes_actualizadas(self) -> str:
        # 'en proceso', 'actualizada', 'asignada' y 'en revisión' representan solicitudes activas / en proceso
        # Tambien se cuentan 'asignada a area' y 'en gestion de area'
        valid_states = ('en proceso', 'actualizada', 'asignada', 'en revisión', 'en revision', 'asignada a area', 'en gestion de area')
        return str(sum(1 for solicitud in (self.solicitudes_filtradas or []) if (str(solicitud.get('estado') or '').strip().lower()) in valid_states))
    
    @rx.var
    def numero_solicitudes_cerradas(self) -> str:
        return str(sum(1 for solicitud in (self.solicitudes_filtradas or []) if (str(solicitud.get('estado') or '').strip().lower()) in ('cerrada', 'finalizada', 'resuelta')))
    
    def _normalize_tipo_solicitud(self, tipo_raw: str) -> str:
        """Normaliza tipos de solicitud a las categorías usadas en los reportes."""
        if not tipo_raw:
            return ""
        tipo = str(tipo_raw).strip().lower()
        if tipo in ("peticion", "petición", "pqr", "solicitud"):
            return "Petición"
        if tipo == "queja":
            return "Queja"
        if tipo == "reclamo":
            return "Reclamo"
        if tipo == "sugerencia":
            return "Sugerencia"
        if "petici" in tipo:
            return "Petición"
        if "queja" in tipo:
            return "Queja"
        if "reclam" in tipo:
            return "Reclamo"
        if "suger" in tipo:
            return "Sugerencia"
        return ""

    @rx.var
    def estadisticas_por_tipo(self) -> dict[str, int]:
        counts = {"Petición": 0, "Queja": 0, "Reclamo": 0, "Sugerencia": 0}
        for solicitud in self.solicitudes_filtradas or []:
            tipo = self._normalize_tipo_solicitud(solicitud.get("tipo_solicitud") or "")
            if tipo in counts:
                counts[tipo] += 1
        return counts

    @rx.var
    def max_registros_tipo(self) -> int:
        values = list(self.estadisticas_por_tipo.values())
        return max(values) if values else 1

    @rx.var
    def monthly_response_times(self) -> list[dict]:
        """Calcula el tiempo promedio de respuesta diario para los últimos 30 días.
        Devuelve lista de dicts: {month: '26 May', dias: float}
        """
        today = date.today()
        start_date = today - timedelta(days=29)
        
        # preparar buckets por día
        buckets: dict[str, list[int]] = {}
        for i in range(30):
            d = start_date + timedelta(days=i)
            buckets[d.strftime("%Y-%m-%d")] = []

        for s in (self.solicitudes_filtradas or []):
            try:
                frp = self._parse_dt(s.get('fecha_respuesta'))
                if not frp:
                    continue
                resp_date = frp.date()
                if resp_date < start_date or resp_date > today:
                    continue
                fr = self._parse_dt(s.get('fecha') or s.get('fecha_radicado'))
                if not fr:
                    continue
                start = fr.date() + timedelta(days=1)
                dias = self._business_days_between(start, resp_date, set())
                
                key = resp_date.strftime("%Y-%m-%d")
                buckets.setdefault(key, []).append(dias)
            except Exception:
                continue

        result = []
        for i in range(30):
            d = start_date + timedelta(days=i)
            key = d.strftime("%Y-%m-%d")
            vals = buckets.get(key, [])
            avg = round(sum(vals) / len(vals), 1) if vals else 0
            label = d.strftime('%d %b')
            result.append({"month": label, "dias": avg})

        return result

    @rx.var
    def compliance_percentage(self) -> int:
        """Calcula cumplimiento: (solicitudes resueltas / total) * 100.
        Resueltas = estado 'Respondida' o 'Cerrada'.
        """
        total = len(self.solicitudes_filtradas or [])
        if total == 0:
            return 0
        closed_states = {"resuelta", "cerrada", "respondida", "finalizada"}
        cerradas = sum(1 for s in (self.solicitudes_filtradas or []) if (s.get('estado') or "").lower() in closed_states)
        return int((cerradas / total) * 100) if total > 0 else 0

    @rx.var
    def compliance_chart_data(self) -> list[dict]:
        total = len(self.solicitudes_filtradas or [])
        closed_states = {"resuelta", "cerrada", "respondida", "finalizada"}
        cerradas = sum(1 for s in (self.solicitudes_filtradas or []) if (s.get('estado') or "").lower() in closed_states)
        no_cerradas = total - cerradas
        return [
            {"name": "Cerradas", "value": cerradas, "fill": "#10b981"},
            {"name": "No cerradas", "value": max(0, no_cerradas), "fill": "#ef4444"},
        ]

    # --- Semáforo: días hábiles y conteos por color ---
    def _parse_dt(self, v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v
        try:
            # ISO format usually works
            return datetime.fromisoformat(v)
        except Exception:
            try:
                from dateutil import parser as _p
                return _p.parse(v)
            except Exception:
                return None

    def _is_business_day(self, d: date, holidays: set):
        return d.weekday() < 5 and d not in holidays

    def _business_days_between(self, start: date, end: date, holidays: set) -> int:
        if end < start:
            return 0
        days = 0
        cur = start
        while cur <= end:
            if self._is_business_day(cur, holidays):
                days += 1
            cur += timedelta(days=1)
        return days

    def _legal_days_for(self, tipo: str, detalle: str | None = None) -> int:
        if not tipo:
            return 15
        t = tipo.lower()
        d = (detalle or "").lower()
        if "consulta" in d or t == "consulta":
            return 30
        if "inform" in d or "copia" in d or "informacion" in d:
            return 10
        if t in ("peticion", "petición", "queja", "reclamo", "sugerencia"):
            return 15
        return 15


    @staticmethod
    def _compute_remaining_for_solicitud(solicitud: dict) -> dict:
        """Computa días restantes y color (fill) para una solicitud dada.
        Usa llaves comunes que retorna `_solicitud_a_dict` como `fecha` y `tipo_solicitud`.
        Retorna dict con `remaining` (int or None) y `fill` (hex color).
        """
        try:
            fecha_raw = solicitud.get("fecha") or solicitud.get("fecha_radicado")
            if not fecha_raw:
                return {"remaining": None, "fill": "gray"}
            # intentar parseo ISO, sino dateutil
            try:
                dt = datetime.fromisoformat(str(fecha_raw))
            except Exception:
                try:
                    from dateutil import parser as _p
                    dt = _p.parse(str(fecha_raw))
                except Exception:
                    return {"remaining": None, "fill": "gray"}

            start = dt.date()
            
            estado_raw = str(solicitud.get("estado") or "").strip().lower()
            closed_states = {"respondida", "respondido", "cerrada", "cerrado", "finalizada", "finalizado", "resuelta", "resuelto"}
            
            ref = date.today()
            if estado_raw in closed_states and solicitud.get("fecha_respuesta"):
                try:
                    ref_dt = datetime.fromisoformat(str(solicitud.get("fecha_respuesta")))
                    ref = ref_dt.date()
                except Exception:
                    try:
                        from dateutil import parser as _p
                        ref = _p.parse(str(solicitud.get("fecha_respuesta"))).date()
                    except Exception:
                        pass

            # Contar días calendario desde la fecha de creación hasta la referencia calculada.
            days = max((ref - start).days, 0)

            tipo = (solicitud.get("tipo_solicitud") or solicitud.get("tipo_pqrs") or "").lower()
            if "consulta" in tipo:
                legal = 30
            elif "inform" in tipo or "copia" in tipo or "informacion" in tipo:
                legal = 10
            elif tipo in ("peticion", "petición", "queja", "reclamo", "sugerencia"):
                legal = 15
            else:
                legal = 15

            remaining = legal - days
            if remaining <= 0:
                fill = "#ef4444"
            elif remaining <= 5:
                fill = "#f59e0b"
            else:
                fill = "#10b981"
            # Representar visualmente el avance hacia el vencimiento.
            if legal > 0:
                width = int(min(max((legal - remaining) / legal * 100, 0), 100))
            else:
                width = 0
            return {"remaining": remaining, "fill": fill, "width": width}
        except Exception:
            return {"remaining": None, "fill": "gray", "width": 0}

    @rx.var
    def semaforo_counts(self) -> dict:
        # Lee solicitudes_filtradas y devuelve conteo por color (dinámico por filtros)
        holidays = set()  # puedes poblar con una consulta a festivos si la tienes
        counts = {"verde": 0, "amarillo": 0, "rojo": 0}
        # Estados que consideramos cerrados/resueltos (normalizados en minúsculas)
        closed_states = {"respondida", "respondido", "respondida", "respondida", "cerrada", "cerrado", "finalizada", "finalizado"}
        for s in (self.solicitudes_filtradas or []):
            estado_raw = (s.get("estado") or "").strip().lower()
            # Si el estado está en la lista de cerrados, lo saltamos; así consideramos activo todo lo demás
            if estado_raw in closed_states:
                continue
            fr = self._parse_dt(s.get("fecha") or s.get("fecha_radicado"))
            if not fr:
                # intentar usar la llave 'fecha_radicado' si existe (compatibilidad)
                fr = self._parse_dt(s.get("fecha_radicado"))
            if not fr:
                continue
            start = fr.date() + timedelta(days=1)
            ref = date.today()
            if s.get("fecha_respuesta"):
                resp = self._parse_dt(s.get("fecha_respuesta"))
                if resp:
                    ref = resp.date()
            used = self._business_days_between(start, ref, holidays)
            # usar `tipo_solicitud` por consistencia con `_solicitud_a_dict`
            legal = self._legal_days_for(s.get("tipo_solicitud"), s.get("tipo_detalle") or s.get("asunto"))
            remaining = legal - used
            if remaining <= 0:
                counts["rojo"] += 1
            elif remaining <= 5:
                counts["amarillo"] += 1
            else:
                counts["verde"] += 1
        return counts

    @rx.var
    def semaforo_chart_data(self) -> list[dict]:
        c = self.semaforo_counts
        return [
            {"name": "Verde", "value": c.get("verde", 0), "fill": "#10b981"},
            {"name": "Amarillo", "value": c.get("amarillo", 0), "fill": "#f59e0b"},
            {"name": "Rojo", "value": c.get("rojo", 0), "fill": "#ef4444"},
        ]

    @rx.var
    def semaforo_total(self) -> int:
        c = self.semaforo_counts
        total_from_counts = int(c.get("verde", 0) + c.get("amarillo", 0) + c.get("rojo", 0))
        # Fallback: si no hay conteos, usar data_grafica_tipo (cantidad)
        fallback = 0
        try:
            for it in (self.data_grafica_tipo or []):
                fallback += int(it.get("cantidad", 0))
        except Exception:
            fallback = 0
        return max(total_from_counts, fallback)

    @rx.var
    def semaforo_bar_data(self) -> list[dict]:
        # Devuelve una lista con un único registro que contiene los valores por color
        c = self.semaforo_counts
        total = int(c.get("verde", 0) + c.get("amarillo", 0) + c.get("rojo", 0))
        if total > 0:
            return [{
                "name": "Semáforo",
                "verde": int(c.get("verde", 0)),
                "amarillo": int(c.get("amarillo", 0)),
                "rojo": int(c.get("rojo", 0)),
            }]
        # Fallback a partir de data_grafica_tipo: sumar todas las solicitudes en verde (representación)
        fallback = 0
        try:
            for it in (self.data_grafica_tipo or []):
                fallback += int(it.get("cantidad", 0))
        except Exception:
            fallback = 0
        return [{"name": "Semáforo", "verde": fallback, "amarillo": 0, "rojo": 0}]
    search_area_query: str = ""
    
    @rx.var
    def top_areas(self) -> list[dict]:
        """Devuelve las áreas responsables por cantidad de solicitudes filtradas (de mayor a menor), filtrables por búsqueda.
        """
        counts = {}
        for s in (self.solicitudes_filtradas or []):
            a = s.get('area_responsable') or 'N/A'
            counts[a] = counts.get(a, 0) + 1
        items = sorted(counts.items(), key=lambda x: x[1], reverse=True)
        
        q = self.search_area_query.strip().lower()
        if q:
            items = [item for item in items if q in str(item[0]).lower()]
            
        return [{"name": name, "total": total} for name, total in items]

    @rx.var
    def solicitudes_por_vencer_data(self) -> list[dict]:
        counts = {
            "Vencidas": 0,
            "1-5 días": 0,
            "6-10 días": 0,
            ">10 días": 0,
        }
        for s in (self.solicitudes_filtradas or []):
            rem = s.get("semaforo_remaining")
            if rem is None:
                counts[">10 días"] += 1
                continue
            if rem <= 0:
                counts["Vencidas"] += 1
            elif rem <= 5:
                counts["1-5 días"] += 1
            elif rem <= 10:
                counts["6-10 días"] += 1
            else:
                counts[">10 días"] += 1
        return [{"name": label, "cantidad": total} for label, total in counts.items()]

    @rx.var
    def solicitudes_por_vencer_total(self) -> int:
        return sum(item.get("cantidad", 0) for item in self.solicitudes_por_vencer_data)

    @rx.var
    def kpi_counts(self) -> dict:
        """Devuelve los 3 KPIs: pendientes (>10 días), en proceso (1-10 días), cerradas (estados cerrados).
        Usa solicitudes filtradas para ser dinámico.
        """
        closed_states = {"respondida", "respondido", "cerrada", "cerrado", "finalizada", "finalizado", "resuelta", "resuelto"}
        pendientes = 0
        en_curso = 0
        cerradas = 0
        for s in (self.solicitudes_filtradas or []):
            estado_raw = str(s.get("estado") or "").strip().lower()
            if estado_raw in closed_states:
                cerradas += 1
                continue
            rem = s.get("semaforo_remaining")
            if rem is None:
                continue
            try:
                rem_i = int(rem)
            except Exception:
                continue
            if rem_i > 10:
                pendientes += 1
            elif 1 <= rem_i <= 10:
                en_curso += 1
            # rem_i <= 0 (vencidas) quedan fuera de estos KPIs según definición
        return {"pendientes": pendientes, "en_curso": en_curso, "cerradas": cerradas}

    @rx.var
    def kpi_pendientes_count(self) -> int:
        return int(self.numero_solicitudes_radicadas)

    @rx.var
    def kpi_en_curso_count(self) -> int:
        return int(self.numero_solicitudes_actualizadas)

    @rx.var
    def kpi_cerradas_count(self) -> int:
        return int(self.numero_solicitudes_cerradas)

    @rx.var
    def solicitudes_filtradas(self) -> list[dict]:
        query = (self.query_solicitud or "").strip().lower()
        tipo = (self.filter_tipo_solicitud or "Todos").lower()
        rango = (self.filter_dias_restantes or "Todos").lower()
        resultados = []
        for solicitud in self.solicitudes or []:
            texto = " ".join(
                str(solicitud.get(field, "") or "")
                for field in ("radicado", "asunto", "descripcion", "creado_por")
            ).lower()
            if query and query not in texto:
                continue
            if tipo not in {"todas", "todos"} and self._normalize_tipo_solicitud(solicitud.get("tipo_solicitud", "")).lower() != tipo:
                continue
            if rango not in {"todas", "todos"}:
                expired = bool(solicitud.get("semaforo_expired", False))
                remaining_raw = solicitud.get("semaforo_remaining", 0)
                try:
                    remaining = int(float(str(remaining_raw).strip() or "0"))
                except Exception:
                    remaining = 0
                if rango == "vencidas":
                    if not expired:
                        continue
                elif rango == "0-3":
                    if expired or remaining < 0 or remaining > 3:
                        continue
                elif rango == "4-10":
                    if expired or remaining < 4 or remaining > 10:
                        continue
                elif rango == "11+":
                    if expired or remaining < 11:
                        continue
            resultados.append(solicitud)
        return resultados

    @rx.var
    def solicitudes_abiertas(self) -> list[dict]:
        closed_states = {"cerrada", "resuelta", "finalizada", "respondida"}
        estado_filtro = (self.filter_estado_solicitud or "Todos").lower()
        res = []
        for s in (self.solicitudes_filtradas or []):
            s_est = str(s.get("estado") or "").strip().lower()
            if s_est in closed_states:
                continue
            if estado_filtro not in {"todas", "todos"} and s_est != estado_filtro:
                continue
            res.append(s)
        return res

    @rx.var
    def semaforo_activas_chart_data(self) -> list[dict]:
        """Semáforo Verde/Amarillo/Rojo solo para solicitudes activas (igual que Excel)."""
        counts = {"Verde": 0, "Amarillo": 0, "Rojo": 0}
        for s in (self.solicitudes_abiertas or []):
            color = _clasificar_semaforo_solicitud(s)
            if color in counts:
                counts[color] += 1
        return [
            {"name": "Verde", "cantidad": counts["Verde"], "fill": "#10b981"},
            {"name": "Amarillo", "cantidad": counts["Amarillo"], "fill": "#f59e0b"},
            {"name": "Rojo", "cantidad": counts["Rojo"], "fill": "#ef4444"},
        ]

    @rx.var
    def semaforo_activas_total(self) -> int:
        return sum(int(item.get("cantidad", 0)) for item in self.semaforo_activas_chart_data)

    @rx.var
    def max_semaforo_activas_cantidad(self) -> int:
        valores = [int(item.get("cantidad", 0)) for item in self.semaforo_activas_chart_data]
        maximo = max(valores) if valores else 0
        return max(maximo, 1)

    @rx.var
    def solicitudes_cerradas_lista(self) -> list[dict]:
        closed_states = {"cerrada", "resuelta", "finalizada", "respondida"}
        return [
            s
            for s in (self.solicitudes_filtradas or [])
            if str(s.get("estado") or "").strip().lower() in closed_states
        ]

    @rx.var
    def documento_nombres_joined(self) -> str:
        return ", ".join(self.documento_nombres or [])
    
    @rx.var
    def documento_nombres_count(self) -> str:
        return str(len(self.documento_nombres or []))

    @rx.var
    def documento_tamano_total(self) -> str:
        total = 0
        for item in self.documentos or []:
            if isinstance(item, dict):
                total += int(item.get("size") or 0)
        return f"{total / (1024 * 1024):.2f} MB"
    
    @rx.var
    def usuarios_registrados_count(self) -> int:
        return len(self.usuarios_registrados or [])
    
    @rx.var
    def solicitud_consultada_adjuntos(self) -> list[dict]:
        docs = self.solicitud_consultada.get("documento_adjuntos", [])
        if isinstance(docs, list) and docs:
            return docs
        return []
    
    
    
    def set_query_solicitud(self, value: str):
        self.query_solicitud = value or ""
    
    def set_filter_tipo_solicitud(self, value: str):
        self.filter_tipo_solicitud = value or "Todos"

    def set_filter_estado_solicitud(self, value: str):
        self.filter_estado_solicitud = value or "Todas"

    def set_filter_dias_restantes(self, value: str):
        self.filter_dias_restantes = value or "Todos"

    def buscar_solicitudes(self):
        self.query_solicitud = (self.query_solicitud or "").strip()

    def set_new_password(self, value: str):
        self.new_password = value
    
    def set_confirm_new_password(self, value: str):
        self.confirm_new_password = value
    
    def borrar_mensajes_de_estado(self):
        self.error_de_registro = ""
        self.succes = ""
        self.error_de_contraseña = ""
        self.succes2 = ""
        
    def validacion_de_entradas(self, require_strong_pw: bool = True) -> bool:
        self.correo_confirmacion_visible = False
        self.correo_confirmacion_mensaje = ""

        if not self.correo or not self.correo.strip():
            self.error_de_registro = "El correo electrónico es obligatorio."
            return False

        if not validar_correo(self.correo):
            self.error_de_registro = "Correo no válido."
            self.correo_confirmacion_visible = True
            self.correo_confirmacion_mensaje = "Correo no válido."
            return False

        if not self.confirmar_correo or not self.confirmar_correo.strip():
            self.error_de_registro = "Debes confirmar tu correo electrónico."
            return False

        if self.correo != self.confirmar_correo:
            self.error_de_registro = "Los correos electrónicos no coinciden."
            return False

        self.correo_confirmacion_visible = True
        self.correo_confirmacion_mensaje = "Correo válido."

        if not self.contraseña or not self.contraseña.strip():
            self.error_de_registro = "La contraseña es obligatoria."
            return False

        if not self.confirmar_contraseña or not self.confirmar_contraseña.strip():
            self.error_de_registro = "Debes confirmar tu contraseña."
            return False

        if require_strong_pw and not cantida_minima_contraseña(self.contraseña):
            self.error_de_registro = "La contraseña debe tener al menos 8 caracteres, incluyendo mayúsculas, minúsculas, números y caracteres especiales."
            return False
        if require_strong_pw and self.contraseña != self.confirmar_contraseña:
            self.error_de_registro = "Las contraseñas no coinciden."
            return False
        return True

    def set_confirmar_correo_match(self, value: bool):
        self.confirmar_correo_match = bool(value)

    def limpiar_formulario_registro(self):
        """Limpia todos los campos del formulario de registro."""
        self.correo = ""
        self.confirmar_correo = ""
        self.contraseña = ""
        self.confirmar_contraseña = ""
        self.tipo_identificacion = ""
        self.numero_identificacion = ""
        self.nombres = ""
        self.apellidos = ""
        self.sexo = ""
        self.direccion = ""
        self.telefono = ""
        self.departamento = ""
        self.ciudad = ""
        self.etnia = ""
        self.persona_vulnerable_registro = ""
        self.acepta_notificaciones = False
        self.acepta_politica_datos = False
        self.correo_validado = False
        self.confirmar_correo_match = False
        self.numero_identificacion_valid = False
        self.nombres_valid = False
        self.apellidos_valid = False
        self.telefono_valid = False
        self.departamento_valid = False
        self.ciudad_valid = False
        self.error_de_registro = ""
        self.correo_confirmacion_visible = False
        self.correo_confirmacion_mensaje = ""
        self.show_password = False
    def validar_campo_simple(self, campo: str) -> bool:
        """Validaciones simples para mostrar iconos de confirmación. Retorna True si el campo parece correcto."""
        val = getattr(self, campo, "")
        ok = False
        if campo == "telefono":
            ok = isinstance(val, str) and len(val) >= 7
        elif campo == "numero_identificacion":
            ok = isinstance(val, str) and len(val) >= 6
        elif campo == "correo":
            ok = validar_correo(val)
        else:
            ok = bool(val and str(val).strip())
        # set dedicated flags for reactivity
        if campo == "telefono":
            self.telefono_valid = ok
        elif campo == "numero_identificacion":
            self.numero_identificacion_valid = ok
        elif campo == "nombres":
            self.nombres_valid = ok
        elif campo == "apellidos":
            self.apellidos_valid = ok
        elif campo == "departamento":
            self.departamento_valid = ok
        elif campo == "ciudad":
            self.ciudad_valid = ok
        return ok

    def validar_correo_accion(self):
        """Acción invocada por el botón 'Validar' junto al correo."""
        self.correo_validado = validar_correo(self.correo)
        if not self.correo_validado:
            self.error_de_registro = "Correo inválido."
        else:
            self.error_de_registro = ""
        return

    # Setters that also validate so we can show inline icons
    def set_and_validate_nombres(self, val: str):
        self.nombres = val
        self.validar_campo_simple("nombres")

    def set_and_validate_apellidos(self, val: str):
        self.apellidos = val
        self.validar_campo_simple("apellidos")

    def set_and_validate_numero_identificacion(self, val: str):
        self.numero_identificacion = val
        self.validar_campo_simple("numero_identificacion")

    def set_and_validate_telefono(self, val: str):
        self.telefono = val
        self.validar_campo_simple("telefono")

    def set_and_validate_departamento(self, val: str):
        self.departamento = val
        self.ciudad = ""  # Limpia la ciudad cuando cambia el departamento
        self.validar_campo_simple("departamento")

    def set_and_validate_ciudad(self, val: str):
        self.ciudad = val
        self.validar_campo_simple("ciudad")

    def set_and_validate_correo(self, val: str):
        """Valida el correo en tiempo real y borra el mensaje si es válido o vacío."""
        self.correo = val or ""
        
        # Si el correo está vacío, borra el mensaje
        if not self.correo.strip():
            self.correo_confirmacion_visible = False
            self.correo_confirmacion_mensaje = ""
            self.error_de_registro = ""
            self.correo_validado = False
            return
        
        # Si es válido, muestra mensaje verde y borra error
        if validar_correo(self.correo):
            self.correo_confirmacion_visible = True
            self.correo_confirmacion_mensaje = "Correo válido."
            self.error_de_registro = ""
            self.correo_validado = True
        else:
            # Si es inválido, muestra mensaje rojo
            self.correo_confirmacion_visible = True
            self.correo_confirmacion_mensaje = "Correo no válido."
            self.error_de_registro = "Correo no válido."
            self.correo_validado = False

    def set_etnia(self, val: str):
        self.etnia = val or ""

    def set_persona_vulnerable_registro(self, val: str):
        self.persona_vulnerable_registro = val or ""

    def set_etnia(self, val: str):
        self.etnia = val or ""

    def set_modal_politica_visible(self, visible: bool):
        self.modal_politica_visible = bool(visible)

    def set_archivo_error_mensaje(self, mensaje: str):
        self.archivo_error_mensaje = mensaje or ""

    def set_correo_confirmacion_visible(self, visible: bool):
        self.correo_confirmacion_visible = bool(visible)

    def set_correo_confirmacion_mensaje(self, mensaje: str):
        self.correo_confirmacion_mensaje = mensaje or ""

    def validar_email(self):
        if validar_correo(self.correo):
            self.correo_confirmacion_visible = True
            self.correo_confirmacion_mensaje = "Correo válido."
        else:
            self.correo_confirmacion_visible = True
            self.correo_confirmacion_mensaje = "Correo no válido."

    def set_descripcion(self, val: str):
        # Guardar descripción y longitud para el contador de caracteres
        self.descripcion = val if val is not None else ""
        # Limitar a 1000 caracteres en la UI
        if len(self.descripcion) > 1000:
            self.descripcion = self.descripcion[:1000]
        self.descripcion_len = len(self.descripcion)

    def set_tipo_solicitud(self, val: str):
        self.tipo_solicitud = val or ""

    def set_persona_vulnerable(self, val: str):
        self.persona_vulnerable = val or ""

    def set_asunto(self, val: str):
        self.asunto = val or ""

    def set_ubicacion(self, val: str):
        self.ubicacion = val or ""

    def set_acepta_politica_solicitud(self, checked: bool):
        self.acepta_politica_solicitud = bool(checked)

    def preconfirmar_politica(self, checked: bool):
        if checked:
            self.modal_politica_visible = True
        else:
            self.acepta_politica_datos = False

    def confirmar_politica(self):
        self.acepta_politica_datos = True
        self.modal_politica_visible = False

    def cancelar_politica(self):
        self.acepta_politica_datos = False
        self.modal_politica_visible = False

    def set_acepta_notificaciones(self, checked: bool):
        self.acepta_notificaciones = bool(checked)

    def _infer_mime_type(self, name: str) -> str:
        ext = os.path.splitext((name or "").lower())[1].lstrip(".")
        if ext in {"png"}:
            return "image/png"
        if ext in {"jpg", "jpeg"}:
            return "image/jpeg"
        if ext in {"webp"}:
            return "image/webp"
        if ext in {"gif"}:
            return "image/gif"
        if ext in {"pdf"}:
            return "application/pdf"
        return "application/octet-stream"

    def _build_file_preview(self, item: Any) -> dict[str, str]:
        name = "adjunto"
        raw_content = ""
        if isinstance(item, dict):
            name = item.get("name") or item.get("filename") or "adjunto"
            if item.get("src"):
                return {
                    "name": str(name),
                    "src": str(item.get("src") or ""),
                    "mime": str(item.get("mime") or self._infer_mime_type(str(name))),
                }
            path_val = str(item.get("path") or "")
            if path_val and _es_imagen(str(name)):
                meta = construir_metadata_adjunto(path_val, str(name))
                if meta.get("preview_src"):
                    return {
                        "name": str(name),
                        "src": meta["preview_src"],
                        "mime": self._infer_mime_type(str(name)),
                    }
            raw_content = str(item.get("content") or item.get("data") or "")
        elif isinstance(item, str):
            name = os.path.basename(item) or "adjunto"
        else:
            name = (
                getattr(item, "name", None)
                or getattr(item, "filename", None)
                or "adjunto"
            )

        mime = self._infer_mime_type(name)
        es_imagen = mime.startswith("image/")
        src = ""
        if es_imagen and raw_content:
            if raw_content.startswith("data:"):
                src = raw_content
            else:
                src = f"data:{mime};base64,{raw_content}"
        return {"name": name, "src": src, "mime": mime}

    def _upload_to_preview_dict(self, upload: rx.UploadFile) -> dict[str, str | int]:
        name = (
            getattr(upload, "name", None)
            or getattr(upload, "filename", None)
            or "adjunto"
        )
        size = int(getattr(upload, "size", 0) or 0)
        path_raw = getattr(upload, "path", None)
        path_val = str(path_raw) if path_raw else ""
        if path_val and not os.path.isfile(path_val):
            resuelto = _resolver_ruta_archivo_existente(path_val)
            if resuelto:
                path_val = resuelto
        mime = self._infer_mime_type(name)
        src = ""
        content_b64 = ""
        try:
            file_obj = getattr(upload, "file", None)
            if file_obj is not None:
                if hasattr(file_obj, "seek"):
                    file_obj.seek(0)
                content_bytes = file_obj.read()
                if content_bytes:
                    content_b64 = base64.b64encode(content_bytes).decode("utf-8")
                    if mime.startswith("image/"):
                        src = f"data:{mime};base64,{content_b64}"
        except Exception:
            content_b64 = ""
        return {"name": str(name), "size": size, "path": path_val, "src": src, "mime": mime, "content": content_b64}

    def set_documento(self, files: list[rx.UploadFile] = None, documento: list[rx.UploadFile] = None, **kwargs):
        """Actualiza adjuntos del ciudadano sin guardar UploadFile en el estado."""
        self.documentos = []
        self.documento_nombres = []
        self.documento = ""
        self.documento_nombre = ""
        self.documento_previews = []
        self.archivo_error_mensaje = ""

        allowed_ext = {"pdf", "png", "jpg", "jpeg"}
        max_files = 3
        max_total_size = 10 * 1024 * 1024
        
        target = files if files is not None else documento
        if target is None:
            return
        archivos = target if isinstance(target, list) else [target]

        if len(archivos) > max_files:
            self.archivo_error_mensaje = "Solo puedes adjuntar hasta 3 archivos."
            return

        normalized: list[dict[str, str | int]] = []
        total_size = 0
        for item in archivos:
            item_data = self._upload_to_preview_dict(item)
            name = str(item_data.get("name") or "adjunto")
            size = int(item_data.get("size") or 0)
            path_val = str(item_data.get("path") or "")
            ext = os.path.splitext(name)[1].lower().lstrip(".")
            if ext not in allowed_ext:
                self.archivo_error_mensaje = "Solo se aceptan archivos PDF, PNG o JPG."
                return
            if size > max_total_size:
                self.archivo_error_mensaje = "Cada archivo no puede superar los 10MB."
                return
            total_size += size

            # Persistir inmediatamente el archivo durante el evento de subida
            os.makedirs(UPLOAD_DIR, exist_ok=True)
            os.makedirs(WEB_UPLOAD_DIR, exist_ok=True)
            nombre_limpio = sanitizar_nombre_archivo(name)
            destino = UPLOAD_DIR / nombre_limpio
            destino_web = WEB_UPLOAD_DIR / nombre_limpio

            guardado_ok = False
            if path_val and os.path.exists(path_val):
                try:
                    shutil.copy2(path_val, destino)
                    shutil.copy2(path_val, destino_web)
                    guardado_ok = True
                except Exception:
                    pass

            if not guardado_ok:
                content_b64 = item_data.get("content")
                if content_b64:
                    try:
                        data = base64.b64decode(content_b64)
                        with open(destino, "wb") as f:
                            f.write(data)
                        with open(destino_web, "wb") as f:
                            f.write(data)
                        guardado_ok = True
                    except Exception:
                        pass

            if not guardado_ok:
                self.archivo_error_mensaje = f"No se pudo guardar el archivo '{name}' en el servidor."
                return

            item_data["path"] = str(destino.resolve())
            normalized.append(item_data)

        if total_size > max_total_size:
            self.archivo_error_mensaje = "La suma de los archivos no puede superar los 10MB."
            return

        self.documentos = normalized
        self.documento_nombres = [
            os.path.basename(str(item.get("path") or ""))
            or sanitizar_nombre_archivo(str(item.get("name") or "adjunto"))
            for item in normalized
        ]
        self.documento_nombre = ", ".join(self.documento_nombres)
        self.documento_previews = [self._build_file_preview(item) for item in normalized]

    def set_editar_solicitud_id(self, id: int):
        self.editar_solicitud_id = id

    def set_eliminar_solicitud_id(self, id: int):
        self.eliminar_solicitud_id = id

    def eliminar_documento(self, index: int):
        # Reconstruir listas para evitar comportamiento reactivo inesperado
        if not (0 <= index < len(self.documentos)):
            return
        nuevos_docs: list[dict[str, str | int]] = []
        nuevos_nombres: list[str] = []
        for i, item in enumerate(self.documentos):
            if i == index:
                continue
            nuevos_docs.append(item)
        for i, nombre in enumerate(self.documento_nombres):
            if i == index:
                continue
            nuevos_nombres.append(nombre)
        self.documentos = nuevos_docs
        self.documento_nombres = nuevos_nombres
        self.documento_nombre = ", ".join(self.documento_nombres)
        self.documento_previews = [self._build_file_preview(item) for item in self.documentos]
        self.archivo_error_mensaje = ""

    def eliminar_documento_por_nombre(self, nombre: str):
        if nombre in self.documento_nombres:
            index = self.documento_nombres.index(nombre)
            self.eliminar_documento(index)

    def quitar_todos_documentos(self):
        self.documento = ""
        self.documentos = []
        self.documento_nombres = []
        self.documento_nombre = ""
        self.documento_previews = []
        self.archivo_error_mensaje = ""

    def confirmar_editar_solicitud(self):
        if self.editar_solicitud_id:
            self.editar_solicitud(self.editar_solicitud_id)
            self.editar_solicitud_id = 0

    def confirmar_eliminar_solicitud(self):
        if self.eliminar_solicitud_id:
            self.eliminar_solicitud(self.eliminar_solicitud_id)
            self.eliminar_solicitud_id = 0

    def set_area_responsable(self, val: str):
        self.area_responsable = val
        if val != "Otros":
            self.area_otro = ""

    def set_area_otro(self, val: str):
        self.area_otro = val

    def set_nuevo_estado(self, val: str):
        self.nuevo_estado = val
        if val != "Asignada a Area":
            self.area_para_asignar = ""

    def set_area_para_asignar(self, val: str):
        self.area_para_asignar = val


    def set_respuesta_solicitud(self, val: str):
        self.respuesta_solicitud = val

    def _limpiar_campos_respuesta_documento(self) -> None:
        self.respuesta_documento = ""
        self.respuesta_documento_nombre = ""
        self.respuesta_documento_content = ""
        self.respuesta_documento_preview_src = ""
        self.respuesta_documento_es_imagen = False

    def _sincronizar_respuesta_documento_legacy(self) -> None:
        """Mantiene campos legacy alineados con el primer adjunto de la lista."""
        if self.respuesta_documentos:
            primero = self.respuesta_documentos[0]
            self.respuesta_documento = str(primero.get("path") or "")
            self.respuesta_documento_nombre = str(primero.get("nombre") or "")
            self.respuesta_documento_content = str(primero.get("content") or "")
            self.respuesta_documento_preview_src = str(primero.get("preview_src") or "")
            self.respuesta_documento_es_imagen = bool(primero.get("es_imagen"))
        else:
            self._limpiar_campos_respuesta_documento()

    def set_respuesta_documento(self, files: list[rx.UploadFile] = None, documento: list[rx.UploadFile] = None, **kwargs):
        """Agrega adjuntos a la respuesta del funcionario (hasta 3) sin recargar el dashboard."""
        self.mensaje_actualizar_estado = ""
        self.respuesta_documento_error = ""
        max_files = 3
        allowed_ext = {"pdf", "png", "jpg", "jpeg"}

        target = files if files is not None else documento
        if target is None:
            return
        archivos = target if isinstance(target, list) else [target]

        nuevos: list[dict[str, Any]] = list(self.respuesta_documentos or [])
        nombres_actuales = {str(d.get("nombre") or "") for d in nuevos}

        for archivo in archivos:
            if archivo is None:
                continue
            if len(nuevos) >= max_files:
                self.respuesta_documento_error = "Solo puedes adjuntar hasta 3 archivos."
                break

            item_data = self._upload_to_preview_dict(archivo)
            name = str(item_data.get("name") or "respuesta_adjunto")
            ext = os.path.splitext(name)[1].lower().lstrip(".")
            if ext not in allowed_ext:
                self.respuesta_documento_error = (
                    f"Tipo de archivo no permitido: .{ext}. Solo se aceptan PDF, PNG o JPG."
                )
                continue

            ruta_guardada = persistir_archivo_en_uploads(
                archivo,
                nombre_preferido=name,
                prefijo="respuesta",
            )
            if not ruta_guardada and item_data.get("content"):
                ruta_guardada = persistir_archivo_en_uploads(
                    {"name": name, "content": item_data["content"]},
                    nombre_preferido=name,
                    prefijo="respuesta",
                )
            if not ruta_guardada:
                self.respuesta_documento_error = (
                    f"No se pudo guardar el archivo '{name}' en el servidor. Intenta de nuevo."
                )
                continue

            basename = os.path.basename(ruta_guardada)
            if basename in nombres_actuales:
                continue
            asegurar_espejo_web(basename)
            meta = construir_metadata_adjunto(ruta_guardada, basename)
            nuevos.append(
                {
                    "path": ruta_guardada,
                    "nombre": basename,
                    "content": str(item_data.get("content") or ""),
                    "preview_src": str(meta.get("preview_src") or item_data.get("src") or ""),
                    "es_imagen": bool(meta.get("es_imagen")),
                }
            )
            nombres_actuales.add(basename)

        self.respuesta_documentos = nuevos
        self._sincronizar_respuesta_documento_legacy()

    def quitar_respuesta_documento_por_nombre(self, nombre: str):
        """Quita un adjunto por nombre sin recargar el dashboard."""
        nombre = str(nombre or "").strip()
        if not nombre:
            return rx.clear_selected_files("upload_respuesta")
        self.respuesta_documentos = [
            d for d in self.respuesta_documentos if str(d.get("nombre") or "") != nombre
        ]
        self._sincronizar_respuesta_documento_legacy()
        self.respuesta_documento_error = ""
        return rx.clear_selected_files("upload_respuesta")

    def quitar_todos_respuesta_documentos(self):
        """Quita todos los adjuntos pendientes sin recargar el dashboard."""
        self.respuesta_documentos = []
        self._limpiar_campos_respuesta_documento()
        self.respuesta_documento_error = ""
        return rx.clear_selected_files("upload_respuesta")

    def quitar_respuesta_documento(self):
        """Alias: quitar todos los adjuntos del modal de estado."""
        return self.quitar_todos_respuesta_documentos()

    def set_asignar_area_mensaje(self, val: str):
        self.asignar_area_mensaje = val

    def set_asignar_area_seleccionada(self, val: str):
        self.asignar_area_seleccionada = val or ""

    def set_asignar_area_nombre(self, val: str):
        self.asignar_area_nombre = val or ""

    def cerrar_editor_estado(self):
        self.editar_estado_id = 0
        self.nuevo_estado = ""
        self.respuesta_solicitud = ""
        self.respuesta_documentos = []
        self._limpiar_campos_respuesta_documento()
        self.respuesta_documento_error = ""
        self.mensaje_actualizar_estado = ""
        return rx.clear_selected_files("upload_respuesta")

    def set_consulta_radicado(self, val: str):
        self.consulta_radicado = val

    def toggle_ayuda_seccion(self, seccion: str):
        self.ayuda_seccion_abierta = "" if self.ayuda_seccion_abierta == seccion else seccion

    def set_registro_seccion_abierta(self, seccion: str):
        seccion = seccion or "cuenta"
        paso = {"cuenta": 1, "identidad": 2, "ubicacion": 3}.get(seccion, 1)
        if paso <= self.registro_paso_habilitado:
            self.registro_seccion_abierta = seccion
            return
        self.error_de_registro = "Completa el paso actual para continuar."

    def _validar_paso_cuenta(self) -> bool:
        self.error_de_registro = ""
        correo = str(self.correo or "").strip()
        if not correo:
            self.error_de_registro = "Completa el correo electrónico."
            return False
        if not validar_correo(correo):
            self.error_de_registro = "Correo no válido."
            return False
        if not str(self.contraseña or "").strip():
            self.error_de_registro = "Completa la contraseña."
            return False
        if not cantida_minima_contraseña(self.contraseña):
            self.error_de_registro = "La contraseña debe tener al menos 8 caracteres, incluyendo mayúsculas, minúsculas, números y caracteres especiales."
            return False
        if not str(self.confirmar_contraseña or "").strip():
            self.error_de_registro = "Confirma la contraseña."
            return False
        if self.contraseña != self.confirmar_contraseña:
            self.error_de_registro = "Las contraseñas no coinciden."
            return False
        return True

    def _validar_paso_identidad(self) -> bool:
        self.error_de_registro = ""
        if not str(self.nombres or "").strip():
            self.error_de_registro = "Completa el campo Nombres."
            return False
        if not str(self.apellidos or "").strip():
            self.error_de_registro = "Completa el campo Apellidos."
            return False
        if not str(self.sexo or "").strip():
            self.error_de_registro = "Selecciona el sexo."
            return False
        if not str(self.tipo_identificacion or "").strip():
            self.error_de_registro = "Selecciona el tipo de identificación."
            return False
        if len(str(self.numero_identificacion or "").strip()) < 6:
            self.error_de_registro = "El número de identificación debe tener al menos 6 caracteres."
            return False
        if len(re.sub(r"\D", "", str(self.telefono or ""))) < 7:
            self.error_de_registro = "Ingresa un teléfono válido con al menos 7 dígitos."
            return False
        return True

    def continuar_a_identidad(self):
        if not self._validar_paso_cuenta():
            return
        self.registro_paso_habilitado = max(self.registro_paso_habilitado, 2)
        self.registro_seccion_abierta = "identidad"
        self.error_de_registro = ""

    def continuar_a_ubicacion(self):
        if not self._validar_paso_identidad():
            return
        self.registro_paso_habilitado = max(self.registro_paso_habilitado, 3)
        self.registro_seccion_abierta = "ubicacion"
        self.error_de_registro = ""

    def actualizar_estado_solicitud(self, files: Any = None):
        """Actualiza el estado de una solicitud con validación para cerrada y guarda historial. HU9."""
        self.mensaje_actualizar_estado = ""
        if not self.editar_estado_id or not self.nuevo_estado:
            self.mensaje_actualizar_estado = "Selecciona un estado válido."
            return
        # HU9: justificación obligatoria para CUALQUIER cambio de estado
        if not str(self.respuesta_solicitud or "").strip():
            self.mensaje_actualizar_estado = "Debes escribir una justificación para el cambio de estado."
            return
        # Si es Asignada a Area, validar que el area est�?© seleccionada
        if self.nuevo_estado == "Asignada a Area" and not str(self.area_para_asignar or "").strip():
            self.mensaje_actualizar_estado = "Selecciona el área a la que se va a asignar la solicitud."
            return
        documentos_respuesta_guardados: list[str] = []
        intento_adjunto = bool(
            self.respuesta_documentos
            or self.respuesta_documento_nombre
            or self.respuesta_documento_content
            or self.respuesta_documento
        )

        def _persistir_fuente(fuente: Any, nombre_preferido: str = "") -> str:
            if not fuente:
                return ""
            try:
                return persistir_archivo_en_uploads(
                    fuente,
                    nombre_preferido=nombre_preferido,
                    prefijo="respuesta",
                ) or ""
            except Exception:
                return ""

        for doc in self.respuesta_documentos or []:
            path = str(doc.get("path") or "")
            nombre_doc = str(doc.get("nombre") or "")
            if path and os.path.isfile(path):
                documentos_respuesta_guardados.append(path)
                asegurar_espejo_web(os.path.basename(path))
                continue
            if doc.get("content") and nombre_doc:
                ruta = _persistir_fuente(
                    {"name": nombre_doc, "content": doc["content"]},
                    nombre_preferido=nombre_doc,
                )
                if ruta:
                    documentos_respuesta_guardados.append(ruta)
                    asegurar_espejo_web(os.path.basename(ruta))

        if not documentos_respuesta_guardados and self.respuesta_documento:
            ruta = _persistir_fuente(
                self.respuesta_documento,
                nombre_preferido=self.respuesta_documento_nombre,
            )
            if ruta:
                documentos_respuesta_guardados.append(ruta)
                asegurar_espejo_web(os.path.basename(ruta))

        if not documentos_respuesta_guardados and files:
            lista = files if isinstance(files, list) else [files]
            for item in lista:
                ruta = _persistir_fuente(item)
                if ruta:
                    documentos_respuesta_guardados.append(ruta)
                    asegurar_espejo_web(os.path.basename(ruta))

        if (
            not documentos_respuesta_guardados
            and self.respuesta_documento_content
            and self.respuesta_documento_nombre
        ):
            ruta = _persistir_fuente(
                {
                    "name": self.respuesta_documento_nombre,
                    "content": self.respuesta_documento_content,
                },
                nombre_preferido=self.respuesta_documento_nombre,
            )
            if ruta:
                documentos_respuesta_guardados.append(ruta)
                asegurar_espejo_web(os.path.basename(ruta))

        if not documentos_respuesta_guardados and (intento_adjunto or files):
            self.mensaje_actualizar_estado = (
                "No se pudo guardar el documento adjunto. Vuelve a subirlo e intenta de nuevo."
            )
            return
        nombres_adjuntos = [os.path.basename(p) for p in documentos_respuesta_guardados]
        estado_nuevo = self.nuevo_estado
        # Guardar el área para cualquier estado (no solo "Asignada a Area")
        area_asignada = str(self.area_para_asignar or "").strip()
        respuesta_enviada = self.respuesta_solicitud or ""
        datos_notificacion: dict[str, Any] | None = None
        try:
            with Session(engine) as session:
                solicitud_obj = session.get(Solicitud, self.editar_estado_id)
                if not solicitud_obj:
                    self.mensaje_actualizar_estado = "Solicitud no encontrada."
                    return
                estado_anterior = solicitud_obj.estado
                solicitud_obj.estado = estado_nuevo
                # Guardar el área seleccionada (si hay una) sin importar el estado
                if area_asignada:
                    solicitud_obj.area_responsable = area_asignada
                if respuesta_enviada:
                    solicitud_obj.respuesta = respuesta_enviada
                closed_states = {"cerrada", "resuelta", "finalizada", "respondida"}
                if str(estado_nuevo or "").strip().lower() in closed_states:
                    solicitud_obj.fecha_respuesta = datetime.now()
                if documentos_respuesta_guardados:
                    tags = ", ".join(nombres_adjuntos)
                    solicitud_obj.respuesta = (solicitud_obj.respuesta or "") + f"\n\n[DOCUMENTO ADJUNTO: {tags}]"
                session.add(solicitud_obj)
                # Combinar área + observaciones en el historial
                obs_historial = respuesta_enviada or None
                if area_asignada:
                    prefix = f"[ÁREA: {area_asignada}]"
                    obs_historial = prefix + ("\n" + obs_historial if obs_historial else "")
                historial = SolicitudEstadoHistorial(
                    solicitud_id=solicitud_obj.id,
                    estado_anterior=estado_anterior,
                    estado_nuevo=estado_nuevo,
                    fecha_cambio=datetime.now(),
                    observaciones=obs_historial,
                    documento_adjunto=", ".join(nombres_adjuntos) if nombres_adjuntos else None,
                )
                session.add(historial)
                session.commit()

                correo_dest = solicitud_obj.creado_por or ""
                nombre_sol = correo_dest
                if solicitud_obj.usuario_id:
                    usuario = session.get(Usuario, solicitud_obj.usuario_id)
                    if usuario:
                        correo_dest = usuario.email or correo_dest
                        nombre_sol = usuario.nombres or nombre_sol

                solicitud_id_actualizada = solicitud_obj.id
                datos_notificacion = {
                    "correo": correo_dest,
                    "nombre": nombre_sol,
                    "radicado": solicitud_obj.radicado,
                    "tipo": solicitud_obj.tipo_solicitud,
                    "estado_anterior": estado_anterior,
                    "estado_nuevo": estado_nuevo,
                    "respuesta": respuesta_enviada,
                    "area_asignada": area_asignada or (solicitud_obj.area_responsable or ""),
                }

            self.mensaje_actualizar_estado = ""
            self.editar_estado_id = 0
            self.nuevo_estado = ""
            self.area_para_asignar = ""
            self.respuesta_solicitud = ""
            self.respuesta_documentos = []
            self._limpiar_campos_respuesta_documento()
            self._refrescar_solicitud_en_lista(solicitud_id_actualizada)
            self.mostrar_toast(f"Estado actualizado a '{estado_nuevo}' correctamente.", "success")

            if datos_notificacion:
                import threading

                threading.Thread(
                    target=_enviar_notificacion_estado_solicitud,
                    args=(dict(datos_notificacion), list(documentos_respuesta_guardados)),
                    daemon=True,
                ).start()
        except Exception as e:
            self.mensaje_actualizar_estado = f"Error actualizando estado: {e}"

    def abrir_editor_estado(self, solicitud_id: int, estado_actual: str):
        """Abre el editor de estado para una solicitud."""
        self.editar_estado_id = solicitud_id
        self.nuevo_estado = estado_actual
        self.respuesta_solicitud = ""
        self.respuesta_documentos = []
        self._limpiar_campos_respuesta_documento()
        self.respuesta_documento_error = ""
        self.mensaje_actualizar_estado = ""
        
        # Cargar el área responsable actual para persistir en el selector
        with Session(engine) as session:
            sol = session.get(Solicitud, solicitud_id)
            if sol and sol.area_responsable:
                self.area_para_asignar = sol.area_responsable
            else:
                self.area_para_asignar = ""

    def abrir_asignar_area(self, solicitud_id: int, area_actual: str):
        """Abre el diálogo para asignar un área a una solicitud con mensaje."""
        self.asignar_area_id = solicitud_id
        self.asignar_area_nombre = area_actual
        self.asignar_area_seleccionada = area_actual or "Atención al Ciudadano"
        self.asignar_area_mensaje = ""
        self.mensaje_asignacion = ""

    def cerrar_asignar_area(self):
        """Cierra el diálogo de asignación de área."""
        self.asignar_area_id = 0
        self.asignar_area_nombre = ""
        self.asignar_area_seleccionada = ""
        self.asignar_area_mensaje = ""
        self.mensaje_asignacion = ""

    def asignar_area_con_mensaje(self):
        """Asigna un área a una solicitud y envía un mensaje al ciudadano."""
        self.mensaje_asignacion = ""
        
        area_a_asignar = self.asignar_area_seleccionada or self.asignar_area_nombre
        if not self.asignar_area_id or not area_a_asignar:
            self.mensaje_asignacion = "Selecciona un área válida."
            return
        
        if not self.asignar_area_mensaje:
            self.mensaje_asignacion = "Escribe un mensaje para el ciudadano."
            return
        
        try:
            with Session(engine) as session:
                solicitud_obj = session.get(Solicitud, self.asignar_area_id)
                if not solicitud_obj:
                    self.mensaje_asignacion = "Solicitud no encontrada."
                    return
                
                estado_anterior = solicitud_obj.estado
                estado_nuevo = "Asignada a Area" if estado_anterior == "Radicada" else estado_anterior
                
                solicitud_obj.estado = estado_nuevo
                solicitud_obj.area_responsable = area_a_asignar
                session.add(solicitud_obj)
                
                # Crear registro en la bitácora / historial
                prefix = f"[ÁREA: {area_a_asignar}]"
                obs_historial = prefix + ("\n" + self.asignar_area_mensaje if self.asignar_area_mensaje else "")
                
                historial = SolicitudEstadoHistorial(
                    solicitud_id=solicitud_obj.id,
                    estado_anterior=estado_anterior,
                    estado_nuevo=estado_nuevo,
                    fecha_cambio=datetime.now(),
                    observaciones=obs_historial,
                    documento_adjunto=None,
                )
                session.add(historial)
                session.commit()
            
            # Enviar notificación por correo al ciudadano
            solicitud_info = None
            for sol in self.solicitudes:
                if sol['id'] == self.asignar_area_id:
                    solicitud_info = sol
                    break
            
            if solicitud_info:
                area_a_asignar = self.asignar_area_seleccionada or self.asignar_area_nombre
                asunto_email = f"Tu solicitud PQRS ha sido asignada a {area_a_asignar}"
                cuerpo_email = f"""
Estimado ciudadano,

Tu solicitud PQRS con número de radicado {solicitud_info['radicado']} ha sido asignada a {area_a_asignar} para su tratamiento.

Detalles de la solicitud:
- Tipo: {solicitud_info['tipo_solicitud']}
- Asunto: {solicitud_info['asunto']}
- Área responsable: {area_a_asignar}

Mensaje del funcionario:
{self.asignar_area_mensaje}

Puedes consultar el estado completo de tu solicitud en nuestro portal web.

Atentamente,
Equipo de Atención al Ciudadano
Sistema PQRS
"""

                # Enviar correo al ciudadano
                enviar_correo_notificacion(solicitud_info['creado_por'], asunto_email, cuerpo_email)
            
            self.mensaje_asignacion = f"Área asignada a {area_a_asignar} y mensaje enviado correctamente."
            self.cargar_solicitudes()
            self.cerrar_asignar_area()
        except Exception as e:
            self.mensaje_asignacion = f"Error asignando área: {e}"

    def _solicitud_a_dict(self, solicitud: Solicitud) -> dict[str, Any]:
        respuesta_text = solicitud.respuesta or ""
        respuesta_documento_basename = None
        if respuesta_text:
            match = re.search(r"\[DOCUMENTO ADJUNTO:\s*([^\]\|]+)\]", respuesta_text)
            if match:
                respuesta_documento_basename = match.group(1).strip()
                respuesta_text = re.sub(r"\s*\[DOCUMENTO ADJUNTO:[^\]]+\]", "", respuesta_text).strip()

        documento_basenames = []
        if solicitud.documento_basename:
            try:
                parsed_names = json.loads(solicitud.documento_basename)
                if isinstance(parsed_names, list):
                    documento_basenames = parsed_names
                else:
                    documento_basenames = [parsed_names]
            except Exception:
                documento_basenames = [solicitud.documento_basename]

        documento_paths = []
        if solicitud.documento:
            try:
                parsed_paths = json.loads(solicitud.documento)
                if isinstance(parsed_paths, list):
                    documento_paths = parsed_paths
                else:
                    documento_paths = [parsed_paths]
            except Exception:
                documento_paths = [solicitud.documento]

        documento_adjuntos = []
        for idx, path in enumerate(documento_paths):
            basename = documento_basenames[idx] if idx < len(documento_basenames) else os.path.basename(str(path).replace("\\", "/"))
            documento_adjuntos.append(construir_metadata_adjunto(str(path), str(basename)))

        documento_basename = documento_adjuntos[0]["basename"] if documento_adjuntos else ""
        documento_href = documento_adjuntos[0]["href"] if documento_adjuntos else ""

        if not respuesta_documento_basename and solicitud.id:
            fallback_hist = _ultimo_adjunto_historial(solicitud.id)
            if fallback_hist:
                respuesta_documento_basename = fallback_hist

        respuesta_meta = (
            construir_metadata_adjunto(respuesta_documento_basename)
            if respuesta_documento_basename
            else {}
        )
        if respuesta_meta.get("basename"):
            respuesta_documento_basename = respuesta_meta["basename"]

        result: dict[str, Any] = {
            "id": solicitud.id,
            "radicado": solicitud.radicado,
            "tipo_solicitud": solicitud.tipo_solicitud,
            "persona_vulnerable": solicitud.persona_vulnerable,
            "asunto": solicitud.asunto,
            "descripcion": solicitud.descripcion,
            "ubicacion": solicitud.ubicacion,
            "area_responsable": solicitud.area_responsable,
            "documento": solicitud.documento,
            "documento_basename": documento_basename,
            "documento_href": documento_href,
            "documento_adjuntos": documento_adjuntos,
            "documento_adjuntos_json": json.dumps(documento_adjuntos),
            "estado": solicitud.estado,
            "respuesta": respuesta_text,
            "respuesta_documento_basename": respuesta_documento_basename or "",
            "respuesta_documento_href": respuesta_meta.get("href", ""),
            "respuesta_documento_es_imagen": respuesta_meta.get("es_imagen", False),
            "respuesta_documento_preview_src": respuesta_meta.get("preview_src", ""),
            "respuesta_documento_existe": respuesta_meta.get("existe", False),
            "fecha": solicitud.fecha.strftime("%Y-%m-%d %H:%M") if isinstance(solicitud.fecha, datetime) else str(solicitud.fecha),
            "creado_por": solicitud.creado_por,
            "usuario_id": solicitud.usuario_id,
            "calificacion_servicio": solicitud.calificacion_servicio,
            "fecha_consulta_ciudadano": solicitud.fecha_consulta_ciudadano.isoformat() if solicitud.fecha_consulta_ciudadano else None,
        }

        try:
            if solicitud.documento:
                parsed = json.loads(solicitud.documento)
                if isinstance(parsed, dict):
                    if "tiempo_respuesta_dias" in parsed:
                        result["tiempo_respuesta_dias"] = parsed.get("tiempo_respuesta_dias")
                    if "fecha_respuesta" in parsed:
                        result["fecha_respuesta"] = parsed.get("fecha_respuesta")
                    if "cumple_plazo" in parsed:
                        result["cumple_plazo"] = parsed.get("cumple_plazo")
        except Exception:
            pass

        return result
        
        

    @rx.var
    def solicitud_consultada_adjuntos(self) -> list[dict[str, str]]:
        docs = self.solicitud_consultada.get("documento_adjuntos", [])
        if not isinstance(docs, list):
            return []

        resultado = []
        for doc in docs:
            if isinstance(doc, dict):
                resultado.append({
                    "basename": str(doc.get("basename", "")),
                    "href": str(doc.get("href", "")),
                    "preview_src": str(doc.get("preview_src", "")),
                    "existe": bool(doc.get("existe", False)),
                })
        return resultado

    @rx.var
    def solicitud_consultada_respuesta_documento_href(self) -> str:
        return str(self.solicitud_consultada.get("respuesta_documento_href") or "")

    @rx.var
    def solicitud_consultada_respuesta_documento_basename(self) -> str:
        return str(self.solicitud_consultada.get("respuesta_documento_basename") or "")

    @rx.var
    def solicitud_consultada_respuesta_documento_preview_src(self) -> str:
        return str(self.solicitud_consultada.get("respuesta_documento_preview_src") or "")

    @rx.var
    def solicitud_consultada_tiene_adjunto_respuesta(self) -> bool:
        basename = self.solicitud_consultada.get("respuesta_documento_basename")
        return bool(basename and str(basename).strip())

    @rx.var
    def solicitud_consultada_adjunto_respuesta_existe(self) -> bool:
        return bool(self.solicitud_consultada.get("respuesta_documento_existe"))

    def _enriquecer_solicitud_dict(self, s: dict[str, Any]) -> dict[str, Any]:
        """Normaliza estado, fecha_respuesta y semáforo para una solicitud en memoria."""
        est = s.get("estado", "")
        area = s.get("area_responsable")
        if est not in ["Radicada", "Asignada a Area", "Solucionada", "Cerrada", "Reabierta", "En Gestion de Area"]:
            if area and str(area).strip() != "":
                s["estado"] = "Asignada a Area"
            else:
                s["estado"] = "Radicada"
        try:
            with engine.connect() as conn:
                row = conn.execute(
                    text("SELECT fecha_respuesta FROM solicitud WHERE id = :id"),
                    {"id": s.get("id")},
                ).fetchone()
                s["fecha_respuesta"] = str(row[0]) if row and row[0] else None
        except Exception:
            s["fecha_respuesta"] = None
        try:
            sem = State._compute_remaining_for_solicitud(s)
            s["semaforo_remaining"] = sem.get("remaining")
            s["semaforo_fill"] = sem.get("fill")
            s["semaforo_width"] = sem.get("width", 0)
            s["semaforo_expired"] = s["semaforo_remaining"] is not None and s["semaforo_remaining"] <= 0
        except Exception:
            s["semaforo_remaining"] = None
            s["semaforo_fill"] = "gray"
            s["semaforo_width"] = 0
            s["semaforo_expired"] = False
        return s

    def _refrescar_solicitud_en_lista(self, solicitud_id: int) -> None:
        """Actualiza solo la solicitud modificada sin recargar todo el dashboard."""
        try:
            with Session(engine) as session:
                obj = session.get(Solicitud, solicitud_id)
                if not obj:
                    return
                actualizado = self._enriquecer_solicitud_dict(self._solicitud_a_dict(obj))
            self.solicitudes = [
                actualizado if s.get("id") == solicitud_id else s
                for s in self.solicitudes
            ]
        except Exception as e:
            print(f"Error refrescando solicitud {solicitud_id}: {e}")

    def cargar_solicitudes(self):
        if not self.es_autenticada:
            self.solicitudes = []
            return
        try:
            with rx.session() as session:
                query = select(Solicitud).order_by(Solicitud.fecha.desc())
                if self.rol_usuario == "ciudadano":
                    # Usar email_actual; si está vacío (sesión antigua) usar correo_usuario como fallback
                    email_filtro = str(self.email_actual or self.correo_usuario or "").strip().lower()
                    if email_filtro:
                        query = query.where(Solicitud.creado_por == email_filtro)
                    else:
                        # Sin email identificado: mostrar lista vacía en lugar de TODAS las solicitudes
                        self.solicitudes = []
                        return
                solicitudes_obj = session.exec(query).all()
                self.solicitudes = [
                    self._enriquecer_solicitud_dict(self._solicitud_a_dict(s))
                    for s in solicitudes_obj
                ]
        except Exception as e:
            print(f"Error cargando solicitudes: {e}")
            self.solicitudes = []

    def hidratar_sesion_ciudadano(self):
        """Sincroniza nombre, email y rol desde BD. Si la sesión es fantasma, hace logout."""
        if not self.es_autenticada:
            self.cargar_solicitudes()
            return

        email = str(self.email_actual or self.correo_usuario or "").strip().lower()
        uid_str = str(self.id_usuario or "0").strip()

        # ── Sesión fantasma: autenticado pero sin ningún identificador ──
        if not email and uid_str in ("0", ""):
            pass  # Sesión fantasma — limpiar y redirigir
            self.es_autentica = "false"
            self.email_actual = ""
            self.correo_usuario = ""
            self.id_usuario = "0"
            self.rol_usuario = ""
            self.nombres = ""
            self.solicitudes = []
            return rx.redirect("/login")

        user = None
        try:
            with rx.session() as session:
                if email:
                    user = session.exec(select(Usuario).where(Usuario.email == email)).first()
                if not user and uid_str not in ("0", ""):
                    try:
                        user = session.exec(select(Usuario).where(Usuario.id == int(uid_str))).first()
                    except Exception:
                        pass
                if user:
                    self.email_actual = user.email
                    self.correo_usuario = user.email
                    self.id_usuario = str(user.id)
                    self.rol_usuario = user.rol or "ciudadano"
                    self.nombres = str(user.nombres or "").strip()

                else:
                    # Usuario no encontrado en BD — sesión inválida, forzar logout

                    self.es_autentica = "false"
                    self.email_actual = ""
                    self.correo_usuario = ""
                    self.id_usuario = "0"
                    self.rol_usuario = ""
                    self.nombres = ""
                    self.solicitudes = []
                    return rx.redirect("/login")
        except Exception as e:
            print(f"[HIDRATAR] Error al hidratar sesión: {e}")
        self.cargar_solicitudes()


    def cargar_datos_funcionario(self):
        self.cargar_solicitudes()
        self.cargar_usuarios()

    def export_reportes_excel(self):
        """Genera un Excel en memoria desde `solicitudes_filtradas` con gráfica incluida."""
        try:
            from datetime import datetime

            data = _sanitizar_datos_exportacion(self.solicitudes_filtradas or [])
            if not data:
                self.mostrar_toast("No hay datos para exportar.", "warning")
                self.mostrar_menu_descarga = False
                return

            try:
                excel_bytes = generar_excel_filtrado_con_grafica(
                    data,
                    self.filter_tipo_solicitud,
                )
                filename = _nombre_archivo_excel_reporte(self.filter_tipo_solicitud)
                download_id = str(uuid.uuid4())

                TEMP_DOWNLOADS[download_id] = {
                    "data": excel_bytes,
                    "filename": filename,
                    "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                }

                self.export_filename = filename
                self.mostrar_toast(
                    f"Excel con gráfica generado. {len(data)} registros ({self.filter_tipo_solicitud}).",
                    "success",
                )
                self.mostrar_menu_descarga = False
                self.export_href = f"/api/download/{download_id}"

            except (ImportError, ModuleNotFoundError):
                self.export_reportes_csv()

        except Exception as e:
            print(f"ERROR CRITICO en export_reportes_excel: {type(e).__name__}: {e}")
            import traceback
            traceback.print_exc()
            self.mostrar_toast(f"Error: {str(e)[:100]}", "error")
            self.mostrar_menu_descarga = False


    def descargar_reporte_excel(self):
        """Descarga el archivo Excel generado"""
        try:
            if not self.export_filename:
                self.mostrar_toast("No hay archivo para descargar.", "warning")
                return
            
            filepath = os.path.join(UPLOAD_DIR, self.export_filename)
            if not os.path.exists(filepath):
                self.mostrar_toast(f"Archivo no encontrado: {self.export_filename}", "error")
                self.export_filename = ""
                self.export_href = ""
                return
            
            # Leer el archivo y preparar para descarga
            with open(filepath, 'rb') as f:
                content = f.read()
            
            # Usar rx.download para forzar la descarga
            return rx.download(
                data=content,
                filename=self.export_filename,
                mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            print(f"ERROR en descargar_reporte_excel: {e}")
            import traceback
            traceback.print_exc()
            self.mostrar_toast(f"Error descargando: {str(e)}", "error")

    def _validar_registro_basico(self) -> None:
        self.error_de_registro = ""
        if not self.validacion_de_entradas():
            return
        required_fields = [
            ("nombres", "Nombre"),
            ("apellidos", "Apellido"),
            ("tipo_identificacion", "Tipo de identificación"),
            ("numero_identificacion", "Número de identificación"),
            ("telefono", "Teléfono"),
            ("direccion", "Dirección"),
            ("departamento", "Departamento"),
            ("ciudad", "Ciudad"),
        ]
        for field_name, label in required_fields:
            if not str(getattr(self, field_name, "")).strip():
                self.error_de_registro = f"Completa el campo obligatorio: {label}."
                return

        if len(re.sub(r"\D", "", str(self.telefono))) < 7:
            self.error_de_registro = "Ingresa un teléfono válido con al menos 7 dígitos."
            return

        if len(str(self.numero_identificacion).strip()) < 6:
            self.error_de_registro = "El número de identificación debe tener al menos 6 caracteres."
            return

        if not self.acepta_politica_datos or not self.acepta_notificaciones:
            self.error_de_registro = "Debes aceptar la política de datos y recibir notificaciones para registrarte."
            return

    def _crear_usuario(self, rol: str, exito_mensaje: str):
        with rx.session() as session:
            correo_normalizado = str(self.correo or "").strip().lower()
            existing_user = session.exec(select(Usuario).where(Usuario.email == correo_normalizado)).first()
            if existing_user:
                self.error_de_registro = "El correo ya está registrado."
                return
            hashed = tiene_password(self.contraseña)
            nuevo_usuario = Usuario(
                email=correo_normalizado,
                Contraseña=hashed,
                rol=rol,
                is_active=True,
                Fecha_de_creacion=datetime.now(),
                tipo_identificacion=str(self.tipo_identificacion or "").strip(),
                numero_identificacion=str(self.numero_identificacion or "").strip(),
                nombres=normalizar_texto(self.nombres),
                apellidos=normalizar_texto(self.apellidos),
                genero=normalizar_texto(self.sexo),
                direccion=str(self.direccion or "").strip(),
                telefono=str(self.telefono or "").strip(),
                departamento=str(self.departamento or "").strip(),
                ciudad=str(self.ciudad or "").strip(),
                etnia=normalizar_texto(self.etnia),
                persona_vulnerable=normalizar_texto(self.persona_vulnerable_registro),
                acepta_notificaciones=bool(self.acepta_notificaciones),
                acepta_politica_datos=bool(self.acepta_politica_datos),
            )
            session.add(nuevo_usuario)
            session.commit()
        print(f"Usuario registrado: {self.correo}")
        # Dispara el correo de bienvenida inmediatamente después del registro.
        enviar_correo_bienvenida(self.correo, self.correo)
        self.succes = exito_mensaje
        self.mostrar_toast("¡Registro exitoso! Bienvenido al sistema.", "success")
        # Obtener el id del usuario recién creado para completar la sesión
        correo_registrado = str(self.correo or "").strip().lower()
        try:
            with rx.session() as session:
                nuevo = session.exec(select(Usuario).where(Usuario.email == correo_registrado)).first()
                if nuevo:
                    self.id_usuario = str(nuevo.id)
                    self.nombres = getattr(nuevo, "nombres", "") or ""
        except Exception as e:
            print(f"No se pudo cargar el id del nuevo usuario: {e}")
        self.limpiar_formulario_registro()
        # Login automático después del registro — asignar TODAS las cookies de sesión
        # igual que en login() para que el filtro de solicitudes funcione correctamente
        self.es_autentica = "true"
        self.email_actual = correo_registrado   # ← clave: el filtro WHERE usa esta cookie
        self.correo_usuario = correo_registrado
        self.rol_usuario = rol
        self.cargar_solicitudes()   # ahora filtra solo las del ciudadano (ninguna por ser nuevo)
        self.cargar_usuarios()

    def signup(self):
        self.borrar_mensajes_de_estado()
        self._validar_registro_basico()
        if self.error_de_registro:
            return
        self._crear_usuario(
            rol="ciudadano",
            exito_mensaje="Registro exitoso. Revisa tu correo para confirmar. Ahora el funcionario puede iniciar sesión.",
        )
        # Redirigir al dashboard del ciudadano después del registro
        if not self.error_de_registro:
            return rx.redirect("/dashboard")

    def signup_funcionario(self):
        self.borrar_mensajes_de_estado()
        if not self.es_autenticada or self.rol_usuario != "funcionario":
            self.error_de_registro = "Solo los funcionarios autenticados pueden registrar nuevos funcionarios."
            return
        self._validar_registro_basico()
        if self.error_de_registro:
            return
        return self._crear_usuario(
            rol="funcionario",
            exito_mensaje="Funcionario registrado con éxito. Ahora puede iniciar sesión con su correo institucional.",
        )

    def login(self, form_data: dict | None = None):
        self.borrar_mensajes_de_estado()
        if form_data:
            self.correo = str(form_data.get("correo") or self.correo or "").strip().lower()
            self.contraseña = str(form_data.get("contraseña") or self.contraseña or "")
        else:
            self.correo = str(self.correo or "").strip().lower()
        if not self.correo or not self.correo.strip():
            self.succes2 = ""
            self.error_de_contraseña = "El correo electrónico es obligatorio."
            return
            
        if not self.contraseña or not self.contraseña.strip():
            self.succes2 = ""
            self.error_de_contraseña = "La contraseña es obligatoria."
            return
        with Session(engine) as session:
            user = session.exec(select(Usuario).where(Usuario.email == self.correo)).first()
            pw_ok = False
            try:
                pw_ok = confirmar_contraseña(self.contraseña, user.Contraseña) if user else False
            except Exception:
                pw_ok = False
            if not user or not pw_ok:
                self.error_de_contraseña = "Correo o contraseña incorrectos."
                color="red"
                self.succes2 = ""
                return
            if not user.is_active:
                self.error_de_contraseña = "La cuenta no está activa."
                color="red"
                self.succes2 = ""
                return
            self.id_usuario = str(user.id)
            self.rol_usuario = user.rol
            self.email_actual = user.email
            self.nombres = getattr(user, 'nombres', 'Ciudadano')  # Extraer nombre si existe en DB
            self.es_autentica = "true"
            self.cargar_solicitudes()
            self.cargar_usuarios()
            self.error_de_contraseña = ""
            self.contraseña = ""
            self.confirmar_contraseña = ""
            self.show_password = False
            self.mostrar_toast("¡Inicio de sesión exitoso! Redirigiendo automáticamente...", "success")
            # Redirigir después de mostrar el toast
            if self.rol_usuario == "funcionario":
                return rx.redirect("/dashboard-funcionario")
            else:
                return rx.redirect("/dashboard")
        

    def redirect_after_login(self):
        if self.rol_usuario == "funcionario":
            return rx.redirect("/dashboard-funcionario")
        return rx.redirect("/dashboard")

    def abrir_logout_confirm(self):
        """Muestra el modal de confirmación de cierre de sesión."""
        self.show_logout_confirm = True

    def cerrar_logout_confirm(self):
        """Cierra el modal de confirmación sin hacer logout."""
        self.show_logout_confirm = False

    def logout(self):
        "cerrar sesion de usuario"
        self.show_logout_confirm = False
        self.id_usuario = "0"
        self.correo = ""
        self.contraseña = ""
        self.confirmar_contraseña = ""
        self.rol_usuario = ""
        self.email_actual = ""
        self.es_autentica = "false"
        self.nombres = ""
        self.show_password = False
        self.succes2 = "Has cerrado sesión exitosamente."
        self.error_de_contraseña = ""
        return rx.redirect("/")

    def change_password(self):
        """Cambiar la contraseña usuario autenticado"""
        self.change_pw_message = ""
        if not self.es_autenticada or not self.id_usuario_num:
            self.change_pw_message = "Debes iniciar sesión para cambiar la contraseña."
            return
        # Validaciones básicas
        if not self.current_password or not self.new_password or not self.confirm_new_password:
            self.change_pw_message = "Completa todos los campos."
            return
        if self.new_password != self.confirm_new_password:
            self.change_pw_message = "La nueva contraseña y su confirmación no coinciden."
            return
        if not cantida_minima_contraseña(self.new_password):
            self.change_pw_message = "La nueva contraseña no cumple los requisitos de seguridad."
            return
        with Session(engine) as session:
            user = session.exec(select(Usuario).where(Usuario.id == self.id_usuario_num)).first()
            if not user:
                self.change_pw_message = "Usuario no encontrado."
                return
            try:
                if not confirmar_contraseña(self.current_password, user.Contraseña):
                    self.change_pw_message = "La contraseña actual es incorrecta."
                    return
            except Exception as e:
                self.change_pw_message = f"Error comprobando contraseña: {e}"
                return
            # Actualizar contraseña
            user.Contraseña = tiene_password(self.new_password)
            session.add(user)
            session.commit()
            self.change_pw_message = "Contraseña cambiada correctamente."
            # Limpiar campos
            self.current_password = ""
            self.new_password = ""
            self.confirm_new_password = ""

    def toggle_show_password(self):
        self.show_password = not self.show_password

    def limpiar_formulario_solicitud(self, keep_message: bool = False):
        self.tipo_solicitud = ""
        self.persona_vulnerable = ""
        self.asunto = ""
        self.descripcion = ""
        self.ubicacion = ""
        self.documento = ""
        self.documentos = []
        self.documento_nombres = []
        self.documento_nombre = ""
        self.documento_previews = []
        self.area_responsable = ""
        self.area_otro = ""
        self.descripcion_len = 0
        self.editar_solicitud_id = 0
        self.pqrs_contacto_email = ""
        self.acepta_politica_solicitud = False
        if not keep_message:
            self.solicitud_mensaje = ""

    def crear_solicitud(self):
        self.solicitud_mensaje = ""
        if not self.tipo_solicitud or not self.asunto or not self.descripcion:
            self.solicitud_mensaje = "Completa los campos obligatorios antes de enviar."
            return

        # Validación de correo para usuarios no registrados
        if not self.es_autenticada:
            if not self.pqrs_contacto_email:
                self.solicitud_mensaje = "Por favor ingresa un correo electrónico para recibir respuesta."
                return
            if not validar_correo(self.pqrs_contacto_email):
                self.solicitud_mensaje = "El correo electrónico ingresado no es válido."
                return
        if not self.area_responsable:
            self.solicitud_mensaje = "Selecciona el área responsable."
            return
        if self.area_responsable == "Otros" and not self.area_otro:
            self.solicitud_mensaje = "Por favor indica el área responsable cuando eliges Otros."
            return
        # Verificar aceptación de política de tratamiento de datos
        if not self.acepta_politica_solicitud:
            self.solicitud_mensaje = "Debes aceptar la Política de Tratamiento de Datos Personales antes de enviar."
            return

        documentos_guardados: list[str] = []
        documento_basenames_guardados: list[str] = []

        def guardar_archivo(item: Any) -> None:
            nombre_pref = ""
            if isinstance(item, dict):
                nombre_pref = str(item.get("name") or item.get("filename") or "")
            ruta = persistir_archivo_en_uploads(
                item,
                nombre_preferido=nombre_pref,
                prefijo="solicitud",
            )
            if not ruta:
                etiqueta = nombre_pref or "adjunto"
                raise ValueError(f"No se pudo guardar el archivo '{etiqueta}' en el servidor.")
            basename = os.path.basename(ruta)
            documentos_guardados.append(basename)
            documento_basenames_guardados.append(basename)

        if self.documentos:
            try:
                os.makedirs(UPLOAD_DIR, exist_ok=True)
                os.makedirs(WEB_UPLOAD_DIR, exist_ok=True)
                for item in self.documentos:
                    guardar_archivo(item)
            except Exception as e:
                self.solicitud_mensaje = f"Error guardando documento: {e}"
                return
        elif self.documento:
            try:
                os.makedirs(UPLOAD_DIR, exist_ok=True)
                os.makedirs(WEB_UPLOAD_DIR, exist_ok=True)
                guardar_archivo(self.documento)
            except Exception as e:

                self.solicitud_mensaje = f"Error guardando documento: {e}"
                return

        if self.editar_solicitud_id:
            try:
                with Session(engine) as session:
                    solicitud_obj = session.get(Solicitud, self.editar_solicitud_id)
                    if not solicitud_obj:
                        self.solicitud_mensaje = "Solicitud no encontrada para editar."
                        return
                    # Obtener persona_vulnerable del usuario autenticado
                    usuario = session.get(Usuario, self.id_usuario_num)
                    persona_vulnerable_valor = usuario.persona_vulnerable if usuario else None
                    
                    solicitud_obj.tipo_solicitud = self.tipo_solicitud
                    solicitud_obj.persona_vulnerable = persona_vulnerable_valor or None
                    solicitud_obj.asunto = self.asunto
                    solicitud_obj.descripcion = self.descripcion
                    solicitud_obj.ubicacion = self.ubicacion or None
                    solicitud_obj.area_responsable = self.area_otro if self.area_responsable == "Otros" else self.area_responsable
                    if documentos_guardados:
                        solicitud_obj.documento = json.dumps(documentos_guardados)
                        solicitud_obj.documento_basename = json.dumps(documento_basenames_guardados)
                    solicitud_obj.estado = "Actualizada"
                    session.add(solicitud_obj)
                    session.commit()
                # Enviar notificación de cambio de estado (si hay correo disponible)
                try:
                    correo_dest = self.email_actual or self.correo
                    if correo_dest and isinstance(correo_dest, str) and "@" in correo_dest:
                        fecha_cambio = datetime.now().strftime("%d/%m/%Y %H:%M")
                        notificar_cambio_estado(
                            nombre_solicitante=getattr(self, 'nombres', correo_dest) or correo_dest,
                            correo_solicitante=correo_dest,
                            numero_solicitud=solicitud_obj.radicado,
                            estado_anterior="(anterior)",
                            estado_nuevo=solicitud_obj.estado,
                            fecha_cambio=fecha_cambio,
                            observaciones="Actualizada desde interfaz",
                            correos_adicionales=None
                        )
                except Exception:
                    pass

                self.solicitud_mensaje = "Solicitud actualizada con éxito."
                self.editar_solicitud_id = 0
                self.limpiar_formulario_solicitud(keep_message=True)
                self.cargar_solicitudes()
                return
            except Exception as e:
                self.solicitud_mensaje = f"Error actualizando solicitud: {e}"
                return

        try:
            with rx.session() as session:
                # Obtener persona_vulnerable del usuario autenticado
                usuario = session.get(Usuario, self.id_usuario_num)
                persona_vulnerable_valor = usuario.persona_vulnerable if usuario else None
                
                solicitud_obj = Solicitud(
                    radicado=f"PQRS-{datetime.now().year}-{uuid.uuid4().hex[:8]}".upper(),
                    tipo_solicitud=self.tipo_solicitud,
                    persona_vulnerable=persona_vulnerable_valor or None,
                    asunto=self.asunto,
                    descripcion=self.descripcion,
                    ubicacion=self.ubicacion or None,
                    area_responsable=self.area_otro if self.area_responsable == "Otros" else self.area_responsable,
                    documento=json.dumps(documentos_guardados) if documentos_guardados else None,
                    documento_basename=json.dumps(documento_basenames_guardados) if documento_basenames_guardados else None,
                    estado="Radicada",
                    fecha=datetime.now(),
                    creado_por=self.email_actual if self.es_autenticada else self.pqrs_contacto_email,
                    usuario_id=self.id_usuario_num if self.id_usuario_num else None,
                )
                session.add(solicitud_obj)
                session.commit()
                radicado_generado = solicitud_obj.radicado
            # Enviar notificación de creación (si hay correo disponible)
            try:
                correo_dest = self.email_actual if self.es_autenticada else self.pqrs_contacto_email
                if correo_dest and isinstance(correo_dest, str) and "@" in correo_dest:
                    fecha_creacion = solicitud_obj.fecha.strftime("%d/%m/%Y %H:%M")
                    fecha_vencimiento = (solicitud_obj.fecha + timedelta(days=15)).strftime("%d/%m/%Y")
                    notificar_solicitud_creada(
                        nombre_solicitante=getattr(self, 'nombres', correo_dest) or correo_dest,
                        correo_solicitante=correo_dest,
                        numero_solicitud=solicitud_obj.radicado,
                        tipo_pqrs=solicitud_obj.tipo_solicitud,
                        fecha_creacion=fecha_creacion,
                        fecha_vencimiento=fecha_vencimiento,
                        correos_adicionales=None
                    )
            except Exception:
                pass

            self.solicitud_mensaje = f"�?? Solicitud enviada con éxito. Radicado: {radicado_generado}"
            self.limpiar_formulario_solicitud(keep_message=True)
            self.cargar_solicitudes()
        except Exception as e:
            self.solicitud_mensaje = f"Error guardando solicitud: {e}"

    def editar_solicitud(self, solicitud_id: int):
        try:
            with Session(engine) as session:
                solicitud_obj = session.get(Solicitud, solicitud_id)
                if solicitud_obj:
                    self.editar_solicitud_id = solicitud_id
                    self.tipo_solicitud = solicitud_obj.tipo_solicitud
                    self.asunto = solicitud_obj.asunto
                    self.descripcion = solicitud_obj.descripcion
                    self.ubicacion = solicitud_obj.ubicacion or ""
                    self.area_responsable = solicitud_obj.area_responsable or ""
                    self.area_otro = solicitud_obj.area_responsable if solicitud_obj.area_responsable and solicitud_obj.area_responsable not in ["Secretaría", "Contabilidad", "Bienestar", "Tesorería", "Atención al Ciudadano"] else ""
                    self.persona_vulnerable = solicitud_obj.persona_vulnerable or ""
                    self.documento = solicitud_obj.documento or ""
                    self.solicitud_mensaje = "Editando solicitud. Actualiza los campos y guarda cambios."
                else:
                    self.solicitud_mensaje = "Solicitud no encontrada."
        except Exception as e:
            self.solicitud_mensaje = f"Error cargando solicitud: {e}"

    mostrar_rastreador: bool = False

    def toggle_rastreador(self):
        """Muestra u oculta el campo de rastreo en el dashboard."""
        self.mostrar_rastreador = not self.mostrar_rastreador

    def consultar_desde_dashboard(self):
        """Ejecuta la consulta y redirige a la página de consulta para ver los resultados."""
        self.consultar_estado_solicitud()
        return rx.redirect("/consultar-estado")

    def cerrar_consulta(self):
        """Limpia los campos de consulta y regresa al inicio."""
        self.consulta_radicado = ""
        self.solicitud_consultada = {}
        self.historial_solicitud = []
        if self.es_autentica:
            return rx.redirect("/dashboard")
        return rx.redirect("/")

    def consultar_estado_solicitud(self):
        """Consulta el estado de una solicitud por número de radicado"""
        self.consulta_mensaje = ""
        self.solicitud_consultada = {}
        
        if not self.consulta_radicado:
            self.consulta_mensaje = "Ingresa un número de radicado válido."
            return
        
        try:
            with Session(engine) as session:
                solicitud = session.exec(
                    select(Solicitud).where(Solicitud.radicado == self.consulta_radicado)
                ).first()
                
                if not solicitud:
                    self.consulta_mensaje = "No se encontró una solicitud con ese número de radicado."
                    return
                
                # HU11: registrar fecha_consulta_ciudadano si la solicitud es Solucionada y aún no se registró
                if str(solicitud.estado or "").strip().lower() == "solucionada" and not solicitud.fecha_consulta_ciudadano:
                    solicitud.fecha_consulta_ciudadano = datetime.now()
                    session.add(solicitud)
                    session.commit()
                    session.refresh(solicitud)

                self.solicitud_consultada = self._solicitud_a_dict(solicitud)
                # HU11: cargar historial de cambios
                self.cargar_historial_solicitud(solicitud.id, session)
                self.consulta_mensaje = "Solicitud encontrada."
                
        except Exception as e:
            self.consulta_mensaje = f"Error consultando solicitud: {e}"

    def cargar_historial_solicitud(self, solicitud_id: int, session=None):
        """Carga el historial de cambios de estado de una solicitud. HU11."""
        try:
            def _query(s):
                items = s.exec(
                    select(SolicitudEstadoHistorial)
                    .where(SolicitudEstadoHistorial.solicitud_id == solicitud_id)
                    .order_by(SolicitudEstadoHistorial.fecha_cambio.desc())
                ).all()
                def _parse_historial(h):
                    obs_raw = h.observaciones or ""
                    area_h = ""
                    obs_clean = obs_raw
                    if obs_raw.startswith("[ÁREA:"):
                        end = obs_raw.find("]")
                        if end != -1:
                            area_h = obs_raw[7:end].strip()
                            obs_clean = obs_raw[end + 1:].lstrip("\n").strip()
                    
                    doc = h.documento_adjunto or ""
                    if not doc:
                        match = re.search(r"\[DOCUMENTO(?: ADJUNTO)?:\s*([^\]\|]+)\]", obs_clean)
                        if match:
                            doc = match.group(1).strip()
                            obs_clean = re.sub(r"\s*\[DOCUMENTO(?: ADJUNTO)?:[^\]]+\]", "", obs_clean).strip()

                    doc_meta = construir_metadata_adjunto(doc) if doc else {}
                    return {
                        "estado_anterior": h.estado_anterior,
                        "estado_nuevo": h.estado_nuevo,
                        "fecha_cambio": h.fecha_cambio.strftime("%d/%m/%Y %H:%M") if isinstance(h.fecha_cambio, datetime) else str(h.fecha_cambio),
                        "observaciones": obs_clean,
                        "area_asignada": area_h,
                        "documento_adjunto": doc_meta.get("basename", doc),
                        "documento_href": doc_meta.get("href", ""),
                        "documento_preview_src": doc_meta.get("preview_src", ""),
                        "documento_existe": doc_meta.get("existe", False),
                    }
                self.historial_solicitud = [_parse_historial(h) for h in items]
            if session:
                _query(session)
            else:
                with Session(engine) as s:
                    _query(s)
        except Exception as e:
            print(f"Error cargando historial: {e}")
            self.historial_solicitud = []

    def enviar_calificacion(self):
        self.calificar_solicitud(self.calificacion_opcion)

    def calificar_solicitud(self, estrellas_str: str):
        """Registra la calificación del ciudadano para una solicitud solucionada. HU11."""
        self.calificacion_mensaje = ""
        try:
            if not estrellas_str:
                self.calificacion_mensaje = "Selecciona una calificación."
                return
            estrellas = int(str(estrellas_str).split(" - ")[0])
        except Exception:
            self.calificacion_mensaje = "Calificación inválida."
            return
        radicado = self.solicitud_consultada.get("radicado", "")
        if not radicado:
            self.calificacion_mensaje = "Primero consulta una solicitud."
            return
        if estrellas < 1 or estrellas > 5:
            self.calificacion_mensaje = "Calificación inválida."
            return
        try:
            with Session(engine) as session:
                sol = session.exec(
                    select(Solicitud).where(Solicitud.radicado == radicado)
                ).first()
                if not sol:
                    self.calificacion_mensaje = "Solicitud no encontrada."
                    return
                if sol.calificacion_servicio is not None:
                    self.calificacion_mensaje = "Ya has calificado este servicio."
                    return
                sol.calificacion_servicio = estrellas
                sol.estado = "Cerrada"
                # Guardar el registro en historial
                observacion_texto = f"Ciudadano calificó el servicio con {estrellas} estrella(s)."
                if self.comentario_calificacion.strip():
                    observacion_texto += f" Comentario: {self.comentario_calificacion.strip()}"
                
                h = SolicitudEstadoHistorial(
                    solicitud_id=sol.id,
                    estado_anterior="Solucionada",
                    estado_nuevo="Cerrada",
                    fecha_cambio=datetime.now(),
                    observaciones=observacion_texto,
                )
                session.add(sol)
                session.add(h)
                session.commit()
            self.calificacion_seleccionada = estrellas
            self.solicitud_consultada = {**self.solicitud_consultada, "calificacion_servicio": estrellas, "estado": "Cerrada"}
            self.cargar_historial_solicitud(self.solicitud_consultada.get("id", 0))
            self.calificacion_mensaje = f"¡Gracias! Has calificado el servicio con {estrellas} estrella(s)."
        except Exception as e:
            self.calificacion_mensaje = f"Error al calificar: {e}"

    def auto_cerrar_solucionadas(self):
        """Cierra automáticamente con calificación 5 las PQRS 'Solucionadas' con >5 días hábiles sin evaluar. HU11."""
        from datetime import date, timedelta
        try:
            with Session(engine) as session:
                solucionadas = session.exec(
                    select(Solicitud)
                    .where(Solicitud.estado == "Solucionada")
                    .where(Solicitud.calificacion_servicio == None)
                    .where(Solicitud.fecha_consulta_ciudadano != None)
                ).all()
                hoy = date.today()
                for sol in solucionadas:
                    fecha_consulta = sol.fecha_consulta_ciudadano.date() if sol.fecha_consulta_ciudadano else None
                    if not fecha_consulta:
                        continue
                    # Calcular días hábiles desde consulta
                    dias_habiles = 0
                    cur = fecha_consulta + timedelta(days=1)
                    while cur <= hoy:
                        if cur.weekday() < 5:  # Lunes a viernes
                            dias_habiles += 1
                        cur += timedelta(days=1)
                    if dias_habiles >= 5:
                        sol.calificacion_servicio = 5
                        sol.estado = "Cerrada"
                        h = SolicitudEstadoHistorial(
                            solicitud_id=sol.id,
                            estado_anterior="Solucionada",
                            estado_nuevo="Cerrada",
                            fecha_cambio=datetime.now(),
                            observaciones="Cierre automático: ciudadano no evaluó en 5 días hábiles. Calificación asignada: 5.",
                        )
                        session.add(sol)
                        session.add(h)
                session.commit()
        except Exception as e:
            print(f"Error en auto_cerrar_solucionadas: {e}")


    def cargar_usuarios(self):
        """Carga la lista de usuarios registrados en el sistema"""
        try:
            with rx.session() as session:
                usuarios = session.exec(select(Usuario)).all()
                self.usuarios_registrados = [
                    {
                        "id": u.id,
                        "email": u.email,
                        "nombres": u.nombres or "",
                        "apellidos": u.apellidos or "",
                        "rol": u.rol,
                        "fecha_creacion": u.Fecha_de_creacion.strftime("%Y-%m-%d") if isinstance(u.Fecha_de_creacion, datetime) else str(u.Fecha_de_creacion),
                        "is_active": "Activo" if u.is_active else "Inactivo",
                    }
                    for u in usuarios
                ]
        except Exception as e:
            print(f"Error cargando usuarios: {e}")
            self.usuarios_registrados = []

    def cambiar_rol_ciudadano_a_funcionario(self):
        """Cambia el rol de un ciudadano a funcionario"""
        self.cambiar_rol_mensaje = ""
        
        if not self.cambiar_rol_email:
            self.cambiar_rol_mensaje = "Ingresa el correo del usuario."
            return
        if not self.confirmar_promocion_rol:
            self.cambiar_rol_mensaje = "Debes confirmar la validación del usuario antes de promover el rol."
            return
        
        try:
            with rx.session() as session:
                usuario = session.exec(
                    select(Usuario).where(Usuario.email == self.cambiar_rol_email)
                ).first()
                
                if not usuario:
                    self.cambiar_rol_mensaje = f"No se encontró usuario con el correo {self.cambiar_rol_email}."
                    return
                
                if usuario.rol == "funcionario":
                    self.cambiar_rol_mensaje = f"El usuario ya es funcionario."
                    return
                
                usuario.rol = "funcionario"
                session.add(usuario)
                session.commit()
                
                # Enviar notificación
                try:
                    asunto = "Rol actualizado - Has sido promovido a Funcionario"
                    cuerpo = """
Estimado usuario,

Te informamos que tu rol en el sistema ha sido actualizado.

Tu nuevo rol: FUNCIONARIO

Con este rol podrás:
- Gestionar solicitudes PQRS
- Asignar áreas responsables
- Actualizar estados de solicitudes
- Ver reportes del sistema

Accede al Dashboard Funcionario con tu correo y contraseña.

Atentamente,
Sistema PQRS
"""
                    enviar_correo_notificacion(self.cambiar_rol_email, asunto, cuerpo)
                except:
                    pass  # No fallar si no se envía el correo
                
                self.cambiar_rol_mensaje = f"�?? Rol del usuario {self.cambiar_rol_email} actualizado a funcionario."
                self.cambiar_rol_email = ""
                self.confirmar_promocion_rol = False
                self.cargar_usuarios()
        except Exception as e:
            self.cambiar_rol_mensaje = f"Error al cambiar rol: {e}"

    def degradar_funcionario_a_ciudadano(self, email: str):
        """Cambia el rol de un funcionario a ciudadano (solo funcionarios autenticados)."""
        self.cambiar_rol_mensaje = ""
        if not self.es_autenticada or self.rol_usuario != "funcionario":
            self.cambiar_rol_mensaje = "Solo funcionarios autenticados pueden cambiar roles."
            return
        email_norm = str(email or "").strip().lower()
        if not email_norm:
            self.cambiar_rol_mensaje = "Correo inválido."
            return
        if self.email_actual and str(self.email_actual).strip().lower() == email_norm:
            self.cambiar_rol_mensaje = "No puedes degradarte a ti mismo."
            return

        try:
            with rx.session() as session:
                usuario = session.exec(select(Usuario).where(Usuario.email == email_norm)).first()
                if not usuario:
                    self.cambiar_rol_mensaje = f"No se encontró usuario con el correo {email_norm}."
                    return
                if usuario.rol != "funcionario":
                    self.cambiar_rol_mensaje = f"El usuario {email_norm} no es funcionario."
                    return

                usuario.rol = "ciudadano"
                session.add(usuario)
                session.commit()

            # Notificación (no bloquear si falla)
            try:
                asunto = "Rol actualizado - Has sido asignado como Ciudadano"
                cuerpo = f"""\
Estimado usuario,

Te informamos que tu rol en el sistema ha sido actualizado.

Tu nuevo rol: CIUDADANO

Atentamente,
Sistema PQRS
"""
                enviar_correo_notificacion(email_norm, asunto, cuerpo)
            except Exception:
                pass

            self.cambiar_rol_mensaje = f"�?? Rol del usuario {email_norm} actualizado a ciudadano."
            self.cargar_usuarios()
        except Exception as e:
            self.cambiar_rol_mensaje = f"Error al cambiar rol: {e}"

    def set_cambiar_rol_email(self, value: str):
        self.cambiar_rol_email = str(value or "").strip().lower()
        self.cambiar_rol_mensaje = ""
        self.confirmar_promocion_rol = False

    def set_confirmar_promocion_rol(self, value: bool):
        self.confirmar_promocion_rol = bool(value)

    def limpiar_form_promocion_rol(self):
        self.cambiar_rol_email = ""
        self.confirmar_promocion_rol = False
        self.cambiar_rol_mensaje = ""


    def set_confirmar_contraseña(self, value: str):
        self.confirmar_contraseña = value

    def set_correo(self, value: str):
        self.correo = value

    def set_contraseña(self, value: str):
        self.contraseña = value

    def set_tipo_identificacion(self, value: str):
        self.tipo_identificacion = value

    def set_genero(self, value: str):
        self.genero = value

    def set_sexo(self, value: str):
        self.sexo = value

    def set_direccion(self, value: str):
        self.direccion = value

    def set_current_password(self, value: str):
        self.current_password = value


def label_requerido(texto: str) -> rx.Component:
    return rx.hstack(
        rx.text(texto, color=rx.color_mode_cond(light="black", dark="white")),
        rx.text("*", color="orange.500"),
        spacing="1",
        align_items="center",
    )

def password_strength_label(password: str) -> rx.Component:
    # This will be rendered conditionally in the UI
    return rx.cond(
        password != "",
        rx.text(
            "Fortaleza de la contraseña: ",
            rx.cond(
                (password.length() >= 8) & (password != ""),
                "Media",
                "Débil"
            ),
            color=rx.cond(
                (password.length() >= 8) & (password != ""),
                "yellow.400",
                "red.400"
            ),
            font_size="xs"
        ),
        rx.box()
    )


def auth_card(title: str, on_submit, show_confirm: bool = False) -> rx.Component:
    text_color = rx.color_mode_cond(light="#1e293b", dark="#f8fafc")
    input_bg = rx.color_mode_cond(light="rgba(255, 255, 255, 0.8)", dark="rgba(30, 41, 59, 0.6)")
    input_border = rx.color_mode_cond(light="rgba(226, 232, 240, 0.8)", dark="rgba(51, 65, 85, 0.8)")
    placeholder_color = rx.color_mode_cond(light="#94a3b8", dark="#64748b")

    input_style = {
        "bg": input_bg,
        "border": f"1px solid {input_border}",
        "color": text_color,
        "size": "2",
        "radius": "large",
        "_placeholder": {"color": placeholder_color},
    }

    confirmar_field = (
        rx.vstack(
            label_requerido("Confirmar Contraseña"),
            rx.input(
                placeholder="Confirmar Contraseña",
                type=rx.cond(State.show_password, "text", "password"),
                value=State.confirmar_contraseña,
                on_change=State.set_confirmar_contraseña,
                width="100%",
                **input_style,
            ),
        )
        if show_confirm
        else rx.box(display="none")
    )

    return rx.box(
        rx.form(
            rx.vstack(
                rx.hstack(
                    rx.heading("CO.", size="6", color=rx.color_mode_cond(light="#1e3a8a", dark="#93c5fd"), font_weight="black"),
                  
                    justify="between",
                    width="100%",
                ),
                rx.heading(title, size="7", color=text_color, font_weight="bold", margin_bottom="0.4em"),
                rx.text(
                    "Completa tus datos para activar tu cuenta y gestionar tus solicitudes.",
                    color=rx.color_mode_cond(light="#475569", dark="#94a3b8"),
                    font_size="sm",
                    width="100%",
                ),
                
                rx.hstack(
                    rx.button("1. Cuenta", type="button", size="2", variant=rx.cond(State.registro_seccion_abierta == "cuenta", "solid", "soft"), color_scheme="blue", on_click=State.set_registro_seccion_abierta("cuenta")),
                    rx.button("2. Identidad", type="button", size="2", variant=rx.cond(State.registro_seccion_abierta == "identidad", "solid", "soft"), color_scheme="blue", on_click=State.set_registro_seccion_abierta("identidad"), is_disabled=State.registro_paso_habilitado < 2),
                    rx.button("3. Ubicación", type="button", size="2", variant=rx.cond(State.registro_seccion_abierta == "ubicacion", "solid", "soft"), color_scheme="blue", on_click=State.set_registro_seccion_abierta("ubicacion"), is_disabled=State.registro_paso_habilitado < 3),
                    width="100%",
                    spacing="2",
                    flex_wrap="wrap",
                ),
                rx.vstack(
                    rx.hstack(
                        rx.text("Progreso del registro", font_size="xs", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8")),
                        rx.spacer(),
                        rx.text(
                            rx.cond(
                                State.registro_seccion_abierta == "cuenta",
                                "33%",
                                rx.cond(
                                    State.registro_seccion_abierta == "identidad",
                                    "66%",
                                    "100%",
                                ),
                            ),
                            font_size="xs",
                            font_weight="semibold",
                            color=rx.color_mode_cond(light="#1d4ed8", dark="#93c5fd"),
                        ),
                        width="100%",
                    ),
                    rx.box(
                        rx.box(
                            height="100%",
                            width=rx.cond(
                                State.registro_seccion_abierta == "cuenta",
                                "33%",
                                rx.cond(
                                    State.registro_seccion_abierta == "identidad",
                                    "66%",
                                    "100%",
                                ),
                            ),
                            bg="linear-gradient(90deg, #2563eb 0%, #0ea5e9 100%)",
                            border_radius="full",
                            transition="width 0.25s ease",
                        ),
                        width="100%",
                        height="8px",
                        border_radius="full",
                        bg=rx.color_mode_cond(light="#dbeafe", dark="rgba(30, 64, 175, 0.28)"),
                        overflow="hidden",
                    ),
                    spacing="1",
                    width="100%",
                ),

                rx.cond(
                    State.registro_seccion_abierta == "cuenta",
                    rx.vstack(
                        rx.vstack(
                            label_requerido("Correo electrónico"),
                            rx.hstack(
                                rx.input(
                                    placeholder="usuario@ejemplo.com",
                                    type="email",
                                    value=State.correo,
                                    on_change=State.set_and_validate_correo,
                                    on_blur=State.validar_correo_accion,
                                    width="100%",
                                    **input_style,
                                ),
                                rx.cond(State.correo_validado, rx.icon("circle-check", color="#10b981", size=20, ml="2"), rx.box()),
                                width="100%",
                            ),
                            rx.cond(State.correo_confirmacion_visible, rx.text(State.correo_confirmacion_mensaje, color=rx.cond(State.correo_validado, "green.500", "red.500"), font_size="sm", mt="1"), rx.box()),
                            width="100%",
                        ),
                        rx.vstack(
                            label_requerido("Contraseña"),
                            rx.hstack(
                                rx.input(
                                    placeholder="�?��?��?��?��?��?��?��?�",
                                    type=rx.cond(State.show_password, "text", "password"),
                                    value=State.contraseña,
                                    on_change=State.set_contraseña,
                                    width="100%",
                                    **input_style,
                                ),
                                rx.button(rx.cond(State.show_password, rx.icon("eye-off", size=18), rx.icon("eye", size=18)), on_click=State.toggle_show_password, variant="soft", size="3", radius="large"),
                                width="100%",
                                spacing="2",
                            ),
                            rx.text("Mínimo 8 caracteres, mayúscula, número y símbolo.", font_size="xs", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8")),
                            password_strength_label(State.contraseña),
                            width="100%",
                        ),
                        confirmar_field,
                        rx.hstack(
                            rx.spacer(),
                            rx.button("Continuar", type="button", size="3", color_scheme="blue", on_click=State.continuar_a_identidad),
                            width="100%",
                        ),
                        spacing="4",
                        width="100%",
                        animation="fadeInUp 0.28s ease",
                        style={
                            "@keyframes fadeInUp": {
                                "0%": {"opacity": "0", "transform": "translateY(8px)"},
                                "100%": {"opacity": "1", "transform": "translateY(0px)"},
                            }
                        },
                    ),
                    rx.box(),
                ),

                rx.cond(
                    State.registro_seccion_abierta == "identidad",
                    rx.vstack(
                        rx.grid(
                            rx.vstack(
                                label_requerido("Nombres"),
                                rx.hstack(
                                    rx.input(placeholder="Tus nombres", value=State.nombres, on_change=State.set_and_validate_nombres, width="100%", **input_style),
                                    rx.cond(State.nombres_valid, rx.icon("circle-check", color="#10b981", size=18, ml="2"), rx.box()),
                                    width="100%",
                                ),
                                width="100%",
                            ),
                            rx.vstack(
                                label_requerido("Apellidos"),
                                rx.hstack(
                                    rx.input(placeholder="Tus apellidos", value=State.apellidos, on_change=State.set_and_validate_apellidos, width="100%", **input_style),
                                    rx.cond(State.apellidos_valid, rx.icon("circle-check", color="#10b981", size=18, ml="2"), rx.box()),
                                    width="100%",
                                ),
                                width="100%",
                            ),
                            template_columns={"base": "1fr", "md": "1fr 1fr"},
                            gap="3",
                            width="100%",
                        ),
                        rx.grid(
                            rx.vstack(
                                rx.text("Sexo", color=text_color, font_weight="medium", font_size="sm"),
                                rx.select(["Femenino", "Masculino", "Prefiero no decirlo"], placeholder="Selecciona", value=State.sexo, on_change=State.set_sexo, **input_style),
                                width="100%",
                            ),
                            rx.vstack(
                                rx.text("Tipo de ID", font_weight="medium", font_size="sm", color=text_color),
                                rx.select(["Cédula", "Pasaporte", "Tarjeta de Identidad"], placeholder="Selecciona", value=State.tipo_identificacion, on_change=State.set_tipo_identificacion, **input_style),
                                width="100%",
                            ),
                            rx.vstack(
                                label_requerido("Número de ID"),
                                rx.hstack(
                                    rx.input(placeholder="123456789", value=State.numero_identificacion, on_change=State.set_and_validate_numero_identificacion, width="100%", **input_style),
                                    rx.cond(State.numero_identificacion_valid, rx.icon("circle-check", color="#10b981", size=20, ml="2"), rx.box()),
                                ),
                                width="100%",
                            ),
                            rx.vstack(
                                rx.text("Teléfono", color=text_color, font_weight="medium", font_size="sm"),
                                rx.hstack(
                                    rx.input(placeholder="Tu teléfono", value=State.telefono, on_change=State.set_and_validate_telefono, width="100%", **input_style),
                                    rx.cond(State.telefono_valid, rx.icon("circle-check", color="#10b981", size=20, ml="2"), rx.box()),
                                ),
                                width="100%",
                            ),
                            template_columns={"base": "1fr", "md": "1fr 1fr"},
                            gap="3",
                            width="100%",
                        ),
                        rx.hstack(
                            rx.button("Atrás", type="button", size="3", variant="soft", on_click=State.set_registro_seccion_abierta("cuenta")),
                            rx.spacer(),
                            rx.button("Continuar", type="button", size="3", color_scheme="blue", on_click=State.continuar_a_ubicacion),
                            width="100%",
                        ),
                        spacing="4",
                        width="100%",
                        animation="fadeInUp 0.28s ease",
                        style={
                            "@keyframes fadeInUp": {
                                "0%": {"opacity": "0", "transform": "translateY(8px)"},
                                "100%": {"opacity": "1", "transform": "translateY(0px)"},
                            }
                        },
                    ),
                    rx.box(),
                ),

                rx.cond(
                    State.registro_seccion_abierta == "ubicacion",
                    rx.vstack(
                        rx.grid(
                            rx.vstack(
                                rx.text("Departamento", color=text_color, font_weight="medium", font_size="sm"),
                                rx.select(
                                    [
                                        "Amazonas", "Antioquia", "Arauca", "Atlántico", "Bolívar", "Boyacá", "Caldas", "Caquetá", "Casanare", "Cauca", "Cesar", "Chocó", "Córdoba",
                                        "Cundinamarca", "Guainía", "Guaviare", "Huila", "La Guajira", "Magdalena", "Meta", "Nariño", "Norte de Santander", "Putumayo", "Quindío", "Risaralda",
                                        "Santander", "Sucre", "Tolima", "Valle del Cauca", "Vaupés", "Vichada",
                                    ],
                                    placeholder="Selecciona",
                                    value=State.departamento,
                                    on_change=State.set_and_validate_departamento,
                                    **input_style,
                                ),
                                width="100%",
                            ),
                            rx.vstack(
                                rx.text("Ciudad", color=text_color, font_weight="medium", font_size="sm"),
                                rx.select(State.ciudades_disponibles, placeholder="Selecciona", value=State.ciudad, on_change=State.set_and_validate_ciudad, is_disabled=State.departamento == "", **input_style),
                                width="100%",
                            ),
                            template_columns={"base": "1fr", "md": "1fr 1fr"},
                            gap="3",
                            width="100%",
                        ),
                        rx.vstack(
                            label_requerido("Dirección"),
                            rx.input(placeholder="Tu dirección", value=State.direccion, on_change=State.set_direccion, width="100%", **input_style),
                            width="100%",
                        ),
                        rx.grid(
                            rx.vstack(
                                rx.text("Etnia", color=text_color, font_weight="medium", font_size="sm"),
                                rx.select(["Ninguna", "Indígena", "Afrocolombiano", "Raizal", "Palenquero", "Gitano/a", "Otro"], placeholder="Selecciona", value=State.etnia, on_change=State.set_etnia, **input_style),
                                width="100%",
                            ),
                            rx.vstack(
                                rx.text("Características especiales", color=text_color, font_weight="medium", font_size="sm"),
                                rx.select(
                                    ["Ninguna", "Habitante de la calle", "No brinda información", "Peligro Inminente", "Periodistas en ejercicio de su actividad", "Primera Infancia", "Veteranos Fuerza Pública", "Víctimas - Conflicto Armado"],
                                    placeholder="Selecciona",
                                    value=State.persona_vulnerable_registro,
                                    on_change=State.set_persona_vulnerable_registro,
                                    **input_style,
                                ),
                                width="100%",
                            ),
                            template_columns={"base": "1fr", "md": "1fr 1fr"},
                            gap="3",
                            width="100%",
                        ),
                        rx.divider(margin_y="2", bg=input_border),
                        rx.vstack(
                            rx.checkbox("Acepto recibir notificaciones por correo", is_checked=State.acepta_notificaciones, on_change=State.set_acepta_notificaciones, color=text_color, size="3"),
                            rx.checkbox(
                                rx.hstack(
                                    rx.link("He leído y acepto la Política de Protección de Datos", href="/politica-privacidad", color="#3b82f6", font_weight="medium"),
                                    rx.text("(Aviso obligatorio)", color=rx.color_mode_cond(light="#94a3b8", dark="#64748b"), font_size="sm"),
                                ),
                                is_checked=State.acepta_politica_datos,
                                on_change=State.preconfirmar_politica,
                                color=text_color,
                                size="3",
                            ),
                            spacing="3",
                            width="100%",
                        ),
                        rx.hstack(
                            rx.button("Atrás", type="button", size="3", variant="soft", on_click=State.set_registro_seccion_abierta("identidad")),
                            rx.spacer(),
                            rx.text("�?ltimo paso", font_size="xs", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8")),
                            width="100%",
                        ),
                        spacing="4",
                        width="100%",
                        animation="fadeInUp 0.28s ease",
                        style={
                            "@keyframes fadeInUp": {
                                "0%": {"opacity": "0", "transform": "translateY(8px)"},
                                "100%": {"opacity": "1", "transform": "translateY(0px)"},
                            }
                        },
                    ),
                    rx.box(),
                ),
                
                rx.cond(
                    State.modal_politica_visible,
                    rx.box(
                        rx.box(
                            rx.box(
                                position="absolute",
                                top="-60px",
                                right="-40px",
                                width="180px",
                                height="180px",
                                bg="rgba(59, 130, 246, 0.35)",
                                border_radius="full",
                                filter="blur(50px)",
                            ),
                            rx.box(
                                position="absolute",
                                bottom="-40px",
                                left="-30px",
                                width="140px",
                                height="140px",
                                bg="rgba(14, 165, 233, 0.25)",
                                border_radius="full",
                                filter="blur(45px)",
                            ),
                            rx.vstack(
                                rx.center(
                                    rx.box(
                                        rx.icon("shield-check", size=32, color="#60a5fa"),
                                        p="4",
                                        bg=rx.color_mode_cond(
                                            light="linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%)",
                                            dark="linear-gradient(135deg, rgba(37, 99, 235, 0.25) 0%, rgba(14, 165, 233, 0.15) 100%)",
                                        ),
                                        border=rx.color_mode_cond(
                                            light="1px solid #93c5fd",
                                            dark="1px solid rgba(96, 165, 250, 0.35)",
                                        ),
                                        border_radius="2xl",
                                        box_shadow=rx.color_mode_cond(
                                            light="0 12px 28px -8px rgba(37, 99, 235, 0.35)",
                                            dark="0 12px 28px -8px rgba(2, 6, 23, 0.8)",
                                        ),
                                    ),
                                    width="100%",
                                ),
                                rx.badge(
                                    "Ley 1581 de 2012",
                                    color_scheme="blue",
                                    variant="soft",
                                    radius="full",
                                    size="1",
                                ),
                                rx.heading(
                                    "Política de Privacidad",
                                    size="6",
                                    color=rx.color_mode_cond(light="#0f172a", dark="#f8fafc"),
                                    font_weight="bold",
                                    text_align="center",
                                    letter_spacing="-0.02em",
                                ),
                                rx.text(
                                    "Al aceptar, confirmas que has leído y comprendido cómo tratamos tus datos personales para la gestión de PQRS.",
                                    color=rx.color_mode_cond(light="#475569", dark="#94a3b8"),
                                    font_size="sm",
                                    line_height="1.6",
                                    text_align="center",
                                ),
                                rx.box(
                                    rx.vstack(
                                        rx.hstack(
                                            rx.icon("lock", size=16, color="#38bdf8"),
                                            rx.text(
                                                "Tus datos se usan solo para trámites y notificaciones.",
                                                font_size="xs",
                                                color=rx.color_mode_cond(light="#334155", dark="#cbd5e1"),
                                            ),
                                            spacing="2",
                                            align_items="center",
                                        ),
                                        rx.hstack(
                                            rx.icon("file-text", size=16, color="#38bdf8"),
                                            rx.link(
                                                "Leer política completa",
                                                href="/politica-privacidad",
                                                font_size="xs",
                                                font_weight="semibold",
                                                color=rx.color_mode_cond(light="#2563eb", dark="#93c5fd"),
                                                _hover={"text_decoration": "underline"},
                                            ),
                                            spacing="2",
                                            align_items="center",
                                        ),
                                        spacing="2",
                                        align_items="start",
                                        width="100%",
                                    ),
                                    width="100%",
                                    p="4",
                                    border_radius="xl",
                                    bg=rx.color_mode_cond(
                                        light="rgba(239, 246, 255, 0.85)",
                                        dark="rgba(15, 23, 42, 0.55)",
                                    ),
                                    border=rx.color_mode_cond(
                                        light="1px solid #bfdbfe",
                                        dark="1px solid rgba(51, 65, 85, 0.8)",
                                    ),
                                ),
                                rx.vstack(
                                    rx.button(
                                        rx.hstack(
                                            rx.icon("check", size=18),
                                            rx.text("Acepto y continúo", font_weight="bold"),
                                            spacing="2",
                                            justify="center",
                                        ),
                                        on_click=State.confirmar_politica,
                                        color_scheme="blue",
                                        size="3",
                                        width="100%",
                                        radius="full",
                                        box_shadow="0 10px 24px -10px rgba(37, 99, 235, 0.65)",
                                    ),
                                    rx.button(
                                        "Cancelar",
                                        on_click=State.cancelar_politica,
                                        variant="ghost",
                                        size="3",
                                        width="100%",
                                        color=rx.color_mode_cond(light="#64748b", dark="#94a3b8"),
                                    ),
                                    spacing="2",
                                    width="100%",
                                ),
                                spacing="4",
                                align_items="center",
                                width="100%",
                                z_index="1",
                            ),
                            position="relative",
                            overflow="hidden",
                            p={"base": "6", "md": "8"},
                            bg=rx.color_mode_cond(
                                light="rgba(255, 255, 255, 0.92)",
                                dark="rgba(15, 23, 42, 0.88)",
                            ),
                            backdrop_filter="blur(24px)",
                            border=rx.color_mode_cond(
                                light="1px solid rgba(147, 197, 253, 0.9)",
                                dark="1px solid rgba(96, 165, 250, 0.28)",
                            ),
                            border_radius="3xl",
                            box_shadow=rx.color_mode_cond(
                                light="0 32px 64px -20px rgba(30, 64, 175, 0.45), 0 0 0 1px rgba(255,255,255,0.5) inset",
                                dark="0 32px 64px -16px rgba(2, 6, 23, 0.95), 0 0 0 1px rgba(255,255,255,0.06) inset",
                            ),
                            width="100%",
                            max_width="480px",
                            animation="modalPop 0.32s cubic-bezier(0.22, 1, 0.36, 1)",
                            style={
                                "@keyframes modalPop": {
                                    "0%": {"opacity": "0", "transform": "scale(0.94) translateY(12px)"},
                                    "100%": {"opacity": "1", "transform": "scale(1) translateY(0px)"},
                                }
                            },
                        ),
                        position="fixed",
                        inset="0",
                        bg="rgba(2, 6, 23, 0.72)",
                        backdrop_filter="blur(10px)",
                        display="flex",
                        align_items="center",
                        justify_content="center",
                        z_index="1000",
                        p="6",
                    ),
                ),
                
                rx.cond(
                    State.error_de_registro != "",
                    rx.box(
                        rx.text(State.error_de_registro, color="#ef4444", font_size="sm", font_weight="medium"),
                        p="3", bg=rx.color_mode_cond(light="#fef2f2", dark="rgba(239, 68, 68, 0.1)"),
                        border_radius="lg",
                        width="100%",
                        animation="flashMessage 5s ease forwards",
                        style={
                            "@keyframes flashMessage": {
                                "0%": {"opacity": "0", "transform": "translateY(-4px)"},
                                "10%": {"opacity": "1", "transform": "translateY(0px)"},
                                "80%": {"opacity": "1", "transform": "translateY(0px)"},
                                "100%": {"opacity": "0", "transform": "translateY(-4px)"},
                            }
                        },
                    ),
                    rx.box(),
                ),
                rx.cond(
                    State.succes != "",
                    rx.box(
                        rx.text(State.succes, color="#10b981", font_size="sm", font_weight="medium"),
                        p="3", bg=rx.color_mode_cond(light="#ecfdf5", dark="rgba(16, 185, 129, 0.1)"),
                        border_radius="lg",
                        width="100%",
                        animation="flashMessage 5s ease forwards",
                        style={
                            "@keyframes flashMessage": {
                                "0%": {"opacity": "0", "transform": "translateY(-4px)"},
                                "10%": {"opacity": "1", "transform": "translateY(0px)"},
                                "80%": {"opacity": "1", "transform": "translateY(0px)"},
                                "100%": {"opacity": "0", "transform": "translateY(-4px)"},
                            }
                        },
                    ),
                    rx.box(),
                ),
                
                rx.hstack(
                    rx.button(
                        "Crear cuenta",
                        type="submit",
                        color_scheme="blue",
                        size="4",
                        radius="large",
                        width={"base": "100%", "md": "300px"},
                        box_shadow="0 8px 20px -8px rgba(37, 99, 235, 0.55)",
                        is_disabled=(State.registro_paso_habilitado < 3) | (State.registro_seccion_abierta != "ubicacion"),
                    ),
                    rx.link("¿Ya tienes una cuenta? Inicia sesión", href="/login", margin_left={"base": "0", "md": "4"}, color="#3b82f6", font_weight="medium"),
                    spacing="6",
                    justify={"base": "center", "md": "start"},
                    width="100%",
                    margin_top="4"
                ),
                
                spacing="4",
                width="100%",
            ),
            on_submit=on_submit,
            width="100%",
        ),
        p={"base": "6", "md": "8"},
        max_width="1100px",
        width="100%",
        bg=rx.color_mode_cond(light="linear-gradient(145deg, #dbeafe 0%, #bfdbfe 100%)", dark="linear-gradient(145deg, #0f172a 0%, #1e3a8a 100%)"),
        border=rx.color_mode_cond(light="1px solid #93c5fd", dark="1px solid #1e40af"),
        border_radius="2xl",
        box_shadow=rx.color_mode_cond(light="0 30px 60px -25px rgba(30, 64, 175, 0.35)", dark="0 30px 60px -25px rgba(2, 6, 23, 0.85)"),
        z_index="1",
        position="relative",
        overflow="hidden",
        style={"borderLeft": "8px solid #2563eb"},
        max_height={"base": "none", "lg": "84vh"},
        overflow_y={"base": "visible", "lg": "auto"},
    )




def nav_icon_link(icon_name: str, text: str, href: str, active: bool = False, display_cond=None) -> rx.Component:
    link_content = rx.link(
        rx.hstack(
            rx.icon(icon_name, size=20, color="rgba(255,255,255,0.85)", transition="transform 0.2s ease"),
            rx.text(
                text,
                class_name="nav-text",
                color="white",
                font_weight="600",
                font_size="13px",
                white_space="nowrap",
                opacity="0",
                max_width="0px",
                overflow="hidden",
                transition="all 0.3s cubic-bezier(0.4, 0, 0.2, 1)",
                margin_left="0px"
            ),
            align_items="center",
            padding="8px",
            border_radius="12px",
            bg="rgba(255,255,255,0.15)" if active else "transparent",
            transition="all 0.2s ease",
            _hover={
                "bg": "rgba(255,255,255,0.25)",
                "& > svg": {
                    "color": "white",
                    "transform": "scale(1.1)"
                },
                "& .nav-text": {
                    "opacity": "1",
                    "max_width": "200px",
                    "margin_left": "8px"
                }
            }
        ),
        href=href,
        text_decoration="none"
    )
    if display_cond is not None:
        return rx.cond(display_cond, link_content, rx.box(display="none"))
    return link_content

def navbar() -> rx.Component:
    return rx.fragment(rx.box(
        rx.hstack(
            # Logo / marca
            rx.link(
                rx.hstack(
                    rx.box(
                        rx.icon("shield-check", size=18, color="#e85d04"),
                        width="36px", height="36px",
                        border_radius="10px",
                        bg="rgba(232,93,4,0.18)",
                        display="flex", align_items="center", justify_content="center", flex_shrink="0",
                    ),
                    rx.text("PQRS", font_size="18px", font_weight="800", color="white", letter_spacing="-0.01em"),
                    spacing="3", align_items="center",
                ),
                href=State.home_url, text_decoration="none", flex_shrink="0",
            ),
            
            rx.box(width="1px", height="24px", bg="rgba(255,255,255,0.2)", margin_x="4"),
            
            # Navegación con iconos dinámicos
            rx.hstack(
                nav_icon_link("home", "Inicio", State.home_url),
                nav_icon_link("file-plus", "Nueva Solicitud", "/solicitudes"),
                
                # Opciones de Funcionario
                nav_icon_link("bar-chart-2", "Reportes", "/reportes", display_cond=State.es_autenticada & (State.rol_usuario == "funcionario")),
                nav_icon_link("users", "Ver Usuarios", "/usuarios", display_cond=State.es_autenticada & (State.rol_usuario == "funcionario")),
                nav_icon_link("user-plus", "Registrar Func.", "/registro-funcionario", display_cond=State.es_autenticada & (State.rol_usuario == "funcionario")),
                nav_icon_link("settings", "Cambiar Rol", "/cambiar-rol", display_cond=State.es_autenticada & (State.rol_usuario == "funcionario")),
                nav_icon_link("help-circle", "Ayuda", "/ayuda-funcionario", display_cond=State.es_autenticada & (State.rol_usuario == "funcionario")),
                
                # Dashboards
                nav_icon_link("layout", "Dashboard Func.", "/dashboard-funcionario", active=True, display_cond=State.es_autenticada & (State.rol_usuario == "funcionario")),
                nav_icon_link("layout", "Mi Panel", "/dashboard", active=True, display_cond=State.es_autenticada & (State.rol_usuario != "funcionario")),
                
                spacing="2",
                align_items="center",
                flex_wrap="nowrap",
                overflow_x="auto"
            ),
            
            rx.spacer(),
            
            # Controles derecha
            rx.hstack(
                rx.color_mode.button(color="rgba(255,255,255,0.8)", _hover={"color": "white"}),
                # Chip animado con icono y nombre del funcionario/usuario
                rx.cond(
                    State.es_autenticada,
                    rx.hstack(
                        rx.box(
                            rx.icon("user-circle", size=20, color="rgba(255,255,255,0.9)"),
                            id="navbar-user-avatar",
                            width="36px",
                            height="36px",
                            border_radius="full",
                            bg="rgba(255,255,255,0.15)",
                            display="flex",
                            align_items="center",
                            justify_content="center",
                            flex_shrink="0",
                            transition="all 0.2s ease",
                        ),
                        rx.box(
                            rx.el.span(
                                State.nombres,
                                id="navbar-user-name",
                                style={
                                    "color": "white",
                                    "font_weight": "700",
                                    "font_size": "13px",
                                    "white_space": "nowrap",
                                    "opacity": "0",
                                    "max_width": "0px",
                                    "overflow": "hidden",
                                    "transition": "all 0.3s cubic-bezier(0.4, 0, 0.2, 1)",
                                    "margin_left": "0px",
                                    "display": "block",
                                    "line_height": "1.2",
                                },
                            ),
                            rx.el.span(
                                State.rol_usuario,
                                id="navbar-user-role",
                                style={
                                    "color": "rgba(255,255,255,0.65)",
                                    "font_weight": "500",
                                    "font_size": "10px",
                                    "white_space": "nowrap",
                                    "opacity": "0",
                                    "max_width": "0px",
                                    "overflow": "hidden",
                                    "transition": "all 0.3s cubic-bezier(0.4, 0, 0.2, 1)",
                                    "margin_left": "0px",
                                    "display": "block",
                                    "text_transform": "capitalize",
                                    "line_height": "1.2",
                                    "margin_top": "1px",
                                },
                            ),
                            display="flex",
                            flex_direction="column",
                            align_items="flex-start",
                        ),
                        align_items="center",
                        id="navbar-user-chip",
                        spacing="2",
                        padding="6px 10px",
                        border_radius="12px",
                        bg="rgba(255,255,255,0.10)",
                        border="1px solid rgba(255,255,255,0.18)",
                        cursor="default",
                        transition="all 0.2s ease",
                    ),
                    rx.box(display="none"),
                ),
                # CSS global para animaciones hover del chip y botón salir
                rx.el.style("""
                    #navbar-user-chip:hover {
                        background: rgba(255,255,255,0.22) !important;
                        border-color: rgba(255,255,255,0.35) !important;
                    }
                    #navbar-user-chip:hover #navbar-user-name {
                        opacity: 1 !important;
                        max-width: 160px !important;
                        margin-left: 4px !important;
                    }
                    #navbar-user-chip:hover #navbar-user-role {
                        opacity: 1 !important;
                        max-width: 160px !important;
                        margin-left: 4px !important;
                    }
                    #navbar-user-chip:hover #navbar-user-avatar {
                        background: rgba(255,255,255,0.28) !important;
                        transform: scale(1.08);
                    }
                    #navbar-logout-btn:hover {
                        background: rgba(239,68,68,0.22) !important;
                        border-color: rgba(239,68,68,0.45) !important;
                    }
                    #navbar-logout-btn:hover #navbar-logout-icon {
                        background: rgba(239,68,68,0.35) !important;
                        transform: scale(1.08) rotate(-8deg);
                    }
                    #navbar-logout-btn:hover #navbar-logout-text {
                        opacity: 1 !important;
                        max-width: 60px !important;
                        margin-left: 8px !important;
                    }
                """),
                # Botón Salir animado con CSS global (sin class_name)
                rx.cond(
                    State.es_autenticada,
                    rx.hstack(
                        rx.box(
                            rx.icon("log-out", size=18, color="rgba(252,165,165,0.95)"),
                            id="navbar-logout-icon",
                            width="34px",
                            height="34px",
                            border_radius="10px",
                            bg="rgba(239,68,68,0.15)",
                            display="flex",
                            align_items="center",
                            justify_content="center",
                            flex_shrink="0",
                            transition="all 0.2s ease",
                        ),
                        rx.el.span(
                            "Salir",
                            id="navbar-logout-text",
                            style={
                                "color": "rgba(252,165,165,1)",
                                "font_weight": "600",
                                "font_size": "13px",
                                "white_space": "nowrap",
                                "opacity": "0",
                                "max_width": "0px",
                                "overflow": "hidden",
                                "transition": "all 0.3s cubic-bezier(0.4, 0, 0.2, 1)",
                                "margin_left": "0px",
                                "display": "inline-block",
                            },
                        ),
                        id="navbar-logout-btn",
                        align_items="center",
                        spacing="0",
                        padding="4px 8px",
                        border_radius="12px",
                        bg="rgba(239,68,68,0.10)",
                        border="1px solid rgba(239,68,68,0.25)",
                        cursor="pointer",
                        transition="all 0.2s ease",
                        on_click=State.abrir_logout_confirm,
                    ),
                    rx.box(display="none"),
                ),
                spacing="4", align_items="center"
            ),
            justify="between", align_items="center", width="100%", max_width="1400px", margin="0 auto"
        ),
        bg=rx.color_mode_cond(light="linear-gradient(90deg, #1e3a8a 0%, #1e40af 100%)", dark="linear-gradient(90deg, #0f172a 0%, #172554 100%)"),
        padding_y="12px", padding_x={"base": "16px", "md": "32px"}, width="100%",
        box_shadow="0 4px 20px -2px rgba(0, 0, 0, 0.2)",
        position="sticky", top="0", z_index="50", border_bottom="1px solid rgba(255,255,255,0.1)"
    ), logout_confirm_modal())


def access_denied_widget(message: str) -> rx.Component:
    # Auto-redirect to login if not authenticated; show access denied if authenticated but wrong role
    return rx.cond(
        rx.State.is_hydrated,
        rx.cond(
            State.es_autenticada,
            rx.vstack(
                rx.heading("Acceso Denegado", size="8", color="red.500"),
                rx.text(message, color="gray.600"),
                rx.vstack(
                    rx.button("Cerrar Sesión", on_click=State.logout, color_scheme="red", width="100%"),
                    rx.link(rx.button("Volver al inicio", color_scheme="blue"), href="/"),
                    spacing="4",
                    align_items="center"
                ),
                spacing="4",
                align_items="center"
            ),
            rx.script("window.location.href = '/login'")
        ),
        rx.center(rx.spinner(size="3", color="#3b82f6"), height="50vh")
    )


def utility_bar() -> rx.Component:
    return rx.hstack(
        rx.link(
            rx.hstack(rx.icon("globe", size=14), rx.text("GOV.CO"), spacing="2", align_items="center"),
            href="/", font_weight="bold", color="rgba(255,255,255,0.9)", text_decoration="none", _hover={"color": "white"}
        ),
        rx.spacer(),
        rx.hstack(
            rx.link(rx.hstack(rx.icon("accessibility", size=14), rx.text("Accesibilidad"), spacing="1"), href="#", font_size="xs", color="rgba(255,255,255,0.8)", text_decoration="none", _hover={"color": "white"}),
            rx.text("|", color="rgba(255,255,255,0.3)"),
            rx.link(rx.hstack(rx.icon("log-in", size=14), rx.text("Inicia sesión"), spacing="1"), href="/login", font_size="xs", color="rgba(255,255,255,0.8)", text_decoration="none", _hover={"color": "white"}),
            rx.text("|", color="rgba(255,255,255,0.3)"),
            rx.link(rx.hstack(rx.icon("user-plus", size=14), rx.text("Regístrate"), spacing="1"), href="/registro", font_size="xs", color="rgba(255,255,255,0.8)", text_decoration="none", _hover={"color": "white"}),
            spacing="4", align_items="center"
        ),
        width="100%", padding_x="24px", padding_y="8px",
        bg=rx.color_mode_cond(light="#0f172a", dark="#020617"),
        border_bottom="1px solid rgba(255,255,255,0.08)"
    )

def logout_confirm_modal():
    """Modal de confirmacion de cierre de sesion."""
    return rx.cond(
        State.show_logout_confirm,
        rx.box(
            rx.box(
                position="fixed", inset="0",
                bg="rgba(0,0,0,0.55)",
                backdrop_filter="blur(4px)",
                z_index="1000",
                on_click=State.cerrar_logout_confirm,
            ),
            rx.center(
                rx.box(
                    rx.vstack(
                        rx.center(
                            rx.box(
                                rx.icon("log-out", size=32, color="#f87171"),
                                width="64px", height="64px",
                                border_radius="full",
                                bg="rgba(239,68,68,0.12)",
                                border="2px solid rgba(239,68,68,0.30)",
                                display="flex",
                                align_items="center",
                                justify_content="center",
                            ),
                            width="100%",
                        ),
                        rx.text(
                            "¿Cerrar sesión?",
                            font_size="22px",
                            font_weight="800",
                            color="white",
                            text_align="center",
                            letter_spacing="-0.02em",
                        ),
                        rx.text(
                            "Tu sesión se cerrará y tendrás que volver a iniciar sesión.",
                            font_size="14px",
                            color="rgba(255,255,255,0.65)",
                            text_align="center",
                            line_height="1.6",
                        ),
                        rx.box(width="100%", height="1px", bg="rgba(255,255,255,0.10)"),
                        rx.hstack(
                            rx.box(
                                rx.hstack(
                                    rx.icon("x", size=16, color="rgba(255,255,255,0.85)"),
                                    rx.text("Cancelar", font_weight="600", font_size="14px", color="rgba(255,255,255,0.9)"),
                                    spacing="2", align_items="center",
                                ),
                                padding="10px 22px",
                                border_radius="12px",
                                bg="rgba(255,255,255,0.08)",
                                border="1px solid rgba(255,255,255,0.15)",
                                cursor="pointer",
                                transition="all 0.2s ease",
                                on_click=State.cerrar_logout_confirm,
                                _hover={
                                    "bg": "rgba(255,255,255,0.16)",
                                    "border": "1px solid rgba(255,255,255,0.30)",
                                    "transform": "translateY(-1px)",
                                },
                            ),
                            rx.box(
                                rx.hstack(
                                    rx.icon("log-out", size=16, color="white"),
                                    rx.text("Sí, salir", font_weight="700", font_size="14px", color="white"),
                                    spacing="2", align_items="center",
                                ),
                                padding="10px 22px",
                                border_radius="12px",
                                bg="linear-gradient(135deg, #dc2626 0%, #b91c1c 100%)",
                                border="1px solid rgba(220,38,38,0.5)",
                                cursor="pointer",
                                transition="all 0.2s ease",
                                box_shadow="0 8px 20px -6px rgba(220,38,38,0.55)",
                                on_click=State.logout,
                                _hover={
                                    "bg": "linear-gradient(135deg, #ef4444 0%, #dc2626 100%)",
                                    "box_shadow": "0 12px 28px -6px rgba(220,38,38,0.75)",
                                    "transform": "translateY(-2px)",
                                },
                            ),
                            spacing="3",
                            justify="center",
                            width="100%",
                        ),
                        spacing="5",
                        align_items="center",
                        width="100%",
                    ),
                    bg="linear-gradient(135deg, rgba(15,23,42,0.97) 0%, rgba(30,27,75,0.97) 100%)",
                    border="1px solid rgba(255,255,255,0.12)",
                    border_radius="24px",
                    padding="36px 32px",
                    width="390px",
                    max_width="90vw",
                    box_shadow="0 30px 60px -20px rgba(0,0,0,0.85)",
                    backdrop_filter="blur(24px)",
                    position="relative",
                    z_index="1002",
                ),
                position="fixed",
                inset="0",
                z_index="1001",
            ),
            position="fixed", inset="0", z_index="1000",
        ),
        rx.box(display="none"),
    )

def toast_notification() -> rx.Component:
    return rx.cond(
        State.toast_visible,
        rx.box(
            rx.hstack(
                rx.box(
                    rx.cond(
                        State.toast_tipo == "success",
                        rx.icon("circle-check", size=22, color="white"),
                        rx.icon("circle-x", size=22, color="white"),
                    ),
                    display="flex",
                    align_items="center",
                    justify_content="center",
                    width="36px",
                    height="36px",
                    border_radius="full",
                    bg=rx.cond(State.toast_tipo == "success", "rgba(255,255,255,0.25)", "rgba(255,255,255,0.25)"),
                    flex_shrink="0",
                ),
                rx.vstack(
                    rx.text(
                        rx.cond(State.toast_tipo == "success", "¡�?xito!", "Error"),
                        font_weight="bold",
                        color="white",
                        font_size="sm",
                        line_height="1.1",
                    ),
                    rx.text(
                        State.toast_mensaje,
                        color="rgba(255,255,255,0.95)",
                        font_size="xs",
                        line_height="1.3",
                    ),
                    spacing="1",
                    align_items="start",
                ),
                rx.spacer(),
                rx.button(
                    rx.icon("x", size=16, color="white"),
                    on_click=State.ocultar_toast,
                    variant="ghost",
                    size="1",
                    _hover={"bg": "rgba(255,255,255,0.15)"},
                    padding="0",
                    min_width="24px",
                    height="24px",
                ),
                spacing="3",
                align_items="center",
                width="100%",
            ),
            position="fixed",
            top="50%",
            left="50%",
            transform="translate(-50%, -50%)",
            z_index="9999",
            min_width="200px",
            max_width="280px",
            padding="10px 12px",
            border_radius="18px",
            bg=rx.cond(
                State.toast_tipo == "success",
                "linear-gradient(135deg, #16a34a, #15803d)",
                "linear-gradient(135deg, #dc2626, #b91c1c)",
            ),
            box_shadow="0 8px 32px rgba(0,0,0,0.22), 0 2px 8px rgba(0,0,0,0.12)",
            style={
                "animation": "slideInToast 0.35s cubic-bezier(0.34, 1.56, 0.64, 1)",
                "@keyframes slideInToast": {
                    "from": {"opacity": "0", "transform": "translateY(24px) scale(0.95)"},
                    "to": {"opacity": "1", "transform": "translateY(0) scale(1)"},
                }
            }
        ),
        rx.box()
    )


def index() -> rx.Component:
    hero_text = rx.color_mode_cond(light="rgba(15, 23, 42, 0.96)", dark="white")
    hero_sub = rx.color_mode_cond(light="rgba(15, 23, 42, 0.72)", dark="rgba(255,255,255,0.75)")
    card_bg = rx.color_mode_cond(light="white", dark="#111827")
    section_bg = rx.color_mode_cond(light="#f8fafc", dark="#020617")
    body_bg = rx.color_mode_cond(light="#f1f5f9", dark="#020617")

    return rx.box(
        rx.color_mode.button(position="top-right"),
        utility_bar(),
        navbar(),
        rx.box(
            rx.container(
                rx.hstack(
                    rx.vstack(
                        rx.text("Plataforma oficial de atención ciudadana", color="white", font_size="sm", bg="#2563eb", padding_x="3", padding_y="2", border_radius="full", mb="4"),
                        rx.heading("Atención PQRS - Enlace 1755", size="8", color="white", line_height="1.1"),
                        rx.text(
                            "Radica, consulta y gestiona tus Peticiones, Quejas, Reclamos y Sugerencias de forma clara, rápida y segura.",
                            color="rgba(255,255,255,0.85)",
                            font_size="lg",
                            max_width="680px"
                        ),
                        rx.hstack(
                            rx.link(rx.button("Radicar PQRS", color_scheme="blue", size="4", width="200px"), href="/solicitudes"),
                            rx.link(rx.button("Consultar Estado", color_scheme="blue", size="4", width="200px"), href="/consultar-estado"),
                            spacing="4",
                            flex_wrap="wrap"
                        ),
                        spacing="6",
                        align_items="start",
                        width="100%",
                        max_width="720px"
                    ),
                    rx.card(
                        rx.vstack(
                            rx.heading("Accesos rápidos", size="5", color="#000000"),
                            rx.link(rx.button("Registrarme", color_scheme="blue", width="100%"), href="/registro"),
                            rx.link(rx.button("Iniciar sesión", variant="solid", color_scheme="gray", width="100%"), href="/login"),
                            rx.link(rx.button("Nueva solicitud", variant="outline", color_scheme="gray", width="100%"), href="/solicitudes"),
                            rx.text("Disponible para ciudadanos que deseen registrar y hacer seguimiento a sus solicitudes.", color="dark", font_size="sm"),
                            spacing="4",
                            align_items="stretch"
                        ),
                        p="6",
                        bg="rgba(255,255,255,0.08)",
                        border="1px solid rgba(255,255,255,0.15)",
                        border_radius="2xl",
                        width="100%",
                        max_width="340px"
                    ),
                    spacing="8",
                    align_items="center",
                    justify="between",
                    flex_wrap="wrap"
                ),
                max_width="1200px",
                padding_y="20",
                padding_x="6"
            ),
            width="100%",
            min_height="520px",
            style={
                "backgroundImage": "linear-gradient(90deg, rgba(15,23,42,0.84), rgba(15,23,42,0.30)), url('/Gemini_Generated_Image_ouyornouyornouyo.png')",
                "backgroundSize": "cover",
                "backgroundPosition": "center",
                "backgroundRepeat": "no-repeat"
            }
        ),

        rx.container(
            rx.vstack(
                rx.vstack(
                    rx.heading("¿Qué deseas hacer hoy?", size="7", color="#0f172a"),
                    rx.text("Accede rápidamente a los servicios principales del sistema.", color="#475569", font_size="md"),
                    spacing="3",
                    align_items="center"
                ),
                rx.hstack(
                    quick_action_card("Radicar PQRS", "Crea una nueva petición, queja, reclamo o sugerencia.", "Ir al formulario", "/solicitudes", "blue"),
                    quick_action_card("Consultar estado", "Revisa el avance y respuesta de tus solicitudes.", "Consultar", "/consultar-estado", "cyan"),
                    quick_action_card("Registro ciudadano", "Crea tu cuenta para gestionar trámites de forma segura.", "Registrarme", "/registro", "green"),
                    quick_action_card("Iniciar sesión", "Accede a tu cuenta y continúa tus gestiones.", "Entrar", "/login", "gray"),
                    spacing="5",
                    justify="center",
                    flex_wrap="wrap"
                ),
                spacing="9",
                align_items="center"
            ),
            max_width="1200px",
            padding_y="20",
            padding_x="6"
        ),

        rx.box(
            rx.container(
                rx.vstack(
                    rx.heading("Atención clara y transparente para la ciudadanía", size="7", color="#0f172a"),
                    rx.text("Este portal facilita la recepción, gestión y seguimiento de solicitudes ciudadanas de manera organizada y accesible.", color="#64748b", font_size="md", text_align="center", max_width="850px"),
                    spacing="4",
                    align_items="center"
                ),
                rx.hstack(
                    info_card("Canal seguro", "Tus datos y solicitudes se gestionan en un entorno controlado."),
                    info_card("Trazabilidad", "Cada solicitud puede registrarse y consultarse con mayor claridad."),
                    info_card("Atención oportuna", "El sistema está pensado para mejorar tiempos y experiencia ciudadana."),
                    spacing="5",
                    justify="center",
                    flex_wrap="wrap"
                ),
                spacing="9",
                align_items="center"
            ),
            width="100%",
            bg=section_bg,
            padding_y="20"
        ),

        rx.container(
            rx.vstack(
                rx.heading("¿Qué significa PQRS?", size="7", color="#0f172a"),
                rx.hstack(
                    pqrs_badge("Petición", "Solicitud respetuosa de información o actuación por parte de la entidad.", "#2563eb"),
                    pqrs_badge("Queja", "Manifestación de inconformidad por la conducta o atención recibida.", "#f59e0b"),
                    pqrs_badge("Reclamo", "Expresión de inconformidad por una prestación deficiente o incumplimiento.", "#ef4444"),
                    pqrs_badge("Sugerencia", "Propuesta o recomendación para mejorar la atención o el servicio.", "#10b981"),
                    spacing="5",
                    justify="center",
                    flex_wrap="wrap"
                ),
                spacing="8",
                align_items="center"
            ),
            max_width="1200px",
            padding_y="20",
            padding_x="6"
        ),

        footer(),
        brand_footer(),
        bg=body_bg,
        width="100%",
        min_height="100vh"
    )


def quick_action_card(title: str, desc: str, button_text: str, href: str, accent: str = "blue") -> rx.Component:
    card_bg = rx.color_mode_cond(light="white", dark="#111827")
    border = rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155")
    text_main = rx.color_mode_cond(light="#0f172a", dark="white")
    text_sec = rx.color_mode_cond(light="#475569", dark="#cbd5e1")

    return rx.card(
        rx.vstack(
            rx.heading(title, size="5", color=text_main),
            rx.text(desc, color=text_sec, font_size="sm"),
            rx.link(rx.button(button_text, color_scheme=accent, width="100%"), href=href),
            spacing="4",
            align_items="start",
            width="100%"
        ),
        bg=card_bg,
        border=border,
        border_radius="2xl",
        p="6",
        width="100%",
        max_width="260px",
        box_shadow="lg"
    )


def info_card(title: str, desc: str) -> rx.Component:
    card_bg = rx.color_mode_cond(light="white", dark="#111827")
    border = rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155")
    text_main = rx.color_mode_cond(light="#0f172a", dark="white")
    text_sec = rx.color_mode_cond(light="#475569", dark="#cbd5e1")

    return rx.card(
        rx.vstack(
            rx.text(title, font_weight="bold", color=text_main, font_size="md"),
            rx.text(desc, color=text_sec, font_size="sm"),
            spacing="3",
            align_items="start"
        ),
        bg=card_bg,
        border=border,
        border_radius="xl",
        p="5",
        width="100%",
        max_width="360px",
        box_shadow="sm"
    )


def pqrs_badge(title: str, desc: str, color: str) -> rx.Component:
    card_bg = rx.color_mode_cond(light="white", dark="#111827")
    border = rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155")
    text_sec = rx.color_mode_cond(light="#475569", dark="#cbd5e1")

    return rx.card(
        rx.vstack(
            rx.box(
                rx.text(title, color="white", font_weight="bold", font_size="sm"),
                bg=color,
                padding_x="3",
                padding_y="2",
                border_radius="full"
            ),
            rx.text(desc, color=text_sec, font_size="sm"),
            spacing="3",
            align_items="start"
        ),
        bg=card_bg,
        border=border,
        border_radius="xl",
        p="5",
        width="100%",
        max_width="360px",
        box_shadow="sm"
    )


def mantenimiento_modal() -> rx.Component:
    card_bg = rx.color_mode_cond(light="white", dark="#111827")
    border = rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155")
    text_main = rx.color_mode_cond(light="#0f172a", dark="#f8fafc")
    text_sec = rx.color_mode_cond(light="#475569", dark="#cbd5e1")
    
    return rx.dialog.root(
        rx.dialog.content(
            rx.vstack(
                rx.hstack(
                    rx.spacer(),
                    rx.dialog.close(
                        rx.icon_button(
                            rx.icon("x", size=18),
                            size="1",
                            variant="ghost",
                            color_scheme="gray",
                            cursor="pointer",
                        )
                    ),
                    width="100%",
                ),
                rx.center(
                    rx.vstack(
                        rx.box(
                            rx.icon("construction", size=48, color="#e85d04"),
                            bg="rgba(232, 93, 4, 0.1)",
                            padding="16px",
                            border_radius="full",
                            display="flex",
                            align_items="center",
                            justify_content="center",
                        ),
                        rx.heading("Enlace en Mantenimiento", size="5", color=text_main, font_weight="700"),
                        rx.text(
                            "Estamos mejorando este servicio para brindarte una mejor experiencia. Vuelve a intentarlo más tarde.",
                            font_size="13px",
                            color=text_sec,
                            text_align="center",
                        ),
                        rx.hstack(
                            rx.icon("clock", size=16, color="#e85d04"),
                            rx.text("Tiempo estimado de restauración: 24 horas", font_size="12px", color=text_sec, font_weight="500"),
                            bg=rx.color_mode_cond(light="#fff7ed", dark="rgba(232, 93, 4, 0.05)"),
                            border="1px solid rgba(232, 93, 4, 0.2)",
                            border_radius="8px",
                            padding_x="3",
                            padding_y="2",
                            width="100%",
                            justify_content="center",
                            margin_top="2",
                        ),
                        spacing="3",
                        align_items="center",
                    ),
                    width="100%",
                ),
                spacing="4",
                align_items="stretch",
            ),
            style={
                "max_width": "420px",
                "padding": "24px",
                "border_radius": "20px",
                "background_color": card_bg,
                "border": border,
                "box_shadow": "0 20px 25px -5px rgba(0, 0, 0, 0.1), 0 10px 10px -5px rgba(0, 0, 0, 0.04)"
            }
        ),
        open=State.mantenimiento_modal_abierto,
        on_open_change=State.set_mantenimiento_modal_abierto,
    )


def footer() -> rx.Component:
    header_color = rx.color_mode_cond(light="black", dark="white")
    text_color = rx.color_mode_cond(light="gray.700", dark="gray.400")
    link_color = rx.color_mode_cond(light="blue.600", dark="blue.300")
    bg_footer = rx.color_mode_cond(light="#f7fafc", dark="#111827")
    border_color = rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #2d3748")

    return rx.box(
        rx.container(
        rx.hstack(
            # Columna 1: Información de la Entidad
            rx.vstack(
                rx.heading("Información de la Entidad", size="6", color=header_color),
                rx.text("Sede Principal: Calle 10 # 5-20, Buenaventura, Valle del Cauca", color=text_color),
                rx.text("Código Postal: 760001", color=text_color),
                rx.text("PBX: (+57) 602 XXX XXXX", color=text_color),
                rx.link(
                    "Correo institucional: atencionalciudadano@empresa.gov.co", 
                    href="mailto:atencionalciudadano@empresa.gov.co",
                    color=link_color
                ),
                rx.link(
                    "Ley 1755 de 2015", 
                    href="https://www.funcionpublica.gov.co/eva/gestornormativo/norma.php?i=65334", 
                    target="_blank",
                    color=link_color
                ),
                rx.text(
                    "Horario de atención presencial: Lunes a Viernes, 7:30 a.m. - 12:00 p.m. y 2:00 p.m. - 5:30 p.m.",
                    color=text_color
                ),
                align_items="start",
            ),
            # Columna 2: Servicio al Ciudadano
            rx.vstack(
                rx.heading("Servicio al Ciudadano", size="6", color=header_color),
                rx.link("Radicar solicitud PQRS", href="/solicitudes", color=link_color),
                rx.link("Consultar estado de solicitud", href="/consultar-estado", color=link_color),
                rx.link("Preguntas Frecuentes (FAQ)", href="#", on_click=State.abrir_mantenimiento, color=link_color),
                rx.link("Tiempos de respuesta (Ley 1755 de 2015)", href="#", on_click=State.abrir_mantenimiento, color=link_color),
                rx.link("Notificaciones por aviso y judiciales", href="#", on_click=State.abrir_mantenimiento, color=link_color),
                rx.link("Política de privacidad y protección de datos", href="/politica-privacidad", color=link_color),
                rx.link("Manual de usuario (Enlace 1755)", href="#", on_click=State.abrir_mantenimiento, color=link_color),
                align_items="start",
            ),
            # Columna 3: Contacto Directo y Redes
            rx.vstack(
                rx.heading("Contacto Directo y Redes", size="6", color=header_color),
                rx.text("Recepción de correspondencia física: Lunes a viernes, 8:00 a.m. a 4:00 p.m.", color=text_color),
                rx.text("Línea gratuita nacional: 01 8000 91XXXX", color=text_color),
                rx.hstack(
                    rx.link("Facebook", href="https://facebook.com", target="_blank", color=link_color),
                    rx.link("X/Twitter", href="https://twitter.com", target="_blank", color=link_color),
                    rx.link("YouTube", href="https://youtube.com", target="_blank", color=link_color),
                    rx.link("LinkedIn", href="https://linkedin.com", target="_blank", color=link_color),
                    spacing="4"
                ),
                rx.text("Sistema gestionado por: Enlace 1755 (Versión 1.0)", font_size="sm", color=text_color),
                align_items="start",
            ),
            spacing="9",
            align_items="start"
        ),
        width="100%",
        padding_top="24px",
        padding_bottom="24px",
        bg=bg_footer,
        border_top=border_color,
        justify="center"
    ),
    mantenimiento_modal(),
    width="100%",
)


def brand_footer() -> rx.Component:
    """Franja inferior con logos institucionales (Universidad del Valle y GOV.CO)."""
    return rx.container(
        rx.hstack(
            rx.image(src="/unival_logo.svg", alt="Universidad del Valle", height="48px"),
            rx.spacer(),
            rx.image(src="/govco_logo.svg", alt="Gobierno de Colombia", height="48px"),
            spacing="6",
            align_items="center",
            justify="center"
        ),
        width="100%",
        padding_top="12px",
        padding_bottom="12px",
        bg="white",
        _dark={"bg": "gray.900", "borderColor": "gray.700"},
        border_top="1px solid #e2e8f0"
    )
 
    

def _ldc_reg(light, dark):
    return rx.color_mode_cond(light=light, dark=dark)

ORANGE        = "#e85d04"
ORANGE_LIGHT  = "#fff7ed"
ORANGE_BORDER = "#fdba74"
NAVY          = "#1e3a8a"
BLUE_ACC      = "#2563eb"

PAGE_BG    = _ldc_reg("#f8fafc", "#0b1120")
CARD_BG    = _ldc_reg("#ffffff", "#1e293b")
CARD_BDR   = _ldc_reg("#e2e8f0", "#334155")
TEXT_MAIN  = _ldc_reg("#1e293b", "#f1f5f9")
TEXT_SUB   = _ldc_reg("#64748b", "#94a3b8")
TEXT_LABEL = _ldc_reg("#374151", "#cbd5e1")
INPUT_BG   = _ldc_reg("#ffffff", "#0f172a")
INPUT_BDR  = _ldc_reg("#d1d5db", "#475569")
INPUT_FOC  = BLUE_ACC
SECT_BG    = _ldc_reg("#f0f7ff", "#0f172a")
DIVIDER    = _ldc_reg("#e2e8f0", "#334155")


def _label(texto: str, required: bool = False) -> rx.Component:
    """Label con asterisco naranja si es requerido."""
    return rx.hstack(
        rx.text(texto, font_size="13px", font_weight="600", color=TEXT_LABEL),
        rx.cond(
            required,
            rx.text("*", color=ORANGE, font_size="13px", font_weight="700"),
            rx.box(),
        ),
        spacing="1",
        align_items="center",
        margin_bottom="4px",
    )


def _input_style() -> dict:
    return dict(
        bg=INPUT_BG,
        border=f"1px solid {INPUT_BDR}",
        border_radius="8px",
        color=TEXT_MAIN,
        font_size="14px",
        height="42px",
        padding_x="12px",
        width="100%",
        _placeholder={"color": TEXT_SUB, "font_size": "13px"},
        _focus={
            "outline": "none",
            "border_color": INPUT_FOC,
            "box_shadow": f"0 0 0 3px rgba(37,99,235,0.12)",
        },
    )


def _select_style() -> dict:
    return dict(
        bg=INPUT_BG,
        border=f"1px solid {INPUT_BDR}",
        border_radius="8px",
        color=TEXT_MAIN,
        font_size="14px",
        _focus={"border_color": INPUT_FOC},
    )


def _input_s() -> dict:
    """Compatibilidad: alias de `_input_style` usado en el código."""
    return _input_style()


def _select_s() -> dict:
    """Compatibilidad: alias de `_select_style` usado en el código."""
    return _select_style()


def _field(label: str, component: rx.Component, required: bool = True) -> rx.Component:
    """Campo completo: label + input."""
    return rx.vstack(
        _label(label, required),
        component,
        spacing="0",
        align_items="start",
        width="100%",
    )


def _field_with_check(
    label: str,
    component: rx.Component,
    is_valid,
    required: bool = True,
) -> rx.Component:
    """Campo con ícono de validación verde a la derecha."""
    return rx.vstack(
        _label(label, required),
        rx.hstack(
            component,
            rx.cond(
                is_valid,
                rx.box(
                    rx.icon("circle-check", size=18, color="#16a34a"),
                    flex_shrink="0",
                ),
                rx.box(width="18px"),
            ),
            spacing="2",
            align_items="center",
            width="100%",
        ),
        spacing="0",
        align_items="start",
        width="100%",
    )


def _section_header(icon: str, title: str, subtitle: str = "") -> rx.Component:
    """Encabezado de sección con línea naranja inferior."""
    return rx.box(
        rx.hstack(
            rx.box(
                rx.icon(icon, size=16, color=ORANGE),
                width="32px", height="32px",
                border_radius="8px",
                bg=ORANGE_LIGHT,
                border=f"1px solid {ORANGE_BORDER}",
                display="flex", align_items="center", justify_content="center",
                flex_shrink="0",
            ),
            rx.vstack(
                rx.text(title, font_size="14px", font_weight="700", color=TEXT_MAIN),
                rx.cond(
                    subtitle != "",
                    rx.text(subtitle, font_size="12px", color=TEXT_SUB),
                    rx.box(),
                ),
                spacing="0", align_items="start",
            ),
            spacing="3", align_items="center",
        ),
        border_bottom=f"2px solid {ORANGE}",
        padding_bottom="10px",
        margin_bottom="16px",
        width="100%",
    )


def _section(icon: str, title: str, content: rx.Component, subtitle: str = "") -> rx.Component:
    """Sección con encabezado y contenido en tarjeta estilizada."""
    return rx.box(
        _section_header(icon, title, subtitle),
        rx.box(
            content,
            bg=CARD_BG,
            border=f"1px solid {CARD_BDR}",
            border_radius="14px",
            padding="18px 20px",
            width="100%",
        ),
        spacing="4",
        width="100%",
    )


def _seccion_cuenta() -> rx.Component:
    """Correo + contraseñas."""
    return rx.box(
        _section_header("lock", "Datos de acceso", "Correo y contraseña para ingresar al sistema"),
        rx.grid(
            rx.box(
                _label("Correo Electrónico", required=True),
                rx.hstack(
                    rx.input(
                        placeholder="tu buzón electrónico",
                        type="email",
                        value=State.correo,
                        on_change=State.set_and_validate_correo,
                        on_blur=State.validar_correo_accion,
                        flex="1",
                        **_input_style(),
                    ),
                    rx.cond(
                        State.correo_validado,
                        rx.box(
                            rx.icon("circle-check", size=18, color="#16a34a"),
                            flex_shrink="0",
                        ),
                        rx.box(width="18px"),
                    ),
                    spacing="2", align_items="center", width="100%",
                ),
                rx.cond(
                    State.correo_confirmacion_visible,
                    rx.text(
                        State.correo_confirmacion_mensaje,
                        color=rx.cond(State.correo_validado, "#16a34a", "#dc2626"),
                        font_size="12px",
                        margin_top="4px",
                    ),
                    rx.box(),
                ),
                grid_column="1 / -1",
            ),
            _field_with_check(
                "Confirmar Correo Electrónico",
                rx.input(
                    placeholder="Repite tu correo electrónico",
                    type="email",
                    value=State.confirmar_correo,
                    on_change=State.set_confirmar_correo,
                    on_blur=lambda: State.set_confirmar_correo_match(State.correo == State.confirmar_correo),
                    **_input_style(),
                ),
                State.confirmar_correo_match & (State.confirmar_correo != ""),
            ),
            rx.vstack(
                _label("Contraseña", required=True),
                rx.hstack(
                    rx.input(
                        placeholder="Contraseña",
                        type=rx.cond(State.show_password, "text", "password"),
                        value=State.contraseña,
                        on_change=State.set_contraseña,
                        flex="1",
                        **_input_style(),
                    ),
                    rx.button(
                        rx.cond(
                            State.show_password,
                            rx.icon("eye_off", size=16, color=TEXT_SUB),
                            rx.icon("eye", size=16, color=TEXT_SUB),
                        ),
                        on_click=State.toggle_show_password,
                        variant="ghost", size="1",
                        _hover={"bg": "transparent"},
                        flex_shrink="0",
                    ),
                    spacing="1", align_items="center", width="100%",
                ),
                spacing="0", align_items="start", width="100%",
            ),
            _field(
                "Confirmar Contraseña",
                rx.input(
                    placeholder="Confirmar Contraseña",
                    type=rx.cond(State.show_password, "text", "password"),
                    value=State.confirmar_contraseña,
                    on_change=State.set_confirmar_contraseña,
                    **_input_style(),
                ),
            ),
            template_columns={"base": "1fr", "md": "1fr 1fr"},
            gap="4",
            width="100%",
        ),
        bg=CARD_BG,
        border=f"1px solid {CARD_BDR}",
        border_radius="14px",
        padding="20px 22px",
        width="100%",
    )


def _seccion_identificacion() -> rx.Component:
    """Tipo ID + número + nombres + apellidos + sexo."""
    return rx.box(
        _section_header("id-card", "Información personal", "Datos de identificación del ciudadano"),
        rx.grid(
            _field(
                "Tipo de Identificación",
                rx.select(
                    ["Cédula", "Pasaporte", "Tarjeta de Identidad"],
                    placeholder="Seleccione...",
                    value=State.tipo_identificacion,
                    on_change=State.set_tipo_identificacion,
                    **_select_style(),
                ),
                required=False,
            ),
            _field_with_check(
                "Número de Identificación",
                rx.input(
                    placeholder="Cédula/NIT",
                    value=State.numero_identificacion,
                    on_change=State.set_and_validate_numero_identificacion,
                    **_input_style(),
                ),
                State.numero_identificacion_valid,
            ),
            _field_with_check(
                "Nombres",
                rx.input(
                    placeholder="Nombre o razón social",
                    value=State.nombres,
                    on_change=State.set_and_validate_nombres,
                    **_input_style(),
                ),
                State.nombres_valid,
            ),
            _field_with_check(
                "Apellidos",
                rx.input(
                    placeholder="Apellidos",
                    value=State.apellidos,
                    on_change=State.set_and_validate_apellidos,
                    **_input_style(),
                ),
                State.apellidos_valid,
            ),
            _field(
                "Género",
                rx.select(
                    ["Femenino", "Masculino", "Prefiero no decirlo"],
                    placeholder="Seleccione...",
                    value=State.sexo,
                    on_change=State.set_sexo,
                    **_select_style(),
                ),
                required=False,
            ),
            _field_with_check(
                "Número de Contacto",
                rx.input(
                    placeholder="un teléfono de contacto",
                    value=State.telefono,
                    on_change=State.set_and_validate_telefono,
                    **_input_style(),
                ),
                State.telefono_valid,
            ),
            template_columns={"base": "1fr", "md": "1fr 1fr 1fr"},
            gap="4",
            width="100%",
        ),
        bg=CARD_BG,
        border=f"1px solid {CARD_BDR}",
        border_radius="14px",
        padding="20px 22px",
        width="100%",
    )


def _seccion_ubicacion() -> rx.Component:
    """Dirección + departamento + ciudad."""
    return rx.box(
        _section_header("map-pin", "Ubicación", "Dirección y municipio de residencia"),
        rx.grid(
            rx.box(
                _label("Dirección", required=True),
                rx.input(
                    placeholder="tu dirección de residencia",
                    value=State.direccion,
                    on_change=State.set_direccion,
                    **_input_style(),
                ),
                grid_column="1 / -1",
            ),
            _field_with_check(
                "Departamento",
                rx.select(
                    [
                        "Amazonas","Antioquia","Arauca","Atlántico","Bogotá D.C.",
                        "Bolívar","Boyacá","Caldas","Caquetá","Casanare","Cauca",
                        "Cesar","Chocó","Córdoba","Cundinamarca","Guainía","Guaviare",
                        "Huila","La Guajira","Magdalena","Meta","Nariño",
                        "Norte de Santander","Putumayo","Quindío","Risaralda",
                        "San Andrés y Providencia","Santander","Sucre","Tolima",
                        "Valle del Cauca","Vaupés","Vichada",
                    ],
                    placeholder="Seleccione...",
                    value=State.departamento,
                    on_change=State.set_and_validate_departamento,
                    **_select_style(),
                ),
                State.departamento_valid,
                required=True,
            ),
            _field_with_check(
                "Ciudad",
                rx.select(
                    State.ciudades_disponibles,
                    placeholder="Seleccione...",
                    value=State.ciudad,
                    on_change=State.set_and_validate_ciudad,
                    is_disabled=State.departamento == "",
                    **_select_style(),
                ),
                State.ciudad_valid,
                required=True,
            ),
            template_columns={"base": "1fr", "md": "1fr 1fr"},
            gap="4",
            width="100%",
        ),
        bg=CARD_BG,
        border=f"1px solid {CARD_BDR}",
        border_radius="14px",
        padding="20px 22px",
        width="100%",
    )


def _seccion_diversidad() -> rx.Component:
    """Etnia + característica de ciudadano."""
    return rx.box(
        _section_header("heart-handshake", "Diversidad e inclusión", "Opcional �?? ayuda a mejorar la atención"),
        rx.grid(
            _field(
                "Etnia",
                rx.select(
                    ["Ninguna","Indígena","Afrocolombiano","Raizal","Palenquero","Gitano/a","Otro"],
                    placeholder="Seleccione...",
                    value=State.etnia,
                    on_change=State.set_etnia,
                    **_select_style(),
                ),
                required=False,
            ),
            _field(
                "Características del ciudadano",
                rx.select(
                    [
                        "Ninguna","Habitante de la calle","No brinda información",
                        "Peligro Inminente","Periodistas en ejercicio de su actividad",
                        "Primera Infancia","Veteranos Fuerza Pública",
                        "Víctimas - Conflicto Armado",
                    ],
                    placeholder="Seleccione...",
                    value=State.persona_vulnerable_registro,
                    on_change=State.set_persona_vulnerable_registro,
                    **_select_style(),
                ),
                required=False,
            ),
            template_columns={"base": "1fr", "md": "1fr 1fr"},
            gap="4",
            width="100%",
        ),
        bg=CARD_BG,
        border=f"1px solid {CARD_BDR}",
        border_radius="14px",
        padding="20px 22px",
        width="100%",
    )


def _seccion_autorizaciones() -> rx.Component:
    """Checkboxes de habeas data y notificaciones."""
    return rx.box(
        _section_header("shield-check", "Autorizaciones", "Política de datos y notificaciones"),
        rx.vstack(
            rx.hstack(
                rx.checkbox(
                    is_checked=State.acepta_notificaciones,
                    on_change=State.set_acepta_notificaciones,
                    color_scheme="orange",
                    size="2",
                ),
                rx.text(
                    "Autorizo de manera expresa que me notifiquen o comuniquen al correo "
                    "electrónico aquí suministrado la respuesta a escritos o solicitudes, "
                    "así como cualquier información relacionada con mis trámites.",
                    font_size="13px",
                    color=TEXT_SUB,
                    line_height="1.6",
                ),
                spacing="3",
                align_items="start",
                width="100%",
                padding="14px 16px",
                bg=SECT_BG,
                border_radius="10px",
                border=f"1px solid {_ldc_reg('#bfdbfe','#1e3a5f')}",
            ),
            rx.hstack(
                rx.checkbox(
                    is_checked=State.acepta_politica_datos,
                    on_change=State.preconfirmar_politica,
                    color_scheme="orange",
                    size="2",
                ),
                rx.hstack(
                    rx.text("He leído y acepto la ", font_size="13px", color=TEXT_SUB),
                    rx.link(
                        "Política de Protección de Datos",
                        href="/politica-privacidad",
                        color=ORANGE,
                        font_size="13px",
                        font_weight="600",
                        text_decoration="none",
                        _hover={"text_decoration": "underline"},
                    ),
                    spacing="0",
                    flex_wrap="wrap",
                ),
                spacing="3",
                align_items="center",
                width="100%",
                padding="14px 16px",
                bg=ORANGE_LIGHT,
                border_radius="10px",
                border=f"1px solid {ORANGE_BORDER}",
            ),
            spacing="3",
            width="100%",
        ),
        bg=CARD_BG,
        border=f"1px solid {CARD_BDR}",
        border_radius="14px",
        padding="20px 22px",
        width="100%",
    )


def _modal_politica() -> rx.Component:
    """Modal de confirmación de política de datos."""
    return rx.cond(
        State.modal_politica_visible,
        rx.box(
            rx.box(
                rx.vstack(
                    rx.hstack(
                        rx.box(
                            rx.icon("shield-check", size=20, color=ORANGE),
                            bg=ORANGE_LIGHT,
                            border_radius="10px",
                            width="40px", height="40px",
                            display="flex", align_items="center", justify_content="center",
                        ),
                        rx.vstack(
                            rx.heading("Confirmación de Política", size="4", color=TEXT_MAIN),
                            rx.text("Política de Protección de Datos Personales", font_size="12px", color=TEXT_SUB),
                            spacing="0", align_items="start",
                        ),
                        spacing="3", align_items="center", width="100%",
                    ),
                    rx.divider(),
                    rx.text(
                        "Al aceptar confirmas que has leído y comprendido cómo se usarán "
                        "tus datos personales para gestionar solicitudes PQRS.",
                        font_size="13px", color=TEXT_SUB, line_height="1.6",
                    ),
                    rx.text(
                        "La aceptación es necesaria para continuar con el registro.",
                        font_size="13px", color=TEXT_MAIN, font_weight="600",
                    ),
                    rx.hstack(
                        rx.button(
                            "Cancelar",
                            on_click=State.cancelar_politica,
                            variant="outline", border_radius="8px", flex="1",
                        ),
                        rx.button(
                            "Aceptar y continuar",
                            on_click=State.confirmar_politica,
                            bg=ORANGE, color="white", border_radius="8px", flex="1",
                            _hover={"bg": "#c2410c"},
                        ),
                        spacing="3", width="100%",
                    ),
                    spacing="4", align_items="stretch", width="100%",
                ),
                bg=CARD_BG,
                border=f"2px solid {ORANGE_BORDER}",
                border_radius="18px",
                padding="28px",
                width="100%",
                max_width="480px",
                box_shadow="0 20px 60px rgba(0,0,0,0.2)",
            ),
            position="fixed", inset="0",
            bg="rgba(15,23,42,0.55)",
            display="flex", align_items="center", justify_content="center",
            z_index="1000", padding="24px",
        ),
    )


def registro_funcionario_page() -> rx.Component:
    """Página de registro de funcionario �?? solo para funcionarios autenticados."""
    acceso_denegado = rx.center(
        rx.vstack(
            rx.icon("shield-x", size=48, color="#ef4444"),
            rx.heading("Acceso Denegado", size="7", color="#ef4444"),
            rx.text("Solo funcionarios autenticados pueden registrar nuevos funcionarios.", color=TEXT_SUB),
            rx.link(rx.button("Ir al Login", color_scheme="blue", border_radius="10px"), href="/login"),
            spacing="4", align_items="center",
        ),
        min_height="80vh",
    )

    return rx.cond(
        State.es_autenticada & (State.rol_usuario == "funcionario"),
        rx.box(
            navbar(),
            _modal_politica(),
            rx.center(
                rx.box(
                    rx.vstack(
                        rx.box(
                            rx.hstack(
                                rx.text("Registrar", font_size="2rem", font_weight="800", color=NAVY),
                                rx.text(" Nuevo Funcionario", font_size="2rem", font_weight="800", color=TEXT_MAIN),
                                spacing="0", flex_wrap="wrap",
                            ),
                            rx.text("Crea la cuenta institucional del nuevo funcionario.", font_size="14px", color=TEXT_SUB),
                            border_bottom=f"3px solid {NAVY}",
                            padding_bottom="16px",
                            margin_bottom="4px",
                            width="100%",
                        ),
                        rx.cond(
                            State.error_de_registro != "",
                            rx.hstack(
                                rx.icon("circle-x", size=16, color="#dc2626"),
                                rx.text(State.error_de_registro, font_size="13px", color="#dc2626"),
                                spacing="2", align_items="center",
                                bg="#fef2f2", border="1px solid #fecaca",
                                border_radius="8px", padding="10px 14px", width="100%",
                            ),
                        ),
                        rx.cond(
                            State.succes != "",
                            rx.hstack(
                                rx.icon("circle-check", size=16, color="#16a34a"),
                                rx.text(State.succes, font_size="13px", color="#16a34a"),
                                spacing="2", align_items="center",
                                bg="#f0fdf4", border="1px solid #bbf7d0",
                                border_radius="8px", padding="10px 14px", width="100%",
                            ),
                        ),
                        _seccion_cuenta(),
                        _seccion_identificacion(),
                        _seccion_ubicacion(),
                        _seccion_diversidad(),
                        _seccion_autorizaciones(),
                        rx.button(
                            rx.hstack(
                                rx.icon("user-check", size=16),
                                rx.text("Registrar Funcionario", font_size="15px", font_weight="600"),
                                spacing="2",
                            ),
                            on_click=State.signup_funcionario,
                            width="100%", height="48px",
                            bg=NAVY, color="white",
                            border_radius="10px",
                            _hover={"bg": "#172554"},
                            transition="background 0.15s ease",
                            is_disabled=~(State.acepta_politica_datos & State.acepta_notificaciones),
                        ),
                        spacing="5", align_items="stretch", width="100%",
                    ),
                    width="100%", max_width="900px",
                    padding={"base": "20px 16px", "md": "32px 40px"},
                ),
                width="100%",
            ),
            bg=PAGE_BG, min_height="100vh", width="100%",
        ),
        acceso_denegado,
    )


def registro_page() -> rx.Component:
    return rx.box(
        navbar(),
        _modal_politica(),

        rx.center(
            rx.box(
                rx.vstack(

                    # �??�?? Encabezado institucional �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    rx.box(
                        rx.vstack(
                            rx.hstack(
                                rx.text(
                                    "Regístrate en el Sistema PQRS" ,
                                    font_size={"base": "1.6rem", "md": "2rem"},
                                    font_weight="800",
                                    color=BLUE_ACC,
                                    letter_spacing="-0.02em",
                                ),
                                rx.text(
                                    "" ,
                                    font_size={"base": "1.6rem", "md": "2rem"},
                                    font_weight="800",
                                    color=TEXT_MAIN,
                                    letter_spacing="-0.02em",
                                ),
                                spacing="0", flex_wrap="wrap",
                            ),
                            rx.text(
                                "Completa el formulario para crear tu cuenta de ciudadano.",
                                font_size="14px", color=TEXT_SUB,
                            ),
                            spacing="2", align_items="start",
                        ),
                        border_bottom=f"3px solid {BLUE_ACC}",
                        padding_bottom="16px",
                        margin_bottom="4px",
                        width="100%",
                    ),

                    # �??�?? Mensajes de error / éxito globales �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    rx.cond(
                        State.error_de_registro != "",
                        rx.hstack(
                            rx.icon("circle-x", size=16, color="#dc2626"),
                            rx.text(State.error_de_registro, font_size="13px", color="#dc2626", font_weight="500"),
                            spacing="2", align_items="center",
                            bg="#fef2f2", border="1px solid #fecaca",
                            border_radius="8px", padding="10px 14px", width="100%",
                        ),
                    ),
                    rx.cond(
                        State.succes != "",
                        rx.hstack(
                            rx.icon("circle-check", size=16, color="#16a34a"),
                            rx.text(State.succes, font_size="13px", color="#16a34a", font_weight="500"),
                            spacing="2", align_items="center",
                            bg="#f0fdf4", border="1px solid #bbf7d0",
                            border_radius="8px", padding="10px 14px", width="100%",
                        ),
                    ),

                    # �??�?? Secciones �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    _seccion_cuenta(),
                    _seccion_identificacion(),
                    _seccion_ubicacion(),
                    _seccion_diversidad(),
                    _seccion_autorizaciones(),

                    # �??�?? Botón enviar �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    rx.vstack(
                        rx.button(
                            rx.hstack(
                                rx.icon("user-plus", size=16),
                                rx.text("Crear mi cuenta", font_size="15px", font_weight="600"),
                                spacing="2",
                            ),
                            on_click=State.signup,
                            width="100%",
                            height="48px",
                            bg=BLUE_ACC,
                            color="white",
                            border_radius="10px",
                            cursor="pointer",
                            _hover={"bg": "#c2410c"},
                            _active={"bg": "#9a3412"},
                            transition="background 0.15s ease",
                            is_disabled=~(State.acepta_politica_datos & State.acepta_notificaciones),
                        ),
                        rx.hstack(
                            rx.text("¿Ya tienes una cuenta?", font_size="13px", color=TEXT_SUB),
                            rx.link(
                                "Inicia sesión",
                                href="/login",
                                color=BLUE_ACC,
                                font_size="13px",
                                font_weight="600",
                                text_decoration="none",
                                _hover={"text_decoration": "underline"},
                            ),
                            spacing="2", justify="center", width="100%",
                        ),
                        spacing="3", align_items="stretch", width="100%",
                    ),

                    spacing="5",
                    align_items="stretch",
                    width="100%",
                ),
                width="100%",
                max_width="900px",
                padding={"base": "20px 16px", "md": "32px 40px"},
            ),
            width="100%",
        ),

        bg=PAGE_BG,
        min_height="100vh",
        width="100%",
    )


def change_password_page() -> rx.Component:
    return rx.box(
        navbar(),
        rx.center(
            rx.card(
                rx.vstack(
                    rx.heading("Cambiar Contraseña", size={"base": "5", "md": "7"}, color=rx.color_mode_cond(light="black", dark="white")),
                    rx.input(placeholder="Contraseña actual", type="password", value=State.current_password, on_change=State.set_current_password, width="100%"),
                    rx.input(placeholder="Nueva contraseña", type="password", value=State.new_password, on_change=State.set_new_password, width="100%"),
                    rx.input(placeholder="Confirmar nueva contraseña", type="password", value=State.confirm_new_password, on_change=State.set_confirm_new_password, width="100%"),
                    rx.button("Cambiar contraseña", on_click=State.change_password, color_scheme="blue", width="100%"),
                    rx.text(State.change_pw_message, color="green.500", font_size="sm")
                ),
                p={"base": "4", "md": "8"},
                max_width={"base": "90%", "md": "560px"},
                width="100%",
            ),
            size="3"
        ),
        bg=rx.color_mode_cond(light="#f8fafc", dark="#0f172a")
    )

def login_page() -> rx.Component:
    return rx.box(
        # Orbes de fondo animados
        rx.box(position="fixed", top="-120px", left="-100px", width="420px", height="420px",
               bg="rgba(37,99,235,0.22)", border_radius="full", filter="blur(110px)", z_index="0",
               style={"animation": "loginOrb1 8s ease-in-out infinite alternate"}),
        rx.box(position="fixed", bottom="-100px", right="-80px", width="380px", height="380px",
               bg="rgba(14,165,233,0.18)", border_radius="full", filter="blur(100px)", z_index="0",
               style={"animation": "loginOrb2 10s ease-in-out infinite alternate"}),
        rx.box(position="fixed", top="40%", left="50%", width="300px", height="300px",
               bg="rgba(99,102,241,0.12)", border_radius="full", filter="blur(90px)", z_index="0",
               style={"animation": "loginOrb1 12s ease-in-out infinite alternate-reverse"}),
        # Boton volver al inicio
        rx.box(
            rx.menu.root(
                rx.menu.trigger(
                    rx.box(
                        rx.icon("home", size=18, color="rgba(255,255,255,0.85)"),
                        width="40px", height="40px",
                        border_radius="full",
                        bg="rgba(255,255,255,0.12)",
                        border="1px solid rgba(255,255,255,0.2)",
                        display="flex", align_items="center", justify_content="center",
                        cursor="pointer",
                        transition="all 0.2s ease",
                        _hover={"bg": "rgba(255,255,255,0.22)", "transform": "scale(1.05)"},
                    )
                ),
                rx.menu.content(
                    rx.menu.item("Ir al Inicio", on_click=rx.redirect("/")),
                    rx.menu.item("Ir a Registro", on_click=rx.redirect("/registro")),
                )
            ),
            position="fixed", top="24px", left="24px", z_index="50"
        ),
        toast_notification(),
        rx.toast.provider(position="top-center", close_button=True, offset="20px"),
        rx.center(
            rx.grid(
                # Panel izquierdo - Info
                rx.box(
                    rx.vstack(
                        # Logo / marca
                        rx.hstack(
                            rx.box(
                                rx.icon("shield-check", size=22, color="#38bdf8"),
                                width="44px", height="44px",
                                border_radius="12px",
                                bg="rgba(56,189,248,0.15)",
                                border="1px solid rgba(56,189,248,0.3)",
                                display="flex", align_items="center", justify_content="center",
                            ),
                            rx.vstack(
                                rx.text("Sistema PQRS", font_size="16px", font_weight="800",
                                        color="white", letter_spacing="-0.01em", line_height="1"),
                                rx.text("Plataforma ciudadana", font_size="11px",
                                        color="rgba(255,255,255,0.55)", font_weight="500", line_height="1"),
                                spacing="1", align_items="start",
                            ),
                            spacing="3", align_items="center",
                        ),
                        rx.box(width="100%", height="1px", bg="rgba(255,255,255,0.10)"),
                        rx.heading(
                            "Gestiona tus solicitudes ciudadanas",
                            size="7",
                            color="white",
                            line_height="1.15",
                            letter_spacing="-0.02em",
                            font_weight="800",
                        ),
                        rx.text(
                            "Consulta estados, tiempos de respuesta y trazabilidad con seguridad y claridad.",
                            color="rgba(255,255,255,0.65)",
                            font_size="15px",
                            line_height="1.6",
                        ),
                        rx.image(
                            src="/pqrs.png",
                            alt="Portal PQRS",
                            width="100%",
                            height="180px",
                            object_fit="cover",
                            border_radius="16px",
                            border="1px solid rgba(255,255,255,0.12)",
                            box_shadow="0 20px 40px -12px rgba(0,0,0,0.55)",
                        ),
                        rx.vstack(
                            rx.hstack(
                                rx.box(rx.icon("shield-check", size=15, color="#38bdf8"),
                                       width="28px", height="28px", border_radius="8px",
                                       bg="rgba(56,189,248,0.12)", border="1px solid rgba(56,189,248,0.25)",
                                       display="flex", align_items="center", justify_content="center"),
                                rx.text("Acceso seguro y protegido", color="rgba(255,255,255,0.78)", font_size="13px"),
                                spacing="3", align_items="center",
                            ),
                            rx.hstack(
                                rx.box(rx.icon("timer", size=15, color="#a78bfa"),
                                       width="28px", height="28px", border_radius="8px",
                                       bg="rgba(167,139,250,0.12)", border="1px solid rgba(167,139,250,0.25)",
                                       display="flex", align_items="center", justify_content="center"),
                                rx.text("Seguimiento de tiempos de respuesta", color="rgba(255,255,255,0.78)", font_size="13px"),
                                spacing="3", align_items="center",
                            ),
                            rx.hstack(
                                rx.box(rx.icon("bell-ring", size=15, color="#34d399"),
                                       width="28px", height="28px", border_radius="8px",
                                       bg="rgba(52,211,153,0.12)", border="1px solid rgba(52,211,153,0.25)",
                                       display="flex", align_items="center", justify_content="center"),
                                rx.text("Notificaciones y trazabilidad completa", color="rgba(255,255,255,0.78)", font_size="13px"),
                                spacing="3", align_items="center",
                            ),
                            spacing="3", align_items="start", width="100%",
                        ),
                        rx.spacer(),
                        rx.text("© 2026 Sistema PQRS · Todos los derechos reservados",
                                color="rgba(255,255,255,0.3)", font_size="11px"),
                        spacing="5", align_items="start", height="100%", width="100%",
                    ),
                    bg="rgba(255,255,255,0.04)",
                    border="1px solid rgba(255,255,255,0.10)",
                    border_radius="28px",
                    p={"base": "7", "md": "8"},
                    backdrop_filter="blur(20px)",
                    display={"base": "none", "lg": "flex"},
                    flex_direction="column",
                    height="100%",
                    box_shadow="inset 0 1px 0 rgba(255,255,255,0.08)",
                ),
                # Panel derecho - Formulario
                rx.box(
                    rx.vstack(
                        # Icono de acceso animado
                        rx.center(
                            rx.box(
                                rx.icon("lock", size=30, color="#60a5fa"),
                                width="68px", height="68px",
                                border_radius="20px",
                                bg="rgba(37,99,235,0.18)",
                                border="1.5px solid rgba(96,165,250,0.35)",
                                display="flex", align_items="center", justify_content="center",
                                box_shadow="0 8px 32px -8px rgba(37,99,235,0.5)",
                                style={"animation": "loginLockPulse 3s ease-in-out infinite"},
                            ),
                            width="100%",
                        ),
                        rx.vstack(
                            rx.heading("Bienvenido de vuelta",
                                       size="7", color="white", text_align="center",
                                       font_weight="800", letter_spacing="-0.02em"),
                            rx.text("Ingresa tus credenciales para acceder al sistema",
                                    color="rgba(255,255,255,0.55)", font_size="14px", text_align="center"),
                            spacing="2", width="100%",
                        ),
                        rx.box(width="100%", height="1px", bg="rgba(255,255,255,0.08)"),
                        rx.form(
                            rx.vstack(
                                # Email field
                                rx.vstack(
                                    rx.hstack(
                                        rx.icon("mail", size=13, color="rgba(255,255,255,0.5)"),
                                        rx.text("Correo electrónico", font_weight="600", font_size="13px",
                                                color="rgba(255,255,255,0.75)"),
                                        spacing="2", align_items="center",
                                    ),
                                    rx.input(
                                        placeholder="usuario@empresa.com",
                                        name="correo",
                                        type="email",
                                        value=State.correo,
                                        on_change=State.set_correo,
                                        width="100%",
                                        size="3",
                                        radius="large",
                                        variant="surface",
                                    ),
                                    spacing="2", width="100%", align_items="start",
                                ),
                                # Password field
                                rx.vstack(
                                    rx.hstack(
                                        rx.icon("key-round", size=13, color="rgba(255,255,255,0.5)"),
                                        rx.text("Contraseña", font_weight="600", font_size="13px",
                                                color="rgba(255,255,255,0.75)"),
                                        spacing="2", align_items="center",
                                    ),
                                    rx.hstack(
                                        rx.input(
                                            placeholder="••••••••",
                                            name="contraseña",
                                            type=rx.cond(State.show_password, "text", "password"),
                                            value=State.contraseña,
                                            on_change=State.set_contraseña,
                                            width="100%",
                                            size="3",
                                            radius="large",
                                            variant="surface",
                                        ),
                                        rx.button(
                                            rx.cond(State.show_password,
                                                    rx.icon("eye-off", size=16),
                                                    rx.icon("eye", size=16)),
                                            on_click=State.toggle_show_password,
                                            type="button",
                                            variant="soft",
                                            size="3",
                                            radius="large",
                                            color_scheme="blue",
                                        ),
                                        width="100%", spacing="2",
                                    ),
                                    spacing="2", width="100%", align_items="start",
                                ),
                                # Error
                                rx.cond(
                                    State.error_de_contraseña != "",
                                    rx.hstack(
                                        rx.icon("circle-x", size=15, color="#f87171"),
                                        rx.text(State.error_de_contraseña, color="#fca5a5",
                                                font_size="13px", font_weight="500"),
                                        spacing="2", align_items="center",
                                        p="3",
                                        bg="rgba(239,68,68,0.12)",
                                        border="1px solid rgba(239,68,68,0.28)",
                                        border_radius="12px",
                                        width="100%",
                                    ),
                                    rx.box(),
                                ),
                                # Submit button
                                rx.button(
                                    rx.hstack(
                                        rx.icon("log-in", size=18, color="white"),
                                        rx.text("Iniciar sesión", font_weight="700", font_size="15px", color="white"),
                                        spacing="2", justify="center", align_items="center",
                                    ),
                                    type="submit",
                                    width="100%",
                                    size="4",
                                    radius="large",
                                    style={
                                        "background": "linear-gradient(135deg, #2563eb 0%, #1d4ed8 50%, #1e40af 100%)",
                                        "border": "1px solid rgba(96,165,250,0.3)",
                                        "cursor": "pointer",
                                        "transition": "all 0.25s cubic-bezier(0.4,0,0.2,1)",
                                        "boxShadow": "0 10px 28px -8px rgba(37,99,235,0.65)",
                                        "paddingTop": "14px",
                                        "paddingBottom": "14px",
                                    },
                                    _hover={
                                        "style": {
                                            "background": "linear-gradient(135deg, #3b82f6 0%, #2563eb 50%, #1d4ed8 100%)",
                                            "boxShadow": "0 16px 36px -8px rgba(37,99,235,0.8)",
                                            "transform": "translateY(-2px)",
                                        }
                                    },
                                ),
                                spacing="4", width="100%",
                            ),
                            on_submit=State.login,
                            width="100%",
                        ),
                        # Divider con texto
                        rx.hstack(
                            rx.box(flex="1", height="1px", bg="rgba(255,255,255,0.10)"),
                            rx.text("o", color="rgba(255,255,255,0.35)", font_size="12px"),
                            rx.box(flex="1", height="1px", bg="rgba(255,255,255,0.10)"),
                            width="100%", align_items="center", spacing="3",
                        ),
                        # Register link
                        rx.center(
                            rx.link(
                                rx.hstack(
                                    rx.icon("user-plus", size=14, color="#60a5fa"),
                                    rx.text("¿No tienes cuenta? Regístrate aquí",
                                            font_size="13px", font_weight="600", color="#60a5fa"),
                                    spacing="2", align_items="center",
                                ),
                                href="/registro",
                                text_decoration="none",
                                _hover={"opacity": "0.8"},
                            ),
                            width="100%",
                        ),
                        rx.text("Al continuar aceptas las políticas de uso y tratamiento de datos.",
                                color="rgba(255,255,255,0.3)", font_size="11px", text_align="center"),
                        spacing="5", width="100%", align_items="stretch",
                    ),
                    bg="rgba(255,255,255,0.05)",
                    border="1px solid rgba(255,255,255,0.12)",
                    border_radius="28px",
                    p={"base": "7", "md": "8"},
                    backdrop_filter="blur(24px)",
                    box_shadow="0 32px 64px -24px rgba(0,0,0,0.7), inset 0 1px 0 rgba(255,255,255,0.10)",
                    style={
                        "@keyframes loginLockPulse": {
                            "0%": {"box_shadow": "0 8px 32px -8px rgba(37,99,235,0.5)"},
                            "50%": {"box_shadow": "0 8px 48px -4px rgba(37,99,235,0.8)"},
                            "100%": {"box_shadow": "0 8px 32px -8px rgba(37,99,235,0.5)"},
                        },
                        "@keyframes loginOrb1": {
                            "from": {"transform": "translate(0, 0) scale(1)"},
                            "to": {"transform": "translate(40px, 30px) scale(1.1)"},
                        },
                        "@keyframes loginOrb2": {
                            "from": {"transform": "translate(0, 0) scale(1)"},
                            "to": {"transform": "translate(-30px, -40px) scale(1.08)"},
                        },
                    },
                ),
                columns={"base": "1", "lg": "1fr 1fr"},
                gap="5",
                width="100%",
                max_width={"base": "95%", "sm": "480px", "lg": "960px"},
                z_index="1",
            ),
            width="100%",
            min_height="100vh",
            position="relative",
            overflow="hidden",
            px="4",
            py={"base": "6", "md": "10"},
        ),
        bg="linear-gradient(135deg, #0f172a 0%, #1e1b4b 50%, #0f172a 100%)",
        width="100%",
        min_height="100vh",
        position="relative",
    )


def politica_privacidad_page() -> rx.Component:
    text_color = rx.color_mode_cond(light="#0f172a", dark="#f8fafc")
    subtext = rx.color_mode_cond(light="#475569", dark="#94a3b8")
    card_bg = rx.color_mode_cond(light="rgba(255, 255, 255, 0.9)", dark="rgba(15, 23, 42, 0.7)")
    card_border = rx.color_mode_cond(light="1px solid rgba(203, 213, 225, 0.75)", dark="1px solid rgba(51, 65, 85, 0.65)")

    def section_card(title: str, icon: str, body: rx.Component) -> rx.Component:
        return rx.box(
            rx.vstack(
                rx.hstack(
                    rx.box(
                        rx.icon(icon, size=18, color="#60a5fa"),
                        p="2.5",
                        border_radius="lg",
                        bg=rx.color_mode_cond(light="rgba(219, 234, 254, 0.9)", dark="rgba(37, 99, 235, 0.18)"),
                        border=rx.color_mode_cond(light="1px solid #bfdbfe", dark="1px solid rgba(96, 165, 250, 0.28)"),
                    ),
                    rx.heading(title, size="4", color=text_color, font_weight="bold"),
                    spacing="3",
                    align_items="center",
                    width="100%",
                ),
                body,
                spacing="3",
                width="100%",
                align_items="start",
            ),
            p={"base": "5", "md": "6"},
            bg=card_bg,
            border=card_border,
            border_radius="2xl",
            backdrop_filter="blur(18px)",
            box_shadow=rx.color_mode_cond(light="0 18px 40px -24px rgba(30, 64, 175, 0.4)", dark="0 18px 40px -24px rgba(2, 6, 23, 0.9)"),
            width="100%",
        )

    return rx.box(
        navbar(),
        rx.center(
            rx.box(
                rx.box(position="absolute", top="-140px", left="-120px", width="360px", height="360px", bg="rgba(37, 99, 235, 0.20)", border_radius="full", filter="blur(95px)", z_index="0"),
                rx.box(position="absolute", bottom="-140px", right="-110px", width="360px", height="360px", bg="rgba(14, 165, 233, 0.18)", border_radius="full", filter="blur(95px)", z_index="0"),
                rx.vstack(
                    rx.box(
                        rx.vstack(
                            rx.badge("Transparencia y protección de datos", color_scheme="blue", variant="soft", radius="full"),
                            rx.heading("Política de Privacidad", size="8", color=text_color, font_weight="bold", letter_spacing="-0.02em"),
                            rx.text(
                                "En esta plataforma tratamos tus datos con responsabilidad, transparencia y seguridad. "
                                "Tu información personal se usa únicamente para gestionar solicitudes PQRS y mejorar el servicio.",
                                color=subtext,
                                font_size="md",
                                line_height="1.6",
                            ),
                            rx.hstack(
                                rx.box(
                                    rx.icon("scale", size=18, color="#93c5fd"),
                                    rx.text("Ley 1581 de 2012", font_size="xs", color=subtext, font_weight="medium"),
                                    spacing="2",
                                    align_items="center",
                                ),
                                rx.box(
                                    rx.icon("clipboard-check", size=18, color="#93c5fd"),
                                    rx.text("Uso limitado a PQRS", font_size="xs", color=subtext, font_weight="medium"),
                                    spacing="2",
                                    align_items="center",
                                ),
                                spacing="4",
                                flex_wrap="wrap",
                                width="100%",
                            ),
                            spacing="3",
                            align_items="start",
                        ),
                        p={"base": "5", "md": "7"},
                        bg=rx.color_mode_cond(light="rgba(255, 255, 255, 0.75)", dark="rgba(15, 23, 42, 0.55)"),
                        border=card_border,
                        border_radius="3xl",
                        backdrop_filter="blur(18px)",
                        box_shadow=rx.color_mode_cond(light="0 30px 60px -30px rgba(30, 64, 175, 0.35)", dark="0 30px 60px -20px rgba(2, 6, 23, 0.9)"),
                        width="100%",
                        z_index="1",
                    ),
                    rx.grid(
                        section_card(
                            "Datos recolectados",
                            "database",
                            rx.vstack(
                                rx.text("Recopilamos datos necesarios para la gestión de solicitudes:", color=subtext, font_size="sm"),
                                rx.vstack(
                                    rx.text("�?� Correo electrónico", color=subtext, font_size="sm"),
                                    rx.text("�?� Identificación (tipo y número)", color=subtext, font_size="sm"),
                                    rx.text("�?� Nombres y apellidos", color=subtext, font_size="sm"),
                                    rx.text("�?� Teléfono y ubicación (departamento/ciudad/dirección)", color=subtext, font_size="sm"),
                                    spacing="1",
                                    align_items="start",
                                ),
                                spacing="2",
                                width="100%",
                                align_items="start",
                            ),
                        ),
                        section_card(
                            "Finalidad del tratamiento",
                            "target",
                            rx.vstack(
                                rx.text("Usamos tus datos para:", color=subtext, font_size="sm"),
                                rx.vstack(
                                    rx.text("�?� Contactarte y notificarte sobre el estado de tu PQRS", color=subtext, font_size="sm"),
                                    rx.text("�?� Radicar y administrar la solicitud en el sistema", color=subtext, font_size="sm"),
                                    rx.text("�?� Generar trazabilidad y auditoría del proceso de atención", color=subtext, font_size="sm"),
                                    spacing="1",
                                    align_items="start",
                                ),
                                spacing="2",
                                width="100%",
                                align_items="start",
                            ),
                        ),
                        section_card(
                            "Derechos del titular",
                            "badge-check",
                            rx.vstack(
                                rx.text(
                                    "Puedes solicitar consulta, actualización, corrección o eliminación de tus datos conforme a la normativa vigente.",
                                    color=subtext,
                                    font_size="sm",
                                    line_height="1.6",
                                ),
                                rx.text(
                                    "También puedes revocar la autorización cuando sea procedente.",
                                    color=subtext,
                                    font_size="sm",
                                ),
                                spacing="2",
                                width="100%",
                                align_items="start",
                            ),
                        ),
                        section_card(
                            "Contacto y solicitudes",
                            "mail",
                            rx.vstack(
                                rx.text("Si necesitas ejercer tus derechos o tienes dudas:", color=subtext, font_size="sm"),
                                rx.hstack(
                                    rx.icon("at-sign", size=16, color="#60a5fa"),
                                    rx.text("Soporte: ", font_size="sm", color=subtext),
                                    rx.text("soporte@empresa.com", font_size="sm", color=rx.color_mode_cond(light="#1d4ed8", dark="#93c5fd"), font_weight="semibold"),
                                    spacing="2",
                                    flex_wrap="wrap",
                                ),
                                spacing="2",
                                width="100%",
                                align_items="start",
                            ),
                        ),
                        columns={"base": "1", "lg": "2"},
                        gap="4",
                        width="100%",
                        z_index="1",
                    ),
                    rx.hstack(
                        rx.link(
                            rx.button(
                                rx.icon("arrow-left", size=18),
                                "Volver",
                                size="3",
                                variant="soft",
                                radius="full",
                                color_scheme="gray",
                            ),
                            href="/registro",
                        ),
                        rx.spacer(),
                        rx.link(
                            rx.button(
                                rx.icon("file-text", size=18),
                                "Ir a registro",
                                size="3",
                                radius="full",
                                color_scheme="blue",
                                box_shadow="0 10px 24px -12px rgba(37, 99, 235, 0.65)",
                            ),
                            href="/registro",
                        ),
                        width="100%",
                        z_index="1",
                        flex_wrap="wrap",
                        spacing="3",
                    ),
                    spacing="5",
                    width="100%",
                    z_index="1",
                ),
                width="100%",
                max_width="1100px",
                position="relative",
                px={"base": "4", "md": "8"},
                py={"base": "8", "md": "12"},
            ),
            width="100%",
            min_height="90vh",
        ),
        bg=rx.color_mode_cond(light="#f8fafc", dark="#0f172a"),
        min_height="100vh",
    )



def dashboard() -> rx.Component:
    # Bento Variables
    text_color = rx.color_mode_cond(light="#0f172a", dark="#f8fafc")
    subtext_color = rx.color_mode_cond(light="#64748b", dark="#94a3b8")
    card_bg = rx.color_mode_cond(light="rgba(255, 255, 255, 0.6)", dark="rgba(30, 41, 59, 0.4)")
    card_border = rx.color_mode_cond(light="1px solid rgba(255, 255, 255, 0.8)", dark="1px solid rgba(51, 65, 85, 0.5)")
    accent_bg = rx.color_mode_cond(light="linear-gradient(135deg, #eff6ff 0%, #dbeafe 100%)", dark="linear-gradient(135deg, rgba(59, 130, 246, 0.1) 0%, rgba(30, 64, 175, 0.2) 100%)")

    # Bloque A: Perfil Izquierdo
    bento_profile = rx.box(
        rx.vstack(
            rx.box(
                rx.heading(rx.cond(State.nombres, State.nombres.to_string()[0:1], "C"), size="9", color="#3b82f6"),
                width="80px", height="80px", border_radius="2xl", bg=rx.color_mode_cond(light="white", dark="#0f172a"),
                display="flex", align_items="center", justify_content="center", box_shadow="0 10px 15px -3px rgba(0,0,0,0.1)",
                margin_bottom="4"
            ),
            rx.heading("¡Hola, ", State.nombres, "!", size="7", color=text_color, font_weight="bold"),
            rx.text("Bienvenido a tu panel digital.", color=subtext_color, font_size="sm"),
            
            rx.divider(margin_y="6", opacity="0.5"),
            
            rx.vstack(
                rx.hstack(rx.icon("file-text", size=18, color="#3b82f6"), rx.text("Total Solicitudes", font_weight="medium"), rx.spacer(), rx.heading(State.solicitudes.length(), size="4"), width="100%"),
                spacing="4", width="100%"
            ),
            
            rx.spacer(),
            rx.box(
                rx.icon("shield-check", size=24, color="#10b981", margin_bottom="2"),
                rx.text("Cuenta verificada y segura.", font_size="xs", color=subtext_color),
                bg=rx.color_mode_cond(light="white", dark="rgba(15, 23, 42, 0.5)"), padding="4", border_radius="xl", width="100%"
            ),
            spacing="2", align_items="start", height="100%", width="100%"
        ),
        p="8", bg=card_bg, backdrop_filter="blur(24px)", border=card_border, border_radius="3xl",
        box_shadow="0 25px 50px -12px rgba(0, 0, 0, 0.05)", height="100%", width="100%", grid_row="span 2"
    )

    # Bloque B: Nueva Solicitud (estilo kpi_metric_card)
    _kpi_card_bg = rx.color_mode_cond(light="#ffffff", dark="#0f1e35")
    bento_action = rx.box(
        rx.vstack(
            rx.hstack(
                rx.box(
                    rx.icon("file-plus-2", size=18, color="#3b82f6"),
                    bg=rx.color_mode_cond(light="#eff6ff", dark="rgba(59,130,246,0.18)"),
                    border_radius="10px",
                    width="38px", height="38px",
                    display="flex", align_items="center", justify_content="center",
                    flex_shrink="0",
                ),
                rx.spacer(),
                width="100%"
            ),
            rx.vstack(
                rx.text("Nueva", font_size="2rem", font_weight="800", color=text_color, line_height="1", letter_spacing="-0.03em"),
                rx.text("Crear Solicitud PQRS", font_size="12px", color=subtext_color, font_weight="500"),
                rx.link(
                    rx.button("Crear ahora ", rx.icon("arrow-right", size=14), size="2", color_scheme="blue", radius="full", margin_top="2"),
                    href="/solicitudes"
                ),
                spacing="1", align_items="start"
            ),
            spacing="3", align_items="start", width="100%"
        ),
        bg=_kpi_card_bg,
        border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
        border_top="3px solid #3b82f6",
        border_radius="14px",
        padding="18px 20px",
        box_shadow=rx.color_mode_cond(light="0 1px 3px rgba(0,0,0,0.05)", dark="0 2px 12px rgba(0,0,0,0.3)"),
        transition="transform 0.18s ease, box-shadow 0.18s ease",
        _hover={"transform": "translateY(-3px)", "box_shadow": "0 8px 24px rgba(59,130,246,0.2)"},
        width="100%"
    )

    # Bloque B2: Rastrea Solicitud (estilo kpi_metric_card)
    bento_consulta = rx.box(
        rx.vstack(
            rx.hstack(
                rx.box(
                    rx.icon("search", size=18, color="#8b5cf6"),
                    bg=rx.color_mode_cond(light="#f5f3ff", dark="rgba(139,92,246,0.18)"),
                    border_radius="10px",
                    width="38px", height="38px",
                    display="flex", align_items="center", justify_content="center",
                    flex_shrink="0",
                ),
                rx.spacer(),
                width="100%"
            ),
            rx.vstack(
                rx.text("Rastrea", font_size="2rem", font_weight="800", color=text_color, line_height="1", letter_spacing="-0.03em"),
                rx.text("Consultar estado de tu solicitud", font_size="12px", color=subtext_color, font_weight="500"),
                rx.hstack(
                    rx.input(
                        placeholder="Número de radicado...",
                        value=State.consulta_radicado,
                        on_change=State.set_consulta_radicado,
                        bg=rx.color_mode_cond(light="#f8fafc", dark="#0f172a"),
                        border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                        size="2", radius="full", width="100%"
                    ),
                    rx.button(rx.icon("arrow-right", size=14), on_click=State.consultar_desde_dashboard, color_scheme="purple", size="2", radius="full"),
                    width="100%", margin_top="2", spacing="2"
                ),
                spacing="1", align_items="start"
            ),
            spacing="3", align_items="start", width="100%"
        ),
        bg=_kpi_card_bg,
        border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
        border_top="3px solid #8b5cf6",
        border_radius="14px",
        padding="18px 20px",
        box_shadow=rx.color_mode_cond(light="0 1px 3px rgba(0,0,0,0.05)", dark="0 2px 12px rgba(0,0,0,0.3)"),
        transition="transform 0.18s ease, box-shadow 0.18s ease",
        _hover={"transform": "translateY(-3px)", "box_shadow": "0 8px 24px rgba(139,92,246,0.2)"},
        width="100%"
    )

    # Bloque C: Bandeja de Entrada (Grid Inferior)
    bento_grid = rx.cond(
        State.solicitudes,
        rx.grid(
            rx.foreach(
                State.solicitudes,
                lambda solicitud: rx.box(
                    rx.vstack(
                        rx.hstack(
                            rx.badge(
                                solicitud["tipo_solicitud"],
                                variant="soft",
                                radius="full",
                                size="2",
                                color_scheme=rx.cond(
                                    solicitud["tipo_solicitud"] == "Petición",
                                    "blue",
                                    rx.cond(
                                        solicitud["tipo_solicitud"] == "Queja",
                                        "orange",
                                        rx.cond(solicitud["tipo_solicitud"] == "Reclamo", "red", "green"),
                                    ),
                                ),
                            ),
                            rx.spacer(),
                            rx.badge(
                                solicitud["estado"],
                                variant="soft",
                                radius="full",
                                size="2",
                                font_weight="medium",
                                color_scheme=rx.cond(
                                    solicitud["estado"] == "Radicada",
                                    "blue",
                                    rx.cond(
                                        solicitud["estado"] == "Cerrada",
                                        "green",
                                        "orange",
                                    ),
                                ),
                            ),
                            width="100%",
                            align_items="center",
                        ),
                        rx.box(rx.text(solicitud["radicado"], font_size="xs", font_family="monospace", color="#94a3b8"), rx.text(solicitud["fecha"], font_size="xs", color="#94a3b8"), display="flex", justify_content="space-between", width="100%", margin_bottom="2"),
                        rx.heading(solicitud["asunto"], size="4", color=text_color, margin_bottom="1", line_height="1.3"),
                        rx.text(solicitud["descripcion"], font_size="sm", color=subtext_color, style={"display": "-webkit-box", "WebkitLineClamp": "2", "WebkitBoxOrient": "vertical", "overflow": "hidden"}),
                        # Mostrar el área responsable si está asignada
                        rx.cond(
                            solicitud.get("area_responsable") != "",
                            rx.hstack(
                                rx.icon("building-2", size=12, color="#f59e0b"),
                                rx.text(
                                    solicitud.get("area_responsable", "Sin área"),
                                    font_size="xs",
                                    color="#f59e0b",
                                    font_weight="medium",
                                ),
                                spacing="1",
                                align_items="center",
                                padding_x="2",
                                padding_y="1",
                                bg=rx.color_mode_cond(light="#fffbeb", dark="rgba(245,158,11,0.15)"),
                                border="1px solid rgba(245,158,11,0.3)",
                                border_radius="md",
                                width="fit-content",
                                margin_top="1",
                            ),
                            rx.box(),
                        ),
                        rx.spacer(),
                        rx.cond(solicitud.get("documento_basename"), rx.hstack(rx.icon("paperclip", size=14, color="#3b82f6"), ui_enlace_archivo(solicitud.get("documento_href", "#"), rx.text("Ver adjunto", color="#3b82f6", font_size="xs"), _hover={"text_decoration": "underline"}), spacing="2", align_items="center", bg=rx.color_mode_cond(light="white", dark="rgba(15,23,42,0.5)"), padding_x="3", padding_y="1.5", border_radius="md", width="100%"), rx.box()),
                        spacing="0", align_items="start", width="100%", height="100%"
                    ),
                    p="6", bg=card_bg, border=card_border, border_radius="2xl", width="100%", height="100%", box_shadow="0 4px 6px -1px rgba(0, 0, 0, 0.05)", backdrop_filter="blur(16px)", _hover={"transform": "translateY(-4px)", "border_color": "rgba(59, 130, 246, 0.5)", "box_shadow": "0 20px 25px -5px rgba(0, 0, 0, 0.1)"}, transition="all 0.3s ease"
                )
            ),
            template_columns={"base": "1fr", "lg": "repeat(2, 1fr)"}, gap="6", width="100%"
        ),
        # Empty State
        rx.center(
            rx.vstack(
                rx.box(rx.icon("layers", size=48, color="#cbd5e1"), bg="rgba(255,255,255,0.5)", padding="6", border_radius="full"),
                rx.heading("Tu bandeja está impecable", size="5", color=text_color),
                rx.text("Aún no hay trámites aquí.", color=subtext_color),
                spacing="3", align_items="center"
            ),
            p="12", bg=card_bg, border=card_border, border_radius="3xl", width="100%", height="100%", min_height="300px", backdrop_filter="blur(16px)"
        )
    )

    return rx.cond(
        rx.State.is_hydrated,
        rx.cond(
            State.es_autenticada & (State.rol_usuario == "ciudadano"),
            rx.box(
                navbar(),
                rx.center(
                    # Fondos holográficos abstractos
                    rx.box(position="absolute", top="-10%", left="0%", width="500px", height="500px", bg="rgba(139, 92, 246, 0.1)", border_radius="full", filter="blur(120px)", z_index="0"),
                    rx.box(position="absolute", bottom="-10%", right="0%", width="600px", height="600px", bg="rgba(59, 130, 246, 0.1)", border_radius="full", filter="blur(150px)", z_index="0"),
                    
                    # Bento Grid Maestro
                    rx.grid(
                        bento_profile,
                        rx.vstack(bento_action, bento_consulta, bento_grid, spacing="6", width="100%", height="100%"),
                        template_columns={"base": "1fr", "xl": "300px 1fr"},
                        gap="6", width="100%", max_width="1400px", padding_x={"base": "4", "md": "8"}, padding_y="12", z_index="1"
                    ),
                    width="100%", min_height="90vh", position="relative", overflow="hidden", align_items="start"
                ),
                bg=rx.color_mode_cond(light="#f8fafc", dark="#0f172a"), width="100%", min_height="100vh"
            ),
            rx.box(
                navbar(),
                rx.center(access_denied_widget("Esta página es solo para ciudadanos."), size="3"),
                bg=rx.color_mode_cond(light="#f8fafc", dark="#0f172a"), min_height="100vh"
            )
        ),
        rx.center(rx.spinner(size="3", color="#3b82f6"), height="50vh")
    )



def ui_boton_descargar_adjunto(href, filename, color_scheme: str = "blue") -> rx.Component:
    """Botón que fuerza la descarga del adjunto al disco local."""
    return rx.cond(
        href != "",
        rx.button(
            rx.hstack(
                rx.icon("download", size=14),
                rx.text("Descargar", font_size="xs", font_weight="medium"),
                spacing="1",
                align_items="center",
            ),
            on_click=rx.download(url=href, filename=filename),
            size="2",
            variant="soft",
            color_scheme=color_scheme,
            cursor="pointer",
        ),
        rx.box(),
    )


def ui_enlace_archivo(href, *children, **props) -> rx.Component:
    """Enlace HTML para ver/descargar adjuntos (no usar rx.link: interpreta /uploads como ruta SPA)."""
    return rx.el.a(
        *children,
        href=href,
        target="_blank",
        rel="noopener noreferrer",
        text_decoration="none",
        **props,
    )


def ui_tarjeta_adjunto(doc) -> rx.Component:
    """Muestra un adjunto con vista previa (imagen) y enlace de descarga."""
    return rx.cond(
        doc["basename"] != "",
        rx.vstack(
            rx.cond(
                doc["preview_src"] != "",
                rx.image(
                    src=doc["preview_src"],
                    alt=doc["basename"],
                    width="100%",
                    max_height="120px",
                    object_fit="contain",
                    border_radius="lg",
                    border="1px solid #bfdbfe",
                ),
                rx.box(),
            ),
            rx.cond(
                doc["href"] != "",
                rx.vstack(
                    rx.text(doc["basename"], font_size="xs", color="#64748b"),
                    ui_boton_descargar_adjunto(doc["href"], doc["basename"]),
                    spacing="1",
                    width="100%",
                    align_items="start",
                ),
                rx.text(
                    "El archivo ya no está en el servidor",
                    font_size="xs",
                    color="#ef4444",
                ),
            ),
            spacing="2",
            width="100%",
            align_items="start",
            p="2",
            bg=rx.color_mode_cond(light="#eff6ff", dark="rgba(59,130,246,0.1)"),
            border="1px solid #bfdbfe",
            border_radius="lg",
        ),
        rx.box(),
    )


def ui_fila_adjunto_respuesta_modal(doc) -> rx.Component:
    """Fila de adjunto en el modal de actualizar estado (funcionario)."""
    return rx.hstack(
        rx.cond(
            doc["es_imagen"] & (doc["preview_src"] != ""),
            rx.image(
                src=doc["preview_src"],
                alt=doc["nombre"],
                width="56px",
                height="56px",
                object_fit="cover",
                border_radius="10px",
                flex_shrink="0",
            ),
            rx.icon("file-check-2", size=18, color="#2563eb"),
        ),
        rx.text(
            doc["nombre"],
            font_size="sm",
            color=rx.color_mode_cond(light="#1d4ed8", dark="#93c5fd"),
            font_weight="600",
            max_width="220px",
            overflow="hidden",
            text_overflow="ellipsis",
            white_space="nowrap",
        ),
        rx.button(
            rx.icon("trash-2", size=14),
            type="button",
            size="2",
            variant="soft",
            color_scheme="red",
            on_click=State.quitar_respuesta_documento_por_nombre(doc["nombre"]),
        ),
        spacing="3",
        align_items="center",
        width="100%",
        p="2",
    )


def ui_tarjeta_adjunto_historial(h) -> rx.Component:
    """Adjunto en una entrada de bitácora / historial."""
    return rx.cond(
        h["documento_existe"],
        rx.vstack(
            rx.cond(
                h["documento_preview_src"] != "",
                rx.image(
                    src=h["documento_preview_src"],
                    alt=h["documento_adjunto"],
                    width="100%",
                    max_height="80px",
                    object_fit="contain",
                    border_radius="md",
                    border="1px solid #bfdbfe",
                    margin_top="1",
                ),
                rx.box(),
            ),
            ui_boton_descargar_adjunto(h["documento_href"], h["documento_adjunto"]),
            spacing="1",
            width="100%",
            align_items="start",
        ),
        rx.text(
            "El archivo ya no está en el servidor",
            font_size="2xs",
            color="#ef4444",
            margin_top="1",
        ),
    )


def kpi_metric_card(title: str, value, icon_name: str, accent_color: str, icon_bg: str) -> rx.Component:
    """Tarjeta KPI reutilizable (reportes y dashboard funcionario)."""
    text_main = rx.color_mode_cond(light="#0f172a", dark="#f0f6ff")
    text_sub = rx.color_mode_cond(light="#64748b", dark="#7ea8c9")
    card_bg = rx.color_mode_cond(light="#ffffff", dark="#0f1e35")
    return rx.box(
        rx.vstack(
            rx.hstack(
                rx.box(
                    rx.icon(icon_name, size=18, color=accent_color),
                    bg=icon_bg,
                    border_radius="10px",
                    width="38px",
                    height="38px",
                    display="flex",
                    align_items="center",
                    justify_content="center",
                    flex_shrink="0",
                ),
                rx.spacer(),
                width="100%",
            ),
            rx.vstack(
                rx.text(
                    value,
                    font_size="2rem",
                    font_weight="800",
                    color=text_main,
                    line_height="1",
                    letter_spacing="-0.03em",
                ),
                rx.text(title, font_size="12px", color=text_sub, font_weight="500"),
                spacing="1",
                align_items="start",
            ),
            spacing="3",
            align_items="start",
            width="100%",
        ),
        bg=card_bg,
        border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
        border_top=f"3px solid {accent_color}",
        border_radius="14px",
        padding="18px 20px",
        box_shadow=rx.color_mode_cond(
            light="0 1px 3px rgba(0,0,0,0.05)",
            dark="0 2px 12px rgba(0,0,0,0.3)",
        ),
        transition="transform 0.18s ease, box-shadow 0.18s ease",
        _hover={
            "transform": "translateY(-3px)",
            "box_shadow": f"0 8px 24px {accent_color}28",
        },
        width="100%",
    )


def dashboard_chart_card(title: str, chart_component, card_bg=None, card_border=None, extra_content=None) -> rx.Component:
    """Tarjeta contenedora de gráficas reutilizable en dashboards."""
    bg = card_bg if card_bg is not None else rx.color_mode_cond(light="rgba(255, 255, 255, 0.85)", dark="rgba(15, 23, 42, 0.7)")
    border = card_border if card_border is not None else rx.color_mode_cond(light="1px solid rgba(255, 255, 255, 0.5)", dark="1px solid rgba(51, 65, 85, 0.5)")
    return rx.box(
        rx.vstack(
            rx.heading(
                title,
                size="5",
                color=rx.color_mode_cond(light="#0f172a", dark="#f8fafc"),
                font_weight="bold",
            ),
            chart_component,
            extra_content if extra_content is not None else rx.box(),
            spacing="4",
            align_items="stretch",
            width="100%",
        ),
        p="5",
        width="100%",
        bg=bg,
        backdrop_filter="blur(24px)",
        border=border,
        border_radius="2xl",
        box_shadow="0 14px 28px -16px rgba(15,23,42,0.2)",
    )


def action_icon_button(icon_name: str, text: str, color_scheme: str = "blue", is_disabled: bool = False) -> rx.Component:
    return rx.button(
        rx.hstack(
            rx.icon(icon_name, size=16, transition="transform 0.2s ease"),
            rx.text(
                text,
                class_name="action-text",
                font_weight="600",
                font_size="12px",
                white_space="nowrap",
                opacity="0",
                max_width="0px",
                overflow="hidden",
                transition="all 0.3s cubic-bezier(0.4, 0, 0.2, 1)",
                margin_left="0px"
            ),
            spacing="0",
            align_items="center",
        ),
        variant="soft",
        size="2",
        radius="full",
        color_scheme=color_scheme,
        is_disabled=is_disabled,
        transition="all 0.2s ease",
        _hover={
            "transform": "translateY(-1px)",
            "& > svg": {
                "transform": "scale(1.1)"
            },
            "& .action-text": {
                "opacity": "1",
                "max_width": "120px",
                "margin_left": "6px"
            }
        }
    )

def modal_vencimiento_solicitudes() -> rx.Component:
    """Modal compartido para listar solicitudes por rango/semáforo de vencimiento."""
    return rx.cond(
        State.vencimiento_modal_abierto,
        rx.box(
            rx.vstack(
                rx.box(
                    rx.hstack(
                        rx.box(
                            rx.icon("layers", color="white", size=28),
                            bg="linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%)",
                            p="3",
                            border_radius="2xl",
                            box_shadow="0 10px 25px -5px rgba(59, 130, 246, 0.5)",
                        ),
                        rx.vstack(
                            rx.heading(
                                State.rango_vencimiento_seleccionado,
                                size="6",
                                font_weight="900",
                                color="white",
                            ),
                            rx.hstack(
                                rx.text("Total:", color="rgba(255,255,255,0.9)", font_size="sm", font_weight="bold"),
                                rx.badge(
                                    State.solicitudes_vencimiento_filtradas.length(),
                                    color_scheme="blue",
                                    variant="solid",
                                    radius="full",
                                ),
                                rx.text(
                                    "solicitudes en este rango.",
                                    color="rgba(255,255,255,0.8)",
                                    font_size="sm",
                                    font_weight="medium",
                                ),
                                spacing="2",
                                align_items="center",
                            ),
                            spacing="1",
                        ),
                        spacing="4",
                        align_items="center",
                        width="100%",
                    ),
                    p="8",
                    bg="linear-gradient(135deg, #0f172a 0%, #1e293b 100%)",
                    border_bottom="1px solid rgba(255,255,255,0.05)",
                    border_top_left_radius="3xl",
                    border_top_right_radius="3xl",
                    width="100%",
                ),
                rx.box(
                    rx.cond(
                        State.solicitudes_vencimiento_filtradas,
                        rx.vstack(
                            rx.table.root(
                                rx.table.header(
                                    rx.table.row(
                                        rx.table.column_header_cell("Radicado"),
                                        rx.table.column_header_cell("Tipo"),
                                        rx.table.column_header_cell("Asunto"),
                                        rx.table.column_header_cell("Área Responsable"),
                                        rx.table.column_header_cell("Días Restantes"),
                                        rx.table.column_header_cell("Acción"),
                                    )
                                ),
                                rx.table.body(
                                    rx.foreach(
                                        State.solicitudes_vencimiento_filtradas,
                                        lambda s: rx.table.row(
                                            rx.table.row_header_cell(
                                                rx.badge(s.get("radicado"), color_scheme="blue", variant="surface", radius="full")
                                            ),
                                            rx.table.cell(s.get("tipo_solicitud")),
                                            rx.table.cell(s.get("asunto")),
                                            rx.table.cell(s.get("area_responsable")),
                                            rx.table.cell(
                                                rx.hstack(
                                                    rx.box(
                                                        width="8px",
                                                        height="8px",
                                                        bg=s.get("semaforo_fill"),
                                                        border_radius="full",
                                                    ),
                                                    rx.text(
                                                        s.get("remaining_str", ""),
                                                        font_weight="semibold",
                                                        color=s.get("semaforo_fill"),
                                                    ),
                                                    spacing="2",
                                                    align_items="center",
                                                )
                                            ),
                                            rx.table.cell(
                                                rx.button(
                                                    rx.icon("eye", size=16),
                                                    on_click=lambda: State.abrir_detalle_solicitud(s.get("id")),
                                                    variant="soft",
                                                    color_scheme="blue",
                                                    size="1",
                                                    radius="full",
                                                )
                                            ),
                                        ),
                                    )
                                ),
                                width="100%",
                                variant="ghost",
                            ),
                            spacing="4",
                            width="100%",
                        ),
                        rx.center(
                            rx.vstack(
                                rx.icon("inbox", size=48, color="gray"),
                                rx.text(
                                    "No se encontraron solicitudes en este rango.",
                                    font_weight="semibold",
                                    color="gray",
                                ),
                                spacing="2",
                                padding_y="8",
                            ),
                            width="100%",
                        ),
                    ),
                    p="6",
                    max_height="400px",
                    overflow_y="auto",
                    width="100%",
                    bg=rx.color_mode_cond(light="rgba(255, 255, 255, 0.4)", dark="rgba(15, 23, 42, 0.4)"),
                ),
                rx.box(
                    rx.button(
                        "Cerrar Ventana",
                        on_click=State.cerrar_vencimiento_modal,
                        bg="linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%)",
                        color="white",
                        size="3",
                        width="100%",
                    ),
                    p="5",
                    border_top=rx.color_mode_cond(light="1px solid rgba(0,0,0,0.05)", dark="1px solid rgba(255,255,255,0.05)"),
                    bg=rx.color_mode_cond(light="rgba(248, 250, 252, 0.4)", dark="rgba(11, 17, 32, 0.4)"),
                    border_bottom_left_radius="3xl",
                    border_bottom_right_radius="3xl",
                    width="100%",
                ),
                spacing="0",
            ),
            p="0",
            border=rx.color_mode_cond(light="1px solid rgba(255,255,255,0.6)", dark="1px solid rgba(255,255,255,0.08)"),
            border_radius="3xl",
            bg=rx.color_mode_cond(light="rgba(255, 255, 255, 0.9)", dark="rgba(15, 23, 42, 0.8)"),
            backdrop_filter="blur(20px)",
            width="100%",
            max_width="850px",
            position="fixed",
            top="50%",
            left="50%",
            transform="translate(-50%, -50%)",
            z_index="1000",
            box_shadow=rx.color_mode_cond(
                light="0 25px 50px -12px rgba(0, 0, 0, 0.25)",
                dark="0 25px 50px -12px rgba(0, 0, 0, 0.5)",
            ),
        ),
    )


def funcionario_dashboard() -> rx.Component:
    text_color = rx.color_mode_cond(light="#0f172a", dark="#f8fafc")
    subtext_color = rx.color_mode_cond(light="#64748b", dark="#94a3b8")
    card_bg = rx.color_mode_cond(light="rgba(255, 255, 255, 0.85)", dark="rgba(15, 23, 42, 0.7)")
    card_border = rx.color_mode_cond(light="1px solid rgba(255, 255, 255, 0.5)", dark="1px solid rgba(51, 65, 85, 0.5)")

    return rx.cond(
        rx.State.is_hydrated,
        rx.cond(
            State.es_autenticada & (State.rol_usuario == "funcionario"),
            rx.box(
                navbar(),
                rx.center(
                    # Fondo Atmosférico
                    rx.box(position="absolute", top="-10%", left="-5%", width="600px", height="600px", bg="rgba(59, 130, 246, 0.15)", border_radius="full", filter="blur(150px)", z_index="0"),
                    rx.box(position="absolute", bottom="-10%", right="-5%", width="700px", height="700px", bg="rgba(139, 92, 246, 0.1)", border_radius="full", filter="blur(150px)", z_index="0"),
                    
                    # Contenedor Principal (GRID BENTO BOX)
                    rx.grid(
                        # ==========================================
                        # COLUMNA IZQUIERDA (1/3) - Perfil y Admin
                        # ==========================================
                        rx.vstack(
                            # Perfil de Funcionario
                            rx.box(
                                rx.vstack(
                                    rx.hstack(
                                        rx.box(rx.icon("shield-check", size=32, color="white"), p="3", bg="linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%)", border_radius="2xl", box_shadow="0 10px 15px -3px rgba(59, 130, 246, 0.4)"),
                                        rx.vstack(
                                            rx.heading("Consola Operativa", size="6", color=text_color, font_weight="bold"),
                                            rx.badge("Admin / Funcionario", color_scheme="blue", variant="soft", radius="full"),
                                            spacing="1", align_items="start"
                                        ),
                                        spacing="4", align_items="center"
                                    ),
                                    rx.divider(margin_y="4", opacity="0.3"),
                                    rx.text("Supervisa, gestiona y da respuesta a las solicitudes ciudadanas desde este panel centralizado.", color=subtext_color, font_size="sm"),
                                    align_items="stretch"
                                ),
                                p="6", width="100%", bg=card_bg, backdrop_filter="blur(24px)", border=card_border, border_radius="3xl", box_shadow="0 20px 40px -15px rgba(0,0,0,0.1)"
                            ),
                            
                            # Modulo de Usuarios Registrados
                            rx.box(
                                rx.vstack(
                                    rx.hstack(
                                        rx.icon("users", size=20, color="#3b82f6"),
                                        rx.heading("Directorio", size="4", color=text_color, font_weight="bold"),
                                        rx.spacer(),
                                        rx.badge(State.usuarios_registrados_count, color_scheme="blue", radius="full"),
                                        width="100%", align_items="center"
                                    ),
                                    rx.divider(margin_y="2", opacity="0.3"),
                                    rx.box(
                                        rx.grid(
                                            rx.foreach(
                                                State.usuarios_registrados[:6],
                                                lambda usuario: rx.box(
                                                    rx.hstack(
                                                        rx.avatar(
                                                            fallback=rx.cond(usuario["rol"] == "funcionario", "FN", "CD"),
                                                            size="2",
                                                            radius="full",
                                                            color_scheme=rx.cond(usuario["rol"] == "funcionario", "blue", "gray"),
                                                        ),
                                                        rx.vstack(
                                                            rx.text(
                                                                usuario["email"],
                                                                font_size="xs",
                                                                font_weight="bold",
                                                                color=text_color,
                                                                no_wrap=True,
                                                                overflow="hidden",
                                                                text_overflow="ellipsis",
                                                                max_width="190px",
                                                            ),
                                                            rx.badge(
                                                                rx.cond(usuario["rol"] == "funcionario", rx.icon("shield", size=12), rx.icon("user", size=12)),
                                                                usuario["rol"],
                                                                color_scheme=rx.cond(usuario["rol"] == "funcionario", "blue", "gray"),
                                                                variant="soft",
                                                                radius="full",
                                                                size="1",
                                                            ),
                                                            spacing="1",
                                                            align_items="start",
                                                        ),
                                                        rx.spacer(),
                                                        rx.hstack(
                                                            action_icon_button("mail", "Contactar", color_scheme="blue", is_disabled=True),
                                                            rx.link(
                                                                action_icon_button("arrow-right", "Detalles", color_scheme="gray"),
                                                                href="/usuarios",
                                                            ),
                                                            spacing="2",
                                                            align_items="center",
                                                        ),
                                                        width="100%",
                                                        align_items="center",
                                                    ),
                                                    p="3",
                                                    border_radius="xl",
                                                    bg=rx.color_mode_cond(light="rgba(248, 250, 252, 0.65)", dark="rgba(30, 41, 59, 0.25)"),
                                                    border=rx.color_mode_cond(light="1px solid rgba(226,232,240,0.9)", dark="1px solid rgba(51,65,85,0.6)"),
                                                    box_shadow=rx.color_mode_cond(light="0 8px 16px -12px rgba(15,23,42,0.25)", dark="0 10px 20px -16px rgba(0,0,0,0.65)"),
                                                    _hover={"transform": "translateY(-1px)"},
                                                    transition="all 0.18s",
                                                ),
                                            ),
                                            columns=rx.breakpoints(initial="1", sm="2"),
                                            spacing="3",
                                            width="100%",
                                        ),
                                        width="100%",
                                        max_height="320px",
                                        overflow_y="auto",
                                        pr="1",
                                    ),
                                    rx.cond(
                                        State.usuarios_registrados_count > 5,
                                        rx.link(rx.button("Ver directorio completo", variant="soft", size="2", width="100%", color_scheme="blue", mt="2"), href="/usuarios")
                                    ),
                                    align_items="stretch"
                                ),
                                p="5", width="100%", bg=card_bg, backdrop_filter="blur(24px)", border=card_border, border_radius="3xl", box_shadow="0 20px 40px -15px rgba(0,0,0,0.1)"
                            ),
                            
                            spacing="6", width="100%"
                        ),

                        # ==========================================
                        # COLUMNA DERECHA (2/3) - Gestion Operativa
                        # ==========================================
                        rx.vstack(
                            # Grid de Metricas (4 Tarjetas KPI modernas)
                            rx.grid(
                                kpi_metric_card(
                                    "Total Solicitudes",
                                    State.numero_solicitudes,
                                    "layers",
                                    "#2563eb",
                                    rx.color_mode_cond(light="#eff6ff", dark="rgba(37, 99, 235, 0.18)"),
                                ),
                                kpi_metric_card(
                                    "Radicadas",
                                    State.numero_solicitudes_radicadas,
                                    "file-plus-2",
                                    "#f59e0b",
                                    rx.color_mode_cond(light="#fffbeb", dark="rgba(245, 158, 11, 0.16)"),
                                ),
                                kpi_metric_card(
                                    "En Proceso",
                                    State.numero_solicitudes_actualizadas,
                                    "activity",
                                    "#0ea5e9",
                                    rx.color_mode_cond(light="#f0fdfe", dark="rgba(14, 165, 233, 0.16)"),
                                ),
                                kpi_metric_card(
                                    "Cerradas",
                                    State.numero_solicitudes_cerradas,
                                    "circle-check",
                                    "#10b981",
                                    rx.color_mode_cond(light="#ecfdf5", dark="rgba(16, 185, 129, 0.14)"),
                                ),
                                columns={"base": "1", "sm": "2", "lg": "4"},
                                gap="4",
                                width="100%",
                                style={"gridTemplateColumns": "repeat(auto-fit, minmax(200px, 1fr))"},
                            ),

                            # Gráfica: estado de vencimiento (solo activas, misma lógica que Excel)
                            dashboard_chart_card(
                                "Estado de Vencimiento (Activas)",
                                rx.box(
                                    rc.bar_chart(
                                        rc.x_axis(data_key="name"),
                                        rc.y_axis(domain=[0, State.max_semaforo_activas_cantidad]),
                                        rc.tooltip(),
                                        rc.bar(
                                            rc.cell(fill="#10b981"),
                                            rc.cell(fill="#f59e0b"),
                                            rc.cell(fill="#ef4444"),
                                            data_key="cantidad",
                                            radius=[6, 6, 0, 0],
                                        ),
                                        data=State.semaforo_activas_chart_data,
                                        width="100%",
                                        height=260,
                                    ),
                                    position="relative",
                                    width="100%",
                                ),
                                card_bg=card_bg,
                                card_border=card_border,
                                extra_content=rx.vstack(
                                    rx.hstack(
                                        rx.badge(
                                            rx.hstack(
                                                rx.icon("layers", size=12),
                                                State.semaforo_activas_total,
                                                spacing="1",
                                                align_items="center",
                                            ),
                                            color_scheme="blue",
                                            variant="soft",
                                            radius="full",
                                        ),
                                        
                                        spacing="2",
                                        align_items="center",
                                        flex_wrap="wrap",
                                    ),
                                    rx.hstack(
                                        rx.button(
                                            "Verde (>5 días)",
                                            size="1",
                                            variant="soft",
                                            color_scheme="green",
                                            radius="full",
                                            on_click=lambda: State.abrir_semaforo_activas_modal("Verde"),
                                        ),
                                        rx.button(
                                            "Amarillo (1-5 días)",
                                            size="1",
                                            variant="soft",
                                            color_scheme="orange",
                                            radius="full",
                                            on_click=lambda: State.abrir_semaforo_activas_modal("Amarillo"),
                                        ),
                                        rx.button(
                                            "Rojo (Vencidas)",
                                            size="1",
                                            variant="soft",
                                            color_scheme="red",
                                            radius="full",
                                            on_click=lambda: State.abrir_semaforo_activas_modal("Rojo"),
                                        ),
                                        spacing="2",
                                        flex_wrap="wrap",
                                        justify_content="center",
                                        width="100%",
                                    ),
                                    spacing="2",
                                    width="100%",
                                ),
                            ),
                            
                            # Command Bar (Buscador y Filtros Cristalinos)
                            rx.box(
                                rx.hstack(
                                    rx.hstack(
                                        rx.icon("search", size=20, color=subtext_color),
                                        rx.input(placeholder="Buscar por radicado, asunto...", value=State.query_solicitud, on_change=State.set_query_solicitud, variant="surface", outline="none", border="none", bg="transparent", _focus={"box_shadow": "none"}, width="250px"),
                                        align_items="center", bg=rx.color_mode_cond(light="rgba(255,255,255,0.5)", dark="rgba(15,23,42,0.5)"), p="2", border_radius="xl", flex="1"
                                    ),
                                    rx.vstack(
                                        rx.text("Estado", font_size="xs", color=subtext_color, font_weight="medium"),
                                        rx.select(
                                            ["Todas", "Radicada", "Asignada a Area", "En Gestion de Area", "Solucionada", "Reabierta", "Cerrada"],
                                            value=State.filter_estado_solicitud,
                                            on_change=State.set_filter_estado_solicitud,
                                            size="3",
                                            radius="full",
                                            variant="soft",
                                            width="170px",
                                        ),
                                        spacing="1",
                                        align_items="start",
                                    ),
                                    rx.vstack(
                                        rx.text("Tipo Solicitud", font_size="xs", color=subtext_color, font_weight="medium"),
                                        rx.select(
                                            ["Todas", "Petición", "Queja", "Reclamo", "Sugerencia"],
                                            value=State.filter_tipo_solicitud,
                                            on_change=State.set_filter_tipo_solicitud,
                                            size="3",
                                            radius="full",
                                            variant="soft",
                                            width="170px",
                                        ),
                                        spacing="1",
                                        align_items="start",
                                    ),
                                    rx.vstack(
                                        rx.text("Días Restantes", font_size="xs", color=subtext_color, font_weight="medium"),
                                        rx.select(
                                            ["Todos", "Vencidas", "0-3", "4-10", "11+"],
                                            value=State.filter_dias_restantes,
                                            on_change=State.set_filter_dias_restantes,
                                            size="3",
                                            radius="full",
                                            variant="soft",
                                            width="170px",
                                            placeholder="Días restantes",
                                        ),
                                        spacing="1",
                                        align_items="start",
                                    ),
                                    rx.button(rx.icon("filter", size=18), "Buscar", on_click=State.buscar_solicitudes, size="3", color_scheme="blue", radius="full", box_shadow="0 4px 10px rgba(59, 130, 246, 0.3)"),
                                    spacing="4", width="100%", align_items="center", flex_wrap="wrap"
                                ),
                                p="4", width="100%", bg=card_bg, backdrop_filter="blur(24px)", border=card_border, border_radius="2xl", margin_top="4"
                            ),
                            
                            # Solicitudes activas (las cerradas van a pestaña aparte)
                            rx.vstack(
                                rx.box(
                                    rx.hstack(
                                        rx.heading("Solicitudes activas", size="5", color=text_color, font_weight="bold"),
                                        rx.spacer(),
                                        rx.badge(State.solicitudes_abiertas.length(), color_scheme="blue", variant="soft", radius="full"),
                                        rx.link(
                                            rx.button(
                                                "Ver cerradas",
                                                size="2",
                                                variant="soft",
                                                color_scheme="green",
                                                radius="full",
                                            ),
                                            href="/dashboard-funcionario-cerradas",
                                        ),
                                        width="100%",
                                    ),
                                    width="100%",
                                    mt="2",
                                ),
                                rx.cond(
                                    State.solicitudes_abiertas,
                                    rx.grid(
                                        rx.foreach(
                                            State.solicitudes_abiertas,
                                            lambda solicitud: rx.box(
                                                rx.vstack(
                                                    # Header: Radicado + Badge de estado
                                                    rx.box(
                                                        rx.hstack(
                                                            rx.hstack(
                                                                rx.icon("hash", size=12, color=rx.color_mode_cond(light="#64748b", dark="#7ea8c9")),
                                                                rx.text(
                                                                    solicitud['radicado'],
                                                                    font_size="xs",
                                                                    font_weight="700",
                                                                    color=rx.color_mode_cond(light="#64748b", dark="#7ea8c9"),
                                                                    letter_spacing="0.02em",
                                                                ),
                                                                spacing="1", align_items="center",
                                                            ),
                                                            rx.spacer(),
                                                            rx.badge(
                                                                rx.cond(solicitud['estado'] == 'Radicada', rx.icon("file-plus-2", size=11),
                                                                    rx.cond(solicitud['estado'] == 'Asignada a Area', rx.icon("building-2", size=11),
                                                                        rx.cond(solicitud['estado'] == 'En Gestion de Area', rx.icon("activity", size=11),
                                                                            rx.cond(solicitud['estado'] == 'Reabierta', rx.icon("refresh-cw", size=11),
                                                                                rx.icon("circle-check", size=11))))),
                                                                solicitud['estado'],
                                                                color_scheme=rx.cond(solicitud['estado'] == 'Radicada', "orange",
                                                                    rx.cond(solicitud['estado'] == 'Asignada a Area', "purple",
                                                                        rx.cond(solicitud['estado'] == 'En Gestion de Area', "blue",
                                                                            rx.cond(solicitud['estado'] == 'Reabierta', "red", "green")))),
                                                                variant="soft", size="1", radius="full",
                                                            ),
                                                            width="100%", align_items="center",
                                                        ),
                                                        padding="12px 16px",
                                                        border_bottom=rx.color_mode_cond(light="1px solid #f1f5f9", dark="1px solid #1e3a5f"),
                                                        width="100%",
                                                    ),
                                                    # Cuerpo: Asunto + Descripción
                                                    rx.box(
                                                        rx.vstack(
                                                            rx.text(
                                                                solicitud['asunto'],
                                                                font_size="sm",
                                                                font_weight="700",
                                                                color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff"),
                                                                line_height="1.35",
                                                                style={"display": "-webkit-box", "WebkitLineClamp": "2", "WebkitBoxOrient": "vertical", "overflow": "hidden"},
                                                            ),
                                                            rx.text(
                                                                solicitud['descripcion'],
                                                                font_size="xs",
                                                                color=rx.color_mode_cond(light="#64748b", dark="#7ea8c9"),
                                                                line_height="1.5",
                                                                style={"display": "-webkit-box", "WebkitLineClamp": "3", "WebkitBoxOrient": "vertical", "overflow": "hidden"},
                                                                min_height="54px",
                                                            ),
                                                            spacing="2", align_items="start", width="100%",
                                                        ),
                                                        padding="12px 16px",
                                                        width="100%",
                                                    ),
                                                    # Footer: Etiquetas + Botones
                                                    rx.box(
                                                        rx.vstack(
                                                            rx.hstack(
                                                                rx.badge(
                                                                    rx.icon("layers", size=11),
                                                                    solicitud['tipo_solicitud'],
                                                                    variant="soft", color_scheme="purple", radius="full", size="1",
                                                                ),
                                                                rx.badge(
                                                                    rx.icon("users", size=11),
                                                                    solicitud.get('area_responsable', 'Sin asignar'),
                                                                    variant="soft", color_scheme="cyan", radius="full", size="1",
                                                                ),
                                                                rx.cond(
                                                                    solicitud['estado'] != 'Cerrada',
                                                                    rx.badge(
                                                                        rx.icon("clock", size=11),
                                                                        rx.cond(solicitud['semaforo_expired'], "Vencida", f"{solicitud['semaforo_remaining']} días"),
                                                                        color_scheme=rx.cond(solicitud['semaforo_fill'] == "green", "green",
                                                                            rx.cond(solicitud['semaforo_fill'] == "orange", "orange", "red")),
                                                                        variant="soft", radius="full", size="1",
                                                                    ),
                                                                    rx.box(),
                                                                ),
                                                                spacing="2", width="100%", flex_wrap="wrap",
                                                            ),
                                                            rx.hstack(
                                                                rx.button(
                                                                    rx.icon("history", size=14), "Historial",
                                                                    on_click=lambda _event, id=solicitud['id']: State.abrir_historial(id),
                                                                    size="1", variant="soft", color_scheme="gray", radius="full",
                                                                    _hover={"transform": "translateY(-1px)"}, transition="all 0.2s ease",
                                                                ),
                                                                rx.button(
                                                                    rx.icon("refresh-cw", size=14), "Estado",
                                                                    on_click=lambda _event, id=solicitud['id'], estado=solicitud['estado']: State.abrir_editor_estado(id, estado),
                                                                    size="1", color_scheme="blue", variant="solid", radius="full",
                                                                    _hover={"transform": "translateY(-1px)"}, transition="all 0.2s ease",
                                                                ),
                                                                rx.button(
                                                                    rx.icon("network", size=14), "Área",
                                                                    on_click=lambda _event, id=solicitud['id'], area=solicitud.get('area_responsable', ''): State.abrir_asignar_area(id, area),
                                                                    size="1", color_scheme="green", variant="soft", radius="full",
                                                                    _hover={"transform": "translateY(-1px)"}, transition="all 0.2s ease",
                                                                ),
                                                                spacing="2", width="100%",
                                                            ),
                                                            spacing="3", width="100%",
                                                        ),
                                                        padding="12px 16px",
                                                        border_top=rx.color_mode_cond(light="1px solid #f1f5f9", dark="1px solid #1e3a5f"),
                                                        bg=rx.color_mode_cond(light="#f8fafc", dark="#0b1120"),
                                                        border_bottom_left_radius="14px",
                                                        border_bottom_right_radius="14px",
                                                        width="100%",
                                                    ),
                                                    spacing="0", align_items="start", width="100%",
                                                ),
                                                bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                                border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                                border_top=rx.cond(solicitud['estado'] == 'Radicada', "3px solid #f59e0b",
                                                    rx.cond(solicitud['estado'] == 'Asignada a Area', "3px solid #8b5cf6",
                                                        rx.cond(solicitud['estado'] == 'En Gestion de Area', "3px solid #2563eb",
                                                            rx.cond(solicitud['estado'] == 'Reabierta', "3px solid #ef4444",
                                                                "3px solid #10b981")))),
                                                border_radius="14px",
                                                width="100%",
                                                overflow="hidden",
                                                box_shadow=rx.color_mode_cond(light="0 1px 3px rgba(0,0,0,0.06)", dark="0 2px 12px rgba(0,0,0,0.3)"),
                                                _hover={"transform": "translateY(-4px)", "box_shadow": rx.color_mode_cond(light="0 8px 24px rgba(37,99,235,0.12)", dark="0 8px 24px rgba(0,0,0,0.5)")},
                                                transition="all 0.22s ease",
                                            )
                                        ),
                                        display="grid",
                                        grid_template_columns=rx.breakpoints(initial="repeat(1, minmax(0, 1fr))", md="repeat(2, minmax(0, 1fr))", lg="repeat(3, minmax(0, 1fr))"),
                                        gap="38px",
                                        padding="38px",
                                        width="100%",
                                    ),
                                    rx.center(
                                        rx.vstack(
                                            rx.icon("inbox", size=64, color=subtext_color, opacity="0.3"),
                                            rx.heading("Sin solicitudes activas", size="5", color=text_color),
                                            rx.text("No hay radicados activos en este filtro.", color=subtext_color),
                                            align_items="center", spacing="4"
                                        ),
                                        min_height="220px", width="100%"
                                    )
                                ),
                                spacing="4",
                                width="100%",
                            ),
                            
                            spacing="4", width="100%"
                        ),
                        
                        template_columns={"base": "1fr", "lg": "340px 1fr"}, gap="6", width="100%", max_width="100%", padding_y="8", padding_x={"base": "2", "md": "4"}, z_index="1"
                    ),
                    
                    # ==========================================
                    # MODALES DE ACCION (Pop-ups Glassmorphism)
                    # ==========================================
                    
                    # Modal: Historial
                    rx.cond(
                        State.historial_modal_abierto,
                        rx.dialog.root(
                            rx.dialog.content(
                                rx.dialog.title("Historial de Estados", display="none"),
                                rx.dialog.description("Historial de cambios de estado", display="none"),
                                rx.vstack(
                                    rx.hstack(rx.icon("history", size=24, color="#3b82f6"), rx.heading("Historial de Estados", size="5"), rx.spacer(), rx.dialog.close(rx.button(rx.icon("x", size=20), on_click=State.cerrar_historial, variant="ghost", color_scheme="gray")), width="100%", align_items="center"),
                                    rx.divider(margin_y="2"),
                                    rx.cond(
                                        State.historial_estados,
                                        rx.vstack(
                                            rx.foreach(
                                                State.historial_estados,
                                                lambda evento: rx.box(
                                                    rx.hstack(
                                                        rx.box(rx.icon("git-commit-horizontal", size=20, color="#10b981"), p="2", bg="rgba(16, 185, 129, 0.1)", border_radius="full"),
                                                        rx.vstack(
                                                            rx.hstack(rx.text("Cambió a:", font_size="sm", color=subtext_color), rx.badge(evento['nuevo'], color_scheme="blue", radius="full"), rx.spacer(), rx.text(evento['fecha'], font_size="xs", color=subtext_color), width="100%", align_items="center"),
                                                            rx.text(rx.cond(evento['obs'] != "", f'"{evento["obs"]}"', "Sin comentarios."), font_size="sm", font_style="italic", color=text_color),
                                                            rx.text(f"Estado anterior: {evento['anterior']}", font_size="xs", color=subtext_color),
                                                            rx.cond(
                                                                evento["documento_adjunto"] != "",
                                                                ui_tarjeta_adjunto_historial(evento),
                                                                rx.box(),
                                                            ),
                                                            spacing="1", width="100%", align_items="start"
                                                        ),
                                                        spacing="4", align_items="start", width="100%"
                                                    ),
                                                    p="4", bg=rx.color_mode_cond(light="#f8fafc", dark="#1e293b"), border_radius="xl", width="100%", margin_bottom="3"
                                                )
                                            ),
                                            width="100%", max_height="400px", overflow_y="auto"
                                        ),
                                        rx.center(rx.text("Aún no hay historial de cambios para esta solicitud.", color=subtext_color), p="6")
                                    ),
                                    spacing="4", width="100%"
                                ),
                                style={"maxWidth": "600px", "borderRadius": "24px", "padding": "24px", "backgroundColor": card_bg, "backdropFilter": "blur(40px)", "border": card_border}
                            ),
                            open=True, on_open_change=State.cerrar_historial
                        )
                    ),
                    
                    # Modal: Editor de Estado (estilo KPI)
                    rx.cond(
                        State.editar_estado_id,
                        rx.fragment(
                            rx.box(
                                position="fixed",
                                top="0",
                                left="0",
                                width="100vw",
                                height="100vh",
                                bg=rx.color_mode_cond(light="rgba(15, 23, 42, 0.45)", dark="rgba(0, 0, 0, 0.65)"),
                                z_index="999",
                                on_click=State.cerrar_editor_estado,
                            ),
                            rx.box(
                                rx.vstack(
                                    rx.box(
                                        rx.hstack(
                                            rx.box(
                                                rx.icon("sliders-horizontal", size=18, color="#2563eb"),
                                                bg=rx.color_mode_cond(light="#eff6ff", dark="rgba(37, 99, 235, 0.18)"),
                                                border_radius="10px",
                                                width="42px",
                                                height="42px",
                                                display="flex",
                                                align_items="center",
                                                justify_content="center",
                                                flex_shrink="0",
                                            ),
                                            rx.vstack(
                                                rx.text(
                                                    "Actualizar estado",
                                                    font_size={"base": "xl", "md": "2rem"},
                                                    font_weight="800",
                                                    color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff"),
                                                    line_height="1",
                                                    letter_spacing="-0.03em",
                                                ),
                                                rx.text(
                                                    "Gestión y seguimiento de la solicitud",
                                                    font_size="12px",
                                                    color=rx.color_mode_cond(light="#64748b", dark="#7ea8c9"),
                                                    font_weight="500",
                                                ),
                                                spacing="1",
                                                align_items="start",
                                            ),
                                            rx.spacer(),
                                            rx.button(
                                                rx.icon("x", size=18),
                                                type="button",
                                                on_click=State.cerrar_editor_estado,
                                                variant="ghost",
                                                color_scheme="gray",
                                                size="2",
                                            ),
                                            spacing="3",
                                            align_items="center",
                                            width="100%",
                                        ),
                                        padding="18px 20px",
                                        border_bottom=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                        width="100%",
                                    ),
                                    rx.box(
                                        rx.vstack(
                                            rx.box(
                                                rx.vstack(
                                                    rx.text(
                                                        "Nuevo estado",
                                                        font_weight="600",
                                                        font_size="sm",
                                                        color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff"),
                                                    ),
                                                    rx.select(
                                                        ["Radicada", "Asignada a Area", "En Gestion de Area", "Solucionada", "Reabierta", "Cerrada"],
                                                        value=State.nuevo_estado,
                                                        on_change=State.set_nuevo_estado,
                                                        required=True,
                                                        width="100%",
                                                        size="3",
                                                    ),
                                                    rx.cond(
                                                        State.nuevo_estado == "Asignada a Area",
                                                        rx.vstack(
                                                            rx.hstack(
                                                                rx.text("Área responsable", font_weight="600", font_size="sm", color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff")),
                                                                rx.badge("Requerido", color_scheme="orange", size="1"),
                                                                spacing="2",
                                                                align_items="center",
                                                            ),
                                                            rx.select(
                                                                State.AREAS_DISPONIBLES,
                                                                placeholder="Selecciona el área...",
                                                                value=State.area_para_asignar,
                                                                on_change=State.set_area_para_asignar,
                                                                width="100%",
                                                                size="3",
                                                                color_scheme="orange",
                                                            ),
                                                            spacing="2",
                                                            width="100%",
                                                            align_items="start",
                                                        ),
                                                    ),
                                                    spacing="3",
                                                    width="100%",
                                                    align_items="stretch",
                                                ),
                                                padding="16px 18px",
                                                border_radius="14px",
                                                bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                                border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                                border_top="3px solid #2563eb",
                                                width="100%",
                                            ),
                                            rx.box(
                                                rx.vstack(
                                                    rx.hstack(
                                                        rx.text("Justificación", font_weight="600", font_size="sm", color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff")),
                                                        rx.badge("Obligatorio", color_scheme="red", size="1"),
                                                        spacing="2",
                                                    ),
                                                    rx.text_area(
                                                        placeholder="Describe el motivo del cambio de estado...",
                                                        value=State.respuesta_solicitud,
                                                        on_change=State.set_respuesta_solicitud,
                                                        rows="3",
                                                        required=True,
                                                        width="100%",
                                                        size="3",
                                                        min_height="88px",
                                                    ),
                                                    spacing="2",
                                                    width="100%",
                                                    align_items="start",
                                                ),
                                                padding="16px 18px",
                                                border_radius="14px",
                                                bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                                border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                                border_top="3px solid #10b981",
                                                width="100%",
                                            ),
                                            rx.box(
                                                rx.vstack(
                                                    rx.text(
                                                        "Documento adjunto (opcional)",
                                                        font_weight="600",
                                                        font_size="sm",
                                                        color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff"),
                                                    ),
                                                    rx.box(
                                                        rx.vstack(
                                                            rx.upload(
                                                                rx.hstack(
                                                                    rx.icon("paperclip", size=16, color=rx.color_mode_cond(light="#64748b", dark="#7ea8c9")),
                                                                    rx.text("PDF, JPG o PNG �?? hasta 3 archivos, máx. 10 MB c/u", font_size="sm", color=rx.color_mode_cond(light="#64748b", dark="#7ea8c9")),
                                                                    spacing="2",
                                                                    align_items="center",
                                                                    justify="center",
                                                                ),
                                                                id="upload_respuesta",
                                                                multiple=True,
                                                                accept={
                                                                    "application/pdf": [".pdf"],
                                                                    "image/png": [".png"],
                                                                    "image/jpeg": [".jpg", ".jpeg"],
                                                                },
                                                                border="none",
                                                                width="100%",
                                                                padding="3",
                                                                on_drop=State.set_respuesta_documento,
                                                            ),
                                                            rx.cond(
                                                                State.respuesta_documento_error != "",
                                                                rx.box(
                                                                    rx.hstack(
                                                                        rx.icon("ban", size=15, color="#ef4444"),
                                                                        rx.text(
                                                                            State.respuesta_documento_error,
                                                                            font_size="sm",
                                                                            color="#ef4444",
                                                                            font_weight="medium",
                                                                        ),
                                                                        spacing="2",
                                                                        align_items="center",
                                                                    ),
                                                                    bg=rx.color_mode_cond(light="#fef2f2", dark="#450a0a"),
                                                                    border=rx.color_mode_cond(light="1px solid #fecaca", dark="1px solid #991b1b"),
                                                                    border_radius="10px",
                                                                    p="3",
                                                                    width="100%",
                                                                ),
                                                            ),
                                                            rx.cond(
                                                                State.respuesta_documentos.length() > 0,
                                                                rx.vstack(
                                                                    rx.foreach(
                                                                        State.respuesta_documentos,
                                                                        ui_fila_adjunto_respuesta_modal,
                                                                    ),
                                                                    rx.button(
                                                                        rx.icon("trash-2", size=14),
                                                                        "Quitar todos los adjuntos",
                                                                        type="button",
                                                                        size="2",
                                                                        variant="outline",
                                                                        color_scheme="red",
                                                                        on_click=State.quitar_todos_respuesta_documentos,
                                                                        width="100%",
                                                                    ),
                                                                    spacing="2",
                                                                    width="100%",
                                                                ),
                                                            ),
                                                            spacing="2",
                                                            width="100%",
                                                        ),
                                                        padding="3",
                                                        border=rx.color_mode_cond(light="2px dashed #cbd5e1", dark="2px dashed #334155"),
                                                        border_radius="12px",
                                                        bg=rx.color_mode_cond(light="#f8fafc", dark="#0b1528"),
                                                        width="100%",
                                                    ),
                                                    spacing="2",
                                                    width="100%",
                                                    align_items="start",
                                                ),
                                                padding="16px 18px",
                                                border_radius="14px",
                                                bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                                border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                                border_top="3px solid #8b5cf6",
                                                width="100%",
                                            ),
                                            rx.cond(
                                                State.mensaje_actualizar_estado,
                                                rx.box(
                                                    rx.text(
                                                        State.mensaje_actualizar_estado,
                                                        color=rx.cond(
                                                            State.mensaje_actualizar_estado.contains("correctamente"),
                                                            "green.700",
                                                            "red.700",
                                                        ),
                                                        font_weight="semibold",
                                                        font_size="sm",
                                                    ),
                                                    bg=rx.cond(
                                                        State.mensaje_actualizar_estado.contains("correctamente"),
                                                        rx.color_mode_cond(light="#d1fae5", dark="#064e3b"),
                                                        rx.color_mode_cond(light="#fee2e2", dark="#450a0a"),
                                                    ),
                                                    border=rx.cond(
                                                        State.mensaje_actualizar_estado.contains("correctamente"),
                                                        rx.color_mode_cond(light="1px solid #a7f3d0", dark="1px solid #065f46"),
                                                        rx.color_mode_cond(light="1px solid #fecaca", dark="1px solid #991b1b"),
                                                    ),
                                                    border_radius="10px",
                                                    p="3",
                                                    width="100%",
                                                ),
                                            ),
                                            spacing="4",
                                            align_items="stretch",
                                            padding="18px 20px",
                                            max_height="65vh",
                                            overflow_y="auto",
                                            width="100%",
                                        ),
                                        width="100%",
                                    ),
                                    rx.box(
                                        rx.hstack(
                                            rx.button(
                                                "Cancelar",
                                                type="button",
                                                on_click=State.cerrar_editor_estado,
                                                variant="soft",
                                                color_scheme="gray",
                                                size="3",
                                                flex="1",
                                            ),
                                            rx.button(
                                                rx.icon("check", size=16),
                                                "Actualizar estado",
                                                type="button",
                                                on_click=State.actualizar_estado_solicitud(
                                                    rx.upload_files(upload_id="upload_respuesta"),
                                                ),
                                                color_scheme="blue",
                                                size="3",
                                                flex="2",
                                            ),
                                            spacing="3",
                                            width="100%",
                                        ),
                                        padding="16px 20px",
                                        border_top=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                        bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                        border_bottom_left_radius="14px",
                                        border_bottom_right_radius="14px",
                                        width="100%",
                                    ),
                                    spacing="0",
                                    width="100%",
                                ),
                                bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                border_top="3px solid #2563eb",
                                border_radius="14px",
                                width="100%",
                                max_width="560px",
                                max_height="92vh",
                                overflow="hidden",
                                position="fixed",
                                top="50%",
                                left="50%",
                                transform="translate(-50%, -50%)",
                                z_index="1000",
                                box_shadow=rx.color_mode_cond(
                                    light="0 8px 24px rgba(37, 99, 235, 0.16)",
                                    dark="0 2px 12px rgba(0,0,0,0.3)",
                                ),
                            ),
                        ),
                    ),
                    
                    # Modal: Asignar Área
                    rx.cond(
                        State.asignar_area_id,
                        rx.fragment(
                            rx.box(
                                position="fixed", top="0", left="0", width="100vw", height="100vh",
                                bg=rx.color_mode_cond(light="rgba(15, 23, 42, 0.45)", dark="rgba(0, 0, 0, 0.65)"),
                                z_index="999",
                                on_click=State.cerrar_asignar_area,
                            ),
                            rx.box(
                                rx.vstack(
                                    # Header
                                    rx.box(
                                        rx.hstack(
                                            rx.box(
                                                rx.icon("network", size=18, color="#10b981"),
                                                bg=rx.color_mode_cond(light="#ecfdf5", dark="rgba(16, 185, 129, 0.18)"),
                                                border_radius="10px",
                                                width="42px", height="42px",
                                                display="flex", align_items="center",
                                                justify_content="center", flex_shrink="0",
                                            ),
                                            rx.vstack(
                                                rx.text("Asignar Área", font_size={"base": "xl", "md": "2rem"}, font_weight="800", color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff"), line_height="1", letter_spacing="-0.03em"),
                                                rx.text("Derivación y direccionamiento de la solicitud", font_size="12px", color=rx.color_mode_cond(light="#64748b", dark="#7ea8c9"), font_weight="500"),
                                                spacing="1", align_items="start",
                                            ),
                                            rx.spacer(),
                                            rx.button(rx.icon("x", size=18), type="button", on_click=State.cerrar_asignar_area, variant="ghost", color_scheme="gray", size="2"),
                                            spacing="3", align_items="center", width="100%",
                                        ),
                                        padding="18px 20px",
                                        border_bottom=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                        width="100%",
                                    ),
                                    # Body
                                    rx.box(
                                        rx.vstack(
                                            # Sección área
                                            rx.box(
                                                rx.vstack(
                                                    rx.text("Área responsable", font_weight="600", font_size="sm", color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff")),
                                                    rx.select(
                                                        ["Secretaría", "Contabilidad", "Bienestar", "Tesorería", "Atención al Ciudadano", "Otros"],
                                                        placeholder="Selecciona el área...",
                                                        value=State.asignar_area_seleccionada,
                                                        on_change=State.set_asignar_area_seleccionada,
                                                        width="100%", size="3",
                                                    ),
                                                    rx.cond(
                                                        State.asignar_area_seleccionada == "Otros",
                                                        rx.vstack(
                                                            rx.text("Especifique otra área", font_weight="600", font_size="sm", color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff")),
                                                            rx.input(placeholder="Escribe el nombre del área...", value=State.asignar_area_nombre, on_change=State.set_asignar_area_nombre, width="100%", size="3"),
                                                            spacing="2", width="100%", align_items="start",
                                                        ),
                                                    ),
                                                    spacing="3", width="100%", align_items="stretch",
                                                ),
                                                padding="16px 18px", border_radius="14px",
                                                bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                                border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                                border_top="3px solid #10b981", width="100%",
                                            ),
                                            # Sección mensaje
                                            rx.box(
                                                rx.vstack(
                                                    rx.hstack(
                                                        rx.text("Mensaje para el ciudadano", font_weight="600", font_size="sm", color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff")),
                                                        rx.badge("Obligatorio", color_scheme="red", size="1"),
                                                        spacing="2",
                                                    ),
                                                    rx.text_area(
                                                        placeholder="Escribe la comunicación que se notificará al ciudadano...",
                                                        value=State.asignar_area_mensaje,
                                                        on_change=State.set_asignar_area_mensaje,
                                                        rows="3", required=True, width="100%", size="3", min_height="88px",
                                                    ),
                                                    spacing="2", width="100%", align_items="start",
                                                ),
                                                padding="16px 18px", border_radius="14px",
                                                bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                                border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                                border_top="3px solid #06b6d4", width="100%",
                                            ),
                                            # Mensaje éxito/error
                                            rx.cond(
                                                State.mensaje_asignacion,
                                                rx.box(
                                                    rx.text(State.mensaje_asignacion, color=rx.cond(State.mensaje_asignacion.contains("correctamente"), "green.700", "red.700"), font_weight="semibold", font_size="sm"),
                                                    bg=rx.cond(State.mensaje_asignacion.contains("correctamente"), rx.color_mode_cond(light="#d1fae5", dark="#064e3b"), rx.color_mode_cond(light="#fee2e2", dark="#450a0a")),
                                                    border=rx.cond(State.mensaje_asignacion.contains("correctamente"), rx.color_mode_cond(light="1px solid #a7f3d0", dark="1px solid #065f46"), rx.color_mode_cond(light="1px solid #fecaca", dark="1px solid #991b1b")),
                                                    border_radius="10px", p="3", width="100%",
                                                ),
                                            ),
                                            spacing="4", align_items="stretch",
                                            padding="18px 20px", max_height="65vh", overflow_y="auto", width="100%",
                                        ),
                                        width="100%",
                                    ),
                                    # Footer
                                    rx.box(
                                        rx.hstack(
                                            rx.button("Cancelar", type="button", on_click=State.cerrar_asignar_area, variant="soft", color_scheme="gray", size="3", flex="1"),
                                            rx.button(rx.icon("check", size=16), "Asignar y Enviar", type="button", on_click=State.asignar_area_con_mensaje, color_scheme="green", size="3", flex="2"),
                                            spacing="3", width="100%",
                                        ),
                                        padding="16px 20px",
                                        border_top=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                        bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                        border_bottom_left_radius="14px", border_bottom_right_radius="14px", width="100%",
                                    ),
                                    spacing="0", width="100%",
                                ),
                                bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                border_top="3px solid #10b981",
                                border_radius="14px",
                                width="100%", max_width="560px", max_height="92vh", overflow="hidden",
                                position="fixed", top="50%", left="50%", transform="translate(-50%, -50%)",
                                z_index="1000",
                                box_shadow=rx.color_mode_cond(light="0 8px 24px rgba(16, 185, 129, 0.16)", dark="0 2px 12px rgba(0,0,0,0.3)"),
                            ),
                        )
                    ),

                    modal_vencimiento_solicitudes(),
                    
                    padding_x={"base": "4", "md": "8"}, width="100%", position="relative", min_height="90vh", align_items="start"
                ),
                bg=rx.color_mode_cond(light="#f1f5f9", dark="#020617"), width="100%", min_height="100vh", style={"margin": "0"}
            ),
            access_denied_widget("Esta página es solo para funcionarios autorizados.")
        ),
        rx.center(rx.spinner(size="3", color="#3b82f6"), height="100vh", bg=rx.color_mode_cond(light="#f1f5f9", dark="#0f172a"))
    )


def funcionario_cerradas_dashboard() -> rx.Component:
    text_color = rx.color_mode_cond(light="#0f172a", dark="#f8fafc")
    subtext_color = rx.color_mode_cond(light="#64748b", dark="#94a3b8")
    card_bg = rx.color_mode_cond(light="rgba(255, 255, 255, 0.85)", dark="rgba(15, 23, 42, 0.7)")
    card_border = rx.color_mode_cond(light="1px solid rgba(255, 255, 255, 0.5)", dark="1px solid rgba(51, 65, 85, 0.5)")

    return rx.cond(
        rx.State.is_hydrated,
        rx.cond(
            State.es_autenticada & (State.rol_usuario == "funcionario"),
            rx.box(
                navbar(),
                rx.box(
                    rx.vstack(
                        rx.hstack(
                            rx.vstack(
                                rx.heading("Solicitudes cerradas", size="7", color=text_color, font_weight="bold"),
                                rx.text("Histórico de radicados finalizados.", color=subtext_color),
                                align_items="start",
                                spacing="1",
                            ),
                            rx.spacer(),
                            rx.link(
                                rx.button("Volver a activas", variant="soft", color_scheme="blue", radius="full"),
                                href="/dashboard-funcionario",
                            ),
                            width="100%",
                            align_items="center",
                        ),
                        rx.divider(margin_y="4", opacity="0.3"),
                        rx.cond(
                            State.solicitudes_cerradas_lista,
                            rx.grid(
                                rx.foreach(
                                    State.solicitudes_cerradas_lista,
                                    lambda solicitud: rx.box(
                                        rx.vstack(
                                                    rx.box(
                                                        rx.hstack(
                                                            rx.badge(solicitud["radicado"], color_scheme="gray", variant="solid", radius="large"),
                                                            rx.spacer(),
                                                            rx.cond(
                                                                solicitud.get("calificacion_servicio") != None,
                                                                rx.hstack(
                                                                    rx.icon("star", color="#f59e0b", size=14),
                                                                    rx.text(solicitud.get("calificacion_servicio", ""), font_weight="bold", color="#f59e0b", font_size="sm"),
                                                                    spacing="1", align_items="center", bg=rx.color_mode_cond(light="rgba(245, 158, 11, 0.15)", dark="rgba(245, 158, 11, 0.2)"), padding="2px 8px", border_radius="full"
                                                                )
                                                            ),
                                                            rx.badge("Cerrada", color_scheme="green", variant="soft", size="2", radius="full"),
                                                            width="100%",
                                                            align_items="center",
                                                        ),
                                                        p="4",
                                                        width="100%",
                                                        bg=rx.color_mode_cond(
                                                            light="linear-gradient(135deg, #ecfdf5 0%, #d1fae5 100%)",
                                                            dark="linear-gradient(135deg, rgba(16,185,129,0.20) 0%, rgba(16,185,129,0.08) 100%)",
                                                        ),
                                                        border_bottom=rx.color_mode_cond(light="1px solid #a7f3d0", dark="1px solid #334155"),
                                                        border_top_left_radius="2xl",
                                                        border_top_right_radius="2xl",
                                            ),
                                                    rx.vstack(
                                                        rx.heading(solicitud["asunto"], size="5", color=text_color, font_weight="bold", line_height="1.25", style={"display": "-webkit-box", "WebkitLineClamp": "2", "WebkitBoxOrient": "vertical", "overflow": "hidden"}),
                                                        rx.text(solicitud["descripcion"], font_size="sm", color=subtext_color, style={"display": "-webkit-box", "WebkitLineClamp": "3", "WebkitBoxOrient": "vertical", "overflow": "hidden"}, min_height="62px"),
                                                        rx.hstack(
                                                            rx.badge(solicitud["tipo_solicitud"], variant="soft", color_scheme="purple", radius="full"),
                                                            rx.badge(solicitud.get("area_responsable", "No asignada"), variant="soft", color_scheme="cyan", radius="full"),
                                                            spacing="2",
                                                            width="100%",
                                                            flex_wrap="wrap",
                                                        ),
                                                        rx.spacer(),
                                                        rx.hstack(
                                                            rx.button(rx.icon("history", size=16), "Historial", on_click=lambda _event, id=solicitud["id"]: State.abrir_historial(id), size="2", variant="soft", color_scheme="gray", radius="full", _hover={"transform": "translateY(-1px)"}, transition="all 0.2s ease"),
                                                            rx.spacer(),
                                                            rx.button("Reabrir / Estado", on_click=lambda _event, id=solicitud["id"], estado=solicitud["estado"]: State.abrir_editor_estado(id, estado), size="2", color_scheme="blue", variant="solid", radius="full", _hover={"transform": "translateY(-1px)", "box_shadow": "0 10px 20px -14px rgba(59,130,246,0.6)"}, transition="all 0.2s ease"),
                                                            width="100%",
                                                            align_items="center",
                                                        ),
                                                        spacing="3",
                                                        align_items="start",
                                                width="100%",
                                                        height="100%",
                                                        p="4",
                                            ),
                                                    spacing="0",
                                            align_items="start",
                                            width="100%",
                                                    height="100%",
                                        ),
                                        bg=rx.color_mode_cond(light="rgba(250,255,251,0.96)", dark="rgba(15,23,42,0.9)"),
                                        border=card_border,
                                                border_radius="2xl",
                                                width="100%",
                                                max_width="380px",
                                                margin_x="auto",
                                                margin_bottom="6",
                                                min_height="340px",
                                                position="relative",
                                                overflow="hidden",
                                                box_shadow="0 16px 28px -14px rgba(0,0,0,0.16)",
                                                _hover={"transform": "translateY(-6px)", "box_shadow": "0 24px 38px -18px rgba(16, 185, 129, 0.28)"},
                                                transition="all 0.25s ease",
                                    ),
                                ),
                                        display="grid",
                                        grid_template_columns=rx.breakpoints(initial="repeat(1, minmax(0, 1fr))", md="repeat(2, minmax(0, 1fr))", lg="repeat(3, minmax(0, 1fr))"),
                                gap="6",
                                width="100%",
                            ),
                            rx.center(
                                rx.vstack(
                                    rx.icon("inbox", size=64, color=subtext_color, opacity="0.3"),
                                    rx.heading("Sin solicitudes cerradas", size="5", color=text_color),
                                    rx.text("No hay cerradas para el filtro actual.", color=subtext_color),
                                    align_items="center",
                                    spacing="4",
                                ),
                                min_height="260px",
                                width="100%",
                            ),
                        ),
                        spacing="4",
                        width="100%",
                        max_width="100%",
                        p={"base": "3", "md": "5"},
                    ),
                    width="100%",
                    bg=rx.color_mode_cond(light="#f8fafc", dark="#0f172a"),
                    min_height="100vh",
                ),
                # Modal: Editor de Estado (Reabrir desde cerradas)
                rx.cond(
                    State.editar_estado_id,
                    rx.fragment(
                        rx.box(
                            position="fixed", top="0", left="0", width="100vw", height="100vh",
                            bg=rx.color_mode_cond(light="rgba(15, 23, 42, 0.45)", dark="rgba(0, 0, 0, 0.65)"),
                            z_index="999",
                            on_click=State.cerrar_editor_estado,
                        ),
                        rx.box(
                            rx.vstack(
                                # Header
                                rx.box(
                                    rx.hstack(
                                        rx.box(
                                            rx.icon("refresh-cw", size=18, color="#f59e0b"),
                                            bg=rx.color_mode_cond(light="#fffbeb", dark="rgba(245, 158, 11, 0.18)"),
                                            border_radius="10px",
                                            width="42px", height="42px",
                                            display="flex", align_items="center",
                                            justify_content="center", flex_shrink="0",
                                        ),
                                        rx.vstack(
                                            rx.text("Reabrir Solicitud", font_size={"base": "xl", "md": "2rem"}, font_weight="800", color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff"), line_height="1", letter_spacing="-0.03em"),
                                            rx.text("Reactivación y cambio de estado del radicado", font_size="12px", color=rx.color_mode_cond(light="#64748b", dark="#7ea8c9"), font_weight="500"),
                                            spacing="1", align_items="start",
                                        ),
                                        rx.spacer(),
                                        rx.button(rx.icon("x", size=18), type="button", on_click=State.cerrar_editor_estado, variant="ghost", color_scheme="gray", size="2"),
                                        spacing="3", align_items="center", width="100%",
                                    ),
                                    padding="18px 20px",
                                    border_bottom=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                    width="100%",
                                ),
                                # Body
                                rx.box(
                                    rx.vstack(
                                        # Sección nuevo estado
                                        rx.box(
                                            rx.vstack(
                                                rx.text("Nuevo estado", font_weight="600", font_size="sm", color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff")),
                                                rx.select(
                                                    ["Reabierta", "Radicada", "Asignada a Area", "En Gestion de Area"],
                                                    value=State.nuevo_estado,
                                                    on_change=State.set_nuevo_estado,
                                                    width="100%", size="3",
                                                ),
                                                rx.cond(
                                                    State.nuevo_estado == "Asignada a Area",
                                                    rx.vstack(
                                                        rx.hstack(
                                                            rx.text("Área responsable", font_weight="600", font_size="sm", color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff")),
                                                            rx.badge("Requerido", color_scheme="orange", size="1"),
                                                            spacing="2", align_items="center",
                                                        ),
                                                        rx.select(
                                                            State.AREAS_DISPONIBLES,
                                                            placeholder="Selecciona el área...",
                                                            value=State.area_para_asignar,
                                                            on_change=State.set_area_para_asignar,
                                                            width="100%", size="3", color_scheme="orange",
                                                        ),
                                                        spacing="2", width="100%", align_items="start",
                                                    ),
                                                ),
                                                spacing="3", width="100%", align_items="stretch",
                                            ),
                                            padding="16px 18px", border_radius="14px",
                                            bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                            border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                            border_top="3px solid #f59e0b", width="100%",
                                        ),
                                        # Sección justificación
                                        rx.box(
                                            rx.vstack(
                                                rx.hstack(
                                                    rx.text("Justificación", font_weight="600", font_size="sm", color=rx.color_mode_cond(light="#0f172a", dark="#f0f6ff")),
                                                    rx.badge("Obligatorio", color_scheme="red", size="1"),
                                                    spacing="2",
                                                ),
                                                rx.text_area(
                                                    placeholder="Describe las razones por las cuales se reactiva y reabre este radicado...",
                                                    value=State.respuesta_solicitud,
                                                    on_change=State.set_respuesta_solicitud,
                                                    rows="3", required=True, width="100%", size="3", min_height="88px",
                                                ),
                                                spacing="2", width="100%", align_items="start",
                                            ),
                                            padding="16px 18px", border_radius="14px",
                                            bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                            border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                            border_top="3px solid #10b981", width="100%",
                                        ),
                                        # Mensaje éxito/error
                                        rx.cond(
                                            State.mensaje_actualizar_estado,
                                            rx.box(
                                                rx.text(State.mensaje_actualizar_estado, color=rx.cond(State.mensaje_actualizar_estado.contains("correctamente"), "green.700", "red.700"), font_weight="semibold", font_size="sm"),
                                                bg=rx.cond(State.mensaje_actualizar_estado.contains("correctamente"), rx.color_mode_cond(light="#d1fae5", dark="#064e3b"), rx.color_mode_cond(light="#fee2e2", dark="#450a0a")),
                                                border=rx.cond(State.mensaje_actualizar_estado.contains("correctamente"), rx.color_mode_cond(light="1px solid #a7f3d0", dark="1px solid #065f46"), rx.color_mode_cond(light="1px solid #fecaca", dark="1px solid #991b1b")),
                                                border_radius="10px", p="3", width="100%",
                                            ),
                                        ),
                                        # Error de archivo
                                        rx.cond(
                                            State.respuesta_documento_error != "",
                                            rx.box(
                                                rx.hstack(
                                                    rx.icon("ban", size=15, color="#ef4444"),
                                                    rx.text(State.respuesta_documento_error, font_size="xs", color="#ef4444", font_weight="medium"),
                                                    spacing="2", align_items="center",
                                                ),
                                                bg=rx.color_mode_cond(light="#fef2f2", dark="rgba(127,29,29,0.25)"),
                                                border=rx.color_mode_cond(light="1px solid #fecaca", dark="1px solid rgba(239,68,68,0.4)"),
                                                border_radius="10px", p="3", width="100%",
                                            ),
                                        ),
                                        spacing="4", align_items="stretch",
                                        padding="18px 20px", max_height="65vh", overflow_y="auto", width="100%",
                                    ),
                                    width="100%",
                                ),
                                # Footer
                                rx.box(
                                    rx.hstack(
                                        rx.button("Cancelar", type="button", on_click=State.cerrar_editor_estado, variant="soft", color_scheme="gray", size="3", flex="1"),
                                        rx.button(
                                            rx.icon("refresh-cw", size=16),
                                            "Reabrir Solicitud",
                                            type="button",
                                            on_click=State.actualizar_estado_solicitud(rx.upload_files(upload_id="upload_respuesta_cerradas")),
                                            color_scheme="orange", size="3", flex="2",
                                        ),
                                        spacing="3", width="100%",
                                    ),
                                    padding="16px 20px",
                                    border_top=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                    bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                                    border_bottom_left_radius="14px", border_bottom_right_radius="14px", width="100%",
                                ),
                                spacing="0", width="100%",
                            ),
                            bg=rx.color_mode_cond(light="#ffffff", dark="#0f1e35"),
                            border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                            border_top="3px solid #f59e0b",
                            border_radius="14px",
                            width="100%", max_width="560px", max_height="92vh", overflow="hidden",
                            position="fixed", top="50%", left="50%", transform="translate(-50%, -50%)",
                            z_index="1000",
                            box_shadow=rx.color_mode_cond(light="0 8px 24px rgba(245, 158, 11, 0.16)", dark="0 2px 12px rgba(0,0,0,0.3)"),
                        ),
                    )
                ),
                # Modal: Historial
                rx.cond(
                    State.historial_modal_abierto,
                    rx.dialog.root(
                        rx.dialog.content(
                            rx.dialog.title("Historial de Estados", display="none"),
                            rx.dialog.description("Historial de cambios de estado", display="none"),
                            rx.vstack(
                                rx.hstack(rx.icon("history", size=24, color="#3b82f6"), rx.heading("Historial de Estados", size="5"), rx.spacer(), rx.dialog.close(rx.button(rx.icon("x", size=20), on_click=State.cerrar_historial, variant="ghost", color_scheme="gray")), width="100%", align_items="center"),
                                rx.divider(margin_y="2"),
                                rx.cond(
                                    State.historial_estados,
                                    rx.vstack(
                                        rx.foreach(
                                            State.historial_estados,
                                            lambda evento: rx.box(
                                                rx.hstack(
                                                    rx.box(rx.icon("git-commit-horizontal", size=20, color="#10b981"), p="2", bg="rgba(16, 185, 129, 0.1)", border_radius="full"),
                                                    rx.vstack(
                                                        rx.hstack(rx.text("Cambió a:", font_size="sm", color=subtext_color), rx.badge(evento['nuevo'], color_scheme="blue", radius="full"), rx.spacer(), rx.text(evento['fecha'], font_size="xs", color=subtext_color), width="100%", align_items="center"),
                                                        rx.text(rx.cond(evento['obs'] != "", f'"{evento["obs"]}"', "Sin comentarios."), font_size="sm", font_style="italic", color=text_color),
                                                        rx.text(f"Estado anterior: {evento['anterior']}", font_size="xs", color=subtext_color),
                                                        rx.cond(
                                                            evento["documento_adjunto"] != "",
                                                            ui_tarjeta_adjunto_historial(evento),
                                                            rx.box(),
                                                        ),
                                                        spacing="1", width="100%", align_items="start"
                                                    ),
                                                    spacing="4", align_items="start", width="100%"
                                                ),
                                                p="4", bg=rx.color_mode_cond(light="#f8fafc", dark="#1e293b"), border_radius="xl", width="100%", margin_bottom="3"
                                            )
                                        ),
                                        width="100%", max_height="400px", overflow_y="auto"
                                    ),
                                    rx.center(rx.text("Aún no hay historial de cambios para esta solicitud.", color=subtext_color), p="6")
                                ),
                                spacing="4", width="100%"
                            ),
                            style={"maxWidth": "600px", "borderRadius": "24px", "padding": "24px", "backgroundColor": card_bg, "backdropFilter": "blur(40px)", "border": card_border}
                        ),
                        open=True, on_open_change=State.cerrar_historial
                    )
                ),
            ),
            rx.box(
                navbar(),
                rx.center(access_denied_widget("Solo funcionarios autenticados pueden acceder a esta función."), size="3"),
                bg=rx.color_mode_cond(light="#f8fafc", dark="#0f172a"),
                min_height="100vh",
            ),
        ),
        rx.center(rx.spinner(size="3", color="#3b82f6"), height="50vh"),
    )


def solicitudes_page() -> rx.Component:
    # navbar() ya está definida en este módulo, no requiere importar desde pqrs
 
    acceso_denegado = rx.center(
        rx.vstack(
            rx.icon("lock", size=48, color=RED_ERR),
            rx.heading("Acceso Denegado", size="7", color=RED_ERR),
            rx.text("Necesitas iniciar sesión para crear una solicitud.", color=TEXT_SUB),
            rx.link(rx.button("Ir al Login", color_scheme="blue", border_radius="10px"), href="/login"),
            spacing="4", align_items="center",
        ),
        min_height="80vh",
    )
 
    contenido = rx.box(
        navbar(),
        rx.center(
            rx.box(
                rx.vstack(
 
                    # �??�?? Encabezado �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    rx.vstack(
                        rx.hstack(
                            rx.box(
                                rx.icon("file-plus", size=22, color="white"),
                                bg=f"linear-gradient(135deg, {NAVY_DARK}, {NAVY})",
                                border_radius="14px",
                                width="50px", height="50px",
                                display="flex", align_items="center", justify_content="center",
                                box_shadow=f"0 6px 20px {NAVY}44",
                                flex_shrink="0",
                            ),
                            rx.vstack(
                                rx.heading(
                                    "Nueva Solicitud PQRS",
                                    size="6", color=TEXT_MAIN,
                                    font_weight="800", letter_spacing="-0.02em",
                                ),
                                rx.text(
                                    "Completa el formulario para radicar tu Petición, Queja, Reclamo o Sugerencia.",
                                    font_size="13px", color=TEXT_SUB,
                                ),
                                spacing="0", align_items="start",
                            ),
                            spacing="3", align_items="center",
                        ),
                        border_bottom=f"2px solid {_ldc_reg('#e2e8f0','#1e3a5f')}",
                        padding_bottom="16px",
                        margin_bottom="4px",
                        width="100%",
                    ),

                    # �??�?? Datos de contacto (Solo para no registrados) �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    rx.cond(
                        ~State.es_autenticada,
                        _section("mail", "Datos de contacto",
                            rx.vstack(
                                _label("Correo electrónico para respuesta", required=True),
                                rx.input(
                                    placeholder="ejemplo@correo.com",
                                    value=State.pqrs_contacto_email,
                                    on_change=State.set_pqrs_contacto_email,
                                    type="email",
                                    **_input_s(),
                                ),
                                rx.text(
                                    "Te enviaremos las actualizaciones y respuesta de tu solicitud a este correo.",
                                    font_size="12px", color=TEXT_SUB,
                                ),
                                spacing="2", align_items="start", width="100%",
                            ),
                        ),
                    ),
 
                    # �??�?? Sección 1: Tipo + Área �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    _section("tag", "Clasificación de la solicitud",
                        rx.grid(
                            rx.vstack(
                                _label("Tipo de Solicitud"),
                                rx.select(
                                    ["Petición", "Queja", "Reclamo", "Sugerencia"],
                                    placeholder="Selecciona el tipo",
                                    value=State.tipo_solicitud,
                                    on_change=State.set_tipo_solicitud,
                                    **_select_s(),
                                ),
                                spacing="0", align_items="start", width="100%",
                            ),
                            rx.vstack(
                                _label("Área Responsable"),
                                rx.select(
                                    ["Secretaría","Contabilidad","Bienestar","Tesorería","Atención al Ciudadano","Otros"],
                                    placeholder="Selecciona el área",
                                    value=State.area_responsable,
                                    on_change=State.set_area_responsable,
                                    **_select_s(),
                                ),
                                spacing="0", align_items="start", width="100%",
                            ),
                            template_columns={"base": "1fr", "md": "1fr 1fr"},
                            gap="4", width="100%",
                        ),
                    ),
 
                    # Campo "otro área"
                    rx.cond(
                        State.area_responsable == "Otros",
                        rx.box(
                            rx.vstack(
                                _label("Especifica el área"),
                                rx.input(
                                    placeholder="Nombre del área responsable",
                                    value=State.area_otro,
                                    on_change=State.set_area_otro,
                                    **_input_s(),
                                ),
                                spacing="0", align_items="start", width="100%",
                            ),
                            bg=CARD_BG, border=f"1px solid {CARD_BDR}",
                            border_radius="14px", padding="18px 20px", width="100%",
                        ),
                    ),
 
                    # �??�?? Sección 2: Asunto + Descripción �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    _section("file-text", "Detalle de la solicitud",
                        rx.vstack(
                            rx.vstack(
                                _label("Asunto"),
                                rx.text_area(
                                    placeholder="Escribe el asunto de tu solicitud...",
                                    value=State.asunto,
                                    on_change=State.set_asunto,
                                    rows="3",
                                    resize="vertical",
                                    min_height="90px",
                                    **_input_s(),
                                ),
                                spacing="0", align_items="start", width="100%",
                            ),
                            rx.vstack(
                                _label("Descripción detallada"),
                                rx.text_area(
                                    placeholder="Escribe aquí todos los detalles de tu solicitud...",
                                    value=State.descripcion,
                                    on_change=State.set_descripcion,
                                    rows="5",
                                    max_length=1000,
                                    resize="vertical",
                                    min_height="130px",
                                    **_input_s(),
                                ),
                                rx.hstack(
                                    rx.spacer(),
                                    rx.text(
                                        State.descripcion_len,
                                        font_size="12px",
                                        color=rx.cond(State.descripcion_len > 900, ORANGE, TEXT_SUB),
                                        font_weight="500",
                                    ),
                                    rx.text(" / 1000 caracteres", font_size="12px", color=TEXT_SUB),
                                    spacing="0", width="100%",
                                ),
                                spacing="1", align_items="start", width="100%",
                            ),
                            spacing="4", width="100%",
                        ),
                    ),
 
                    # �??�?? Sección 3: Adjuntos �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    _section("paperclip", "Documento adjunto (opcional)",
                        rx.vstack(
                            rx.text(
                                "Puedes adjuntar hasta 3 archivos PDF, PNG o JPG (máx. 10 MB en total).",
                                font_size="12px", color=TEXT_SUB, line_height="1.6",
                            ),
 
                            # �??�?? Zona de carga: widget JS autocontenido �??�??�??�??�??�??
                            rx.script("""
window.__pqrsUpload = window.__pqrsUpload || (function(){
    function init(root){
        if(root.__pqrsInit) return;
        root.__pqrsInit = true;
 
        const MAX  = 10*1024*1024, MAX_N = 3;
        const OK   = ['pdf','png','jpg','jpeg'];
        let   list = [];
 
        const zone  = root.querySelector('[data-zone]');
        const inp   = root.querySelector('[data-inp]');
        const lbl   = root.querySelector('[data-lbl]');
        const meta  = root.querySelector('[data-meta]');
        const barW  = root.querySelector('[data-barw]');
        const bar   = root.querySelector('[data-bar]');
        const usedL = root.querySelector('[data-used]');
        const freeL = root.querySelector('[data-free]');
        const rows  = root.querySelector('[data-rows]');
        const errEl = root.querySelector('[data-err]');
 
        function fmt(b){ return (b/1048576).toFixed(2)+' MB'; }
        function ext(n){ return n.split('.').pop().toLowerCase(); }
        function total(){ return list.reduce((a,f)=>a+f.size,0); }
 
        function showErr(msg){
            errEl.textContent=msg; errEl.style.display='block';
        }

        function clearErr(){
            errEl.textContent=''; errEl.style.display='none';
        }
 
        function render(){
            const used=total(), free=Math.max(0,MAX-used);
            const pct=Math.min(100,(used/MAX)*100);
 
            // barra
            barW.style.display = list.length?'block':'none';
            bar.style.width    = pct+'%';
            const col = pct>=95?'#ef4444':pct>=75?'#f59e0b':'#10b981';
            bar.style.background=col;
            usedL.textContent = fmt(used)+' usado';
            freeL.textContent = fmt(free)+' libres';
            freeL.style.color = col;
 
            // zona label
            if(list.length){
                lbl.textContent = list.map(f=>f.name).join(', ');
                lbl.style.color = '#f0f6ff';
                meta.textContent= list.length+' archivo(s)';
            } else {
                lbl.textContent = 'Arrastra y suelta archivos aquí o haz clic para explorar';
                lbl.style.color = '#94a3b8';
                meta.textContent= 'PDF, PNG, JPG · máx. 10 MB · hasta 3 archivos';
            }
 
            // filas
            rows.innerHTML='';
            list.forEach(function(f,i){
                const d=document.createElement('div');
                d.style.cssText='display:flex;align-items:center;gap:8px;'+
                    'padding:7px 0;border-bottom:1px solid #1e3a5f;';
                d.innerHTML=
                    '<svg width="13" height="13" viewBox="0 0 24 24" fill="none"'+
                    ' stroke="#3b82f6" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16'+
                    'a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/>'+
                    '</svg>'+
                    '<span style="flex:1;font-size:12px;color:#f0f6ff;">'+
                        f.name+' <em style="color:#64748b;font-style:normal;">('+fmt(f.size)+')</em>'+
                    '</span>'+
                    '<button type="button" style="background:none;border:none;cursor:pointer;'+
                    'color:#ef4444;font-size:14px;padding:2px 6px;" data-i="'+i+'">�??</button>';
                rows.appendChild(d);
            });
            rows.querySelectorAll('[data-i]').forEach(function(btn){
                btn.onclick=function(e){ e.preventDefault(); list.splice(+btn.dataset.i,1); render(); };
            });
            if(list.length){
                const clearAll=document.createElement('button');
                clearAll.type='button';
                clearAll.textContent='Quitar todos los adjuntos';
                clearAll.style.cssText='margin-top:8px;width:100%;padding:6px 10px;border:1px solid #991b1b;'+
                    'border-radius:8px;background:#450a0a;color:#fca5a5;font-size:12px;cursor:pointer;';
                clearAll.onclick=function(e){
                    e.preventDefault();
                    list=[];
                    clearErr();
                    inp.value='';
                    render();
                };
                rows.appendChild(clearAll);
            }
        }
 
        inp.addEventListener('change',function(){
            const inc=Array.from(this.files);
            clearErr();
            for(const f of inc){
                if(!OK.includes(ext(f.name))){
                    showErr('�? Formato ".' + ext(f.name) + '" no permitido. Solo se aceptan PDF, PNG o JPG.');
                    this.value=''; return;
                }
                if(f.size>MAX){
                    showErr('�? "'+f.name+'" supera el límite de 10 MB.');
                    this.value=''; return;
                }
                if(list.some(function(e){return e.name===f.name&&e.size===f.size;})) continue;
                if(list.length>=MAX_N){
                    showErr('�? Solo puedes adjuntar hasta '+MAX_N+' archivos.');
                    this.value=''; return;
                }
                if(total()+f.size>MAX){
                    showErr('�? Agregar "'+f.name+'" superaría el límite. Libre: '+fmt(MAX-total()));
                    this.value=''; return;
                }
                list.push({name:f.name,size:f.size});
            }
            this.value='';
            render();
        });
 
        // drag & drop
        zone.addEventListener('dragover',function(e){
            e.preventDefault();
            zone.style.borderColor='#2563eb';
            zone.style.background='#0f2744';
        });
        zone.addEventListener('dragleave',function(){
            zone.style.borderColor='#334155';
            zone.style.background='#0a1628';
        });
        zone.addEventListener('drop',function(e){
            e.preventDefault();
            zone.style.borderColor='#334155';
            zone.style.background='#0a1628';
            // simular change con los archivos soltados
            const dt=new DataTransfer();
            Array.from(e.dataTransfer.files).forEach(function(f){ dt.items.add(f); });
            inp.files=dt.files;
            inp.dispatchEvent(new Event('change'));
        });
    }
 
    // observar DOM para inicializar cuando el widget aparezca
    const obs=new MutationObserver(function(){
        document.querySelectorAll('[data-pqrs-upload]').forEach(init);
    });
    obs.observe(document.body,{childList:true,subtree:true});
    document.querySelectorAll('[data-pqrs-upload]').forEach(init);
    return {init:init};
})();
"""),
                            rx.el.div(
                                # Zona de carga
                                rx.el.div(
                                    rx.el.div(
                                        rx.el.svg(
                                            rx.el.path(d="M16 16 L12 12 L8 16"),
                                            rx.el.path(d="M20.39 18.39A5 5 0 0 0 18 9h-1.26A8 8 0 1 0 3 16.3"),
                                            width="22", height="22", viewBox="0 0 24 24",
                                            fill="none", stroke="#3b82f6",
                                            stroke_width="2", stroke_linecap="round",
                                            style={"flexShrink":"0"},
                                        ),
                                        rx.el.div(
                                            rx.el.div(
                                                "Arrastra y suelta archivos aquí o haz clic para explorar",
                                                data_lbl=True,
                                                style={"fontSize":"13px","color":"#94a3b8","fontWeight":"500"},
                                            ),
                                            rx.el.div(
                                                "PDF, PNG, JPG · máx. 10 MB · hasta 3 archivos",
                                                data_meta=True,
                                                style={"fontSize":"11px","color":"#64748b","marginTop":"3px"},
                                            ),
                                        ),
                                        style={"display":"flex","alignItems":"center","gap":"12px"},
                                    ),
                                    rx.el.input(
                                        data_inp=True,
                                        type="file",
                                        accept="application/pdf,image/png,image/jpeg",
                                        multiple=True,
                                        style={
                                            "position":"absolute","inset":"0",
                                            "width":"100%","height":"100%",
                                            "opacity":"0","cursor":"pointer",
                                        },
                                    ),
                                    data_zone=True,
                                    style={
                                        "position":"relative","border":"2px dashed #334155",
                                        "borderRadius":"12px","padding":"16px 20px",
                                        "background":"#0a1628","cursor":"pointer",
                                        "transition":"border-color .15s,background .15s",
                                        "width":"100%",
                                    },
                                    onmouseenter="this.style.borderColor='#2563eb';this.style.background='#0f2744'",
                                    onmouseleave="this.style.borderColor='#334155';this.style.background='#0a1628'",
                                ),
                                # Barra de almacenamiento
                                rx.el.div(
                                    rx.el.div(
                                        rx.el.span("0.00 MB usado", data_used=True,
                                                   style={"fontSize":"11px","color":"#94a3b8"}),
                                        rx.el.span("10.00 MB libres", data_free=True,
                                                   style={"fontSize":"11px","fontWeight":"700","color":"#10b981"}),
                                        style={"display":"flex","justifyContent":"space-between","marginBottom":"4px"},
                                    ),
                                    rx.el.div(
                                        rx.el.div(
                                            data_bar=True,
                                            style={
                                                "height":"100%","width":"0%","borderRadius":"6px",
                                                "background":"#10b981",
                                                "transition":"width .35s ease,background .35s ease",
                                            },
                                        ),
                                        style={
                                            "background":"#1e3a5f","borderRadius":"6px",
                                            "height":"7px","overflow":"hidden",
                                        },
                                    ),
                                    data_barw=True,
                                    style={"display":"none","marginTop":"10px"},
                                ),
                                # Lista de archivos
                                rx.el.div(data_rows=True, style={"marginTop":"6px"}),
                                # Error
                                rx.el.div(
                                    data_err=True,
                                    style={
                                        "display":"none","marginTop":"8px",
                                        "padding":"8px 14px",
                                        "background":"#2d0000","border":"1px solid #991b1b",
                                        "borderRadius":"8px","fontSize":"12px","color":"#ef4444",
                                    },
                                ),
                                data_pqrs_upload=True,
                                style={"width":"100%"},
                            ),
 
                            # Error desde State (validaciones servidor)
                            rx.cond(
                                State.archivo_error_mensaje != "",
                                rx.hstack(
                                    rx.icon("alert-circle", size=14, color=RED_ERR),
                                    rx.text(State.archivo_error_mensaje, font_size="12px", color=RED_ERR),
                                    spacing="2", align_items="center",
                                ),
                            ),
                            spacing="3", width="100%",
                        ),
                    ),
 
                    # �??�?? Autorización �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    rx.hstack(
                        rx.checkbox(
                            is_checked=State.acepta_politica_solicitud,
                            on_change=State.set_acepta_politica_solicitud,
                            color_scheme="blue",
                            size="2",
                        ),
                        rx.hstack(
                            rx.text("He leído y acepto la ", font_size="13px", color=TEXT_SUB),
                            rx.link(
                                "Política de Tratamiento de Datos Personales",
                                href="/politica-privacidad",
                                color=BLUE_ACC,
                                font_size="13px",
                                font_weight="600",
                                text_decoration="none",
                                _hover={"text_decoration": "underline"},
                            ),
                            spacing="0", flex_wrap="wrap",
                        ),
                        spacing="3", align_items="center",
                        bg=_ldc_reg("#eff6ff","#0a1628"),
                        border=f"1px solid {_ldc_reg('#bfdbfe','#1e3a5f')}",
                        border_radius="10px",
                        padding="12px 16px",
                        width="100%",
                    ),
 
                    # �??�?? Botón enviar �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    rx.button(
                        rx.hstack(
                            rx.icon("send", size=16),
                            rx.text("Enviar Solicitud", font_size="15px", font_weight="600"),
                            spacing="2",
                        ),
                        on_click=State.crear_solicitud,
                        width="100%",
                        height="48px",
                        bg=rx.cond(
                            State.acepta_politica_solicitud,
                            f"linear-gradient(135deg, {NAVY_DARK}, {NAVY})",
                            _ldc_reg("#e2e8f0","#1e293b"),
                        ),
                        color=rx.cond(State.acepta_politica_solicitud, "white", TEXT_SUB),
                        border_radius="12px",
                        cursor=rx.cond(State.acepta_politica_solicitud, "pointer", "not-allowed"),
                        box_shadow=rx.cond(
                            State.acepta_politica_solicitud,
                            f"0 4px 16px {NAVY}44",
                            "none",
                        ),
                        _hover=rx.cond(
                            State.acepta_politica_solicitud,
                            {"opacity": "0.92", "transform": "translateY(-1px)"},
                            {},
                        ),
                        transition="all 0.15s ease",
                        is_disabled=~State.acepta_politica_solicitud,
                    ),

                    # �??�?? Mensaje resultado �??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??�??
                    rx.cond(
                        State.solicitud_mensaje != "",
                        rx.hstack(
                            rx.icon(
                                rx.cond(State.solicitud_mensaje.contains("éxito"), "circle-check", "circle-x"),
                                size=16,
                                color=rx.cond(State.solicitud_mensaje.contains("éxito"), GREEN_OK, RED_ERR),
                            ),
                            rx.text(
                                State.solicitud_mensaje, font_size="13px", font_weight="500",
                                color=rx.cond(State.solicitud_mensaje.contains("éxito"), GREEN_OK, RED_ERR),
                            ),
                            spacing="2", align_items="center",
                            bg=rx.cond(State.solicitud_mensaje.contains("éxito"),
                                       _ldc_reg("#f0fdf4","#002818"), _ldc_reg("#fef2f2","#2d0000")),
                            border=rx.cond(State.solicitud_mensaje.contains("éxito"),
                                           "1px solid #bbf7d0", "1px solid #fecaca"),
                            border_radius="10px", padding="12px 16px", width="100%",
                        ),
                    ),
 
                    spacing="4",
                    align_items="stretch",
                    width="100%",
                ),
                width="100%",
                max_width="720px",
                padding={"base": "20px 16px", "md": "36px 40px"},
            ),
            width="100%",
        ),
        bg=PAGE_BG, min_height="100vh", width="100%",
    )
 
    return contenido

NAVY      = "#1e3a8a"
NAVY_DARK = "#172554"
BLUE_ACC  = "#2563eb"
GREEN_OK  = "#16a34a"
ORANGE    = "#f59e0b"
RED_ERR   = "#ef4444"
VIOLET    = "#8b5cf6"
 
def consultar_estado_page() -> rx.Component:
    text_color = rx.color_mode_cond(light="#0f172a", dark="#f8fafc")
    subtext_color = rx.color_mode_cond(light="#64748b", dark="#94a3b8")
    card_bg = rx.color_mode_cond(light="rgba(255, 255, 255, 0.85)", dark="rgba(30, 41, 59, 0.7)")
    card_border = rx.color_mode_cond(light="1px solid rgba(255, 255, 255, 0.9)", dark="1px solid rgba(51, 65, 85, 0.5)")

    def estado_badge(estado: str) -> rx.Component:
        return rx.badge(
            estado,
            color_scheme=rx.match(
                estado,
                ("Radicada", "orange"),
                ("Asignada a Area", "blue"),
                ("En Gestion de Area", "purple"),
                ("Solucionada", "teal"),
                ("Reabierta", "yellow"),
                ("Cerrada", "green"),
                "gray"
            ),
            size="2",
            radius="full",
            variant="surface",
        )

    def timeline_step(title: str, is_active: bool, icon_name: str) -> rx.Component:
        color = rx.cond(is_active, "#2563eb", rx.color_mode_cond(light="#cbd5e1", dark="#475569"))
        bg_color = rx.cond(is_active, "rgba(37,99,235,0.1)", rx.color_mode_cond(light="#f8fafc", dark="rgba(15,23,42,0.4)"))
        border_color = rx.cond(is_active, "#2563eb", rx.color_mode_cond(light="#e2e8f0", dark="#334155"))
        return rx.vstack(
            rx.box(
                rx.icon(icon_name, size=16, color=color),
                bg=bg_color,
                border=f"2px solid {border_color}",
                border_radius="full",
                width="36px", height="36px",
                display="flex", align_items="center", justify_content="center",
                box_shadow=rx.cond(is_active, "0 4px 10px rgba(37,99,235,0.2)", "none"),
            ),
            rx.text(title, font_size="12px", font_weight="bold", color=rx.cond(is_active, text_color, subtext_color)),
            spacing="1",
            align_items="center"
        )

    def timeline_connector(is_active: bool) -> rx.Component:
        color = rx.cond(is_active, "#2563eb", rx.color_mode_cond(light="#cbd5e1", dark="#475569"))
        return rx.box(
            height="2px",
            flex="1",
            bg=color,
            margin_bottom="18px",
        )

    return rx.box(
        navbar(),
        rx.center(
            rx.box(
                rx.vstack(
                    # Header de la tarjeta con botón cerrar
                    rx.hstack(
                        rx.vstack(
                            rx.heading("Consultar Estado de Solicitud", size="6", color=text_color, font_weight="bold"),
                            rx.text("Rastrea tu Petición, Queja, Reclamo o Sugerencia.", color=subtext_color, font_size="sm"),
                            align_items="start", spacing="1"
                        ),
                        rx.spacer(),
                        rx.icon_button(
                            "x",
                            variant="soft",
                            color_scheme="gray",
                            radius="full",
                            size="3",
                            on_click=State.cerrar_consulta,
                            cursor="pointer",
                        ),
                        width="100%",
                        align_items="center",
                        padding_bottom="4",
                        border_bottom=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155")
                    ),
                    
                    # Formulario de consulta
                    rx.box(
                        rx.text("Número de Radicado", font_weight="semibold", font_size="sm", color=text_color, margin_bottom="2"),
                        rx.hstack(
                            rx.input(
                                placeholder="Ej: PQRS-2024-abc12345",
                                value=State.consulta_radicado,
                                on_change=State.set_consulta_radicado,
                                bg=rx.color_mode_cond(light="white", dark="#0f172a"),
                                border=rx.color_mode_cond(light="1px solid #cbd5e1", dark="1px solid #334155"),
                                radius="large",
                                size="3",
                                width="100%",
                                box_shadow="inset 0 1px 2px rgba(0,0,0,0.05)"
                            ),
                            rx.button(
                                "Buscar",
                                icon="search",
                                on_click=State.consultar_estado_solicitud,
                                color_scheme="blue",
                                size="3",
                                radius="large",
                                box_shadow="0 4px 6px -1px rgba(59, 130, 246, 0.3)"
                            ),
                            width="100%",
                            spacing="3"
                        ),
                        width="100%",
                        margin_y="4"
                    ),
                    
                    # Mensaje de resultado
                    rx.cond(
                        State.consulta_mensaje,
                        rx.box(
                            rx.text(
                                State.consulta_mensaje,
                                color=rx.cond(
                                    State.consulta_mensaje.contains("encontrada") & ~State.consulta_mensaje.contains("No se encontró"),
                                    "green.700",
                                    "red.700"
                                ),
                                font_weight="medium",
                                font_size="sm"
                            ),
                            p="3",
                            border_radius="lg",
                            border=rx.cond(
                                State.consulta_mensaje.contains("encontrada") & ~State.consulta_mensaje.contains("No se encontró"),
                                "1px solid #bbf7d0", "1px solid #fecaca"
                            ),
                            bg=rx.cond(
                                State.consulta_mensaje.contains("encontrada") & ~State.consulta_mensaje.contains("No se encontró"),
                                rx.color_mode_cond(light="#f0fdf4", dark="rgba(22,163,74,0.1)"),
                                rx.color_mode_cond(light="#fef2f2", dark="rgba(239,68,68,0.1)")
                            ),
                            width="100%",
                            margin_bottom="4"
                        )
                    ),
                    
                    # Detalles de la solicitud
                    rx.cond(
                        State.solicitud_consultada,
                        rx.box(
                            rx.vstack(
                                # Cabecera de Radicado Premium
                                rx.hstack(
                                    rx.hstack(
                                        rx.box(
                                            rx.icon("file-text", size=18, color="#2563eb"),
                                            bg=rx.color_mode_cond(light="#eff6ff", dark="rgba(37,99,235,0.1)"),
                                            border_radius="lg",
                                            width="36px", height="36px",
                                            display="flex", align_items="center", justify_content="center",
                                        ),
                                        rx.vstack(
                                            rx.text("Detalle de la Solicitud", font_weight="bold", font_size="md", color=text_color),
                                            rx.text(State.solicitud_consultada.get("radicado", ""), font_family="monospace", font_size="sm", color="#2563eb", font_weight="bold"),
                                            spacing="0", align_items="start"
                                        ),
                                        spacing="3", align_items="center"
                                    ),
                                    rx.spacer(),
                                    estado_badge(State.solicitud_consultada.get("estado", "")),
                                    width="100%", align_items="center", padding_bottom="3",
                                    border_bottom=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155")
                                ),
                                
                                # Timeline de Estados Premium
                                rx.box(
                                    rx.hstack(
                                        timeline_step("Radicada", True, "file-text"),
                                        timeline_connector(State.solicitud_consultada.get("estado") != "Radicada"),
                                        timeline_step(
                                            "En Trámite",
                                            State.solicitud_consultada.get("estado") != "Radicada",
                                            "activity"
                                        ),
                                        timeline_connector(
                                            (State.solicitud_consultada.get("estado") == "Solucionada") |
                                            (State.solicitud_consultada.get("estado") == "Cerrada")
                                        ),
                                        timeline_step(
                                            "Solucionada",
                                            (State.solicitud_consultada.get("estado") == "Solucionada") |
                                            (State.solicitud_consultada.get("estado") == "Cerrada"),
                                            "circle-check"
                                        ),
                                        timeline_connector(State.solicitud_consultada.get("estado") == "Cerrada"),
                                        timeline_step(
                                            "Cerrada",
                                            State.solicitud_consultada.get("estado") == "Cerrada",
                                            "archive"
                                        ),
                                        spacing="0", align_items="center", width="100%"
                                    ),
                                    padding_y="4",
                                    width="100%"
                                ),
                                
                                # Datos de la Solicitud
                                rx.grid(
                                    rx.box(
                                        rx.text("Tipo de Solicitud", font_size="xs", color=subtext_color, text_transform="uppercase", font_weight="bold"),
                                        rx.text(State.solicitud_consultada.get("tipo_solicitud", ""), font_size="sm", color=text_color, font_weight="medium", margin_top="1")
                                    ),
                                    rx.box(
                                        rx.text("Área Responsable", font_size="xs", color=subtext_color, text_transform="uppercase", font_weight="bold"),
                                        rx.text(State.solicitud_consultada.get("area_responsable", "Por asignar"), font_size="sm", color=text_color, font_weight="medium", margin_top="1")
                                    ),
                                    rx.box(
                                        rx.text("Fecha de Registro", font_size="xs", color=subtext_color, text_transform="uppercase", font_weight="bold"),
                                        rx.text(State.solicitud_consultada.get("fecha", ""), font_size="sm", color=text_color, margin_top="1")
                                    ),
                                    rx.box(
                                        rx.text("Asunto", font_size="xs", color=subtext_color, text_transform="uppercase", font_weight="bold"),
                                        rx.text(State.solicitud_consultada.get("asunto", ""), font_size="sm", color=text_color, font_weight="medium", margin_top="1")
                                    ),
                                    template_columns={"base": "repeat(1, 1fr)", "sm": "repeat(2, 1fr)"},
                                    gap="4",
                                    width="100%",
                                    bg=rx.color_mode_cond(light="#f8fafc", dark="#0f172a"),
                                    p="4",
                                    border_radius="xl",
                                    border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e293b")
                                ),
                                
                                # Descripción
                                rx.cond(
                                    State.solicitud_consultada.get("descripcion") != "",
                                    rx.vstack(
                                        rx.text("Descripción de la Solicitud", font_weight="semibold", font_size="xs", color=subtext_color, text_transform="uppercase"),
                                        rx.box(
                                            rx.text(State.solicitud_consultada.get("descripcion", ""), font_size="sm", color=text_color, line_height="1.6"),
                                            bg=rx.color_mode_cond(light="#f8fafc", dark="#0f172a"),
                                            border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e293b"),
                                            border_radius="xl",
                                            p="4",
                                            width="100%"
                                        ),
                                        spacing="1", width="100%", align_items="start", margin_top="2"
                                    )
                                ),
                                
                                # Respuesta del Funcionario
                                rx.cond(
                                    State.solicitud_consultada.get("respuesta") != "",
                                    rx.vstack(
                                        rx.hstack(
                                            rx.icon("message-circle", size=14, color="#10b981"),
                                            rx.text("Respuesta del Funcionario", font_weight="bold", font_size="xs", color="#10b981", text_transform="uppercase"),
                                            spacing="2", align_items="center"
                                        ),
                                        rx.box(
                                            rx.text(State.solicitud_consultada.get("respuesta", ""), font_size="sm", color=text_color, line_height="1.6"),
                                            bg=rx.color_mode_cond(light="#ecfdf5", dark="rgba(16,185,129,0.1)"),
                                            border="1px solid #a7f3d0",
                                            border_left="4px solid #10b981",
                                            border_radius="xl",
                                            p="4",
                                            width="100%"
                                        ),
                                        spacing="2", width="100%", align_items="start", margin_top="4"
                                    )
                                ),
                                rx.cond(
                                    State.solicitud_consultada_tiene_adjunto_respuesta,
                                    rx.vstack(
                                        rx.hstack(
                                            rx.icon("paperclip", size=14, color="#10b981"),
                                            rx.text("Adjunto de respuesta (funcionario)", font_weight="semibold", font_size="xs", color="#10b981", text_transform="uppercase"),
                                            spacing="2", align_items="center",
                                        ),
                                        rx.cond(
                                            State.solicitud_consultada_adjunto_respuesta_existe,
                                            rx.vstack(
                                                rx.cond(
                                                    State.solicitud_consultada_respuesta_documento_preview_src != "",
                                                    rx.image(
                                                        src=State.solicitud_consultada_respuesta_documento_preview_src,
                                                        alt="Adjunto de respuesta",
                                                        width="100%",
                                                        max_height="120px",
                                                        object_fit="contain",
                                                        border_radius="lg",
                                                        border="1px solid #a7f3d0",
                                                    ),
                                                    rx.box(),
                                                ),
                                                ui_boton_descargar_adjunto(
                                                    State.solicitud_consultada_respuesta_documento_href,
                                                    State.solicitud_consultada_respuesta_documento_basename,
                                                    color_scheme="green",
                                                ),
                                                spacing="2",
                                                width="100%",
                                                align_items="start",
                                            ),
                                            rx.text(
                                                "El archivo ya no está en el servidor",
                                                font_size="xs",
                                                color="#ef4444",
                                            ),
                                        ),
                                        spacing="2", width="100%", align_items="start", margin_top="2",
                                    ),
                                ),
                                # Adjuntos
                                rx.cond(
                                    State.solicitud_consultada.get("documento_adjuntos"),
                                    rx.vstack(
                                        rx.hstack(
                                            rx.icon("paperclip", size=14, color="#3b82f6"),
                                            rx.text("Documentos Adjuntos", font_weight="semibold", font_size="xs", color=subtext_color, text_transform="uppercase"),
                                            spacing="2", align_items="center"
                                        ),
                                        rx.vstack(
                                            rx.foreach(
                                                State.solicitud_consultada_adjuntos,
                                                lambda doc: ui_tarjeta_adjunto(doc),
                                            ),
                                            spacing="2", width="100%"
                                        ),
                                        spacing="2", width="100%", align_items="start", margin_top="4"
                                    )
                                ),
                                
                                # Calificación Obligatoria para HU11
                                rx.cond(
                                    (State.solicitud_consultada.get("estado") == "Solucionada") | (State.solicitud_consultada.get("estado") == "Cerrada"),
                                    rx.box(
                                        rx.vstack(
                                            rx.hstack(
                                                rx.icon("star", color="#10b981", size=24),
                                                rx.heading("¡Califica tu Solicitud!", size="5", color="#10b981"),
                                                align_items="center", spacing="2"
                                            ),
                                            rx.text("Por favor, califica nuestro servicio para cerrar definitivamente el caso.", font_size="sm", color=text_color),
                                            rx.text("Nota: Si no calificas en 5 días hábiles, se cerrará automáticamente con 5 estrellas.", font_size="xs", color=subtext_color),
                                            
                                            rx.cond(
                                                State.solicitud_consultada.get("calificacion_servicio") == None,
                                                rx.vstack(
                                                    rx.text_area(
                                                        placeholder="Deja un comentario opcional sobre el servicio...",
                                                        value=State.comentario_calificacion,
                                                        on_change=State.set_comentario_calificacion,
                                                        width="100%", size="2", margin_top="2"
                                                    ),
                                                    rx.vstack(
                                                        rx.select(
                                                            ["5 - Excelente", "4 - Muy Bueno", "3 - Bueno", "2 - Regular", "1 - Malo"],
                                                            placeholder="Selecciona una calificación",
                                                            on_change=State.set_calificacion_opcion,
                                                            size="3", radius="large", width="100%"
                                                        ),
                                                        rx.button(
                                                            "Enviar Calificación",
                                                            on_click=State.enviar_calificacion,
                                                            color_scheme="green", size="3", radius="large", width="100%"
                                                        ),
                                                        width="100%", margin_top="2", spacing="3"
                                                    ),
                                                    width="100%"
                                                ),
                                                rx.text(f"Ya calificaste este servicio con {State.solicitud_consultada.get('calificacion_servicio')} estrella(s). ¡Gracias!", color="green.600", font_weight="bold", margin_top="2")
                                            ),
                                            rx.cond(
                                                State.calificacion_mensaje,
                                                rx.text(State.calificacion_mensaje, color="blue.500", font_size="sm", font_weight="medium", margin_top="2")
                                            ),
                                            align_items="start",
                                            spacing="2",
                                            width="100%"
                                        ),
                                        p="5",
                                        border="1px solid #6ee7b7",
                                        bg=rx.color_mode_cond(light="#ecfdf5", dark="rgba(16,185,129,0.1)"),
                                        border_radius="xl",
                                        width="100%",
                                        margin_y="6",
                                        box_shadow="0 4px 6px -1px rgba(16, 185, 129, 0.1)"
                                    )
                                ),

                                # Bitácora HU11
                                rx.box(
                                    rx.vstack(
                                        rx.hstack(
                                            rx.icon("git-commit-horizontal", size=14, color="#8b5cf6"),
                                            rx.text("Bitácora", font_size="sm", font_weight="bold", color=text_color),
                                            rx.text("· historial de seguimiento", font_size="2xs", color=subtext_color),
                                            spacing="2",
                                            align_items="center",
                                            width="100%",
                                        ),
                                        rx.cond(
                                            State.historial_solicitud.length() > 0,
                                            rx.vstack(
                                                rx.foreach(
                                                    State.historial_solicitud,
                                                    lambda h: rx.box(
                                                        rx.vstack(
                                                            rx.hstack(
                                                                rx.badge(h["estado_anterior"], color_scheme="gray", variant="soft", size="1"),
                                                                rx.icon("arrow-right", size=10, color=subtext_color),
                                                                rx.badge(h["estado_nuevo"], color_scheme="purple", variant="soft", size="1"),
                                                                rx.spacer(),
                                                                rx.text(h["fecha_cambio"], font_size="2xs", color=subtext_color),
                                                                width="100%",
                                                                align_items="center",
                                                            ),
                                                            rx.cond(
                                                                h["area_asignada"] != "",
                                                                rx.text(h["area_asignada"], font_size="2xs", color="#f59e0b", font_weight="600"),
                                                                rx.box(),
                                                            ),
                                                            rx.cond(
                                                                h["observaciones"] != "",
                                                                rx.text(
                                                                    h["observaciones"],
                                                                    font_size="xs",
                                                                    color=text_color,
                                                                    white_space="pre-wrap",
                                                                    line_height="1.4",
                                                                ),
                                                                rx.box(),
                                                            ),
                                                            rx.cond(
                                                                h["documento_adjunto"] != "",
                                                                ui_tarjeta_adjunto_historial(h),
                                                                rx.box(),
                                                            ),
                                                            spacing="1",
                                                            width="100%",
                                                            align_items="start",
                                                        ),
                                                        p="3",
                                                        border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                                        border_left="2px solid #8b5cf6",
                                                        border_radius="lg",
                                                        width="100%",
                                                    ),
                                                ),
                                                spacing="2",
                                                width="100%",
                                                max_height="220px",
                                                overflow_y="auto",
                                            ),
                                            rx.center(
                                                rx.text("No hay historial disponible aún.", font_size="xs", color=subtext_color),
                                                p="4",
                                                width="100%",
                                            ),
                                        ),
                                        spacing="2",
                                        width="100%",
                                        align_items="start",
                                    ),
                                    p="3",
                                    border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #1e3a5f"),
                                    border_radius="lg",
                                    margin_top="2",
                                    width="100%",
                                ),

                            spacing="0",
                                align_items="start",
                                width="100%"
                            ),
                            width="100%",
                            margin_top="2",
                            animation="fadeIn 0.5s ease-out"
                        )
                    ),
                    
                    spacing="0",
                    align_items="start",
                    width="100%"
                ),
                bg=card_bg,
                backdrop_filter="blur(16px)",
                border=card_border,
                border_radius="3xl",
                box_shadow="0 25px 50px -12px rgba(0, 0, 0, 0.15)",
                width="100%",
                max_width="750px",
                p={"base": "6", "md": "8"},
                margin_y="10",
                style={
                    "@keyframes fadeIn": {
                        "from": {"opacity": "0", "transform": "translateY(10px)"},
                        "to": {"opacity": "1", "transform": "translateY(0)"}
                    }
                }
            ),
            size="3",
            width="100%",
            padding_x="4"
        ),
        bg=rx.color_mode_cond(
            light="linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%)", 
            dark="linear-gradient(135deg, #0f172a 0%, #020617 100%)"
        ),
        min_height="100vh",
        width="100%"
    )


def reportes_page() -> rx.Component:
    def chart_card(title: str, chart_component, width="100%", extra_content=None):
        return rx.box(
            rx.vstack(
                rx.heading(title, size="5", color=rx.color_mode_cond(light="#1e293b", dark="#f8fafc"), font_weight="bold"),
                chart_component,
                extra_content if extra_content else rx.box(),
                spacing="4",
                align_items="center",
                width="100%"
            ),
            p="6",
            bg=rx.color_mode_cond(light="#ffffff", dark="#1e293b"),
            border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155"),
            border_radius="3xl",
            box_shadow=rx.color_mode_cond(light="0 4px 6px -1px rgba(0, 0, 0, 0.05)", dark="none"),
            width=width,
            _hover={"box_shadow": rx.color_mode_cond(light="0 10px 15px -3px rgba(0, 0, 0, 0.1)", dark="0 10px 15px -3px rgba(0, 0, 0, 0.5)"), "transform": "translateY(-2px)", "border_color": rx.color_mode_cond(light="#93c5fd", dark="#3b82f6")},
            transition="all 0.3s ease"
        )

    content = rx.box(
        navbar(),
        rx.center(
            rx.vstack(
                # Encabezado y Command Center
                rx.hstack(
                    rx.vstack(
                        rx.heading("Centro de Reportes", size="8", color=rx.color_mode_cond(light="#0f172a", dark="#f8fafc"), font_weight="bold", letter_spacing="-0.02em"),
                        rx.text("Análisis avanzado y métricas de solicitudes", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8"), font_size="lg"),
                        align_items="start",
                        spacing="1"
                    ),
                    rx.spacer(),
                    # Barra de Filtros y Descarga
                    rx.box(
                        rx.hstack(
                            rx.hstack(
                                rx.icon("calendar", size=16, color=rx.color_mode_cond(light="#64748b", dark="#94a3b8")),
                                rx.select(
                                    ["�?ltimos 12 meses", "�?ltimos 6 meses", "Este año"],
                                    placeholder="Rango",
                                    variant="soft", radius="full", size="2"
                                ),
                                spacing="2", align_items="center"
                            ),
                            rx.divider(orientation="vertical", height="20px", border_color=rx.color_mode_cond(light="#e2e8f0", dark="#334155")),
                            rx.hstack(
                                rx.icon("filter", size=16, color=rx.color_mode_cond(light="#64748b", dark="#94a3b8")),
                                rx.select(
                                    ["Todos", "Petición", "Queja", "Reclamo", "Sugerencia"],
                                    placeholder="Tipo",
                                    value=State.filter_tipo_solicitud,
                                    on_change=State.set_filter_tipo_solicitud,
                                    variant="soft", radius="full", size="2"
                                ),
                                spacing="2", align_items="center"
                            ),
                            rx.divider(orientation="vertical", height="20px", border_color=rx.color_mode_cond(light="#e2e8f0", dark="#334155")),
                            rx.box(
                                rx.button(
                                    rx.icon("download", size=16),
                                    "Exportar",
                                    on_click=State.toggle_menu_descarga,
                                    color_scheme="blue",
                                    variant="solid",
                                    radius="full",
                                    box_shadow="0 4px 6px -1px rgba(59, 130, 246, 0.3)",
                                ),
                                rx.cond(
                                    State.mostrar_menu_descarga,
                                    rx.vstack(
                                        rx.button(rx.icon("file-spreadsheet", size=16), "Excel + Gráfica", on_click=State.descargar_excel_y_abrir, color_scheme="green", width="100%", size="2", variant="soft", justify="start"),
                                        rx.button(rx.icon("file-text", size=16), "CSV", on_click=State.descargar_csv_y_abrir, color_scheme="cyan", width="100%", size="2", variant="soft", justify="start"),
                                        spacing="2",
                                        position="absolute",
                                        top="120%",
                                        right="0",
                                        bg=rx.color_mode_cond(light="rgba(255, 255, 255, 0.9)", dark="rgba(30, 41, 59, 0.9)"),
                                        backdrop_filter="blur(12px)",
                                        border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155"),
                                        border_radius="xl",
                                        p="2",
                                        box_shadow="xl",
                                        z_index="10",
                                        width="140px"
                                    ),
                                    rx.box()
                                ),
                                position="relative"
                            ),
                            spacing="4",
                            align_items="center"
                        ),
                        bg=rx.color_mode_cond(light="rgba(248, 250, 252, 0.6)", dark="rgba(15, 23, 42, 0.4)"),
                        backdrop_filter="blur(16px)",
                        border=rx.color_mode_cond(light="1px solid rgba(226, 232, 240, 0.8)", dark="1px solid rgba(30, 41, 59, 0.8)"),
                        border_radius="full",
                        padding="2",
                        padding_x="5",
                        box_shadow=rx.color_mode_cond(light="0 4px 6px -1px rgba(0,0,0,0.05)", dark="0 4px 6px -1px rgba(0,0,0,0.5)")
                    ),
                    width="100%",
                    align_items="end",
                ),
                
                # Fila de KPIs (Top Row)
                rx.grid(
                    kpi_metric_card(
                        "Total Solicitudes",
                        State.numero_solicitudes,
                        "layers",
                        "#2563eb",
                        rx.color_mode_cond(light="#eff6ff", dark="rgba(37, 99, 235, 0.18)"),
                    ),
                    kpi_metric_card(
                        "Radicadas",
                        State.numero_solicitudes_radicadas,
                        "file-plus-2",
                        "#f59e0b",
                        rx.color_mode_cond(light="#fffbeb", dark="rgba(245, 158, 11, 0.16)"),
                    ),
                    kpi_metric_card(
                        "En Proceso",
                        State.kpi_en_curso_count,
                        "activity",
                        "#0ea5e9",
                        rx.color_mode_cond(light="#f0fdfe", dark="rgba(14, 165, 233, 0.16)"),
                    ),
                    kpi_metric_card(
                        "Cerradas",
                        State.kpi_cerradas_count,
                        "circle-check",
                        "#10b981",
                        rx.color_mode_cond(light="#ecfdf5", dark="rgba(16, 185, 129, 0.14)"),
                    ),
                    columns={"base": "1", "sm": "2", "lg": "4"},
                    spacing="6",
                    width="100%",
                    margin_y="6",
                    style={"gridTemplateColumns": "repeat(auto-fit, minmax(200px, 1fr))"},
                ),

                # Contenedor de Layout Horizontal Responsivo (Gráficas + Sidebar Premium)
                rx.flex(
                    # Columna Izquierda: Gráficas y Distribución por Área
                    rx.vstack(
                        # Rejilla de Gráficas (2 Columnas)
                        rx.grid(
                            # 1. Cumplimiento (Dona)
                            chart_card(
                                "Nivel de Cumplimiento",
                                rc.pie_chart(
                                    rc.tooltip(),
                                    rc.legend(layout="horizontal", vertical_align="bottom", align="center"),
                                    rc.pie(
                                        rc.cell(fill="#10b981"),
                                        rc.cell(fill="#ef4444"),
                                        data=State.compliance_chart_data,
                                        data_key="value",
                                        name_key="name",
                                        cx="50%",
                                        cy="50%",
                                        outer_radius=100,
                                        inner_radius=65,
                                        label=True,
                                    ),
                                    width="100%",
                                    height=300,
                                ),
                                extra_content=rx.center(
                                    rx.hstack(
                                        rx.text(State.compliance_percentage, font_size="4xl", font_weight="bold", color="#10b981"),
                                        rx.text("%", font_size="xl", font_weight="bold", color="#10b981"),
                                    ),
                                    margin_top="-150px",
                                    margin_bottom="110px",
                                    pointer_events="none"
                                )
                            ),

                            # 2. Tipos de Solicitud (Barras)
                            chart_card(
                                "Volumen por Tipo",
                                rc.bar_chart(
                                    rc.x_axis(data_key="name"),
                                    rc.y_axis(domain=[0, State.max_registros_tipo]),
                                    rc.tooltip(),
                                    rc.bar(data_key="cantidad", fill="#3b82f6", radius=[6, 6, 0, 0]),
                                    data=State.data_grafica_tipo,
                                    width="100%",
                                    height=300,
                                )
                            ),

                            # 3. Tiempos de Respuesta (Línea)
                            chart_card(
                                "Tiempos Promedio de Respuesta",
                                rc.line_chart(
                                    rc.x_axis(data_key="month"),
                                    rc.y_axis(),
                                    rc.tooltip(),
                                    rc.line(type="monotone", data_key="dias", stroke="#8b5cf6", stroke_width=4, dot={"r": 5, "fill": "#8b5cf6", "stroke": "#ffffff", "strokeWidth": 2}),
                                    data=State.monthly_response_times,
                                    width="100%",
                                    height=300,
                                ),
                                extra_content=rx.box(),
                            ),

                            # 4. Solicitudes por vencer (Barras semáforo)
                            chart_card(
                                "Estado de Vencimiento",
                                rx.box(
                                    # Gráfico principal
                                    rc.bar_chart(
                                        rc.x_axis(data_key="name"),
                                        rc.y_axis(),
                                        rc.tooltip(),
                                        rc.bar(
                                            rc.cell(fill="#ef4444"),
                                            rc.cell(fill="#f59e0b"),
                                            rc.cell(fill="#f59e0b"),
                                            rc.cell(fill="#10b981"),
                                            data_key="cantidad",
                                            radius=[6, 6, 0, 0]
                                        ),
                                        data=State.solicitudes_por_vencer_data,
                                        width="100%",
                                        height=300,
                                    ),
                                    # Capa superpuesta con botones invisibles sobre cada barra para capturar clicks
                                    rx.hstack(
                                        rx.box(width="25%", height="100%", cursor="pointer", on_click=lambda: State.abrir_vencimiento_modal("Vencidas")),
                                        rx.box(width="25%", height="100%", cursor="pointer", on_click=lambda: State.abrir_vencimiento_modal("1-5 días")),
                                        rx.box(width="25%", height="100%", cursor="pointer", on_click=lambda: State.abrir_vencimiento_modal("6-10 días")),
                                        rx.box(width="25%", height="100%", cursor="pointer", on_click=lambda: State.abrir_vencimiento_modal(">10 días")),
                                        position="absolute",
                                        top="0",
                                        left="10%",  # offset for y-axis
                                        width="90%",
                                        height="85%", # offset for x-axis
                                        z_index="10",
                                        opacity="0"
                                    ),
                                    position="relative",
                                    width="100%"
                                ),
                                extra_content=rx.center(
                                    rx.hstack(
                                        rx.button("Vencidas", size="1", variant="soft", color_scheme="red", radius="full", on_click=lambda: State.abrir_vencimiento_modal("Vencidas")),
                                        rx.button("1-5 días", size="1", variant="soft", color_scheme="orange", radius="full", on_click=lambda: State.abrir_vencimiento_modal("1-5 días")),
                                        rx.button("6-10 días", size="1", variant="soft", color_scheme="yellow", radius="full", on_click=lambda: State.abrir_vencimiento_modal("6-10 días")),
                                        rx.button(">10 días", size="1", variant="soft", color_scheme="green", radius="full", on_click=lambda: State.abrir_vencimiento_modal(">10 días")),
                                        spacing="2",
                                        flex_wrap="wrap",
                                        justify_content="center"
                                    ),
                                    width="100%",
                                    margin_top="2"
                                )
                            ),
                            columns={"base": "1", "lg": "2"},
                            spacing="6",
                            width="100%",
                        ),

                        spacing="6",
                        flex="1",
                        width="100%",
                        align_items="stretch"
                    ),

                    # Columna Derecha: Sidebar con Tarjetas Premium
                    rx.vstack(
                        # 1. Resumen numérico
                        rx.box(
                            rx.vstack(
                                rx.hstack(
                                    rx.icon("activity", size=16, color="#2563eb"),
                                    rx.text("Resumen de Actividad", font_size="sm", font_weight="700", color=rx.color_mode_cond(light="#1e293b", dark="#f8fafc")),
                                    spacing="2", align_items="center"
                                ),
                                rx.divider(border_color=rx.color_mode_cond(light="#e2e8f0", dark="#334155")),
                                rx.hstack(
                                    rx.text("Solicitudes Totales", font_size="13px", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8"), flex="1"),
                                    rx.text(State.numero_solicitudes, font_size="13px", font_weight="700", color=rx.color_mode_cond(light="#1e293b", dark="#f8fafc")),
                                    width="100%", align_items="center",
                                    padding_y="2",
                                    border_bottom=f"1px solid {rx.color_mode_cond(light='#f1f5f9', dark='#1e293b')}"
                                ),
                                rx.hstack(
                                    rx.text("Tiempo Promedio Cierre", font_size="13px", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8"), flex="1"),
                                    rx.text("10 min", font_size="13px", font_weight="700", color=rx.color_mode_cond(light="#1e293b", dark="#f8fafc")),
                                    width="100%", align_items="center",
                                    padding_y="2"
                                ),
                                spacing="3", align_items="stretch", width="100%"
                            ),
                            p="5",
                            bg=rx.color_mode_cond(light="#ffffff", dark="#1e293b"),
                            border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155"),
                            border_radius="2xl",
                            box_shadow="0 4px 6px -1px rgba(0,0,0,0.05)",
                            width="100%"
                        ),

                        # 2. Áreas Destacadas
                        rx.box(
                            rx.vstack(
                                rx.hstack(
                                    rx.icon("building-2", size=16, color="#8b5cf6"),
                                    rx.text("Áreas con más solicitudes", font_size="sm", font_weight="700", color=rx.color_mode_cond(light="#1e293b", dark="#f8fafc")),
                                    spacing="2", align_items="center"
                                ),
                                rx.divider(border_color=rx.color_mode_cond(light="#e2e8f0", dark="#334155")),
                                rx.input(
                                    placeholder="Filtrar área...",
                                    value=State.search_area_query,
                                    on_change=State.set_search_area_query,
                                    variant="surface",
                                    radius="full",
                                    size="1",
                                    width="100%"
                                ),
                                rx.box(
                                    rx.vstack(
                                        rx.foreach(
                                            State.top_areas,
                                            lambda row: rx.hstack(
                                                rx.box(
                                                    rx.icon("users", size=14, color=rx.color_mode_cond(light="#64748b", dark="#94a3b8")),
                                                    p="1",
                                                    bg=rx.color_mode_cond(light="#f1f5f9", dark="rgba(255,255,255,0.05)"),
                                                    border_radius="md",
                                                    flex_shrink="0"
                                                ),
                                                rx.text(row.get('name'), font_size="13px", font_weight="medium", color=rx.color_mode_cond(light="#334155", dark="#cbd5e1"), flex="1", overflow="hidden", text_overflow="ellipsis", white_space="nowrap"),
                                                rx.badge(row.get('total'), color_scheme="purple", variant="soft", radius="full", size="1"),
                                                width="100%",
                                                padding_y="2",
                                                border_bottom=f"1px solid {rx.color_mode_cond(light='#f1f5f9', dark='#1e293b')}",
                                                align_items="center",
                                                spacing="2"
                                            )
                                        ),
                                        spacing="0", width="100%"
                                    ),
                                    max_height="240px",
                                    overflow_y="auto",
                                    width="100%",
                                    padding_right="1"
                                ),
                                spacing="3", align_items="stretch", width="100%"
                            ),
                            p="5",
                            bg=rx.color_mode_cond(light="#ffffff", dark="#1e293b"),
                            border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155"),
                            border_radius="2xl",
                            box_shadow="0 4px 6px -1px rgba(0,0,0,0.05)",
                            width="100%"
                        ),

                        # 3. Estados Resumen
                        rx.box(
                            rx.vstack(
                                rx.hstack(
                                    rx.icon("layers", size=16, color="#f59e0b"),
                                    rx.text("Estados actuales", font_size="sm", font_weight="700", color=rx.color_mode_cond(light="#1e293b", dark="#f8fafc")),
                                    spacing="2", align_items="center"
                                ),
                                rx.divider(border_color=rx.color_mode_cond(light="#e2e8f0", dark="#334155")),
                                rx.vstack(
                                    rx.foreach(
                                        State.data_grafica_estado,
                                        lambda row: rx.hstack(
                                            rx.box(
                                                width="8px", height="8px",
                                                border_radius="full",
                                                bg=rx.match(
                                                    row.get("name"),
                                                    ("Radicada", "#f59e0b"),
                                                    ("Asignada a Area", "#3b82f6"),
                                                    ("En Gestion de Area", "#8b5cf6"),
                                                    ("Solucionada", "#10b981"),
                                                    ("Reabierta", "#eab308"),
                                                    ("Cerrada", "#16a34a"),
                                                    "#64748b"
                                                ),
                                                flex_shrink="0"
                                            ),
                                            rx.text(row.get("name", ""), font_size="13px", color=rx.color_mode_cond(light="#334155", dark="#cbd5e1"), flex="1"),
                                            rx.text(row.get("cantidad", 0), font_size="13px", font_weight="700", color=rx.color_mode_cond(light="#1e293b", dark="#f8fafc")),
                                            spacing="3", align_items="center",
                                            padding_y="2",
                                            border_bottom=f"1px solid {rx.color_mode_cond(light='#f1f5f9', dark='#1e293b')}",
                                            width="100%"
                                        )
                                    ),
                                    spacing="0", width="100%"
                                ),
                                spacing="3", align_items="stretch", width="100%"
                            ),
                            p="5",
                            bg=rx.color_mode_cond(light="#ffffff", dark="#1e293b"),
                            border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155"),
                            border_radius="2xl",
                            box_shadow="0 4px 6px -1px rgba(0,0,0,0.05)",
                            width="100%"
                        ),
                        spacing="6",
                        width={"base": "100%", "lg": "320px"},
                        flex_shrink="0",
                        align_items="stretch"
                    ),
                    flex_direction={"base": "column", "lg": "row"},
                    spacing="6",
                    width="100%",
                    align_items="start",
                ),

                spacing="6",
                width="100%",
                max_width="1400px",
                padding_y="8"
            ),
            width="100%",
        ),
        
        # Modal de Vencimiento de Solicitudes
        rx.cond(
            State.vencimiento_modal_abierto,
            rx.box(
                rx.vstack(
                    rx.box(
                        # Orbes Decorativos Translúcidos
                        rx.box(position="absolute", top="-40px", left="-20%", width="150px", height="150px", 
                               bg=rx.cond(State.rango_vencimiento_seleccionado == "Vencidas", "rgba(239, 68, 68, 0.6)", 
                                          rx.cond(State.rango_vencimiento_seleccionado == "1-5 días", "rgba(245, 158, 11, 0.6)", "rgba(16, 185, 129, 0.6)")), 
                               border_radius="full", filter="blur(40px)"),
                        
                        rx.hstack(
                            rx.box(
                                rx.icon("layers", color="white", size=28),
                                bg="linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%)",
                                p="3",
                                border_radius="2xl",
                                box_shadow="0 10px 25px -5px rgba(59, 130, 246, 0.5)",
                            ),
                            rx.vstack(
                                rx.heading(
                                    rx.cond(
                                        State.rango_vencimiento_seleccionado == "Vencidas",
                                        "Solicitudes Vencidas",
                                        rx.cond(
                                            State.rango_vencimiento_seleccionado == "1-5 días",
                                            "Críticas (1-5 días)",
                                            rx.cond(
                                                State.rango_vencimiento_seleccionado == "6-10 días",
                                                "Por Vencer (6-10 días)",
                                                "A Salvo (>10 días)"
                                            )
                                        )
                                    ),
                                    size="6", font_weight="900", background_image="linear-gradient(90deg, #ffffff, #e2e8f0)", background_clip="text", color="transparent"
                                ),
                                rx.hstack(
                                    rx.text(
                                        "Total:",
                                        color="rgba(255,255,255,0.9)",
                                        font_size="sm",
                                        font_weight="bold"
                                    ),
                                    rx.badge(
                                        State.solicitudes_vencimiento_filtradas.length(),
                                        color_scheme="blue",
                                        variant="solid",
                                        radius="full"
                                    ),
                                    rx.text(
                                        "solicitudes filtradas en este rango.",
                                        color="rgba(255,255,255,0.8)",
                                        font_size="sm",
                                        font_weight="medium"
                                    ),
                                    spacing="2",
                                    align_items="center"
                                ),
                                spacing="1"
                            ),
                            spacing="4",
                            align_items="center",
                            position="relative",
                            z_index="2",
                            width="100%"
                        ),
                        p="8",
                        bg="linear-gradient(135deg, #0f172a 0%, #1e293b 100%)",
                        border_bottom="1px solid rgba(255,255,255,0.05)",
                        border_top_left_radius="3xl",
                        border_top_right_radius="3xl",
                        position="relative",
                        overflow="hidden",
                        width="100%"
                    ),
                    
                    # Contenido / Tabla de Solicitudes
                    rx.box(
                        rx.cond(
                            State.solicitudes_vencimiento_filtradas,
                            rx.vstack(
                                rx.table.root(
                                    rx.table.header(
                                        rx.table.row(
                                            rx.table.column_header_cell("Radicado"),
                                            rx.table.column_header_cell("Tipo"),
                                            rx.table.column_header_cell("Asunto"),
                                            rx.table.column_header_cell("Área Responsable"),
                                            rx.table.column_header_cell("Días Restantes"),
                                            rx.table.column_header_cell("Acción"),
                                        )
                                    ),
                                    rx.table.body(
                                        rx.foreach(
                                            State.solicitudes_vencimiento_filtradas,
                                            lambda s: rx.table.row(
                                                rx.table.row_header_cell(
                                                    rx.badge(s.get("radicado"), color_scheme="blue", variant="surface", radius="full")
                                                ),
                                                rx.table.cell(s.get("tipo_solicitud")),
                                                rx.table.cell(s.get("asunto")),
                                                rx.table.cell(s.get("area_responsable")),
                                                rx.table.cell(
                                                    rx.hstack(
                                                        rx.box(
                                                            width="8px",
                                                            height="8px",
                                                            bg=s.get("semaforo_fill"),
                                                            border_radius="full"
                                                        ),
                                                        rx.text(
                                                            s.get("remaining_str", ""),
                                                            font_weight="semibold",
                                                            color=s.get("semaforo_fill")
                                                        ),
                                                        spacing="2",
                                                        align_items="center"
                                                    )
                                                ),
                                                rx.table.cell(
                                                    rx.button(
                                                        rx.icon("eye", size=16),
                                                        on_click=lambda: State.abrir_detalle_solicitud(s.get("id")),
                                                        variant="soft",
                                                        color_scheme="blue",
                                                        size="1",
                                                        radius="full"
                                                    )
                                                )
                                            )
                                        )
                                    ),
                                    width="100%",
                                    variant="ghost"
                                ),
                                spacing="4",
                                width="100%"
                            ),
                            # Estado Vacío
                            rx.center(
                                rx.vstack(
                                    rx.icon("inbox", size=48, color="gray"),
                                    rx.text("No se encontraron solicitudes pendientes en este rango.", font_weight="semibold", color="gray"),
                                    spacing="2",
                                    padding_y="8"
                                ),
                                width="100%"
                            )
                        ),
                        p="6",
                        max_height="400px",
                        overflow_y="auto",
                        width="100%",
                        bg=rx.color_mode_cond(light="rgba(255, 255, 255, 0.4)", dark="rgba(15, 23, 42, 0.4)")
                    ),
                    
                    # Botón de Cerrar
                    rx.box(
                        rx.button(
                            "Cerrar Ventana",
                            on_click=State.cerrar_vencimiento_modal,
                            bg="linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%)",
                            color="white",
                            size="3",
                            width="100%",
                            box_shadow="0 4px 14px 0 rgba(59,130,246,0.39)",
                            _hover={"transform": "translateY(-2px)", "box_shadow": "0 6px 20px rgba(59,130,246,0.5)"},
                            transition="all 0.2s"
                        ),
                        p="5",
                        border_top=rx.color_mode_cond(light="1px solid rgba(0,0,0,0.05)", dark="1px solid rgba(255,255,255,0.05)"),
                        bg=rx.color_mode_cond(light="rgba(248, 250, 252, 0.4)", dark="rgba(11, 17, 32, 0.4)"),
                        border_bottom_left_radius="3xl",
                        border_bottom_right_radius="3xl",
                        width="100%"
                    ),
                    spacing="0",
                ),
                p="0",
                border=rx.color_mode_cond(light="1px solid rgba(255,255,255,0.6)", dark="1px solid rgba(255,255,255,0.08)"),
                border_radius="3xl",
                bg=rx.color_mode_cond(light="rgba(255, 255, 255, 0.9)", dark="rgba(15, 23, 42, 0.8)"),
                backdrop_filter="blur(20px)",
                width="100%",
                max_width="850px",
                position="fixed",
                top="50%",
                left="50%",
                transform="translate(-50%, -50%)",
                z_index="1000",
                box_shadow=rx.color_mode_cond(light="0 25px 50px -12px rgba(0, 0, 0, 0.25), 0 0 0 1px rgba(0,0,0,0.05)", dark="0 25px 50px -12px rgba(0, 0, 0, 0.5), 0 0 0 1px rgba(255,255,255,0.1)")
            )
        ),
        
        # Modal de Detalles de la Solicitud Seleccionada
        rx.cond(
            State.detalle_solicitud_modal_abierto,
            rx.box(
                # Fondo oscuro semitransparente detrás del modal
                rx.box(
                    position="fixed",
                    top="0",
                    left="0",
                    width="100vw",
                    height="100vh",
                    bg="rgba(15, 23, 42, 0.75)",
                    backdrop_filter="blur(10px)",
                    z_index="1050",
                    on_click=State.cerrar_detalle_solicitud
                ),
                # Contenido del Modal de Detalles
                rx.box(
                    rx.vstack(
                        # Encabezado con Gradiente Premium
                        rx.box(
                            rx.vstack(
                                rx.hstack(
                                    rx.icon("info", size=24, color="#3b82f6"),
                                    rx.heading(
                                        f"Detalles de Solicitud: {State.solicitud_consultada.get('radicado', '')}",
                                        size="5",
                                        color="#ffffff"
                                    ),
                                    rx.spacer(),
                                    rx.button(
                                        rx.icon("x", size=20),
                                        on_click=State.cerrar_detalle_solicitud,
                                        variant="ghost",
                                        color="#ffffff",
                                        _hover={"bg": "rgba(255,255,255,0.1)"}
                                    ),
                                    width="100%",
                                    align_items="center"
                                ),
                                spacing="1",
                                align_items="start"
                            ),
                            p="6",
                            bg="linear-gradient(135deg, #0f172a 0%, #1e293b 100%)",
                            border_bottom="1px solid rgba(255,255,255,0.05)",
                            border_top_left_radius="3xl",
                            border_top_right_radius="3xl",
                            position="relative",
                            overflow="hidden",
                            width="100%"
                        ),
                        
                        # Cuerpo del Modal (Scrollable)
                        rx.vstack(
                            rx.grid(
                                rx.vstack(
                                    rx.text("Número de Radicado:", font_weight="bold", color=rx.color_mode_cond(light="#475569", dark="#94a3b8"), font_size="xs", text_transform="uppercase"),
                                    rx.text(State.solicitud_consultada.get("radicado", ""), color=rx.color_mode_cond(light="#0f172a", dark="#ffffff"), font_size="sm", font_weight="semibold")
                                ),
                                rx.vstack(
                                    rx.text("Tipo de Solicitud:", font_weight="bold", color=rx.color_mode_cond(light="#475569", dark="#94a3b8"), font_size="xs", text_transform="uppercase"),
                                    rx.text(State.solicitud_consultada.get("tipo_solicitud", ""), color=rx.color_mode_cond(light="#0f172a", dark="#ffffff"), font_size="sm", font_weight="semibold")
                                ),
                                rx.vstack(
                                    rx.text("Estado Actual:", font_weight="bold", color=rx.color_mode_cond(light="#475569", dark="#94a3b8"), font_size="xs", text_transform="uppercase"),
                                    rx.badge(
                                        State.solicitud_consultada.get("estado", ""),
                                        color_scheme=rx.cond(
                                            State.solicitud_consultada.get("estado") == "Radicada",
                                            "orange",
                                            rx.cond(
                                                State.solicitud_consultada.get("estado") == "Actualizada",
                                                "blue",
                                                "green"
                                            )
                                        ),
                                        radius="full",
                                        variant="solid"
                                    )
                                ),
                                rx.vstack(
                                    rx.text("Fecha de Radicación:", font_weight="bold", color=rx.color_mode_cond(light="#475569", dark="#94a3b8"), font_size="xs", text_transform="uppercase"),
                                    rx.text(State.solicitud_consultada.get("fecha", ""), color=rx.color_mode_cond(light="#0f172a", dark="#ffffff"), font_size="sm")
                                ),
                                rx.vstack(
                                    rx.text("Área Responsable:", font_weight="bold", color=rx.color_mode_cond(light="#475569", dark="#94a3b8"), font_size="xs", text_transform="uppercase"),
                                    rx.text(State.solicitud_consultada.get("area_responsable", "No asignada"), color=rx.color_mode_cond(light="#0f172a", dark="#ffffff"), font_size="sm")
                                ),
                                rx.vstack(
                                    rx.text("Creado Por:", font_weight="bold", color=rx.color_mode_cond(light="#475569", dark="#94a3b8"), font_size="xs", text_transform="uppercase"),
                                    rx.text(State.solicitud_consultada.get("creado_por", ""), color=rx.color_mode_cond(light="#0f172a", dark="#ffffff"), font_size="sm")
                                ),
                                rx.cond(
                                    State.solicitud_consultada.get("calificacion_servicio") != None,
                                    rx.vstack(
                                        rx.text("Calificación:", font_weight="bold", color=rx.color_mode_cond(light="#475569", dark="#94a3b8"), font_size="xs", text_transform="uppercase"),
                                        rx.hstack(
                                            rx.icon("star", color="#f59e0b", size=16),
                                            rx.text(State.solicitud_consultada.get("calificacion_servicio", ""), font_weight="bold", color="#f59e0b", font_size="sm"),
                                            spacing="1", align_items="center"
                                        )
                                    )
                                ),
                                template_columns={"base": "1fr", "sm": "repeat(2, 1fr)", "md": "repeat(3, 1fr)"},
                                gap="4",
                                width="100%",
                                p="4",
                                border_radius="2xl",
                                bg=rx.color_mode_cond(light="#f1f5f9", dark="rgba(255,255,255,0.02)"),
                                border="1px dashed rgba(128,128,128,0.2)"
                            ),
                            
                            rx.vstack(
                                rx.text("Asunto:", font_weight="bold", color=rx.color_mode_cond(light="#475569", dark="#94a3b8"), font_size="xs", text_transform="uppercase"),
                                rx.text(State.solicitud_consultada.get("asunto", ""), color=rx.color_mode_cond(light="#0f172a", dark="#ffffff"), font_weight="bold", font_size="md"),
                                width="100%",
                                align_items="start"
                            ),
                            
                            # Descripción
                            rx.vstack(
                                rx.text("Descripción Detallada:", font_weight="bold", color=rx.color_mode_cond(light="#475569", dark="#94a3b8"), font_size="xs", text_transform="uppercase"),
                                rx.box(
                                    rx.text(State.solicitud_consultada.get("descripcion", ""), color=rx.color_mode_cond(light="#1e293b", dark="#cbd5e1"), font_size="sm", white_space="pre-wrap"),
                                    p="4",
                                    border=rx.color_mode_cond(light="1px solid #cbd5e0", dark="1px solid rgba(255,255,255,0.08)"),
                                    border_radius="2xl",
                                    bg=rx.color_mode_cond(light="#f8fafc", dark="rgba(15,23,42,0.6)"),
                                    width="100%"
                                ),
                                width="100%",
                                align_items="start"
                            ),
                            
                            # Respuesta
                            rx.cond(
                                State.solicitud_consultada.get("respuesta"),
                                rx.vstack(
                                    rx.hstack(
                                        rx.icon("message_square_plus", size=18, color="#10b981"),
                                        rx.text("Respuesta del Funcionario:", font_weight="bold", color="#10b981", font_size="xs", text_transform="uppercase"),
                                        spacing="2"
                                    ),
                                    rx.box(
                                        rx.text(State.solicitud_consultada.get("respuesta", ""), color=rx.color_mode_cond(light="#14532d", dark="#a7f3d0"), font_size="sm", white_space="pre-wrap"),
                                        p="4",
                                        border="1px solid #10b981",
                                        border_radius="2xl",
                                        bg=rx.color_mode_cond(light="#f0fff4", dark="rgba(16,185,129,0.08)"),
                                        width="100%"
                                    ),
                                    width="100%",
                                    align_items="start"
                                )
                            ),
                            rx.cond(
                                State.solicitud_consultada_tiene_adjunto_respuesta,
                                rx.vstack(
                                    rx.hstack(
                                        rx.icon("paperclip", size=16, color="#10b981"),
                                        rx.text("Adjunto de respuesta (funcionario):", font_weight="bold", color="#10b981", font_size="xs", text_transform="uppercase"),
                                        spacing="2",
                                    ),
                                    rx.cond(
                                        State.solicitud_consultada_adjunto_respuesta_existe,
                                        rx.vstack(
                                            rx.cond(
                                                State.solicitud_consultada_respuesta_documento_preview_src != "",
                                                rx.image(
                                                    src=State.solicitud_consultada_respuesta_documento_preview_src,
                                                    alt="Adjunto de respuesta",
                                                    width="100%",
                                                    max_height="120px",
                                                    object_fit="contain",
                                                    border_radius="lg",
                                                ),
                                                rx.box(),
                                            ),
                                            ui_boton_descargar_adjunto(
                                                State.solicitud_consultada_respuesta_documento_href,
                                                State.solicitud_consultada_respuesta_documento_basename,
                                                color_scheme="green",
                                            ),
                                            spacing="2",
                                            width="100%",
                                            align_items="start",
                                        ),
                                        rx.text("El archivo ya no está en el servidor", font_size="xs", color="#ef4444"),
                                    ),
                                    width="100%",
                                    align_items="start",
                                    spacing="2",
                                ),
                            ),
                            # Adjuntos Ciudadano
                            rx.cond(
                                State.solicitud_consultada.get("documento_adjuntos"),
                                rx.vstack(
                                    rx.hstack(
                                        rx.icon("paperclip", size=16, color="#3b82f6"),
                                        rx.text("Documentos Adjuntos por Ciudadano:", font_weight="bold", color=rx.color_mode_cond(light="#475569", dark="#94a3b8"), font_size="xs", text_transform="uppercase"),
                                        spacing="2"
                                    ),
                                    rx.vstack(
                                        rx.foreach(
                                            State.solicitud_consultada_adjuntos,
                                            lambda doc: ui_tarjeta_adjunto(doc),
                                        ),
                                        align_items="start",
                                        spacing="2",
                                        width="100%"
                                    ),
                                    width="100%",
                                    align_items="start"
                                )
                            ),
                            rx.cond(
                                State.historial_solicitud.length() > 0,
                                rx.vstack(
                                    rx.hstack(
                                        rx.icon("history", size=16, color="#8b5cf6"),
                                        rx.text("Bitácora", font_weight="bold", font_size="xs", color="#8b5cf6", text_transform="uppercase"),
                                        spacing="2",
                                    ),
                                    rx.foreach(
                                        State.historial_solicitud,
                                        lambda h: rx.box(
                                            rx.vstack(
                                                rx.hstack(
                                                    rx.badge(h["estado_anterior"], size="1", variant="soft"),
                                                    rx.icon("arrow-right", size=12),
                                                    rx.badge(h["estado_nuevo"], size="1", color_scheme="purple", variant="soft"),
                                                    rx.spacer(),
                                                    rx.text(h["fecha_cambio"], font_size="2xs", color="#64748b"),
                                                    width="100%",
                                                    align_items="center",
                                                ),
                                                rx.cond(
                                                    h["observaciones"] != "",
                                                    rx.text(h["observaciones"], font_size="xs", white_space="pre-wrap"),
                                                    rx.box(),
                                                ),
                                                rx.cond(
                                                    h["documento_adjunto"] != "",
                                                    ui_tarjeta_adjunto_historial(h),
                                                    rx.box(),
                                                ),
                                                spacing="2",
                                                width="100%",
                                            ),
                                            p="3",
                                            border="1px solid #e2e8f0",
                                            border_radius="lg",
                                            width="100%",
                                            margin_bottom="2",
                                        ),
                                    ),
                                    width="100%",
                                    align_items="start",
                                    spacing="2",
                                ),
                            ),
                            p="6",
                            max_height="450px",
                            overflow_y="auto",
                            width="100%",
                            spacing="4"
                        ),
                        
                        # Botón de Cerrar del Modal Detalles
                        rx.box(
                            rx.button(
                                "Cerrar Detalles",
                                on_click=State.cerrar_detalle_solicitud,
                                bg="linear-gradient(135deg, #475569 0%, #334155 100%)",
                                color="#ffffff",
                                size="3",
                                width="100%",
                                box_shadow="0 4px 14px 0 rgba(100,116,139,0.3)",
                                _hover={"transform": "translateY(-2px)", "box_shadow": "0 6px 20px rgba(100,116,139,0.4)"},
                                transition="all 0.2s"
                            ),
                            p="5",
                            border_top=rx.color_mode_cond(light="1px solid rgba(0,0,0,0.05)", dark="1px solid rgba(255,255,255,0.05)"),
                            bg=rx.color_mode_cond(light="#f8fafc", dark="rgba(11, 17, 32, 0.4)"),
                            border_bottom_left_radius="3xl",
                            border_bottom_right_radius="3xl",
                            width="100%"
                        ),
                        spacing="0",
                    ),
                    p="0",
                    border=rx.color_mode_cond(light="1px solid #cbd5e1", dark="1px solid rgba(255,255,255,0.08)"),
                    border_radius="3xl",
                    bg=rx.color_mode_cond(light="#ffffff", dark="#111827"),
                    width="95%",
                    max_width="720px",
                    position="fixed",
                    top="50%",
                    left="50%",
                    transform="translate(-50%, -50%)",
                    z_index="1100",
                    box_shadow="0 30px 60px -15px rgba(0,0,0,0.5)"
                )
            )
        ),
        width="100%",
        min_height="100vh",
        bg=rx.color_mode_cond(light="#f8fafc", dark="#0f172a")
    )
    # Restringir acceso: solo funcionarios y administradores
    return rx.cond(
        State.es_autenticada & ((State.rol_usuario == "funcionario") | (State.rol_usuario == "administrador")),
        content,
        rx.box(
            navbar(),
            rx.center(
                access_denied_widget("Solo funcionarios autenticados pueden ver esta página."),
                size="3"
            )
        )
    )

def usuarios_page() -> rx.Component:
    return rx.cond(
        State.es_autenticada & (State.rol_usuario == "funcionario"),
        rx.box(
            navbar(),
            rx.box(
                rx.vstack(
                    rx.hstack(
                        rx.vstack(
                            rx.heading("Directorio de Personal", size="7", color=rx.color_mode_cond(light="#0f172a", dark="#f8fafc"), font_weight="bold", letter_spacing="-0.02em"),
                            rx.text("Gestión y control de cuentas de usuarios registrados", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8"), font_size="md"),
                            align_items="start",
                            spacing="1"
                        ),
                        rx.spacer(),
                        rx.box(
                            rx.icon("users", size=24, color=rx.color_mode_cond(light="#2563eb", dark="#60a5fa")),
                            p="3",
                            bg=rx.color_mode_cond(light="#eff6ff", dark="rgba(96, 165, 250, 0.1)"),
                            border_radius="xl"
                        ),
                        width="100%",
                        align_items="center"
                    ),
                    
                    rx.divider(margin_y="4", border_color=rx.color_mode_cond(light="#e2e8f0", dark="#334155")),

                    rx.box(
                        rx.cond(
                            State.usuarios_registrados,
                            rx.vstack(
                                rx.table.root(
                                    rx.table.header(
                                        rx.table.row(
                                            rx.table.column_header_cell("Usuario"),
                                            rx.table.column_header_cell("Contacto"),
                                            rx.table.column_header_cell("Rol", align="center"),
                                            rx.table.column_header_cell("Registro"),
                                            rx.table.column_header_cell("Estado", align="center"),
                                            rx.table.column_header_cell("Acciones", align="center"),
                                        ),
                                    ),
                                    rx.table.body(
                                        rx.foreach(
                                            State.usuarios_registrados,
                                            lambda usuario: rx.table.row(
                                                rx.table.cell(
                                                    rx.hstack(
                                                        rx.avatar(fallback=rx.cond(usuario["rol"] == "funcionario", "FN", "CD"), size="3", radius="full", color_scheme=rx.cond(usuario["rol"] == "funcionario", "blue", "gray")),
                                                        rx.vstack(
                                                            rx.text(f"{usuario['nombres']} {usuario['apellidos']}", font_weight="semibold", font_size="sm", color=rx.color_mode_cond(light="#1e293b", dark="#f8fafc")),
                                                            rx.text(f"ID: {usuario['id']}", font_size="xs", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8")),
                                                            spacing="0", align_items="start"
                                                        ),
                                                        spacing="3", align_items="center"
                                                    )
                                                ),
                                                rx.table.cell(
                                                    rx.hstack(
                                                        rx.icon("mail", size=14, color=rx.color_mode_cond(light="#64748b", dark="#94a3b8")),
                                                        rx.text(usuario["email"], font_size="sm", color=rx.color_mode_cond(light="#475569", dark="#cbd5e1")),
                                                        spacing="2", align_items="center"
                                                    )
                                                ),
                                                rx.table.cell(
                                                    rx.badge(
                                                        rx.cond(usuario["rol"] == "funcionario", rx.icon("shield", size=12), rx.icon("user", size=12)),
                                                        usuario["rol"],
                                                        color_scheme=rx.cond(usuario["rol"] == "funcionario", "blue", "gray"),
                                                        variant="soft",
                                                        radius="full",
                                                        size="2",
                                                    ),
                                                    align="center"
                                                ),
                                                rx.table.cell(rx.text(usuario["fecha_creacion"], font_size="sm", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8"))),
                                                rx.table.cell(
                                                    rx.hstack(
                                                        rx.box(width="8px", height="8px", border_radius="full", bg=rx.cond(usuario["is_active"] == "True", "#10b981", rx.cond(usuario["is_active"], "#10b981", "#ef4444"))),
                                                        rx.text(rx.cond(usuario["is_active"] == "True", "Activo", rx.cond(usuario["is_active"], "Activo", "Inactivo")), font_size="sm", color=rx.color_mode_cond(light="#475569", dark="#cbd5e1")),
                                                        spacing="2", align_items="center", justify="center"
                                                    ),
                                                    align="center"
                                                ),
                                                rx.table.cell(
                                                    rx.cond(
                                                        usuario["rol"] == "funcionario",
                                                        rx.button(
                                                            "Degradar a ciudadano",
                                                            size="2",
                                                            variant="soft",
                                                            color_scheme="red",
                                                            on_click=State.degradar_funcionario_a_ciudadano(usuario["email"]),
                                                            is_disabled=State.email_actual.to_string().lower() == usuario["email"].to_string().lower(),
                                                        ),
                                                        rx.box(),
                                                    ),
                                                    align="center",
                                                ),
                                                align_items="center",
                                                _hover={"bg": rx.color_mode_cond(light="#f8fafc", dark="rgba(30, 41, 59, 0.5)")},
                                                transition="background 0.2s"
                                            )
                                        )
                                    ),
                                    width="100%",
                                    size="3",
                                    variant="surface"
                                ),
                                rx.hstack(
                                    rx.text(f"Total de registros: {State.usuarios_registrados_count}", font_weight="medium", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8"), font_size="sm"),
                                    rx.spacer(),
                                    width="100%",
                                    padding_top="4"
                                ),
                                spacing="4",
                                width="100%"
                            ),
                            rx.vstack(
                                rx.icon("users", size=48, color=rx.color_mode_cond(light="#cbd5e1", dark="#475569")),
                                rx.text("No hay usuarios registrados en el sistema.", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8"), font_size="lg", font_weight="medium"),
                                spacing="4",
                                padding_y="12",
                                align_items="center"
                            )
                        ),
                        p="6",
                        border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155"),
                        border_radius="2xl",
                        bg=rx.color_mode_cond(light="rgba(255, 255, 255, 0.8)", dark="rgba(15, 23, 42, 0.6)"),
                        backdrop_filter="blur(16px)",
                        box_shadow=rx.color_mode_cond(light="0 4px 6px -1px rgba(0, 0, 0, 0.05)", dark="0 10px 15px -3px rgba(0, 0, 0, 0.5)"),
                        width="100%",
                        overflow_x="auto"
                    ),
                    
                    spacing="6",
                    align_items="stretch",
                    width="100%"
                ),
                bg=rx.color_mode_cond(light="white", dark="#1a202c"),
                max_width="100%",
                p=rx.breakpoints(initial="4", md="6"),
                box_shadow="2xl",
                border_radius="3xl",
                width="100%",
                margin_y="4"
            ),
            width="100%",
            max_width="100%",
            p=rx.breakpoints(initial="2", md="3"),
            bg=rx.color_mode_cond(light="#f8fafc", dark="#0f172a"),
            min_height="100vh"
        ),
        rx.box(
            navbar(),
            rx.center(
                access_denied_widget("Solo funcionarios autenticados pueden ver esta página."),
                size="3"
            )
        )
    )


def cambiar_rol_page() -> rx.Component:
    text_main = rx.color_mode_cond(light="#1e293b", dark="#f1f5f9")
    text_sub = rx.color_mode_cond(light="#64748b", dark="#94a3b8")
    card_bg = rx.color_mode_cond(light="white", dark="#1e293b")
    card_bdr = rx.color_mode_cond(light="#e2e8f0", dark="#334155")
    page_bg = rx.color_mode_cond(light="#f8fafc", dark="#0f172a")
    input_bg = rx.color_mode_cond(light="#f8fafc", dark="#0f172a")
    input_bdr = rx.color_mode_cond(light="#e2e8f0", dark="#1e293b")
    blue_acc = "#2563eb"
    navy = "#1e3a8a"
    navy_dark = "#172554"
    red_err = "#ef4444"
    green = "#10b981"

    acceso_denegado = rx.center(
        rx.vstack(
            rx.icon("shield-x", size=48, color=red_err),
            rx.heading("Acceso Denegado", size="7", color=red_err),
            rx.text("Solo funcionarios autenticados pueden acceder.", color=text_sub),
            rx.link(rx.button("Ir al Login", color_scheme="blue", border_radius="10px"), href="/login"),
            spacing="4", align_items="center",
        ),
        min_height="80vh",
    )
 
    contenido = rx.box(
        navbar(),
        rx.center(
            rx.box(
                rx.vstack(
                    # Encabezado con ícono
                    rx.vstack(
                        rx.box(
                            rx.icon("user-cog", size=28, color="white"),
                            bg=f"linear-gradient(135deg, {navy_dark}, {navy})",
                            border_radius="16px",
                            width="60px", height="60px",
                            display="flex", align_items="center", justify_content="center",
                            box_shadow=f"0 8px 24px {navy}55",
                        ),
                        rx.heading(
                            "Cambiar Rol de Usuario",
                            size="6", color=text_main,
                            font_weight="800", letter_spacing="-0.02em",
                            text_align="center",
                        ),
                        rx.text(
                            "Promueve a un ciudadano al rol de funcionario del sistema",
                            font_size="14px", color=text_sub, text_align="center",
                        ),
                        spacing="3", align_items="center", width="100%",
                    ),
                    # Tarjeta formulario
                    rx.box(
                        rx.vstack(
                            # Aviso informativo
                            rx.hstack(
                                rx.box(
                                    rx.icon("info", size=15, color=blue_acc),
                                    width="30px", height="30px",
                                    border_radius="8px",
                                    bg=rx.color_mode_cond(light="#eff6ff", dark="#0f2744"),
                                    display="flex", align_items="center", justify_content="center",
                                    flex_shrink="0",
                                ),
                                rx.text(
                                    "Al promover un ciudadano, tendrá acceso al panel de "
                                    "funcionario y recibirá una notificación por correo.",
                                    font_size="13px", color=text_sub, line_height="1.6",
                                ),
                                spacing="3", align_items="start",
                                bg=rx.color_mode_cond(light="#eff6ff", dark="#0a1628"),
                                border=rx.color_mode_cond(light="1px solid #bfdbfe", dark="1px solid #1e3a5f"),
                                border_radius="10px",
                                padding="12px 16px",
                                width="100%",
                            ),
                            # Campo correo
                            rx.vstack(
                                rx.hstack(
                                    rx.text("Correo del usuario a promover", font_size="13px", font_weight="600", color=text_main),
                                    rx.text("*", color="#e85d04", font_size="13px"),
                                    spacing="1",
                                ),
                                rx.box(
                                    rx.hstack(
                                        rx.icon("mail", size=16, color=text_sub),
                                        rx.input(
                                            placeholder="usuario@ejemplo.com",
                                            value=State.cambiar_rol_email,
                                            on_change=State.set_cambiar_rol_email,
                                            type="email",
                                            border="none",
                                            bg="transparent",
                                            color=text_main,
                                            font_size="14px",
                                            flex="1",
                                            _focus={"outline": "none"},
                                        ),
                                        spacing="2", align_items="center", width="100%",
                                    ),
                                    bg=input_bg,
                                    border=f"1.5px solid {input_bdr}",
                                    border_radius="10px",
                                    padding="10px 14px",
                                    width="100%",
                                    _focus_within={
                                        "border_color": blue_acc,
                                        "box_shadow": "0 0 0 3px rgba(37,99,235,0.12)",
                                    },
                                ),
                                spacing="2", align_items="start", width="100%",
                            ),
                            
                            # Checkbox para PQRS-1 backend
                            rx.checkbox(
                                "Confirmo validación de identidad y autorización.",
                                is_checked=State.confirmar_promocion_rol,
                                on_change=State.set_confirmar_promocion_rol,
                                color=text_sub,
                                size="2",
                            ),

                            # Mensaje resultado
                            rx.cond(
                                State.cambiar_rol_mensaje != "",
                                rx.hstack(
                                    rx.icon(
                                        rx.cond(State.cambiar_rol_mensaje.contains("�??"), "circle-check", "circle-x"),
                                        size=16,
                                        color=rx.cond(State.cambiar_rol_mensaje.contains("�??"), green, red_err),
                                    ),
                                    rx.text(
                                        State.cambiar_rol_mensaje,
                                        font_size="13px", font_weight="500",
                                        color=rx.cond(State.cambiar_rol_mensaje.contains("�??"), green, red_err),
                                    ),
                                    spacing="2", align_items="center",
                                    bg=rx.cond(State.cambiar_rol_mensaje.contains("�??"), rx.color_mode_cond(light="#f0fdf4", dark="#002818"), rx.color_mode_cond(light="#fef2f2", dark="#2d0000")),
                                    border=rx.cond(State.cambiar_rol_mensaje.contains("�??"), "1px solid #bbf7d0", "1px solid #fecaca"),
                                    border_radius="8px",
                                    padding="10px 14px",
                                    width="100%",
                                ),
                            ),
                            # Botón
                            rx.button(
                                rx.hstack(
                                    rx.icon("arrow-up-circle", size=16),
                                    rx.text("Promover a Funcionario", font_size="14px", font_weight="600"),
                                    spacing="2",
                                ),
                                on_click=State.cambiar_rol_ciudadano_a_funcionario,
                                width="100%", height="46px",
                                bg=f"linear-gradient(135deg, {navy_dark}, {navy})",
                                color="white",
                                border_radius="10px",
                                box_shadow=f"0 4px 14px {navy}44",
                                _hover={"opacity": "0.92", "transform": "translateY(-1px)"},
                                transition="all 0.15s ease",
                                is_disabled=(State.cambiar_rol_email == "") | (~State.confirmar_promocion_rol),
                            ),
                            spacing="4", align_items="stretch", width="100%",
                        ),
                        bg=card_bg,
                        border=f"1px solid {card_bdr}",
                        border_radius="18px",
                        padding="28px 30px",
                        box_shadow=rx.color_mode_cond(light="0 4px 24px rgba(0,0,0,0.07)", dark="0 4px 32px rgba(0,0,0,0.4)"),
                        width="100%",
                    ),
                    spacing="6", align_items="center", width="100%",
                ),
                width="100%", max_width="480px",
                padding={"base": "24px 16px", "md": "48px 24px"},
            ),
            width="100%", min_height="90vh",
        ),
        bg=page_bg, min_height="100vh", width="100%",
    )
 
    return rx.cond(
        State.es_autenticada & (State.rol_usuario == "funcionario"),
        contenido, acceso_denegado,
    )


def ayuda_funcionario_page() -> rx.Component:
    def acordeon_item(titulo: str, key: str, icono: str, contenido: rx.Component) -> rx.Component:
        abierto = State.ayuda_seccion_abierta == key
        return rx.box(
            rx.vstack(
                rx.button(
                    rx.hstack(
                        rx.hstack(
                            rx.box(
                                rx.icon(icono, size=18, color=rx.color_mode_cond(light="#2563eb", dark="#60a5fa")),
                                p="2",
                                bg=rx.color_mode_cond(light="#eff6ff", dark="rgba(96, 165, 250, 0.12)"),
                                border_radius="lg",
                            ),
                            rx.text(titulo, font_weight="bold", font_size="lg"),
                            spacing="3",
                            align_items="center",
                        ),
                        rx.spacer(),
                        rx.icon(rx.cond(abierto, "chevron_up", "chevron_down"), size=20),
                        width="100%",
                        align_items="center",
                    ),
                    on_click=State.toggle_ayuda_seccion(key),
                    variant="ghost",
                    width="100%",
                    justify="start",
                    p="4",
                ),
                rx.cond(
                    abierto,
                    rx.box(contenido, px="6", pb="6", width="100%"),
                    rx.box(),
                ),
                spacing="0",
                width="100%",
                align_items="start",
            ),
            bg=rx.color_mode_cond(light="rgba(255,255,255,0.95)", dark="rgba(30,41,59,0.9)"),
            border=rx.color_mode_cond(light="1px solid #dbeafe", dark="1px solid #334155"),
            border_radius="2xl",
            box_shadow=rx.color_mode_cond(light="0 10px 20px -12px rgba(0,0,0,0.12)", dark="0 10px 25px -15px rgba(0,0,0,0.7)"),
            width="100%",
        )

    hero = rx.box(
        rx.vstack(
            rx.hstack(
                rx.box(
                    rx.icon("circle_help", size=38, color="white"),
                    p="4",
                    bg="linear-gradient(135deg, #2563eb 0%, #7c3aed 100%)",
                    border_radius="2xl",
                    box_shadow="0 10px 25px -10px rgba(59,130,246,0.7)",
                ),
                rx.vstack(
                    rx.heading("Centro de Ayuda para Funcionarios", size="9", color="white", font_weight="bold"),
                    rx.text(
                        "Todo lo que necesitas para operar PQRS con rapidez, trazabilidad y calidad de servicio.",
                        color="rgba(255,255,255,0.9)",
                        font_size="lg",
                    ),
                    spacing="2",
                    align_items="start",
                ),
                spacing="4",
                width="100%",
                align_items="center",
            ),
            rx.grid(
                rx.box(
                    rx.text("Estado sugerido", font_size="sm", color="rgba(255,255,255,0.8)"),
                    rx.text("Radicada �?? En Proceso �?? Cerrada", font_weight="bold", color="white", font_size="lg"),
                ),
                rx.box(
                    rx.text("Tiempo de respuesta", font_size="sm", color="rgba(255,255,255,0.8)"),
                    rx.text("Métrica en días hábiles reales", font_weight="bold", color="white", font_size="lg"),
                ),
                rx.box(
                    rx.text("Objetivo", font_size="sm", color="rgba(255,255,255,0.8)"),
                    rx.text("Cerrar más rápido y mejor", font_weight="bold", color="white", font_size="lg"),
                ),
                columns={"base": "1", "md": "3"},
                spacing="4",
                width="100%",
            ),
            spacing="5",
            width="100%",
        ),
        p={"base": "6", "md": "8"},
        border_radius="3xl",
        bg="linear-gradient(120deg, #0f172a 0%, #1e3a8a 45%, #6d28d9 100%)",
        width="100%",
    )

    return rx.cond(
        State.es_autenticada & (State.rol_usuario == "funcionario"),
        rx.box(
            navbar(),
            rx.box(
                rx.box(
                    rx.vstack(
                        hero,
                        rx.vstack(
                        acordeon_item(
                            "1. Flujo recomendado de atención",
                            "estados",
                            "list_checks",
                            rx.vstack(
                                rx.text("Radicada: se recibió la solicitud y aún no ha sido gestionada.", font_size="lg"),
                                rx.text("En Proceso: ya existe análisis, contacto o trabajo activo sobre el caso.", font_size="lg"),
                                rx.text("Cerrada: se entregó respuesta final y, de ser necesario, adjuntos de soporte.", font_size="lg"),
                                rx.text("Consejo: evita cierres vacíos; documenta siempre la solución.", font_size="lg", font_weight="medium", color=rx.color_mode_cond(light="#475569", dark="#cbd5e1")),
                                spacing="2",
                                width="100%",
                                align_items="start",
                            ),
                        ),
                        acordeon_item(
                            "2. Respuestas y documentos adjuntos",
                            "adjuntos",
                            "paperclip",
                            rx.vstack(
                                rx.text("En 'Actualizar estado' redacta una respuesta clara y orientada al ciudadano.", font_size="lg"),
                                rx.text("Si adjuntas evidencia (PDF/imagen), el sistema la envía por correo al cerrar.", font_size="lg"),
                                rx.text("Nombra archivos de forma descriptiva para facilitar auditorías y seguimiento.", font_size="lg"),
                                spacing="2",
                                width="100%",
                                align_items="start",
                            ),
                        ),
                        acordeon_item(
                            "3. Reportes y lectura de métricas",
                            "reportes",
                            "chart_line",
                            rx.vstack(
                                rx.text("La gráfica de tiempos usa solicitudes cerradas con fecha real de respuesta.", font_size="lg"),
                                rx.text("Tooltip 'dias: X' significa promedio de días hábiles en esa fecha.", font_size="lg"),
                                rx.text("Usa filtros por tipo y exportaciones (Excel/CSV) para informes ejecutivos.", font_size="lg"),
                                spacing="2",
                                width="100%",
                                align_items="start",
                            ),
                        ),
                        acordeon_item(
                            "4. Administración de usuarios",
                            "usuarios",
                            "users",
                            rx.vstack(
                                rx.text("Consulta roles y estado de cuentas desde 'Ver Usuarios'.", font_size="lg"),
                                rx.text("Puedes degradar de funcionario a ciudadano cuando corresponda.", font_size="lg"),
                                rx.text("Por seguridad, no está permitido degradar tu propia cuenta.", font_size="lg"),
                                spacing="2",
                                width="100%",
                                align_items="start",
                            ),
                        ),
                        acordeon_item(
                            "5. Buenas prácticas operativas",
                            "buenas_practicas",
                            "shield_check",
                            rx.vstack(
                                rx.text("Prioriza vencidas y críticas con el semáforo de vencimientos.", font_size="lg"),
                                rx.text("Mantén respuestas concretas, respetuosas y con lenguaje ciudadano.", font_size="lg"),
                                rx.text("Antes de cerrar: valida ortografía, adjuntos y consistencia del estado.", font_size="lg"),
                                spacing="2",
                                width="100%",
                                align_items="start",
                            ),
                        ),
                        spacing="4",
                        width="100%",
                        ),
                        # Bloque de cierre para ocupar altura y mantener estética profesional
                        rx.box(
                            rx.grid(
                                rx.box(
                                    rx.text("SLA recomendado", font_size="sm", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8")),
                                    rx.text("Responder en < 5 días hábiles", font_weight="bold", font_size="lg"),
                                    p="4",
                                    border_radius="xl",
                                    bg=rx.color_mode_cond(light="#f8fafc", dark="rgba(15,23,42,0.55)"),
                                    border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155"),
                                ),
                                rx.box(
                                    rx.text("Calidad de respuesta", font_size="sm", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8")),
                                    rx.text("Clara, completa y verificable", font_weight="bold", font_size="lg"),
                                    p="4",
                                    border_radius="xl",
                                    bg=rx.color_mode_cond(light="#f8fafc", dark="rgba(15,23,42,0.55)"),
                                    border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155"),
                                ),
                                rx.box(
                                    rx.text("Seguimiento", font_size="sm", color=rx.color_mode_cond(light="#64748b", dark="#94a3b8")),
                                    rx.text("Revisar vencimientos cada día", font_weight="bold", font_size="lg"),
                                    p="4",
                                    border_radius="xl",
                                    bg=rx.color_mode_cond(light="#f8fafc", dark="rgba(15,23,42,0.55)"),
                                    border=rx.color_mode_cond(light="1px solid #e2e8f0", dark="1px solid #334155"),
                                ),
                                columns={"base": "1", "md": "3"},
                                spacing="4",
                                width="100%",
                            ),
                            width="100%",
                            margin_top="auto",
                            pt="6",
                        ),
                        spacing="6",
                        width="100%",
                        max_width="1650px",
                        min_height="calc(100vh - 130px)",
                        p={"base": "3", "md": "4"},
                    ),
                    width="100%",
                    bg=rx.color_mode_cond(light="rgba(255,255,255,0.7)", dark="rgba(15,23,42,0.45)"),
                    border=rx.color_mode_cond(light="1px solid #dbeafe", dark="1px solid #1e293b"),
                    border_radius="3xl",
                    box_shadow=rx.color_mode_cond(light="0 16px 40px -24px rgba(0,0,0,0.2)", dark="0 18px 45px -25px rgba(0,0,0,0.8)"),
                    margin_x={"base": "6px", "md": "14px"},
                ),
                width="100%",
                bg=rx.color_mode_cond(light="#f1f5f9", dark="#0f172a"),
                min_height="100vh",
                padding_top={"base": "6px", "md": "10px"},
                padding_bottom={"base": "12px", "md": "18px"},
            ),
        ),
        rx.box(
            navbar(),
            rx.center(access_denied_widget("Solo funcionarios autenticados pueden acceder a esta función."), size="3"),
        ),
    )

app = rx.App()
app.add_page(index, route="/", title="Inicio - Sistema PQRS", on_load=State.redirigir_si_autenticado)
app.add_page(registro_page, route="/registro", title="Registro de Ciudadano", on_load=State.redirigir_si_autenticado)
app.add_page(registro_funcionario_page, route="/registro-funcionario", title="Registro de Funcionario", on_load=State.redirigir_si_autenticado)
app.add_page(login_page, route="/login", title="Iniciar Sesión", on_load=State.redirigir_si_autenticado)
app.add_page(solicitudes_page, route="/solicitudes", title="Nueva Solicitud PQRS")
app.add_page(change_password_page, route="/cambiar-contrasena", title="Cambiar Contraseña")
app.add_page(dashboard, route="/dashboard", title="Panel de Ciudadano", on_load=State.hidratar_sesion_ciudadano)
app.add_page(funcionario_dashboard, route="/dashboard-funcionario", title="Panel de Funcionario", on_load=State.cargar_datos_funcionario)
app.add_page(funcionario_cerradas_dashboard, route="/dashboard-funcionario-cerradas", title="Solicitudes Cerradas", on_load=State.cargar_datos_funcionario)
app.add_page(usuarios_page, route="/usuarios", title="Gestión de Usuarios", on_load=State.cargar_usuarios)
app.add_page(cambiar_rol_page, route="/cambiar-rol", title="Cambiar Rol de Usuario")
app.add_page(ayuda_funcionario_page, route="/ayuda-funcionario", title="Ayuda para Funcionarios")
app.add_page(consultar_estado_page, route="/consultar-estado", title="Consultar Estado de Solicitud")
app.add_page(politica_privacidad_page, route="/politica-privacidad", title="Política de Privacidad")
app.add_page(reportes_page, route="/reportes", title="Reportes PQRS", on_load=State.cargar_solicitudes)

if app._api is not None:
    app._api.mount(
        "/assets/uploads",
        StaticFiles(directory=str(UPLOAD_DIR), check_dir=False),
        name="assets_uploads",
    )
    app._api.mount(
        "/uploads",
        StaticFiles(directory=str(UPLOAD_DIR), check_dir=False),
        name="uploads",
    )

    from starlette.responses import FileResponse, Response

    async def serve_uploaded_file(filename: str):
        nombre = sanitizar_nombre_archivo(filename)
        ruta = _resolver_ruta_archivo_existente(nombre)
        if not ruta:
            return Response(
                content=b"Archivo no encontrado",
                status_code=404,
                media_type="text/plain",
            )
        asegurar_espejo_web(nombre)
        return FileResponse(ruta, filename=Path(ruta).name)

    async def download_file(download_id: str):
        try:
            if download_id not in TEMP_DOWNLOADS:
                return Response(
                    content=b"Archivo no encontrado o expirado",
                    status_code=404,
                    media_type="text/plain",
                )

            file_data = TEMP_DOWNLOADS[download_id]
            filename = file_data.get("filename", "descargar.bin")
            data = file_data.get("data", b"")
            mime = file_data.get("mime", "application/octet-stream")

            del TEMP_DOWNLOADS[download_id]

            return Response(
                content=data,
                media_type=mime,
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        except Exception as e:
            print(f"Error en download_file: {e}")
            return Response(
                content=b"Error al descargar el archivo",
                status_code=500,
                media_type="text/plain",
            )

    for _path, _handler in (
        ("/download/{download_id}", download_file),
        ("/api/download/{download_id}", download_file),
        ("/uploads/{filename}", serve_uploaded_file),
        ("/assets/uploads/{filename}", serve_uploaded_file),
    ):
        try:
            app._api.add_api_route(_path, _handler, methods=["GET"])
        except Exception:
            try:
                from starlette.routing import Route
                app._api.routes.insert(0, Route(_path, endpoint=_handler, methods=["GET"]))
            except Exception as route_err:
                print(f"No se pudo registrar ruta {_path}: {route_err}")


